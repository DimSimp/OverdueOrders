from __future__ import annotations

ZONE_ORDER = ["String Room", "Back Area", "Out Front", "Picks"]


def generate_picking_list(
    orders: list,
    pick_zones: dict,
) -> list[dict]:
    """
    Group line items by SKU across all orders, sum quantities.

    Returns [{sku, description, qty, zone}, ...] sorted by zone then SKU.
    Unassigned SKUs (no zone or unrecognised zone) appear last.
    """
    sku_map: dict[str, dict] = {}
    for order in orders:
        for li in order.line_items:
            sku = li.sku
            if not sku:
                continue
            name = (
                getattr(li, "product_name", None)
                or getattr(li, "title", None)
                or ""
            )
            if sku not in sku_map:
                sku_map[sku] = {
                    "sku": sku,
                    "description": name,
                    "qty": 0,
                    "zone": pick_zones.get(sku, ""),
                }
            sku_map[sku]["qty"] += li.quantity

    zone_rank = {z: i for i, z in enumerate(ZONE_ORDER)}

    def _sort_key(item: dict) -> tuple:
        zone = item["zone"]
        rank = zone_rank.get(zone, len(ZONE_ORDER))  # unassigned goes last
        return (rank, item["sku"].lower())

    return sorted(sku_map.values(), key=_sort_key)


def export_picking_list_xlsx(items: list[dict], output_path: str) -> None:
    """Write XLSX with zone section headers and page breaks between zones.

    Columns: SKU | Description | QTY
    Each zone section is preceded by an empty row (except the first) and starts
    with a bold coloured header row.  Data rows have a thin border on all sides.
    A manual page break is inserted before the empty row of each new section.
    All columns fit on one A4 portrait page.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.pagebreak import Break

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Picking List"

    # Fixed column widths
    SKU_WIDTH = 20
    QTY_WIDTH = 6
    # A4 portrait with ~0.7" margins gives ~94 character-units usable width.
    # Description gets whatever remains after the fixed columns.
    _A4_USABLE = 94
    max_desc_len = max((len(item["description"]) for item in items), default=10)
    desc_width = min(max_desc_len + 2, _A4_USABLE - SKU_WIDTH - QTY_WIDTH)
    desc_width = max(desc_width, 20)  # never narrower than 20

    ws.column_dimensions["A"].width = SKU_WIDTH
    ws.column_dimensions["B"].width = desc_width
    ws.column_dimensions["C"].width = QTY_WIDTH

    # Page number in the top-right header — size 20
    ws.oddHeader.right.text = "&P of &N"
    ws.oddHeader.right.size = 20

    # Border style used on every data cell
    thin = Side(style="thin")
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column header row
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col, text in enumerate(["SKU", "Description", "QTY"], start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = Font(bold=True, size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    current_row = 2
    current_zone: str | None = None
    zone_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    for item in items:
        zone = item["zone"] or "Unassigned"

        if zone != current_zone:
            if current_zone is not None:
                # Page break on the last data row of the previous section
                ws.row_breaks.append(Break(id=current_row - 1))
                # Empty spacer row between sections
                current_row += 1

            # Zone section header
            cell = ws.cell(row=current_row, column=1, value=zone)
            cell.font = Font(bold=True, size=13)
            cell.fill = zone_fill
            cell.alignment = Alignment(horizontal="left")
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=3,
            )
            current_row += 1
            current_zone = zone

        # Data row — bordered cells
        for col, value in enumerate([item["sku"], item["description"], item["qty"]], start=1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = data_border
        current_row += 1

    # Print settings — fit all columns to one A4 portrait page
    ws.print_area = f"A1:C{current_row - 1}"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0   # allow as many pages tall as needed
    ws.page_setup.orientation = "portrait"

    wb.save(output_path)
