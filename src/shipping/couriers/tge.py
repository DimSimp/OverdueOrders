from __future__ import annotations

import base64
import logging
import math
import uuid
from datetime import datetime
from pathlib import Path

import requests
from requests.auth import HTTPBasicAuth

from src.shipping.base_courier import BaseCourier
from src.shipping.models import BookingResult, Quote, ShipmentRequest, next_business_day

log = logging.getLogger("courier.tge")

# ── API endpoints ─────────────────────────────────────────────────────────────
RATE_URL = "https://api.teamglobalexp.com:6930/gateway/TollMessageRateEnquiryRestService/1.0/tom/rateEnquiry"
MANIFEST_URL = "https://api.teamglobalexp.com:6930/gateway/TollMessageManifestRestService/1.0/tom/receiveManifest"
PRINT_URL = "https://au-print-prod-apigw.internal.myteamge.com/printDocument"


class TGECourier(BaseCourier):
    name = "Team Global Express"
    code = "tge"

    def __init__(self, config: dict):
        super().__init__(config)
        self._username = config.get("username", "")
        self._password = config.get("password", "")
        self._account_code = config.get("account_code", "")
        # Print API credentials
        self._api_key = config.get("api_key", "")
        self._mytoll_identity = config.get("mytoll_identity", "")
        self._mytoll_token = config.get("mytoll_token", "")
        # Optional mutual-TLS cert for print API
        self._cert_file = config.get("cert_file", "")
        self._key_file = config.get("key_file", "")
        # Optional surcharges Excel (postcode → extra charge)
        self._surcharges_file = config.get("surcharges_file", "")
        # PostgreSQL for SSCC / ShipmentID sequencing
        self._db_host = config.get("db_host", "")
        self._db_port = int(config.get("db_port", 5432))
        self._db_name = config.get("db_name", "postgres")
        self._db_user = config.get("db_user", "")
        self._db_password = config.get("db_password", "")

    # ── Availability ──────────────────────────────────────────────────────────

    def is_available(self, request: ShipmentRequest) -> bool:
        if not (self._username and self._password and self._account_code):
            return False
        # TGE is road freight only — not used for express/overnight orders
        if request.shipping_type == "Express":
            return False
        return True

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _auth(self) -> HTTPBasicAuth:
        return HTTPBasicAuth(self._username, self._password)

    def _json_headers(self) -> dict:
        return {"Content-Type": "application/json", "Accept": "application/json"}

    def _print_headers(self) -> dict:
        return {
            "Content-Type": "application/json",
            "Accept": "application/json",
            "x-api-key": self._api_key,
            "callId": "SCARLETTMUSIC",
            "x-mytoll-identity": self._mytoll_identity,
            "Channel": "CES",
            "x-mytoll-token": self._mytoll_token,
        }

    def _toll_header(self, doc_type: str, source_code: str = "SCAR") -> dict:
        now = datetime.now()
        return {
            "MessageVersion": "3.1",
            "MessageIdentifier": str(uuid.uuid4()),
            "CreateTimestamp": f"{now.strftime('%Y-%m-%d')}T{now.strftime('%H:%M:%S')}+10:00",
            "DocumentType": doc_type,
            "Environment": "PRD",
            "MessageSender": "SCARLETTMUSIC",
            "MessageReceiver": "TOLL",
            "SourceSystemCode": source_code,
        }

    def _get_surcharge(self, postcode: str) -> float:
        """Load postcode surcharge from Excel file, if configured."""
        if not self._surcharges_file:
            return 0.0
        if not Path(self._surcharges_file).exists():
            log.debug("TGE surcharges file not found: %s", self._surcharges_file)
            return 0.0
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self._surcharges_file, read_only=True, data_only=True)
            sheet = wb[wb.sheetnames[0]]
            for row in range(2, sheet.max_row + 1):
                cell_val = sheet[f"A{row}"].value
                if cell_val is None:
                    continue
                if str(cell_val).strip() == postcode.strip():
                    surcharge_val = sheet[f"E{row}"].value
                    wb.close()
                    return float(surcharge_val) if surcharge_val else 0.0
            wb.close()
        except Exception as exc:
            log.warning("Failed to load TGE surcharges: %s", exc)
        return 0.0

    def _calc_additional_costs(self, request: ShipmentRequest) -> float:
        """Calculate per-item weight/dimension surcharges + insurance."""
        additional = 0.0
        for pkg in request.packages:
            dims = sorted([pkg.width_cm, pkg.length_cm, pkg.height_cm])
            vol = pkg.volume_m3
            # Heavy item surcharge
            if pkg.weight_kg > 35:
                additional += 77.55
            # Oversized or very heavy
            if pkg.weight_kg > 35 or dims[0] > 180 or dims[1] > 180 or dims[2] > 180 or vol > 0.7:
                additional += 50.0
            elif pkg.weight_kg > 30 or dims[0] > 60 or dims[1] > 80 or dims[2] > 120 or vol > 0.7:
                additional += 12.0
        # Insurance (once per shipment)
        if request.order_value < 500:
            additional += 6.95
        else:
            additional += request.order_value * 0.02
        # Postcode-based road express surcharge
        additional += self._get_surcharge(request.receiver.postcode)
        return additional

    def _build_rate_items(self, request: ShipmentRequest) -> list:
        """Build ShipmentItem list for the rate enquiry payload."""
        items = []
        for pkg in request.packages:
            items.append({
                "Commodity": {
                    "CommodityCode": "Z",
                    "CommodityDescription": "ALL FREIGHT",
                },
                "ShipmentItemTotals": {"ShipmentItemCount": "1"},
                "Dimensions": {
                    "Width": str(math.ceil(pkg.width_cm)),
                    "Length": str(math.ceil(pkg.length_cm)),
                    "Height": str(math.ceil(pkg.height_cm)),
                    "Volume": str(round(pkg.volume_m3, 4)),
                    "Weight": str(pkg.weight_kg),
                },
            })
        return items

    @staticmethod
    def _calc_sscc_check_digit(sscc_17: str) -> int:
        """GS1 SSCC check digit for a 17-digit string."""
        # Odd-index positions (0-based: 0,2,4,6,8,10,12,14,16) × 3
        odd_sum = (
            int(sscc_17[0]) + int(sscc_17[2]) + int(sscc_17[4]) +
            int(sscc_17[6]) + int(sscc_17[8]) + int(sscc_17[10]) +
            int(sscc_17[12]) + int(sscc_17[14]) + int(sscc_17[16])
        ) * 3
        # Even-index positions (0-based: 1,3,5,7,9,11,13,15)
        even_sum = (
            int(sscc_17[1]) + int(sscc_17[3]) + int(sscc_17[5]) +
            int(sscc_17[7]) + int(sscc_17[9]) + int(sscc_17[11]) +
            int(sscc_17[13]) + int(sscc_17[15])
        )
        total = odd_sum + even_sum
        rounded = math.ceil(total / 10) * 10
        return rounded - total

    def _get_db_connection(self):
        import psycopg2
        return psycopg2.connect(
            host=self._db_host,
            port=self._db_port,
            database=self._db_name,
            user=self._db_user,
            password=self._db_password,
            connect_timeout=10,
        )

    # ── Quote ─────────────────────────────────────────────────────────────────

    def get_quote(self, request: ShipmentRequest) -> list:
        next_day = next_business_day()
        pickup_time = f"{next_day.strftime('%Y-%m-%d')}T09:00:00+10:00"
        additional_costs = self._calc_additional_costs(request)

        toll_message = {
            "@version": "3.1",
            "@encoding": "utf-8",
            "TollMessage": {
                "Header": self._toll_header("RateEnquiry", "SCAR"),
                "RateEnquiry": {
                    "Request": {
                        "BusinessID": "IPEC",
                        "SystemFields": {"PickupDateTime": pickup_time},
                        "ShipmentService": {
                            "ServiceCode": "X",
                            "ServiceDescription": "Road Express",
                        },
                        "ShipmentFlags": {"ExtraServiceFlag": "true"},
                        "ShipmentFinancials": {
                            "ExtraServicesAmount": {
                                "Currency": "AUD",
                                "Value": str(round(float(request.order_value))),
                            }
                        },
                        "BillToParty": {"AccountCode": self._account_code},
                        "ConsignorParty": {
                            "PhysicalAddress": {
                                "Suburb": request.sender.city.upper(),
                                "StateCode": request.sender.state.upper(),
                                "PostalCode": request.sender.postcode,
                                "CountryCode": "AU",
                            }
                        },
                        "ConsigneeParty": {
                            "PhysicalAddress": {
                                "Suburb": request.receiver.city,
                                "StateCode": request.receiver.state,
                                "PostalCode": request.receiver.postcode,
                                "CountryCode": "AU",
                            }
                        },
                        "ShipmentItems": {
                            "ShipmentItem": self._build_rate_items(request)
                        },
                    }
                },
            },
        }

        last_error = ""
        for attempt in range(6):
            if attempt > 0:
                toll_message["TollMessage"]["Header"]["MessageIdentifier"] = str(uuid.uuid4())
            try:
                r = requests.post(
                    url=RATE_URL,
                    auth=self._auth(),
                    json=toll_message,
                    headers=self._json_headers(),
                    timeout=30,
                )
                response = r.json()
                # Check for API-level timeout error
                try:
                    err_msg = (
                        response["TollMessage"]["ErrorMessages"]
                        ["ErrorMessage"][0]["ErrorMessage"]
                    )
                    if "timeout" in err_msg.lower():
                        last_error = f"Timeout: {err_msg}"
                        continue
                    last_error = err_msg
                    break
                except (KeyError, IndexError, TypeError):
                    pass  # No error — proceed to parse price

                price = float(
                    response["TollMessage"]["RateEnquiry"]["Response"]
                    ["TotalChargeAmount"]["Value"]
                )
                total = round(price + additional_costs, 2)
                return [Quote(
                    courier_name=self.name,
                    courier_code=self.code,
                    service_name="Road Express",
                    price=total,
                    estimated_days="1–5 business days",
                    raw_response=response,
                )]
            except Exception as exc:
                last_error = str(exc)
                break

        return [Quote(
            courier_name=self.name,
            courier_code=self.code,
            service_name="Road Express",
            price=0,
            estimated_days="",
            error=last_error or "Unknown error",
        )]

    # ── Booking ───────────────────────────────────────────────────────────────

    def book(self, request: ShipmentRequest, quote=None) -> BookingResult:
        """Create manifest + retrieve label PDF via TGE APIs."""
        log.info("Booking TGE shipment for order %s", request.order_id)

        receiver = request.receiver
        sender = request.sender

        # Sanitise name / address fields
        rec_name = receiver.name.replace("'", "").replace("\u2019", "")
        address1 = (receiver.street1 or "").replace("'", "").replace("\u2019", "")
        address2 = (receiver.street2 or address1).replace("'", "").replace("\u2019", "")
        suburb = receiver.city.replace("'", "").replace("\u2019", "")
        phone = (receiver.phone or "").replace(" ", "")

        # ── Get ShipmentID / SSCC sequence from DB ────────────────────────────
        if not self._db_host or not self._db_user:
            return BookingResult(
                courier_name=self.name, tracking_number="", label_pdf=None,
                booking_reference="",
                error="Database not configured for TGE — add db_host/db_user to config.",
            )
        try:
            conn = self._get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT MAX(ShipmentID), MAX(SSCC) FROM Toll;")
            row = cursor.fetchone()
            max_shipment_id = int(row[0]) if row and row[0] is not None else 0
            max_sscc = int(row[1]) if row and row[1] is not None else 10000000000000000  # seed
        except Exception as exc:
            return BookingResult(
                courier_name=self.name, tracking_number="", label_pdf=None,
                booking_reference="",
                error=f"Database connection failed: {exc}",
            )

        # ── Timestamps ────────────────────────────────────────────────────────
        now = datetime.now()
        current_time = f"{now.strftime('%Y-%m-%d')}T{now.strftime('%H:%M:%S')}+10:00"
        next_day = next_business_day()
        pickup_time = f"{next_day.strftime('%Y-%m-%d')}T09:00:00+10:00"

        # ── Build shipment payload ─────────────────────────────────────────────
        order_value_str = str(round(float(request.order_value)))

        # The ShipmentID used in the manifest is the current max (before per-item increment)
        shipment_id_str = str(max_shipment_id)

        shipment_base = {
            "ShipmentFinancials": {
                "ExtraServicesAmount": {"Currency": "AUD", "Value": order_value_str}
            },
            "ShipmentID": shipment_id_str,
            "CreateDateTime": current_time,
            "ConsigneeParty": {
                "PartyName": rec_name,
                "PhysicalAddress": {
                    "AddressLine1": address1,
                    "AddressLine2": address2,
                    "Suburb": suburb,
                    "PostalCode": receiver.postcode,
                    "StateCode": receiver.state,
                    "CountryName": "Australia",
                    "CountryCode": "AU",
                },
                "Contact": {
                    "Name": rec_name,
                    "Phone": {"Number": phone},
                },
            },
            "BillToParty": {
                "AccountCode": self._account_code,
                "PartyName": "SCARLETTMUSIC",
                "PhysicalAddress": {
                    "AddressLine1": sender.street1,
                    "Suburb": sender.city,
                    "PostalCode": sender.postcode,
                    "StateCode": sender.state,
                    "CountryName": "Australia",
                    "CountryCode": "AU",
                },
                "Contact": {
                    "Name": "Kyal Scarlett",
                    "Phone": {"Number": sender.phone or "0393185751"},
                },
            },
            "ShipmentItemCollection": {"ShipmentItem": []},
        }

        # print_payload accumulates items with full dimensions (needed for label API)
        print_items: list = []
        db_inserts: list = []

        for pkg in request.packages:
            max_shipment_id += 1
            max_sscc += 1

            sscc_17 = str(max_sscc).zfill(17)
            if len(sscc_17) != 17:
                conn.close()
                return BookingResult(
                    courier_name=self.name, tracking_number="", label_pdf=None,
                    booking_reference="",
                    error=f"SSCC sequence out of range (got {max_sscc}); check Toll DB.",
                )
            check_digit = self._calc_sscc_check_digit(sscc_17)
            full_sscc = "00" + sscc_17 + str(check_digit)

            sscc_id_entry = [{"Value": full_sscc, "SchemeName": "SSCC"}]

            # Item for manifest call
            manifest_item = {
                "IDs": {"ID": sscc_id_entry},
                "ShipmentItemTotals": {"ShipmentItemCount": "1"},
                "ShipmentService": {
                    "ServiceCode": "X",
                    "ServiceDescription": "ROAD EXPRESS",
                    "ShipmentProductCode": "1",
                },
                "Description": "Carton",
                "Dimensions": {
                    "Volume": str(round(pkg.volume_m3, 4)),
                    "Weight": str(round(pkg.weight_kg)),
                },
                "References": {
                    "Reference": [
                        {"ReferenceType": "ConsignorItemReference", "ReferenceValue": request.order_id},
                        {"ReferenceType": "ConsigneeItemReference", "ReferenceValue": request.order_id},
                    ]
                },
            }
            shipment_base["ShipmentItemCollection"]["ShipmentItem"].append(manifest_item)

            # Item for print call (includes full dimensions)
            print_items.append({
                "IDs": {"ID": sscc_id_entry},
                "ShipmentItemTotals": {"ShipmentItemCount": "1"},
                "ShipmentService": {
                    "ServiceCode": "X",
                    "ServiceDescription": "ROAD EXPRESS",
                    "ShipmentProductCode": "1",
                },
                "Description": "Carton",
                "Dimensions": {
                    "Volume": str(round(pkg.volume_m3, 4)),
                    "Weight": str(round(pkg.weight_kg, 1)),
                    "Length": math.ceil(pkg.length_cm),
                    "Width": math.ceil(pkg.width_cm),
                    "Height": math.ceil(pkg.height_cm),
                },
                "References": {
                    "Reference": [
                        {"ReferenceType": "ConsignorItemReference", "ReferenceValue": request.order_id},
                        {"ReferenceType": "ConsigneeItemReference", "ReferenceValue": request.order_id},
                    ]
                },
            })

            db_inserts.append((
                max_shipment_id, max_sscc, rec_name, address1, address2, suburb,
                receiver.postcode, receiver.state, request.order_id, phone,
                1, full_sscc,
                math.ceil(pkg.length_cm), math.ceil(pkg.width_cm),
                round(pkg.weight_kg, 1), math.ceil(pkg.height_cm),
                round(pkg.volume_m3, 4),
            ))

        # ── Save items to DB ──────────────────────────────────────────────────
        try:
            for row_data in db_inserts:
                cursor.execute(
                    "INSERT INTO Toll(ShipmentID, SSCC, Name, Address1, Address2, Suburb, "
                    "Postcode, State, Reference, phone_number, item_count, current_sscc, "
                    "length, width, weight, height, volume) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);",
                    row_data,
                )
            conn.commit()
            conn.close()
        except Exception as exc:
            try:
                conn.rollback()
                conn.close()
            except Exception:
                pass
            return BookingResult(
                courier_name=self.name, tracking_number="", label_pdf=None,
                booking_reference="",
                error=f"DB insert failed: {exc}",
            )

        # ── Manifest API ──────────────────────────────────────────────────────
        manifest_header = self._toll_header("Manifest", "XH56")
        manifest_header["CreateTimestamp"] = current_time

        manifest_message = {
            "@version": "3.1",
            "@encoding": "utf-8",
            "TollMessage": {
                "Header": manifest_header,
                "Manifest": {
                    "BusinessID": "IPEC",
                    "CreateDateTime": current_time,
                    "DatePeriodCollection": {
                        "DatePeriod": [{"DateType": "DespatchDate", "DateTime": pickup_time}]
                    },
                    "ConsignorParty": {
                        "PartyName": "Scarlett Music",
                        "PhysicalAddress": {
                            "AddressLine1": sender.street1,
                            "Suburb": sender.city.upper(),
                            "PostalCode": sender.postcode,
                            "StateCode": sender.state.upper(),
                            "CountryName": "Australia",
                            "CountryCode": "AU",
                        },
                        "Contact": {
                            "Name": "Michael Demetre",
                            "Phone": {"Number": "1800688586"},
                        },
                    },
                    "ShipmentCollection": {"Shipment": [shipment_base]},
                },
            },
        }

        try:
            r = requests.post(
                url=MANIFEST_URL,
                auth=self._auth(),
                json=manifest_message,
                headers=self._json_headers(),
                timeout=60,
            )
            manifest_resp = r.json()
            response_id = str(
                manifest_resp["TollMessage"]["ResponseMessages"]
                ["ResponseMessage"][0]["ResponseID"]["Value"]
            )
            if response_id != "200":
                err = manifest_resp["TollMessage"]["ErrorMessages"]["ErrorMessage"][0]["ErrorMessage"]
                return BookingResult(
                    courier_name=self.name, tracking_number="", label_pdf=None,
                    booking_reference="",
                    error=f"TGE manifest API error: {err}",
                )
        except Exception as exc:
            return BookingResult(
                courier_name=self.name, tracking_number="", label_pdf=None,
                booking_reference="",
                error=f"TGE manifest request failed: {exc}",
            )

        tracking = shipment_id_str
        log.info("TGE manifest OK: ShipmentID=%s", tracking)

        # ── Print / Label API ─────────────────────────────────────────────────
        now2 = datetime.now()
        current_time2 = now2.strftime("%Y-%m-%dT%H:%M:%SZ")
        pickup_time2 = next_day.strftime("%Y-%m-%dT09:00:00Z")

        total_volume = sum(pkg.volume_m3 for pkg in request.packages)
        total_weight = sum(pkg.weight_kg for pkg in request.packages)

        print_header = {
            "MessageVersion": "1.0",
            "MessageIdentifier": str(uuid.uuid4()),
            "CreateTimestamp": current_time2,
            "Environment": "PRD",
            "MessageSender": "SCARLETTMUSIC",
            "MessageReceiver": "Toll",
            "SourceSystemCode": "XH56",
        }

        print_payload = {
            "BusinessID": "IPEC",
            "PrintDocumentType": "Label",
            "PrintSettings": {
                "IsLabelThermal": "false",
                "IsZPLRawResponseRequired": "false",
                "PDF": {
                    "IsPDFA4": "false",
                    "PDFSettings": {"StartQuadrant": "1"},
                },
            },
            "ConsignorParty": {
                "Contact": {
                    "Name": "Kyal",
                    "Phone": {"Number": sender.phone or "0393185751"},
                },
                "PartyName": "Scarlett Music",
                "PhysicalAddress": {
                    "AddressLine1": sender.street1,
                    "AddressType": "Business",
                    "CountryCode": "AU",
                    "PostalCode": sender.postcode,
                    "StateCode": sender.state.upper(),
                    "Suburb": sender.city.upper(),
                },
            },
            "CreateDateTime": current_time2,
            "ShipmentCollection": {
                "Shipment": [
                    {
                        "BillToParty": {"AccountCode": self._account_code},
                        "ConsigneeParty": {
                            "Contact": {"Name": rec_name, "Phone": {"Number": phone}},
                            "PartyName": rec_name,
                            "PhysicalAddress": {
                                "AddressType": "Business",
                                "AddressLine1": address1,
                                "AddressLine2": address2,
                                "Suburb": suburb,
                                "PostalCode": receiver.postcode,
                                "StateCode": receiver.state,
                                "CountryName": "Australia",
                                "CountryCode": "AU",
                            },
                        },
                        "CreateDateTime": current_time2,
                        "DatePeriodCollection": {
                            "DatePeriod": [
                                {"DateTime": pickup_time2, "DateType": "DespatchDate"}
                            ]
                        },
                        "Orders": {"Order": [{}]},
                        "References": {
                            "Reference": [
                                {
                                    "ReferenceType": "ShipmentReference1",
                                    "ReferenceValue": request.order_id,
                                }
                            ]
                        },
                        "ShipmentID": tracking,
                        "ShipmentItemCollection": {"ShipmentItem": print_items},
                        "ShipmentTotals": {
                            "MiscellaneousItemCount": 0,
                            "Volume": {"UOM": "m3", "Value": round(total_volume, 4)},
                            "Weight": {"UOM": "kg", "Value": round(total_weight, 1)},
                        },
                    }
                ]
            },
        }

        print_message = {
            "TollMessage": {
                "Header": print_header,
                "Print": print_payload,
            }
        }

        cert = None
        if self._cert_file and self._key_file:
            if Path(self._cert_file).exists() and Path(self._key_file).exists():
                cert = (self._cert_file, self._key_file)
            else:
                log.warning("TGE cert files configured but not found — attempting print without cert")

        try:
            r = requests.post(
                url=PRINT_URL,
                auth=self._auth(),
                json=print_message,
                headers=self._print_headers(),
                cert=cert,
                timeout=60,
            )
            print_resp = r.json()
            pdf_b64 = (
                print_resp["TollMessage"]["ResponseMessages"]
                ["ResponseMessage"][0]["ResponseMessage"]
            )
            label_pdf = base64.b64decode(pdf_b64)
            log.info("TGE label downloaded: %d bytes", len(label_pdf))
        except Exception as exc:
            log.error("TGE label download failed: %s", exc)
            return BookingResult(
                courier_name=self.name,
                tracking_number=tracking,
                label_pdf=None,
                booking_reference=tracking,
                error=f"Booked (tracking: {tracking}) but label failed: {exc}",
            )

        return BookingResult(
            courier_name=self.name,
            tracking_number=tracking,
            label_pdf=label_pdf,
            booking_reference=tracking,
        )
