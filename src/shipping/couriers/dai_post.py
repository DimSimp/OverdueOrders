from __future__ import annotations

import base64
import logging
import time
from datetime import datetime
from pathlib import Path

import requests
from requests.auth import HTTPBasicAuth

from src.shipping.base_courier import BaseCourier
from src.shipping.models import BookingResult, Quote, ShipmentRequest

log = logging.getLogger("courier.dai_post")

MAX_WEIGHT_KG = 22
MAX_ORDER_VALUE = 200

# Weight brackets: (max_weight, column_letter)
# Columns B-K in the rates sheet
WEIGHT_BRACKETS = [
    (0.5, "B"),
    (1.0, "C"),
    (1.5, "D"),
    (2.0, "D"),
    (2.5, "E"),
    (3.0, "E"),
    (3.5, "F"),
    (4.0, "F"),
    (4.5, "G"),
    (5.0, "G"),
    (7.0, "H"),
    (10.0, "I"),
    (15.0, "J"),
    (22.0, "K"),
]

# Map column letters to 0-based column indices
_COL_INDEX = {chr(c): c - ord("A") for c in range(ord("A"), ord("L") + 1)}


class DaiPostCourier(BaseCourier):
    name = "DAI Post"
    code = "dai_post"

    def __init__(self, config: dict):
        super().__init__(config)
        self._postcodes_file = config.get("postcodes_file", "")
        self._rates_file = config.get("rates_file", "")
        self._username = config.get("username", "")
        self._password = config.get("password", "")
        self._account = config.get("account", "SCA")
        self._origin_terminal = config.get("origin_terminal", "TME")

    def is_available(self, request: ShipmentRequest) -> bool:
        if not self._postcodes_file or not self._rates_file:
            return False
        # Single parcel only
        if len(request.packages) > 1:
            return False
        # Max $200 order value
        if request.order_value > MAX_ORDER_VALUE:
            return False
        # Max weight
        if request.packages and request.packages[0].weight_kg > MAX_WEIGHT_KG:
            return False
        # Express not supported
        if request.shipping_type == "Express":
            return False
        return True

    def get_quote(self, request: ShipmentRequest) -> list[Quote]:
        try:
            import openpyxl
        except ImportError:
            return [Quote(
                courier_name=self.name, courier_code=self.code,
                service_name="", price=0, estimated_days="",
                error="openpyxl library not installed",
            )]

        postcodes_path = Path(self._postcodes_file)
        rates_path = Path(self._rates_file)

        if not postcodes_path.exists():
            return [Quote(
                courier_name=self.name, courier_code=self.code,
                service_name="", price=0, estimated_days="",
                error=f"Postcodes file not found: {postcodes_path}",
            )]
        if not rates_path.exists():
            return [Quote(
                courier_name=self.name, courier_code=self.code,
                service_name="", price=0, estimated_days="",
                error=f"Rates file not found: {rates_path}",
            )]

        pkg = request.packages[0]
        weight = round(pkg.weight_kg, 2)
        postcode = request.receiver.postcode.strip()

        try:
            pc_wb = openpyxl.load_workbook(str(postcodes_path), read_only=True, data_only=True)
            pc_sheet = pc_wb[pc_wb.sheetnames[0]]

            # Find zone for postcode
            zone = None
            for row in pc_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                sheet_pc = str(row[0] or "").strip()
                if len(sheet_pc) == 3:
                    sheet_pc = f"0{sheet_pc}"
                if sheet_pc == postcode:
                    zone = str(row[1] or "").strip()
                    break
            pc_wb.close()

            if not zone:
                return [Quote(
                    courier_name=self.name, courier_code=self.code,
                    service_name="", price=0, estimated_days="",
                    error=f"Postcode {postcode} not found in DAI Post zones",
                )]

            # Determine weight column
            col_letter = None
            for max_w, col in WEIGHT_BRACKETS:
                if weight <= max_w:
                    col_letter = col
                    break

            if col_letter is None:
                return [Quote(
                    courier_name=self.name, courier_code=self.code,
                    service_name="", price=0, estimated_days="",
                    error=f"Weight {weight}kg exceeds DAI Post limit",
                )]

            col_idx = _COL_INDEX[col_letter]

            # Try "Sheet" first, then "Sheet2" (DAI Post rolls rates to Sheet2 on update)
            rates_wb = openpyxl.load_workbook(str(rates_path), read_only=True, data_only=True)
            _preferred = [n for n in ["Sheet", "Sheet2"] if n in rates_wb.sheetnames]
            _sheet_names = _preferred or list(rates_wb.sheetnames)

            price = None
            for _sheet_name in _sheet_names:
                for row in rates_wb[_sheet_name].iter_rows(min_row=2, values_only=True):
                    rate_zone = str(row[0] or "").strip()
                    if rate_zone.lower() == zone.lower():
                        try:
                            price = float(row[col_idx]) * 1.1  # +10% GST
                        except (ValueError, TypeError, IndexError):
                            pass
                        break
                if price is not None:
                    break
            rates_wb.close()

            if price is None:
                return [Quote(
                    courier_name=self.name, courier_code=self.code,
                    service_name="", price=0, estimated_days="",
                    error=f"No rate found for zone '{zone}' at {weight}kg",
                )]

            # PO Box surcharge
            addr = (request.receiver.street1 + " " + request.receiver.street2).lower()
            if "po box" in addr or "p.o. box" in addr:
                price += 0.80

            return [Quote(
                courier_name=self.name,
                courier_code=self.code,
                service_name="Standard Parcel",
                price=round(price, 2),
                estimated_days="3-7 business days",
                raw_response={"zone": zone, "weight": weight},
            )]

        except Exception as exc:
            return [Quote(
                courier_name=self.name, courier_code=self.code,
                service_name="", price=0, estimated_days="",
                error=str(exc),
            )]

    def book(self, request: ShipmentRequest, quote=None) -> BookingResult:
        """Book a DAI Post shipment and return tracking number + label PDF."""
        if not self._username or not self._password:
            return BookingResult(
                courier_name=self.name, tracking_number="", label_pdf=None,
                booking_reference="", error="DAI Post credentials not configured",
            )

        receiver = request.receiver
        pkg = request.packages[0] if request.packages else None
        weight = round(pkg.weight_kg, 2) if pkg else 1.0
        order_value = request.order_value or 0
        signature = "1" if round(order_value) > 200 else "0"
        now = datetime.now()
        current_time = f"{now.strftime('%Y-%m-%d')}T{now.strftime('%H:%M:%S')}"
        # Use Unix timestamp as job number (unique per call, no DB required)
        job_number = int(time.time())

        payload = {
            "shipment": {
                "service": "Parcel Right",
                "labelformat": "PDF",
                "account": self._account,
                "datetime": current_time,
                "reference": f"{request.order_id} {receiver.name}",
                "jobnumber": job_number,
                "signature": signature,
                "value": str(round(order_value)),
                "currency": "AUD",
                "uom": "kg",
                "weight": weight,
                "originterminal": self._origin_terminal,
                "shipper": {
                    "name": "Scarlett Music",
                    "attention": "Kyal",
                    "addr1": "286-288 Ballarat Rd",
                    "addr2": "",
                    "city": "Footscray",
                    "state": "VIC",
                    "country": "AU",
                    "postal": "3011",
                    "phone": "+61 03 8256 3460",
                    "email": "kyal@scarlettmusic.com.au",
                },
                "consignee": {
                    "name": receiver.name,
                    "attention": receiver.name,
                    "addr1": receiver.street1,
                    "addr2": receiver.street2 or "",
                    "city": receiver.city,
                    "state": receiver.state,
                    "country": "AU",
                    "postal": receiver.postcode,
                    "phone": receiver.phone or "",
                    "email": receiver.email or "",
                },
                "item": [{
                    "description": f"Scarlett Music Order {request.order_id}",
                    "qty": "1",
                    "unit": "pc",
                    "value": str(round(order_value)),
                }],
            }
        }

        log.info("Booking DAI Post for order %s (jobnumber=%d, weight=%.2fkg)",
                 request.order_id, job_number, weight)
        log.debug("DAI Post payload: %s", payload)

        try:
            resp = requests.post(
                "https://daiglobaltrack.com/prod/serviceconnect",
                auth=HTTPBasicAuth(self._username, self._password),
                json=payload,
                timeout=20,
            )
            log.debug("DAI Post response HTTP %d: %s", resp.status_code, resp.text[:500])
            resp.raise_for_status()
            data = resp.json()
        except Exception as exc:
            log.error("DAI Post booking request failed: %s", exc)
            return BookingResult(
                courier_name=self.name, tracking_number="", label_pdf=None,
                booking_reference="", error=f"DAI Post API error: {exc}",
            )

        try:
            tracking_number = data["shipmentresponse"]["tracknbr"]
            label_b64 = data["shipmentresponse"]["label"]
            label_pdf = base64.b64decode(label_b64)
            log.info("DAI Post booked: tracking=%s  label=%d bytes", tracking_number, len(label_pdf))
        except (KeyError, TypeError) as exc:
            log.error("DAI Post response missing expected fields: %s  response=%s", exc, data)
            return BookingResult(
                courier_name=self.name, tracking_number="", label_pdf=None,
                booking_reference="", error=f"Unexpected DAI Post response: {exc}",
            )

        return BookingResult(
            courier_name=self.name,
            tracking_number=str(tracking_number),
            label_pdf=label_pdf,
            booking_reference=str(job_number),
        )

    def cancel_shipment(self, tracking_number: str, **kwargs) -> tuple[bool, str]:
        """Cancel a DAI Post shipment by tracking number.

        Returns (success: bool, message: str).
        """
        if not self._username or not self._password:
            return False, "DAI Post credentials not configured"

        payload = {"cancelshipment": {"tracknbr": str(tracking_number)}}
        log.info("Cancelling DAI Post shipment: %s", tracking_number)

        try:
            resp = requests.post(
                "https://daiglobaltrack.com/prod/serviceconnect",
                auth=HTTPBasicAuth(self._username, self._password),
                json=payload,
                timeout=20,
            )
            log.debug("DAI Post cancel HTTP %d: %s", resp.status_code, resp.text[:500])
            resp.raise_for_status()
            data = resp.json()
            log.info("DAI Post cancel response: %s", data)
            return True, str(data)
        except Exception as exc:
            log.error("DAI Post cancel failed: %s", exc)
            return False, str(exc)
