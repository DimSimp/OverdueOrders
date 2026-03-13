from __future__ import annotations

from itertools import groupby
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.data_processor import MatchedOrder

# ── Column definitions ────────────────────────────────────────────────────────

# Summary sheet: hierarchical formatted view (no Arrived column — * prefix on SKU serves that purpose)
_SUMMARY_COLS   = ["SKU", "Qty", "Description", "Notes"]
_SUMMARY_WIDTHS = [20,     6,    45,            40]
_SUMMARY_LAST   = len(_SUMMARY_COLS)

# Data sheet: flat pre-formatted data
_DATA_COLS   = ["Platform", "Order No", "SKU", "Description", "Arrived", "Qty", "Notes"]
_DATA_WIDTHS = [15,          18,         22,    55,            10,        6,     45]

# ── Styles ────────────────────────────────────────────────────────────────────

_PLATFORM_FILL  = PatternFill("solid", fgColor="1F3864")  # dark navy
_ORDER_FILL     = PatternFill("solid", fgColor="D6DCE4")  # light grey-blue
_HEADER_FILL    = PatternFill("solid", fgColor="4472C4")  # medium blue
_THIN = Side(style="thin")

_PLATFORM_FONT  = Font(bold=True, color="FFFFFF", size=12)
_ORDER_FONT     = Font(bold=True, size=11)
_HEADER_FONT    = Font(bold=True, color="FFFFFF", size=11)
_NORMAL_FONT    = Font(size=11)


def _platform_sort_key(platform: str) -> tuple:
    pl = platform.lower()
    if pl == "website":
        return (0, platform)
    if pl == "ebay":
        return (2, platform)
    return (1, platform)


# ── Public entry point ────────────────────────────────────────────────────────

def export_to_xlsx(
    matched_orders: list[MatchedOrder],
    output_dir: str,
) -> str:
    """
    Export matched orders to a formatted .xlsx file with two sheets:

    Summary — hierarchical view grouped by Platform → Order No → items.
              Platform headers are navy, order headers are grey.
              Items with an invoice match are highlighted green.
              A blank row separates each order.

    Data    — flat pre-formatted table (one row per item) suitable for
              manually building a pivot table in Excel.

    Returns the absolute path of the created file.
    Raises ValueError if no orders to export.
    """
    if not matched_orders:
        raise ValueError("No matched orders to export.")

    sorted_orders = sorted(
        matched_orders,
        key=lambda m: (_platform_sort_key(m.platform), m.order_id),
    )

    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, sorted_orders)

    ws_data = wb.create_sheet("Data")
    _write_data_sheet(ws_data, sorted_orders)

    out_path = Path(output_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    from datetime import date
    filename = f"{date.today().isoformat()} matched orders.xlsx"
    filepath = out_path / filename
    wb.save(filepath)

    return str(filepath.resolve())


# ── Sheet writers ─────────────────────────────────────────────────────────────

def _write_summary_sheet(ws, sorted_orders: list[MatchedOrder]) -> None:
    for col_idx, width in enumerate(_SUMMARY_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Header row
    for col_idx, header in enumerate(_SUMMARY_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    row = 2
    for platform, platform_items in groupby(sorted_orders, key=lambda m: m.platform):
        # Platform header — spans all columns
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=_SUMMARY_LAST,
        )
        cell = ws.cell(row=row, column=1, value=f"  {platform.upper()}")
        cell.font = _PLATFORM_FONT
        cell.fill = _PLATFORM_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        platform_list = list(platform_items)
        for order_id, order_items in groupby(platform_list, key=lambda m: m.order_id):
            items = list(order_items)
            order_start_row = row

            # Order header — spans all columns
            ws.merge_cells(
                start_row=row, start_column=1,
                end_row=row,   end_column=_SUMMARY_LAST,
            )
            cell = ws.cell(row=row, column=1, value=f"    {order_id}")
            cell.font = _ORDER_FONT
            cell.fill = _ORDER_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 18
            row += 1

            for m in items:
                # Col 1: SKU (* prefix if arrived), Col 2: Qty,
                # Col 3: Description, Col 4: Notes
                sku = f"*{m.sku}" if m.is_invoice_match else m.sku
                vals = [sku, m.quantity, m.description, m.notes]
                for col_idx, val in enumerate(vals, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.font = _NORMAL_FONT
                    cell.alignment = Alignment(
                        vertical="center",
                        wrap_text=(col_idx in (3, 4)),  # Description and Notes wrap
                    )
                row += 1

            order_end_row = row - 1
            _apply_order_borders(ws, order_start_row, order_end_row)

            # Blank separator after each order
            row += 1

    # Print setup: A4 portrait, fit all columns to one page wide, auto height
    ws.page_setup.paperSize   = "9"          # A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0           # unlimited rows
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.90           # extra space for header
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3

    # Page number top-right in size-20 font
    ws.oddHeader.right.text = "&20&P"
    ws.evenHeader.right.text = "&20&P"
    ws.page_setup.differentOddEven = False


def _apply_order_borders(ws, start_row: int, end_row: int) -> None:
    """
    Apply a thin border around the order block (cols A-C) and a separate
    thin border around the Notes column (col D), so deleting Notes doesn't
    remove the border from the rest of the order.
    """
    notes_col = _SUMMARY_LAST  # col D (4)
    main_last = notes_col - 1  # col C (3)

    for r in range(start_row, end_row + 1):
        for c in range(1, notes_col + 1):
            cell = ws.cell(row=r, column=c)
            existing = cell.border
            top = existing.top
            bottom = existing.bottom
            left = existing.left
            right = existing.right

            # Main block (cols 1-3)
            if c <= main_last:
                if c == 1:
                    left = _THIN
                if c == main_last:
                    right = _THIN
                if r == start_row:
                    top = _THIN
                if r == end_row:
                    bottom = _THIN

            # Notes column (col 4) — independent border box
            if c == notes_col:
                left = _THIN
                right = _THIN
                if r == start_row:
                    top = _THIN
                if r == end_row:
                    bottom = _THIN

            cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def _write_data_sheet(ws, sorted_orders: list[MatchedOrder]) -> None:
    for col_idx, width in enumerate(_DATA_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Header row
    for col_idx, header in enumerate(_DATA_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    row = 2
    for (platform, order_id), items in groupby(
        sorted_orders, key=lambda m: (m.platform, m.order_id)
    ):
        for m in items:
            sku  = f"*{m.sku}" if m.is_invoice_match else m.sku
            vals = [platform, order_id, sku, m.description,
                    "*" if m.is_invoice_match else "", m.quantity, m.notes]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = _NORMAL_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (4, 7)))
            row += 1
        # Blank separator between orders
        row += 1
