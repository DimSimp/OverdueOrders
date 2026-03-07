import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
import holidays, datetime, requests, json, pprint, psycopg2, os
from psycopg2 import OperationalError
from ebaysdk.exception import ConnectionError
from ebaysdk.trading import Connection as Trading
from rich.console import Console
import time, pprint, pyperclip, traceback
from openpyxl import Workbook

import smtplib, ssl, imaplib, time, pyperclip, re, sys, requests, json
import ftplib

from openpyxl import Workbook
from mailmerge import MailMerge
from docx2pdf import convert
from pdfnup import generateNup
from PyPDF2 import PdfFileMerger, PdfFileReader, _reader

GetItemheaders = {'Content-Type': 'application/json', 'Accept': 'application/json', 'NETOAPI_ACTION': 'GetItem',
                  'NETOAPI_USERNAME': 'kyalAPI', 'NETOAPI_KEY': 'HSZ3EyI9ki6sseqRIkwy6tL6yFNPEUyM', }


filename = 'Morning_Inventory_Report.xlsx'
filename2 = 'Afternoon_Inventory_Report.xlsx'

print("Logging into FTP....")


pass_condition = ''

while True:

    if pass_condition == 'true':
        break

    pass_condition = ''
    try:
      ftp = ftplib.FTP("ftp.drivehq.com")
      pass_condition = 'true'
      break

    except:
        continue
ftp.login("kyaldabomb", "D5es4stu!")

print("Getting Morning / Afternoon Report files...")

ftp.retrbinary("RETR " + filename, open(rf"C:\Python\{filename}", 'wb').write)
ftp.retrbinary("RETR " + filename2, open(rf"C:\Python\{filename2}", 'wb').write) ####will need to change location to documents

ftp.quit()



###Need to download excel files from FTP######

df1 = pd.read_excel(r"C:\Python\Morning_Inventory_Report.xlsx", header=None)
df2 = pd.read_excel(r"C:\Python\Afternoon_Inventory_Report.xlsx", header=None)

print('Comparing the two sheets...')

# Find the differences
df_diff = pd.concat([df1, df2]).drop_duplicates(keep=False)

# Write the differences to an Excel file
df_diff.to_excel(r"C:\Python\Inventory_Report_test.xlsx", index=False)

# Create dictionaries for morning and afternoon quantities
morning_qty = df1.set_index(0)[8].to_dict()
afternoon_qty = df2.set_index(0)[8].to_dict()

items_ordered_final = []

# Iterate through the differences
for _, row in df_diff.iterrows():
    sku = row[0]

    afternoon_quantity = afternoon_qty.get(sku, 0)
    morning_quantity = morning_qty.get(sku, 0)

    try:
        qty_difference = float(afternoon_quantity) - float(morning_quantity)
    except:
        continue
    if qty_difference > 0:
        items_ordered_final.append([sku, qty_difference, row[18]])

print(f"Found {len(items_ordered_final)} items with increased quantity.")

# If you need to see the items:
for item in items_ordered_final:
    print(f"SKU: {item[0]}, Increased by: {item[1]}, Supplier: {item[2]}")

#print(items_ordered_final)  #####Just the orders that arrived, AKA stock that has INCREASED in quantity
###### GOT ALL THE ITEMS THAT

print('Got all items ordered today...')



############################################ PART 2 ########################################
        ##################### Comparing the items ordered in to overdue orders   ########################

api = Trading(appid="KyalScar-Awaiting-PRD-aca833663-f54c920e", timeout=120, config_file=None,
              devid="5c97c2a8-eae7-49cc-ae81-c3ed1bbab335", certid="PRD-ca8336639ad2-75af-45bc-992a-d0bc",
              token="v^1.1#i^1#r^1#f^0#I^3#p^3#t^Ul4xMF84OkUzOEU3NzVFQzA3NDUwNjUyMkY0NTc4QjlENjRFRDVFXzFfMSNFXjI2MA==")

ONE_DAY = datetime.timedelta(days=1)  ###### Getting next business day for pickups
HOLIDAYS_AU = holidays.AU(prov='VIC')


def create_connection(db_name, db_user, db_password, db_host, db_port):
    connection = None
    try:
        connection = psycopg2.connect(
            database=db_name,
            user=db_user,
            password=db_password,
            host=db_host,
            port=db_port,
        )
        print("Connection to PostgreSQL DB successful")
    except OperationalError as e:
        print(f"The error '{e}' occurred")
    return connection


connection = create_connection(
    "postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432"
)


def overdue_date():
    two_previous_days = datetime.date.today() - ONE_DAY
    return two_previous_days


while True:
    try:

        response = api.execute('GetMyeBaySelling', {
            "SoldList": {"Include": True, "IncludeNotes": True, "DurationInDays": 7,
                         "Pagination": {"EntriesPerPage": 200}}})
        unformatteddic = response.dict()
        paginationresult = int(unformatteddic['SoldList']['PaginationResult']['TotalNumberOfPages'])
        break
    except:
        continue
SoldList = {}
soldlistnumber = 0

#for yy in range(int(paginationresult)):

for yy in range(int(1)):

    pass_condition_list = ''

    while True:

        if pass_condition_list == 'true':
            break

        try:
            response = api.execute('GetMyeBaySelling', {
                "SoldList": {"Include": True, "IncludeNotes": True, "DurationInDays": 10,
                             "OrderStatusFilter": "AwaitingShipment",
                             "Pagination": {"EntriesPerPage": 200, "PageNumber": int(yy) + 1}}})
            unformatteddic = response.dict()
            # pprint.pprint(unformatteddic['SoldList']['OrderTransactionArray'])

            for t in unformatteddic['SoldList']['OrderTransactionArray']['OrderTransaction']:
                SoldList.update({soldlistnumber: t})
                soldlistnumber = soldlistnumber + 1

            pass_condition_list = 'true'

            break
        except:
            continue


overdue = overdue_date()

overdue_string = overdue.strftime("%Y-%m-%d 23:59:59")

overdue = datetime.datetime.strptime(overdue_string, '%Y-%m-%d %H:%M:%S')

messaged_orders = []

overdue_orders = []

tod = datetime.datetime.now()
d = datetime.timedelta(days=120)
a = tod - d
date_placed = a.strftime("%Y-%m-%d 00:00:00")

#####Will look for overdue orders placed from 60 days ago to 2 business days ago
## Need to add , 'kogan', 'mydeal', 'ozsale', 'Amazon AU', 'catch', 'ebay' to sales channcel
data = {'Filter': {'DatePlacedFrom': date_placed,
                   'DatePlacedTo': overdue_string,
                   'SalesChannel': ['website', 'ebay', 'amazon au', 'catch', 'kogan', 'mydeal', 'ozsale', 'everydaymarket', 'control panel', 'BigW', 'quote'], 'OrderStatus': ['Pick', 'On Hold', 'On Hold', 'Uncommited'],
                   'OutputSelector': ["ID", "GrandTotal",
                                      'OrderLine.ProductName',
                                      'OrderLine.Quantity', 'StickyNotes', 'OrderLine.UnitPrice', 'ShippingOption',
                                      'ShipAddress', 'SalesChannel', 'PurchaseOrderNumber', 'Email',
                                      'eBay.eBayUsername', 'OrderLine.eBay.DatePaid', 'KitComponents']}}

headers = {'Content-Type': 'application/json', 'Accept': 'application/json', 'NETOAPI_ACTION': 'GetOrder',
           'NETOAPI_USERNAME': 'kyalAPI', 'NETOAPI_KEY': 'HSZ3EyI9ki6sseqRIkwy6tL6yFNPEUyM', }

r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=headers, json=data)

response = r.text
response = json.loads(response)

#pprint.pprint(response)

# VVVVVVVVVVVVVVVVVVVVVVVV   ITERATING THROUGH ITEMS TO FIND KITTED SKUS VVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVV

total_items = []

for x in response['Order']:
    orderlinelen = len(x['OrderLine'])

    for y in range(orderlinelen):

        total_items.append(x['OrderLine'][y]['SKU'])

GetItemdata = {'Filter': {'SKU': total_items,
                              'OutputSelector': ["Name", 'KitComponents', 'PriceGroups']}}

r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=GetItemheaders,
                                      json=GetItemdata, timeout=1000)

kitted_response = r.text
kitted_response = json.loads(kitted_response)

for x in kitted_response['Item']:
    if len(x['KitComponents'][0]['KitComponent']) < 1:  ###Not a kitted item
        continue
    else:
        parent_sku = x['SKU']
        kitted_skus = []

        if isinstance(x['KitComponents'][0]['KitComponent'], dict) is True:
            component_sku = x['KitComponents'][0]['KitComponent']['ComponentSKU']
            kitted_skus.append(component_sku)
        else:
            for k_skus in x['KitComponents'][0]['KitComponent']:
                component_sku =k_skus['ComponentSKU']
                quantity = k_skus['AssembleQuantity']
                kitted_skus.append(component_sku)

        for yy in response['Order']:
            orderlinelen = len(yy['OrderLine'])

            for yyy in range(orderlinelen):
                product_sku = yy['OrderLine'][yyy]['SKU']
                if product_sku == parent_sku:
                    for kitted_items in kitted_skus:
                        yy['OrderLine'].append({'Quantity': 1,
                                                'ProductName': f'Child SKU of {parent_sku} bundle',
                                                'SKU': kitted_items})

# ^^^^^^^^^^^^^^^^^^   ITERATING THROUGH ITEMS TO FIND KITTED SKUS ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^

print('Comparing items ordered today with overdue orders...')

for x in response['Order']:
    sales_channel = x['SalesChannel']
    order_number = x['OrderID']

    if order_number == 'N359311':
        print('bing')
        print('bong')


    if order_number in messaged_orders:
        continue

    order_failed = ''



    if sales_channel.lower() != 'ebay':  ######### adding the orders that are NOT ebay to the orders that need to be messaged list
        sticky_note = x['StickyNotes']
        orderlinelen = len(x['OrderLine'])

        sticky_notes_total = sticky_note

        if sticky_note == '':
            pass

        shipping_method = ''

        if 'express' in x['ShippingOption'].lower():

            shipping_method = 'express'

        if 'click' in x['ShippingOption'].lower():

            shipping_method = 'Click & Collect'

        if 'express' in shipping_method.lower() or 'click' in shipping_method.lower():
           #order_number = f'[{shipping_method.upper()}] {order_number}'
           x['OrderID'] = f'[{shipping_method.upper()}] {order_number}'

        if isinstance(sticky_note, list):

            has_arrived = 'false'

            sticky_notes_total = ''
            counter = 0

            for y in sticky_note:

                if 'customer' in y['Title'].lower() or 'messaged' in y['Title'].lower() or 'happy' in y['Title'].lower():
                    pass

                else:

                    counter+=1

                    sticky_notes_total = f'''{sticky_notes_total}Note {str(counter)}
Title: {y['Title']}

Description: {y['Description']}

'''

                if 'ware' in y['Description'].lower() or 'fact' in y['Description'].lower():
                    has_arrived = 'true'

                elif 'drop' in y['Description'].lower():
                    order_failed = 'true'

            if has_arrived == 'true':
                x['Total_Notes'] = sticky_notes_total

                overdue_orders.append(x)
                continue


        else:

            if sticky_note == '':
                pass
            elif 'ware' in sticky_note['Description'].lower() or 'fact' in sticky_note['Description'].lower():

                sticky_notes_total = f'''Title: {sticky_note['Title']}

Description: {sticky_note['Description']}'''
                x['Total_Notes'] = sticky_notes_total

                overdue_orders.append(x)
                continue

            elif 'drop' in sticky_note['Description'].lower():
                continue

        if order_failed == 'true':
            continue

        sticky_note_length = len(sticky_note)

        has_arrived = 'false'

        for y in range(orderlinelen):

            product_sku = x['OrderLine'][y]['SKU']
            product_title = x['OrderLine'][y]['ProductName']


            for arrived_items in items_ordered_final:

                if product_sku == arrived_items[0]:
                    has_arrived = 'true'
                    x['OrderLine'][y]['ProductName'] = f'[JA] {product_title}'
                    x['OrderLine'][y]['Supplier'] = arrived_items[2]

        sticky_notes_total = ''
        counter = 0
        if isinstance(sticky_note, list):
            for y in sticky_note:

                if 'customer' in y['Title'].lower() or 'messaged' in y['Title'].lower() or 'happy' in y['Title'].lower():
                    pass

                else:

                    counter += 1

                    sticky_notes_total = f'''{sticky_notes_total}Note {str(counter)}
Title: {y['Title']}

Description: {y['Description']}'''

        elif sticky_note == '':
            pass

        else:

            if 'customer' in sticky_note['Title'].lower() or 'messaged' in sticky_note['Title'].lower() or 'happy' in sticky_note['Title'].lower():
                pass

            else:

                sticky_notes_total = f'''Title: {sticky_note['Title']}

Description: {sticky_note['Description']}'''

        if has_arrived == 'true': ##if one item matches, will append the whole order and assume it's doable
            x['Total_Notes'] = sticky_notes_total
            overdue_orders.append(x)


    else:

        date_paid = x['OrderLine'][0]['eBay']['DatePaid']
        orderlinelen = len(x['OrderLine'])


        date_paid = datetime.datetime.strptime(date_paid, '%Y-%m-%d %H:%M:%S')

        if date_paid > overdue:
            continue

        # elif date_paid < overdue:
        #     print('bong')

        order_failed = 'false'
        ultimate_pass = 'false'
        ebay_order_number = x['PurchaseOrderNumber']

        response = api.execute('GetOrders', {"OrderIDArray": {"OrderID": ebay_order_number}})
        unformatteddic = response.dict()

        try:
            shipped_time = unformatteddic['OrderArray']['Order'][0]['ShippedTime']
            continue
        except:
            pass
############## CHECKING STICKY NOTES FIRST BEFORE CHECKING IF ITEM HAS ARRIVES
        try:

            for xx in unformatteddic['OrderArray']['Order'][0]['TransactionArray']['Transaction']:
                new_orders_orderline = xx['OrderLineItemID']

                shipping_method = ''

                if float(xx['ActualShippingCost']['value']) > 0:

                    shipping_method = 'express'


                    order_number = f'[{shipping_method.upper()}] {order_number}'

                for t in SoldList:

                    passcondition = 'failed'

                    Stickydic = {}

                    if 'Transaction' in SoldList[t]:
                        try:

                            orderlinelen = 1

                            soldlist_orderline = SoldList[t]['Transaction']['OrderLineItemID']

                            if soldlist_orderline == new_orders_orderline:

                                passcondition = 'passed'

                                item_id = SoldList[t]['Transaction']['Item']['ItemID']

                                x.update({"eBayItemID": f"{item_id}"})

                                try:

                                    sticky_note = SoldList[t]['Transaction']['Item']['PrivateNotes']

                                except KeyError:
                                    sticky_note = ''

                                if 'ware' in sticky_note.lower() or 'fact' in sticky_note.lower():

                                    overdue_orders.append(x)
                                    ultimate_pass = 'true'
                                    continue

                                elif 'drop' in sticky_note.lower():
                                    order_failed = 'true'

                                # p.update({'Sticky Note': sticky_note})
                                # preFinalList.append(p)
                        except TypeError:
                            continue


                        else:
                            continue

                    elif 'Order' in SoldList[t]:

                        orderline_length = len(SoldList[t]['Order']['TransactionArray']['Transaction'])

                        for multitransactions in SoldList[t]['Order']['TransactionArray'][
                            'Transaction']:  ###  Checking Multi Order Lines

                            soldlist_orderline = multitransactions['OrderLineItemID']

                            if soldlist_orderline == new_orders_orderline:

                                item_id = multitransactions['Item']['ItemID']

                                x.update({"eBayItemID": f"{item_id}"})

                                try:
                                    sticky_note = multitransactions['Item']['PrivateNotes']
                                except KeyError:
                                    sticky_note = ''

                                if 'ware' in sticky_note.lower() or 'fact' in sticky_note.lower():

                                    overdue_orders.append(x)
                                    ultimate_pass = 'true'
                                    continue

                                elif 'drop' in sticky_note.lower():
                                    order_failed = 'true'
        except:
            order_failed = 'true'
        if order_failed == 'true' or ultimate_pass == 'true':
            continue

        has_arrived = 'false'

        for y in range(orderlinelen):

            product_sku = x['OrderLine'][y]['SKU']

            product_title = x['OrderLine'][y]['ProductName']





            for arrived_items in items_ordered_final:
                if product_sku == arrived_items[0]:
                    has_arrived = 'true'
                    x['OrderLine'][y]['ProductName'] = f'[JA] {product_title}'
                    x['OrderLine'][y]['Supplier'] = arrived_items[2]

        if has_arrived == 'true': ##if one item matches, will append the whole order and assume it's doable

            try:
                x['Total_Notes'] = sticky_note
            except:
                x['Total_Notes'] = ''

            overdue_orders.append(x)
            continue




        # pprint.pprint(unformatteddic)

        ###################PART 3###########################

print('Creating Excel Picksheet...')

wb = Workbook()
wsMasterPickSheet = wb.create_sheet('MasterPickSheet')
wsMasterPickSheet['A1'] = 'Order Number'
wsMasterPickSheet['B1'] = 'Sales Channel'
wsMasterPickSheet['C1'] = 'Item Title'
wsMasterPickSheet['D1'] = 'SKU'
wsMasterPickSheet['E1'] = 'Quantity'
wsMasterPickSheet['F1'] = 'Supplier'
wsMasterPickSheet['G1'] = 'Sticky notes'


for orders in overdue_orders:
    sales_channel = orders['SalesChannel']

    try:

        sticky_note = orders['Total_Notes']

    except:
        sticky_note = ''

    # if orders['PurchaseOrderNumber'] == '':
    #
    #     order_id = orders['OrderID']
    #
    # else:
    #     order_id = orders['PurchaseOrderNumber']

    if 'ebay' in sales_channel.lower():
        order_id = orders['PurchaseOrderNumber']

    else:
        order_id = orders['ID']

    for items in orders['OrderLine']:
        item_title = items['ProductName']
        item_sku = items['SKU']
        item_quantity = items['Quantity']

        try:
            supplier = items['Supplier']

        except:
            supplier = 'NA'


        maxrow = wsMasterPickSheet.max_row
        maxrow = maxrow + 1
        wsMasterPickSheet['A' + str(maxrow)].value = order_id
        wsMasterPickSheet['B' + str(maxrow)].value = sales_channel
        wsMasterPickSheet['C' + str(maxrow)].value = item_title
        wsMasterPickSheet['D' + str(maxrow)].value = item_sku
        wsMasterPickSheet['E' + str(maxrow)].value = item_quantity
        wsMasterPickSheet['F' + str(maxrow)].value = supplier
        try:
            wsMasterPickSheet['G' + str(maxrow)].value = sticky_note
        except:
            wsMasterPickSheet['G' + str(maxrow)].value = ''



wb.save(r'\\SERVER\Python\Envelope Templates\MasterPickSheet.xlsx')

template = r"\\SERVER\Python\Envelope Templates\Label Template JAN 2022.docx"

document = MailMerge(template)
# print(document.get_merge_fields())

document.merge_templates(overdue_orders, separator='continuous_section')

document.write(r"\\SERVER\Python\Envelope Templates\Labels.docx")

convert(r"\\SERVER\Python\Envelope Templates\Labels.docx",
        r"\\SERVER\Python\Envelope Templates\Labels_Intermediate.pdf")

generateNup(r"\\SERVER\Python\Envelope Templates\Labels_Intermediate.pdf", 4,
            r"\\SERVER\Python\Envelope Templates\Labels.pdf")


os.startfile(r'\\SERVER\Python\Envelope Templates\MasterPickSheet.xlsx')
#os.startfile(r"\\SERVER\Python\Envelope Templates\Labels.pdf")





