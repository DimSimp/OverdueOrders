from openpyxl import Workbook
import openpyxl, requests, pprint, json, time, holidays, datetime, math, re, sys, psycopg2, shutil, PyPDF2, base64
from psycopg2 import OperationalError
from requests.auth import HTTPBasicAuth
import os.path, zeep, base64, uuid
from zeep.transports import Transport
import code128
from requests import Session
from zeep.plugins import HistoryPlugin
from lxml import etree
import xmltodict
from send2trash import send2trash
from mailmerge import MailMerge
from docx2pdf import convert
from pdfnup import generateNup
from oauthlib.oauth2 import BackendApplicationClient
from requests_oauthlib import OAuth2Session
from PyPDF2 import PdfFileMerger, PdfFileReader

import urllib  # Stuff for Google API
import json
import requests
from urllib.parse import urlparse
from time import sleep
import pandas as pd
import ctypes.wintypes
from ebaysdk.exception import ConnectionError
from ebaysdk.trading import Connection as Trading
import concurrent.futures

try:
    from pil import Image, ImageDraw, ImageFont
    from pil import ImageOps

except ModuleNotFoundError as err:
    from PIL import Image, ImageDraw, ImageFont
    from PIL import ImageOps

from pdf2image import convert_from_path, convert_from_bytes
from rich.console import Console

console = Console(force_terminal=True)

while True:

    password = input('''Before thy pass, enter thy name of the big grey cat

    Enter Response:''')

    if password.lower() == 'chester':
        break

    else:
        print('Fool, password incorrect. Try Again.')
        time.sleep(1)
        continue

CSIDL_PERSONAL = 5  # My Documents
SHGFP_TYPE_CURRENT = 0  # Get current, not default value
buf = ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_PERSONAL, None, SHGFP_TYPE_CURRENT, buf)
documents_folder = buf.value
imagelocations = documents_folder + r"\Python"


def get_google_address():  # inital definiton for google places api
    global suburb
    global postal_code
    dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
    qstr = urllib.parse.urlencode(dict1)
    URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
    URL = URL + qstr
    response = urllib.request.urlopen(URL)
    data = json.load(response)
    # pprint.pprint(data['candidates'][0]['place_id'])
    placeid = data['candidates'][0]['place_id']
    payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
    r = requests.get(
        'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
    # pprint.pprint(r.json())
    data = r.json()
    # pprint.pprint (data['result']['address_components'])
    for t in range(20):
        if data['result']['address_components'][t]['types'][0] == 'locality':
            suburb = data['result']['address_components'][t]['long_name']
            break

        else:
            continue

    for t in range(20):
        if data['result']['address_components'][t]['types'][0] == 'postal_code':
            postal_code = data['result']['address_components'][t]['long_name']
            break

        else:
            continue


############## STUFF FOR SQL SERVER ##############################

def create_connection(db_name, db_user, db_password, db_host, db_port):
    connection = None
    try:
        connection = psycopg2.connect(
            database=db_name,
            user=db_user,
            password=db_password,
            host=db_host,
            port=db_port,
            keepalives=1,
            keepalives_idle=30,
            keepalives_interval=10,
            keepalives_count=5
        )
    # print("Connection to PostgreSQL DB successful")
    except OperationalError as e:
        print(f"The error '{e}' occurred")
    return connection


def execute_read_query(connection, query):
    cursor = connection.cursor()
    result = None
    try:
        cursor.execute(query)
        result = cursor.fetchall()
        return result
    except OperationalError as e:
        print(f"The error '{e}' occurred")


def execute_sku_query(connection, item_check):
    needs_weight_check = 'false'
    cursor1 = connection.cursor()
    result = None
    try:
        result = cursor1.execute(item_check)
        result = cursor1.fetchall()

        if result == []:
            needs_weight_check = 'true'
            return needs_weight_check

        else:
            return needs_weight_check

    except OperationalError as e:
        print(e)
        sys.exit()


def add_item(sku, location):
    add_item = f"INSERT INTO item_data (sku, item_location) VALUES ('{SKU}', '{location}'); COMMIT;"
    cursor = connection.cursor()
    cursor.execute(add_item, item)


def execute_insert_query(connection, item_check):
    cursor1 = connection.cursor()
    result = None
    try:
        result = cursor1.execute(item_check)
        # result = cursor1.fetchall()

    except OperationalError as e:
        print(e)
        sys.exit()


def item_insert(sku):
    item_check = f"INSERT INTO sendle_weight_test(item_number) VALUES('{sku}');COMMIT;"
    execute_insert_query(connection, item_check)


def item_check(sku):
    item_check = f"SELECT 1 FROM sendle_weight_test WHERE item_number = '{sku}';"
    results = execute_sku_query(connection, item_check)
    return results


def item_thickness_check(sku):
    item_check = f"SELECT 1 FROM fastwaythicknesstest WHERE item_number = '{sku}';"
    results = execute_sku_query(connection, item_check)
    return results


def item_thickness_insert(sku):
    item_check = f"INSERT INTO fastwaythicknesstest(item_number) VALUES('{sku}');COMMIT;"
    execute_insert_query(connection, item_check)


connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")

history = HistoryPlugin()
session = Session()
transport = Transport(session=session)

#######NEED TO UNCOMMENT BELOW IN PROD

wsdl = 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS?wsdl'
# wsdl = 'http://triton.alliedexpress.com.au:8080/ttws-ejb/TTWS'

############## STUFF FOR SQL SERVER ##############################

client = BackendApplicationClient(client_id='fw-fl2-MEL0900146-d2f1bfc5a108')
oauth = OAuth2Session(client=client)
token = oauth.fetch_token(token_url='https://identity.fastway.org/connect/token',
                          client_id='fw-fl2-MEL0900146-d2f1bfc5a108',
                          client_secret='4c6483bb-5994-4773-b263-e0dfb7b29edf', scope='fw-fl2-api-au')

bearer_token = token['access_token']
bearer_token = 'bearer ' + bearer_token  # Authorization token

base_url = 'https://api.myfastway.com.au'

Fastway_Headers = {"Authorization": bearer_token,
                   "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
                   "Accept-Encoding": "*",
                   "Connection": "keep-alive"}  # Authorization Header

#######^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^  Fastway Authorization ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^###############

freightster_headers = {'Content-Type': 'application/json', 'Accept': 'application/json',
                       'Authorization': 'S1OL4341HUTIDPKRS0LMOZ9QHANPC1L5',
                       "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
                       "Accept-Encoding": "*",
                       "Connection": "keep-alive"}

#######^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^  Freighster Headers ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^###############

wb = openpyxl.load_workbook(r"\\SERVER\Python\Freightster.xlsx")
sheet = wb['Freightster']

#######^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^  Freightster loading pricing zone worksheet ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^###############

today = datetime.date.today()
if today.isoweekday() in set((6, 7)):
    today += datetime.timedelta(days=today.isoweekday() % 5)
next_day_allied = str(today.day) + '/' + str(today.month) + '/' + str(today.year) + " 10:00:00"

history = HistoryPlugin()
session = Session()
transport = Transport(session=session)
wsdl = 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS?wsdl'

try:
    allied_client = zeep.Client(wsdl=wsdl, transport=transport, plugins=[history])

    allied_client.transport.session.proxies = {
        # Utilize for all http/https connections
        'http': 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS', }
    allied_account = allied_client.service.getAccountDefaults('755cf13abb3934695f03bd4a75cfbca7', "SCAMUS", "VIC",
                                                              "AOE")

except:
    allied_failed = 'true'

ONE_DAY = datetime.timedelta(days=1)  ###### Getting next business day for pickups
HOLIDAYS_AU = holidays.AU(prov='VIC')


def next_business_day():
    next_day = datetime.date.today() + ONE_DAY
    while next_day.weekday() in holidays.WEEKEND or next_day in HOLIDAYS_AU:
        next_day += ONE_DAY
    return next_day


next_day = next_business_day()

next_cp_day = next_day.strftime('%Y-%m-%d')
next_day = next_day.strftime('%d/%m/%Y')

next_cp_day = f"{next_cp_day} 04:00 PM"

cp_auth = '113153878:CBBC1CAD2F335C9CEEA1D0BC63C7056ADBC9C36DD54A9A4530A4380D9AAC5FE0'
### This is for sandbox, when ready, replace 2nd half with CBBC1CAD2F335C9CEEA1D0BC63C7056ADBC9C36DD54A9A4530A4380D9AAC5FE0

cp_auth_encoded = cp_auth.encode()
cp_auth_encoded = base64.b64encode(cp_auth_encoded)
cp_auth_encoded = cp_auth_encoded.decode()

cp_headers = {'Host': 'api.couriersplease.com.au',
              'Accept': 'application/json',
              'Content-Type': 'application/json',
              'Authorization': f'Basic {cp_auth_encoded}',
              'Content-Length': '462',
              "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
              "Accept-Encoding": "*",
              "Connection": "keep-alive"}

accountnumber = '0007805312'
username = '966afe1c-c07b-4902-b1d7-77ad1aac0915'
secret = 'x428429fabeb99f420f1'

minimum_tracking_amount = 30

wb = Workbook()
wsMini = wb.active
wsMini.title = "Mini."
# wsMini = wb.create_sheet('Mini.')
wsDevil = wb.create_sheet('Devil.')
wsLabel = wb.create_sheet('Label.')
wsAustraliaPost = wb.create_sheet('AustraliaPost')
wsSendle = wb.create_sheet('Sendle')
wsTransdirect = wb.create_sheet('Transdirect')
wsMasterPickSheet = wb.create_sheet('MasterPickSheet')

wsMini['A1'] = wsDevil['A1'] = wsLabel['A1'] = 'ID'
wsMini['B1'] = wsDevil['B1'] = wsLabel['B1'] = 'Buyer Name'
wsMini['C1'] = wsDevil['C1'] = wsLabel['C1'] = 'Buyer Email'
wsMini['D1'] = wsDevil['D1'] = wsLabel['D1'] = 'Buyer Address 1'
wsMini['E1'] = wsDevil['E1'] = wsLabel['E1'] = 'Buyer Address 2'
wsMini['F1'] = wsDevil['F1'] = wsLabel['F1'] = 'Buyer City'
wsMini['G1'] = wsDevil['G1'] = wsLabel['G1'] = 'Buyer State'
wsMini['H1'] = wsDevil['H1'] = wsLabel['H1'] = 'Buyer Postcode'
wsMini['I1'] = wsDevil['I1'] = wsLabel['I1'] = 'Item Title'
wsMini['J1'] = wsDevil['J1'] = wsLabel['J1'] = 'Custom Label'
wsMini['K1'] = wsDevil['K1'] = wsLabel['K1'] = 'Quantity'
wsMini['L1'] = wsDevil['L1'] = wsLabel['L1'] = 'Total Price'
wsMini['M1'] = wsDevil['M1'] = wsLabel['M1'] = 'Postage Service'
wsMini['N1'] = wsDevil['N1'] = wsLabel['N1'] = 'Item Note / Postage Type'
wsMini['O1'] = wsDevil['O1'] = wsLabel['O1'] = 'Item Number (eBay)'
wsMini['P1'] = wsDevil['P1'] = wsLabel['P1'] = 'Height (cm)'
wsMini['Q1'] = wsDevil['Q1'] = wsLabel['Q1'] = 'Length (cm)'
wsMini['R1'] = wsDevil['R1'] = wsLabel['R1'] = 'Width (cm)'
wsMini['S1'] = wsDevil['S1'] = wsLabel['S1'] = 'Weight (kg)'
wsMini['T1'] = wsDevil['T1'] = wsLabel['T1'] = 'Phone'
wsMini['U1'] = wsDevil['U1'] = wsLabel['U1'] = 'Company'

wsAustraliaPost['A1'] = 'C_CONSIGNMENT_ID'  ##IGNORED, LEAVE BLANK
wsAustraliaPost['B1'] = 'C_CONSIGNEE_EMAIL'  ### Email for reciever
wsAustraliaPost['C1'] = 'C_EMAIL_NOTIFICATION'  ### Email notifications on? Always make 'Y'
wsAustraliaPost['D1'] = 'A_ACTUAL_CUBIC_WEIGHT'  ### Formula for this is length (cm) x width (cm) x height (cm) / 6000
wsAustraliaPost['E1'] = 'A_LENGTH'  ### Length (cm)
wsAustraliaPost['F1'] = 'A_WIDTH'  ### Width (cm)
wsAustraliaPost['G1'] = 'A_HEIGHT'  ##Height (cm)
wsAustraliaPost['H1'] = 'G_DESCRIPTION'  ### Scarlett music order xxx
wsAustraliaPost['I1'] = 'G_WEIGHT'  ### Weight in kg
wsAustraliaPost['J1'] = 'C_CHARGE_CODE'  ### Always 7D55 (Standard eParcel)
wsAustraliaPost['K1'] = 'C_CONSIGNEE_NAME'  ### Customer Name
wsAustraliaPost['L1'] = 'C_CONSIGNEE_BUSINESS_NAME'  ###Company
wsAustraliaPost['M1'] = 'C_CONSIGNEE_ADDRESS_1'  ### AddressLine1
wsAustraliaPost['N1'] = 'C_CONSIGNEE_ADDRESS_2'  ###AddressLine2
wsAustraliaPost['O1'] = 'C_CONSIGNEE_SUBURB'  ###SUBURB
wsAustraliaPost['P1'] = 'C_CONSIGNEE_STATE_CODE'  ##Short State Code
wsAustraliaPost['Q1'] = 'C_CONSIGNEE_COUNTRY_CODE'  ##AU
wsAustraliaPost['R1'] = 'C_CONSIGNEE_PHONE_NUMBER'  ##Phone
wsAustraliaPost['S1'] = 'C_PHONE_PRINT_REQUIRED'  ##Show phone number, always 'Y'
wsAustraliaPost['T1'] = 'C_SIGNATURE_REQUIRED'  # Y always i guess
wsAustraliaPost['U1'] = 'C_REF'  ##Reference / Order ID
wsAustraliaPost['V1'] = 'C_REF_PRINT_REQUIRED'  ##Show reference, always Y
wsAustraliaPost['W1'] = 'C_CONSIGNEE_POSTCODE'  ##Postcode

wsAustraliaPost['A2'] = 'IGNORED'
wsAustraliaPost['B2'] = 'MANDATORY/OPTIONAL REFER TO GUIDE'
wsAustraliaPost['C2'] = 'OPTIONAL'
wsAustraliaPost['D2'] = 'MANDATORY'
wsAustraliaPost['E2'] = 'MANDATORY/OPTIONAL REFER TO GUIDE'
wsAustraliaPost['F2'] = 'MANDATORY/OPTIONAL REFER TO GUIDE'
wsAustraliaPost['G2'] = 'MANDATORY/OPTIONAL REFER TO GUIDE'
wsAustraliaPost['H2'] = 'MANDATORY'
wsAustraliaPost['I2'] = 'MANDATORY/OPTIONAL REFER TO GUIDE'
wsAustraliaPost['J2'] = 'MANDATORY'
wsAustraliaPost['K2'] = 'MANDATORY/OPTIONAL REFER TO GUIDE'
wsAustraliaPost['L2'] = 'OPTIONAL'
wsAustraliaPost['M2'] = 'MANDATORY/OPTIONAL REFER TO GUIDE'
wsAustraliaPost['N2'] = 'OPTIONAL'
wsAustraliaPost['O2'] = 'MANDATORY/OPTIONAL REFER TO GUIDE'
wsAustraliaPost['P2'] = 'MANDATORY/OPTIONAL REFER TO GUIDE'
wsAustraliaPost['Q2'] = 'MANDATORY/OPTIONAL REFER TO GUIDE'
wsAustraliaPost['R2'] = 'MANDATORY/OPTIONAL REFER TO GUIDE'
wsAustraliaPost['S2'] = 'OPTIONAL'
wsAustraliaPost['T2'] = 'OPTIONAL'
wsAustraliaPost['U2'] = 'OPTIONAL'
wsAustraliaPost['V2'] = 'OPTIONAL'
wsAustraliaPost['W2'] = 'MANDATORY/OPTIONAL REFER TO GUIDE'

wsSendle['A1'] = 'sender_name'
wsSendle['B1'] = 'sender_company'
wsSendle['C1'] = 'sender_address_line1'
wsSendle['D1'] = 'sender_address_line2'
wsSendle['E1'] = 'sender_suburb'
wsSendle['F1'] = 'sender_state_name'
wsSendle['G1'] = 'sender_postcode'
wsSendle['H1'] = 'sender_country'
wsSendle['I1'] = 'sender_contact_number'
wsSendle['J1'] = 'pickup_instructions'  ##At the music shop, open 9am-6pm. Parking is best across the road at The Palms.
wsSendle['K1'] = 'receiver_name'
wsSendle['L1'] = 'receiver_email'
wsSendle['M1'] = 'receiver_company'
wsSendle['N1'] = 'receiver_address_line1'
wsSendle['O1'] = 'receiver_address_line2'
wsSendle['P1'] = 'receiver_suburb'
wsSendle['Q1'] = 'receiver_state_name'
wsSendle['R1'] = 'receiver_postcode'
wsSendle['S1'] = 'receiver_country'
wsSendle['T1'] = 'receiver_contact_number'
wsSendle['U1'] = 'delivery_instructions'
wsSendle['V1'] = 'pickup_date'
wsSendle['W1'] = 'kilogram_weight'
wsSendle['X1'] = 'cubic_metre_volume'  ### l x w x h (in meters!)
wsSendle['Y1'] = 'description'
wsSendle['Z1'] = 'customer_reference'

wsTransdirect['A1'] = 'Order ID'  ##
wsTransdirect['B1'] = 'Goods Description'  ##Scarlett Music order number xxxx
wsTransdirect['C1'] = 'Carrier'  # Eg Fastway
wsTransdirect['D1'] = 'Pickup Date'  # Use the next_day variable
wsTransdirect['E1'] = 'Address Type'  # Always go residential I think just to be safe
wsTransdirect['F1'] = 'Receiver Name'
wsTransdirect[
    'G1'] = 'Email Address'  # Will have to make sure if there's no email, default to info@scarlettmusic.com.au
wsTransdirect[
    'H1'] = 'Phone Number'  # will have to make sure to parse out the '+' and make sure it's the rigt amount of digits
wsTransdirect['I1'] = 'Business Name'  # Company variable (?)
wsTransdirect['J1'] = 'Street Address'
wsTransdirect['K1'] = 'Suburb/Town'
wsTransdirect['L1'] = 'Postcode'
wsTransdirect['M1'] = 'State'
wsTransdirect['N1'] = 'Country'  # 'AU' always
wsTransdirect['O1'] = 'Weight'  # kg
wsTransdirect['P1'] = 'Length'  # cm
wsTransdirect['Q1'] = 'Width'  # cm
wsTransdirect['R1'] = 'Height'  # cm
wsTransdirect['S1'] = 'Quantity'  # Always 1 i think?
wsTransdirect['T1'] = 'Packaging'  # Always 'Carton'
wsTransdirect['U1'] = 'Alcohol/Glass/Liquids'  # IGNORE

wsMasterPickSheet['A1'] = 'Item Title'
wsMasterPickSheet['B1'] = 'SKU'
wsMasterPickSheet['C1'] = 'Quantity'

# wb.save(r'C:\Python Stuff\OrderAPItesting\Orders.xlsx')

######## Creating basic excel template

if os.path.isfile(r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Transdirect.csv') == True:
    send2trash(r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Transdirect.csv')

if os.path.isfile(r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Sendle.csv') == True:
    send2trash(r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Sendle.csv')

if os.path.isfile(r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\AustraliaPost.csv') == True:
    send2trash(r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\AustraliaPost.csv')

if os.path.isfile(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\Labels.docx") == True:
    send2trash(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\Labels.docx")

if os.path.isfile(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Devilopes.docx") == True:
    send2trash(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Devilopes.docx")

if os.path.isfile(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Minilopes.docx") == True:
    send2trash(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Minilopes.docx")

if os.path.isfile(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Labels.pdf") == True:
    send2trash(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Labels.pdf")

if os.path.isfile(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\Labels_Intermediate.pdf") == True:
    send2trash(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\Labels_Intermediate.pdf")

label_location = r'C:\Users\lorel\Documents\Python\DailyLabels'  ### where to store fastway labels

folder = label_location

try:
    cursor = connection.cursor()
except:
    connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
    cursor = connection.cursor()

cursor.execute(f"TRUNCATE dailyorders; COMMIT;")
cursor.close()

try:
    for filename in os.listdir(folder):
        file_path = os.path.join(folder, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print('Failed to delete %s. Reason: %s' % (file_path, e))
except:
    pass

GetOrderdata = {
    'Filter': {'SalesChannel': ['website', 'kogan', 'mydeal', 'ozsale', 'catch', 'Amazon AU', 'EverydayMarket', "Amazon Seller A", "BigW"], 'OrderStatus': ['Pick'],
               'OutputSelector': ["ID", "ShippingOption", "Email", "GrandTotal", 'OrderLine.ProductName',
                                  'OrderLine.Quantity', ' OrderLine.Weight', 'OrderLine.Cubic',
                                  'OrderLine.ExtraOptions', 'StickyNotes', 'ShipAddress', 'OrderLine.WarehouseName',
                                  'SalesChannel', 'DatePlaced', 'ShippingTotal']}}

GetOrderheaders = {'Content-Type': 'application/json', 'Accept': 'application/json', 'NETOAPI_ACTION': 'GetOrder',
                   'NETOAPI_USERNAME': 'kyalAPI', 'NETOAPI_KEY': 'HSZ3EyI9ki6sseqRIkwy6tL6yFNPEUyM',
                   "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
                   "Accept-Encoding": "*",
                   "Connection": "keep-alive"}

r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=GetOrderheaders, json=GetOrderdata)

#######^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^  Neto Get Order Info ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^#####################

GetItemheaders = {'Content-Type': 'application/json', 'Accept': 'application/json', 'NETOAPI_ACTION': 'GetItem',
                  'NETOAPI_USERNAME': 'kyalAPI', 'NETOAPI_KEY': 'HSZ3EyI9ki6sseqRIkwy6tL6yFNPEUyM',
                  "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
                  "Accept-Encoding": "*",
                  "Connection": "keep-alive"}

#######^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^  Neto Get Item Info ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^###############

UpdateItemheaders = {'Content-Type': 'application/json', 'Accept': 'application/json', 'NETOAPI_ACTION': 'UpdateItem',
                     'NETOAPI_USERNAME': 'kyalAPI', 'NETOAPI_KEY': 'HSZ3EyI9ki6sseqRIkwy6tL6yFNPEUyM',
                     "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
                     "Accept-Encoding": "*",
                     "Connection": "keep-alive"}

#######^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^  Neto Update Item Info ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^###############

AusPostQuoteHeader = {'Content-Type': 'application/json', 'Accept': 'application/json', 'Account-Number': '0007805312'}

username = '966afe1c-c07b-4902-b1d7-77ad1aac0915'
secret = 'x428429fabeb99f420f1'
#######^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^  Aus Post quote info ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^###############


orderresponse = r.text
orderresponse = json.loads(orderresponse)

################################### JANUARY 2021 UPDATE START #####################################
for order in orderresponse['Order']:  # Makes all the OrderIDs lists
    OrderID = order['OrderID']
    OrderIDList = {'OrderID': [OrderID]}
    order.update(OrderIDList)

for order in range(len(orderresponse[
                           'Order'])):  # Itereates throgh each order and 'merges' orders with the same address + customer name. Deletes Duplicate.
    sub_order_minus = 0
    try:
        order_id = orderresponse['Order'][order]['OrderID']

        for sub_order in range(len(orderresponse['Order'])):
            sub_order -= sub_order_minus
            if orderresponse['Order'][sub_order]['OrderID'] == order_id:
                pass

            else:
                if orderresponse['Order'][order]['ShipFirstName'] == orderresponse['Order'][sub_order][
                    'ShipFirstName'] and orderresponse['Order'][order]['ShipLastName'] == \
                        orderresponse['Order'][sub_order]['ShipLastName'] and orderresponse['Order'][order][
                    'ShipStreetLine1'] == orderresponse['Order'][sub_order]['ShipStreetLine1']:
                    ###Found two orders going to the same name and address

                    # Need to append orderlines, order_id, and stickynotes(? Maybe don't merge if there's stickynotes)
                    if orderresponse['Order'][order]['StickyNotes'] == '' and orderresponse['Order'][sub_order][
                        'StickyNotes'] == '' and orderresponse['Order'][order][
                        'ShippingOption'] != 'Express Shipping' and orderresponse['Order'][sub_order][
                        'ShippingOption'] != 'Express Shipping':
                        orderresponse['Order'][order]['OrderLine'].extend(
                            orderresponse['Order'][sub_order]['OrderLine'])
                        orderresponse['Order'][order]['OrderID'].extend(orderresponse['Order'][sub_order]['OrderID'])
                        del orderresponse['Order'][sub_order]
                        sub_order_minus += 1

    except IndexError:
        pass

for orders in orderresponse['Order']:
    pre_string_order_id = ''

    for orderid in orders['OrderID']:
        pre_string_order_id = pre_string_order_id + ' + ' + orderid

    pre_string_order_id = pre_string_order_id[3:]

    string_order_id = {'String_OrderID': pre_string_order_id}

    orders.update(string_order_id)

## want to make a string form of srn "N192387 / N189037 / N237894"

# pprint.pprint(orderresponse)

print('Getting Neto orders...')

item_counter = 0

numNetoOrders = len(orderresponse['Order'])

OrdersJSON = {}
OrdersJSON['Minilope'] = []
OrdersJSON['Devilope'] = []
OrdersJSON['Label'] = []
OrdersJSON['Satchel'] = []

mergeMinilopes = []
mergeDevilopes = []
mergeLabel = []

# print(len(response['Order']))
# print(len(response['Order'][0]['OrderLine']))

for x in orderresponse['Order']:

    item_counter += 1

    Book = 'False'

    if 'Book' in x['ShippingOption']:
        continue

    for t in x[
        'OrderLine']:  ##########Just doing by best to filter out the books, don't want to do the nuclear option to remove everything with 'book' in the title
        if 'Hal Leonard' in t['WarehouseName'] or 'Devirra' in t['WarehouseName'] or 'softcover book' in t[
            'ProductName'].lower() or 'book/cd' in t['ProductName'].lower() or 'alfred' in t[
            'ProductName'].lower() or 'piano safari' in t['ProductName'].lower():
            Book = 'True'

    if Book == 'True':
        continue

    if x['StickyNotes'] != '':
        continue

    if 'Express' in x['ShippingOption']:
        continue

    if 'Click' in x['ShippingOption']:
        continue

    SalesChannel = x['SalesChannel']

    date_placed = x['DatePlaced']

    OrderID = x['OrderID']

    if SalesChannel == 'Website':
        item_counter_string = str(item_counter) + ' (Website)'

    if SalesChannel == 'Kogan':

        item_counter_string = str(item_counter) + ' (Kogan)'

        for yy in range(len(OrderID)):
            OrderID[yy] = OrderID[yy] + ' (Kogan)'

    if SalesChannel.lower() == 'mydeal':

        item_counter_string = str(item_counter) + ' (MyDeal)'

        for yy in range(len(OrderID)):
            OrderID[yy] = OrderID[yy] + ' (MyDeal)'

    if SalesChannel.lower() == 'everydaymarket':

        item_counter_string = str(item_counter) + ' (EverydayMarket)'

        for yy in range(len(OrderID)):
            OrderID[yy] = OrderID[yy] + ' (EverydayMarket)'

    if SalesChannel.lower() == 'ozsale':
        item_counter_string = str(item_counter) + ' (OzSale)'

        for yy in range(len(OrderID)):
            OrderID[yy] = OrderID[yy] + ' (OzSale)'

    if SalesChannel.lower() == 'catch':
        item_counter_string = str(item_counter) + ' (Catch)'

        for yy in range(len(OrderID)):
            OrderID[yy] = OrderID[yy] + ' (Catch)'

    if SalesChannel.lower() == 'amazon au':
        item_counter_string = str(item_counter) + ' (Amazon)'

        for yy in range(len(OrderID)):
            OrderID[yy] = OrderID[yy] + ' (Amazon)'

    if SalesChannel.lower() == 'amazon seller a':
        item_counter_string = str(item_counter) + ' (Amazon)'

        for yy in range(len(OrderID)):
            OrderID[yy] = OrderID[yy] + ' (Amazon)'


    if SalesChannel.lower() == 'bigw':
        item_counter_string = str(item_counter) + ' (Big W)'

        for yy in range(len(OrderID)):
            OrderID[yy] = OrderID[yy] + ' (Big W)'

    OrderLineNum = len(x['OrderLine'])
    City = x['ShipCity']
    FirstName = x['ShipFirstName']
    LastName = x['ShipLastName']
    String_OrderID = x['String_OrderID']

    try:

        Phone = x['ShipPhone'].replace('+', '')

    except KeyError:
        Phone = '0417557472'

    if Phone == '':
        Phone = '0417557472'

    Postcode = x['ShipPostCode']
    State = x['ShipState']

    if 'victoria' in State.strip().lower() or 'v.i.c' in State.strip().lower() or 'vic' in State.strip().lower():
        State = 'VIC'
    if 'new south wales' in State.strip().lower() or 'n.s.w' in State.strip().lower() or 'nsw' in State.strip().lower():
        State = 'NSW'
    if 'tasmania' in State.strip().lower() or 't.a.s' in State.strip().lower() or 'tas' in State.strip().lower():
        State = 'TAS'
    if 'queensland' in State.strip().lower() or 'q.l.d' in State.strip().lower() or 'qld' in State.strip().lower():
        State = 'QLD'
    if 'northern territory' in State.strip().lower() or 'n.t' in State.strip().lower() or 'nt' in State.strip().lower():
        State = 'NT'
    if 'western' in State.strip().lower() or 'w.a' in State.strip().lower() or 'wa' in State.strip().lower():
        State = 'WA'
    if 'south australia' in State.strip().lower() or 's.a' in State.strip().lower() or 'sa' in State.strip().lower():
        State = 'SA'
    if 'australian capital' in State.strip().lower() or 'a.c.t' in State.strip().lower() or 'act' in State.strip().lower():
        State = 'ACT'
    if 'tasmania' in State.strip().lower() or 'tas' in State.strip().lower() or 't.a.s' in State.strip().lower():
        State = 'TAS'

    try:

        Email = x['Email']

    except KeyError:
        Email = 'info@scarlettmusic.com.au'

    if Email == '':
        Email = 'info@scarlettmusic.com.au'

    TotalPrice = x['GrandTotal']
    AddressLine1 = x['ShipStreetLine1']

    if 'ShipStreetLine2' in x:

        AddressLine2 = x['ShipStreetLine2']

    else:
        AddressLine2 = ''

    if 'ShipCompany' in x:

        Company = x['ShipCompany']

    else:
        Company = ''

    ###### TIME TO PUT 1 Liner's in their appropriate place

    if int(OrderLineNum) == 1 and int(x['OrderLine'][0]['Quantity']) < 2:

        SKU = x['OrderLine'][0]['SKU']
        Quantity = x['OrderLine'][0]['Quantity']

        if SKU == 'ECG25':
            print('bing')
            print('bong')

        passcondition = 'failed'

        while passcondition == 'failed':

            newPostageType = ''

            GetItemdata = {'Filter': {'SKU': str(SKU),
                                      'OutputSelector': ["Name", 'PrimarySupplier', 'ShippingHeight', 'ShippingLength',
                                                         'ShippingWidth', 'ShippingWeight', 'Misc06', 'Categories',
                                                         'CategoryID', 'WarehouseQuantity']}}

            try:
                r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=GetItemheaders,
                                  json=GetItemdata)

                response = r.text
                response = json.loads(response)
            # pprint.pprint(response)
            except:
                continue

            try:

                for xx in response['Item'][0]['WarehouseQuantity']:
                    if '15' or '4' in xx['WarehouseID']:
                        passcondition = 'true'

                        continue

            except IndexError:
                passcondition = 'true'
                continue

            if response['Item'][0]['Misc06'].lower() == 'drop ship':
                passcondition = 'true'
                continue

            if response['Item'][0]['Misc06'] == 'e-parcel':
                response['Item'][0]['Misc06'] = ''

            ItemName = response['Item'][0]['Name']

            ############################## Checking if Neto order has a postage type. If not, prompts user to add#####################

            if response['Item'][0]['Misc06'] == '' or response['Item'][0]['Misc06'] == 'n':

                while True:

                    PostageType = input(
                        '''\n\n\n\n\n\nPostage type not found for ''' + ItemName + ' (SKU : ' + SKU + ', VALUE : $' + TotalPrice + ''').

                    Which of the following best suits it?

                    1) Minilope
                    2) Devilope
                    3) Popelope
                    4) Satchel / eParcel
                    5) Dropship
                    6) I don't know! (Makes no permanent change. Will just add to Label)
                    7) Skip this order!

                    Enter Response:''')

                    if PostageType == str(1):

                        newPostageType = 'Minilope'

                        UpdateItemData = {'Item': {'SKU': SKU, 'Misc06': newPostageType}}

                        r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=UpdateItemheaders,
                                          json=UpdateItemData)

                        if float(TotalPrice) > minimum_tracking_amount:
                            newPostageType = 'Satchel'
                            length = 18
                            width = 23
                            height = 4
                            weight = 0.25

                            newPostageType = 'Satchel'

                            UpdateItemData = {'Item': {'SKU': SKU, 'Misc06': newPostageType,
                                                       'ShippingHeight': height / 100, 'ShippingLength': length / 100,
                                                       'ShippingWidth': width / 100, 'ShippingWeight': weight}}

                            r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI',
                                              headers=UpdateItemheaders,
                                              json=UpdateItemData)

                            # OrdersJSON['Orders'].append({'OrderID': OrderID, 'BuyerName' : FirstName + ' ' + LastName,
                            #                              'Email' : Email, 'AddressLine1' : AddressLine1, 'AddressLine2': AddressLine2, 'City': City, 'State' : State,
                            # 'Postcode': Postcode, 'ItemName': ItemName, 'SKU': SKU, 'Quantity': Quantity, 'TotalPrice': TotalPrice, 'PostageType': newPostageType,
                            #                              'height' : height, 'length': length, 'width': width, 'weight': weight})
                            #

                        break


                    elif PostageType == str(2):

                        newPostageType = 'Devilope'

                        UpdateItemData = {'Item': {'SKU': SKU, 'Misc06': newPostageType}}

                        r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=UpdateItemheaders,
                                          json=UpdateItemData)

                        if float(TotalPrice) > minimum_tracking_amount:
                            newPostageType = 'Satchel'
                            length = 18
                            width = 23
                            height = 4
                            weight = 0.25

                            UpdateItemData = {'Item': {'SKU': SKU, 'Misc06': newPostageType,
                                                       'ShippingHeight': height / 100, 'ShippingLength': length / 100,
                                                       'ShippingWidth': width / 100, 'ShippingWeight': weight}}

                            r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI',
                                              headers=UpdateItemheaders,
                                              json=UpdateItemData)

                        break

                    elif PostageType == str(3):

                        newPostageType = 'Label'

                        UpdateItemData = {'Item': {'SKU': SKU, 'Misc06': newPostageType}}

                        r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=UpdateItemheaders,
                                          json=UpdateItemData)

                        if float(TotalPrice) > minimum_tracking_amount:
                            newPostageType = 'Satchel'
                            length = 18
                            width = 23
                            height = 4
                            weight = 0.25

                            UpdateItemData = {'Item': {'SKU': SKU, 'Misc06': newPostageType,
                                                       'ShippingHeight': height / 100, 'ShippingLength': length / 100,
                                                       'ShippingWidth': width / 100, 'ShippingWeight': weight}}

                            r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI',
                                              headers=UpdateItemheaders,
                                              json=UpdateItemData)

                        break

                    elif PostageType == str(4):

                        newPostageType = 'Satchel'

                        response['Item'][0]['Misc06'] = 'satchel'

                        UpdateItemData = {'Item': {'SKU': SKU, 'Misc06': newPostageType}}

                        r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=UpdateItemheaders,
                                          json=UpdateItemData)

                        break


                    elif PostageType == str(5):

                        newPostageType = 'Drop Ship'

                        UpdateItemData = {'Item': {'SKU': SKU, 'Misc06': newPostageType}}

                        r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=UpdateItemheaders,
                                          json=UpdateItemData)

                        break

                    elif PostageType == str(6):
                        newPostageType = 'Label'
                        response['Item'][0]['Misc06'] = 'Label'

                        break

                    elif PostageType == str(7):
                        newPostageType = 'Skip'

                        break

                    else:
                        print('\nValid number not found. Try again. \n')
                        time.sleep(1)
                        continue

            if response['Item'][0]['Misc06'].lower() == 'satchel' and float(
                    response['Item'][0]['ShippingWeight']) <= 0.5:
                sendle_weight_results = item_check(SKU)

                item_thickness_results = item_thickness_check(SKU)

                if sendle_weight_results == 'true':

                    while True:
                        sendleweighttest = input(f'''

Hold up there cowboy - any chance the following package would be less than 250g?

{ItemName} (SKU: {SKU})

1) Ohhh boy yessir!
2) No
3) I don't know!

Enter response here:''')

                        if sendleweighttest == str(1):
                            weight = 0.25
                            UpdateItemData = {'Item': {'SKU': SKU,
                                                       'ShippingWeight': weight}}

                            r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI',
                                              headers=UpdateItemheaders,
                                              json=UpdateItemData)
                            item_insert(SKU)
                            break

                        elif sendleweighttest == str(2):

                            item_insert(SKU)
                            break

                        elif sendleweighttest == str(3):

                            newPostageType = 'Label'
                            response['Item'][0]['Misc06'] = 'Label'

                            break

                        else:
                            print('Valid input not found, try again')
                            time.sleep(1)
                            continue

                if item_thickness_results == 'true' and newPostageType != str(6):
                    while True:
                        sendleweighttest = console.input(f'''

        ___
     __|___|__
      ('o_o')                             
      _\~-~/_    ______.                  STOP RIGHT THERE. 
     //\__/\ \ ~(_]---'                   
    / )O  O( .\/_)                        I AM A [red]TERRORIST[/].
    \ \    / \_/            
    )/_|  |_\\                             WOULD THE FOLLOWING PACKAGE BE [red]LESS THAN 3cm THICK[/]?
   // /(\/)\ \\     
   /_/      \_\\      
  (_||      ||_)      
    \| |__| |/
     | |  | |                             
     | |  | |                            
     |_|  |_|                             
     /_\  /_\\                            


Item: {ItemName}   

SKU: {SKU}           

                                          1) Yes Mr Terrorist.
                                          2) No sir   
                                          3) I don't know!
                                          Enter response here:''')

                        if sendleweighttest == str(1):
                            height = 21
                            width = 21
                            length = 3
                            UpdateItemData = {'Item': {'SKU': SKU,
                                                       'ShippingHeight': height / 100, 'ShippingLength': length / 100,
                                                       'ShippingWidth': width / 100}}

                            r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI',
                                              headers=UpdateItemheaders,
                                              json=UpdateItemData)
                            item_thickness_insert(SKU)
                            break

                        elif sendleweighttest == str(2):

                            item_thickness_insert(SKU)
                            break

                        elif sendleweighttest == str(3):

                            newPostageType = 'Label'
                            response['Item'][0]['Misc06'] = 'Label'

                            break

                        else:
                            print('Valid input not found, try again')
                            time.sleep(1)
                            continue


            elif response['Item'][0]['Misc06'].lower() == 'satchel' and response['Item'][0][
                'ShippingHeight'] == '0.000':

                while True:

                    newPostageType = 'Satchel'
                    thickness_check = 'false'

                    ShippingDimensions = input('''
                    Shipping dimensions not found for ''' + ItemName + ' (SKU : ' + SKU + ''').

                    Which of the following best suits it?

                    1) Bubble Mailer (<250g)
                    2) Bubble Mailer (>250g)
                    3) Other (will need to measure)
                    4) I don't know!

                    Enter Response:''')

                    if ShippingDimensions == str(1) or ShippingDimensions == str(2):

                        if ShippingDimensions == str(1):

                            while True:
                                sendleweighttest = input(f'''

                                               Would this package be less than 3cm thick?

                                               1) Yes
                                               2) No
                                               3) I don't know!

                                               Enter response here:''')

                                if sendleweighttest == str(1):
                                    thickness_test_result = 'true'
                                    break
                                elif sendleweighttest == str(2) or sendleweighttest == str(3):
                                    thickness_test_result = 'false'
                                    break

                                else:
                                    print('Valid input not found, try again')
                                    time.sleep(1)
                                    continue

                        weight = 0.5

                        if ShippingDimensions == str(1):
                            weight = 0.25

                        if sendleweighttest == str(3):
                            newPostageType = 'Label'
                            response['Item'][0]['Misc06'] = 'Label'

                            break

                        if thickness_test_result == 'true':
                            height = 21
                            width = 21
                            length = 3
                        else:
                            length = 18
                            width = 23
                            height = 4

                        UpdateItemData = {
                            'Item': {'SKU': SKU, 'ShippingHeight': height / 100, 'ShippingLength': length / 100,
                                     'ShippingWidth': width / 100, 'ShippingWeight': weight}}

                        r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=UpdateItemheaders,
                                          json=UpdateItemData)

                        break


                    elif ShippingDimensions == str(3):

                        weight = float(input('What is the weight? (kg)'))
                        length = float(input('What is the length? (cm)'))
                        width = float(input('What is the width? (cm)'))
                        height = float(input('What is the height? (cm)'))

                        UpdateItemData = {
                            'Item': {'SKU': SKU, 'ShippingHeight': height / 100, 'ShippingLength': length / 100,
                                     'ShippingWidth': width / 100, 'ShippingWeight': weight}}

                        r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=UpdateItemheaders,
                                          json=UpdateItemData)

                        break

                    elif ShippingDimensions == str(4):
                        newPostageType = 'Label'
                        response['Item'][0]['Misc06'] = 'Label'

                        break

                    else:
                        continue

            GetItemdata = {'Filter': {'SKU': str(SKU),
                                      'OutputSelector': ["Name", 'ShippingHeight', 'ShippingLength', 'ShippingWidth',
                                                         'ShippingWeight', 'Misc06']}}

            r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=GetItemheaders,
                              json=GetItemdata)

            response = r.text
            response = json.loads(response)

            PostageType = response['Item'][0]['Misc06']

            if response['Item'][0]['Misc06'] == 'e-parcel':
                response['Item'][0]['Misc06'] = 'Satchel'

            if newPostageType == 'Drop Ship' or 'skip' in newPostageType.lower():
                break

            if newPostageType == 'Label':
                PostageType = 'Label'
                response['Item'][0]['Misc06'] = 'Label'

            #  print(response['Item'][0]['ShippingHeight'])

            weight = float(response['Item'][0]['ShippingWeight'])
            length = float(response['Item'][0]['ShippingLength']) * 100
            width = float(response['Item'][0]['ShippingWidth']) * 100
            height = float(response['Item'][0]['ShippingHeight']) * 100

            weight = round(weight, 2)
            length = round(length, 2)
            width = round(width, 2)
            height = round(height, 2)

            if 'mini' in PostageType.lower():
                OrdersJSON['Minilope'].append(
                    {'OrderID': OrderID, 'SalesRecordNumber': OrderID, 'BuyerName': FirstName + ' ' + LastName,
                     'Email': Email, 'AddressLine1': AddressLine1,
                     'AddressLine2': AddressLine2, 'City': City, 'State': State,
                     'Postcode': Postcode, 'ItemName': ItemName, 'SKU': SKU,
                     'Quantity': Quantity, 'TotalPrice': TotalPrice,
                     'PostageType': newPostageType,
                     'height': height, 'length': length, 'width': width, 'weight': weight,
                     'Orderlines': OrderLineNum, 'Company': Company, 'Phone': Phone,
                     'String_OrderID': String_OrderID, 'String_SalesRecordNumber': String_OrderID,
                     'SalesChannel': SalesChannel, 'item_count': str(item_counter),
                     'item_count_string': item_counter_string, 'date_placed': date_placed})

                passcondition = 'true'

            if 'devil' in PostageType.lower():
                OrdersJSON['Devilope'].append(
                    {'OrderID': OrderID, 'SalesRecordNumber': OrderID, 'BuyerName': FirstName + ' ' + LastName,
                     'Email': Email, 'AddressLine1': AddressLine1,
                     'AddressLine2': AddressLine2, 'City': City, 'State': State,
                     'Postcode': Postcode, 'ItemName': ItemName, 'SKU': SKU,
                     'Quantity': Quantity, 'TotalPrice': TotalPrice,
                     'PostageType': newPostageType,
                     'height': height, 'length': length, 'width': width, 'weight': weight,
                     'Orderlines': OrderLineNum, 'Company': Company, 'Phone': Phone, 'String_OrderID': String_OrderID,
                     'String_SalesRecordNumber': String_OrderID, 'SalesChannel': SalesChannel,
                     'item_count': str(item_counter), 'item_count_string': item_counter_string,
                     'date_placed': date_placed})

                passcondition = 'true'

            if 'label' in PostageType.lower():
                OrdersJSON['Label'].append(
                    {'OrderID': OrderID, 'SalesRecordNumber': OrderID, 'BuyerName': FirstName + ' ' + LastName,
                     'Email': Email, 'AddressLine1': AddressLine1,
                     'AddressLine2': AddressLine2, 'City': City, 'State': State,
                     'Postcode': Postcode, 'ItemName': ItemName, 'SKU': SKU,
                     'Quantity': Quantity, 'TotalPrice': TotalPrice,
                     'PostageType': newPostageType,
                     'height': height, 'length': length, 'width': width, 'weight': weight,
                     'Orderlines': OrderLineNum, 'Company': Company, 'Phone': Phone,
                     'String_OrderID': String_OrderID, 'String_SalesRecordNumber': String_OrderID,
                     'SalesChannel': SalesChannel, 'item_count': str(item_counter),
                     'item_count_string': item_counter_string, 'date_placed': date_placed})

                passcondition = 'true'
                # pprint.pprint(OrdersJSON)

                break  ######SEPERATE THIS INTO MINI DEVIL LABEL SATCHEL EG OrdersJSON[Satchel}

            if 'drop' in PostageType.lower() or 'skip' in PostageType.lower():
                passcondition = 'true'
                break
            # pprint.pprint(OrdersJSON)

            if 'satchel' in PostageType.lower() and response['Item'][0]['ShippingHeight'] != '0.000':
                OrdersJSON['Satchel'].append(
                    {'OrderID': OrderID, 'SalesRecordNumber': OrderID, 'BuyerName': FirstName + ' ' + LastName,
                     'Email': Email, 'AddressLine1': AddressLine1,
                     'AddressLine2': AddressLine2,
                     'City': City, 'State': State,
                     'Postcode': Postcode, 'ItemName': ItemName, 'SKU': SKU,
                     'Quantity': Quantity, 'TotalPrice': TotalPrice,
                     'PostageType': newPostageType,
                     'height': height, 'length': length, 'width': width, 'weight': weight,
                     'Orderlines': OrderLineNum, 'Company': Company, 'Phone': Phone,
                     'String_OrderID': String_OrderID, 'String_SalesRecordNumber': String_OrderID,
                     'SalesChannel': SalesChannel, 'item_count': str(item_counter),
                     'item_count_string': item_counter_string, 'date_placed': date_placed})

                passcondition = 'true'

                # pprint.pprint(OrdersJSON)
                break

            continue
    # print(x)

    if int(OrderLineNum) > 1 or int(x['OrderLine'][0]['Quantity']) > 1:

        passcondition = 'false'

        while passcondition == 'false':

            #  pprint.pprint(x)

            namemulti = {}
            SKUmulti = {}
            quantitymulti = {}

            for y in range(OrderLineNum):
                namemulti['Name (Item {0})'.format(y + 1)] = x['OrderLine'][y]['ProductName']
                SKUmulti['SKU (Item {0})'.format(y + 1)] = x['OrderLine'][y]['SKU']
                quantitymulti['Quantity (Item {0})'.format(y + 1)] = x['OrderLine'][y]['Quantity']

            PostageType = 'label'

            if 'label' in PostageType.lower():
                OrdersJSON['Label'].append(
                    {'OrderID': OrderID, 'SalesRecordNumber': OrderID, 'BuyerName': FirstName + ' ' + LastName,
                     'Email': Email, 'AddressLine1': AddressLine1,
                     'AddressLine2': AddressLine2, 'City': City, 'State': State,
                     'Postcode': Postcode, 'ItemName': namemulti, 'SKU': SKUmulti,
                     'Quantity': quantitymulti, 'TotalPrice': TotalPrice,
                     'PostageType': PostageType,
                     'height': 'NA', 'length': 'NA', 'width': 'NA', 'weight': 'NA',
                     'Orderlines': OrderLineNum, 'Company': Company, 'Phone': Phone,
                     'String_OrderID': String_OrderID, 'String_SalesRecordNumber': String_OrderID,
                     'SalesChannel': SalesChannel, 'item_count': str(item_counter),
                     'item_count_string': item_counter_string,
                     'date_placed': date_placed})

                passcondition = 'true'
                # pprint.pprint(OrdersJSON)

                break  ######SEPERATE THIS INTO MINI DEVIL LABEL SATCHEL EG OrdersJSON[Satchel}

            if PostageType == 'Drop Ship':
                passcondition = 'true'
                #  pprint.pprint(OrdersJSON)
                break

                #  pprint.pprint(OrdersJSON)

###########This is all the neto orders (In theory) added to the JSON file (OrdersJSON)

api = Trading(appid="KyalScar-Awaiting-PRD-aca833663-f54c920e", timeout=1000, config_file=None,
              devid="5c97c2a8-eae7-49cc-ae81-c3ed1bbab335", certid="PRD-ca8336639ad2-75af-45bc-992a-d0bc",
              token="v^1.1#i^1#r^1#f^0#I^3#p^3#t^Ul4xMF84OkUzOEU3NzVFQzA3NDUwNjUyMkY0NTc4QjlENjRFRDVFXzFfMSNFXjI2MA==")

GetOrderdata = {'Filter': {'SalesChannel': 'ebay', 'OrderStatus': ['Pick', 'On Hold'],
                           'OutputSelector': ["ID", "ShippingOption", "Email", "GrandTotal", 'OrderLine.ProductName',
                                              'OrderLine.Quantity', ' OrderLine.Weight', 'OrderLine.Cubic',
                                              'OrderLine.ExtraOptions', 'StickyNotes', 'ShipAddress',
                                              'PurchaseOrderNumber', 'OrderLine.WarehouseName', 'DatePlaced',
                                              'ShippingTotal']}}

GetOrderheaders = {'Content-Type': 'application/json', 'Accept': 'application/json', 'NETOAPI_ACTION': 'GetOrder',
                   'NETOAPI_USERNAME': 'kyalAPI', 'NETOAPI_KEY': 'HSZ3EyI9ki6sseqRIkwy6tL6yFNPEUyM', }

r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=GetOrderheaders, json=GetOrderdata)

orderresponse = r.text
orderresponse = json.loads(orderresponse)
pprint.pprint(orderresponse)

print(f"\n=== DIAGNOSTIC: Neto returned {len(orderresponse.get('Order', []))} eBay orders (before filtering) ===\n")

for orders in orderresponse['Order']:
    string_order_id = {'String_OrderID': orders['OrderID']}

    orders.update(string_order_id)

# Duplicate orders are merged, OrderId's are merged into list form, string of all SRNs are stored in form "N192387 / N189037 / N237894"
################################### JANUARY 2021 UPDATE END #####################################

eBayJSON = []
_diag_book_shipping = 0
_diag_book_product = 0
_diag_sticky_notes = 0
_diag_express = 0
_diag_shipping_cost = 0
_diag_passed = 0

for x in orderresponse['Order']:

    Book = 'False'

    if 'Book' in x['ShippingOption']:
        _diag_book_shipping += 1
        continue

    for t in x[
        'OrderLine']:  ##########Just doing by best to filter out the books, don't want to do the nuclear option to remove everything with 'book' in the title
        if 'Hal Leonard' in t['WarehouseName'] or 'Devirra' in t['WarehouseName'] or 'softcover book' in t[
            'ProductName'].lower() or 'book/cd' in t['ProductName'].lower() or 'alfred' in t[
            'ProductName'].lower() or 'piano safari' in t['ProductName'].lower() or 'essential elements' in t[
            'ProductName'].lower() or 'ameb' in t['ProductName'].lower():
            Book = 'True'

    if Book == 'True':
        _diag_book_product += 1
        continue

    if x['StickyNotes'] != '':  ####WE DO NOT HAVE STICKY NOTES YET
        _diag_sticky_notes += 1
        continue

    if 'Express' in x['ShippingOption']:  #### CANNOT DETERMINE EXPRESS VIA NETO, WILL NEED TO CHECK WITH EBAY
        _diag_express += 1
        continue

    item_counter += 1
    item_counter_string = f'{str(item_counter)} eBay'

    OrderID = x['OrderID']
    OrderLineNum = len(x['OrderLine'])
    City = x['ShipCity']
    FirstName = x['ShipFirstName']
    LastName = x['ShipLastName']
    EbayOrderID = x['PurchaseOrderNumber']
    String_OrderID = x['String_OrderID']
    shipping_total = x['ShippingTotal']
    if float(shipping_total) > 30:
        _diag_shipping_cost += 1
        continue

    date_placed = x['DatePlaced']

    try:

        Phone = x['ShipPhone'].replace('+61', '')
        Phone = Phone.replace(' ', '')
        Phone = Phone.replace('+', '')


    except KeyError:
        Phone = '0417557472'

    if Phone == '':
        Phone = '0417557472'

    try:
        float(Phone)

    except ValueError:
        Phone = '0417557472'

    Postcode = x['ShipPostCode']
    State = x['ShipState']

    if 'victoria' in State.strip().lower() or 'v.i.c' in State.strip().lower() or 'vic' in State.strip().lower():
        State = 'VIC'
    if 'new south wales' in State.strip().lower() or 'n.s.w' in State.strip().lower() or 'nsw' in State.strip().lower():
        State = 'NSW'
    if 'tasmania' in State.strip().lower() or 't.a.s' in State.strip().lower() or 'tas' in State.strip().lower():
        State = 'TAS'
    if 'queensland' in State.strip().lower() or 'q.l.d' in State.strip().lower() or 'qld' in State.strip().lower():
        State = 'QLD'
    if 'northern territory' in State.strip().lower() or 'n.t' in State.strip().lower() or 'nt' in State.strip().lower():
        State = 'NT'
    if 'western' in State.strip().lower() or 'w.a' in State.strip().lower() or 'wa' in State.strip().lower():
        State = 'WA'
    if 'south australia' in State.strip().lower() or 's.a' in State.strip().lower() or 'sa' in State.strip().lower():
        State = 'SA'
    if 'australian capital' in State.strip().lower() or 'a.c.t' in State.strip().lower() or 'act' in State.strip().lower():
        State = 'ACT'
    if 'tasmania' in State.strip().lower() or 'tas' in State.strip().lower() or 't.a.s' in State.strip().lower():
        State = 'TAS'

    try:

        Email = x['Email']

    except KeyError:
        Email = 'info@scarlettmusic.com.au'

    if Email == '':
        Email = 'info@scarlettmusic.com.au'

    TotalPrice = x['GrandTotal']
    AddressLine1 = x['ShipStreetLine1']

    if 'ShipStreetLine2' in x:

        AddressLine2 = x['ShipStreetLine2']

    else:
        AddressLine2 = ''

    if 'ShipCompany' in x:

        Company = x['ShipCompany']

    else:
        Company = ''

    # print('Orderline Number: ' + str(OrderLineNum))
    # print(EbayOrderID)

    ##NO QUANTITY, SKU, ITEMNAME, POSTAGE TYPE, ITEM DIMENSIONS

    SKUmulti = {}
    quantitymulti = {}

    for y in range(OrderLineNum):  ####Sorting all the variables that could *potentially* be multiline

        SKUmulti['SKU (Item {0})'.format(y + 1)] = x['OrderLine'][y]['SKU']
        quantitymulti['Quantity (Item {0})'.format(y + 1)] = x['OrderLine'][y]['Quantity']

    eBayJSON.append({'OrderID': OrderID, 'BuyerName': FirstName + ' ' + LastName,
                     'Email': Email, 'AddressLine1': AddressLine1,
                     'AddressLine2': AddressLine2, 'City': City, 'State': State,
                     'Postcode': Postcode, 'TotalPrice': TotalPrice,
                     'Orderlines': OrderLineNum, 'SKU': SKUmulti, 'Quantity': quantitymulti,
                     'PurchaseOrderNumber': EbayOrderID, 'Company': Company, 'Phone': Phone,
                     'String_OrderID': String_OrderID, 'SalesChannel': 'eBay', 'item_count': str(item_counter),
                     'item_count_string': item_counter_string,
                     'date_placed': date_placed})

_diag_passed = len(eBayJSON)
print(f"\n=== DIAGNOSTIC: Filter summary ===")
print(f"  Filtered by 'Book' in ShippingOption: {_diag_book_shipping}")
print(f"  Filtered by book warehouse/product:   {_diag_book_product}")
print(f"  Filtered by StickyNotes not empty:     {_diag_sticky_notes}")
print(f"  Filtered by 'Express' in shipping:     {_diag_express}")
print(f"  Filtered by shipping_total > $2:       {_diag_shipping_cost}")
print(f"  PASSED all filters:                    {_diag_passed}")
print(f"  TOTAL:                                 {_diag_book_shipping + _diag_book_product + _diag_sticky_notes + _diag_express + _diag_shipping_cost + _diag_passed}")
print(f"=================================\n")

numdic = len(eBayJSON)
print(str(numdic) + ' potential eBay orders found.')
print('Getting eBay orders. This may take a minute.')
# pprint.pprint(eBayJSON)

####################################### BELOW IS THE NEW TEST CODE ##############################
potential_purchase_orders = []

for x in eBayJSON:
    potential_purchase_orders.append(x['PurchaseOrderNumber'])

# pprint.pprint(potential_purchase_orders)

try:

    potential_purchase_orders.remove('13-06996-93029')
    potential_purchase_orders.remove('02-07290-90095')
    potential_purchase_orders.remove('26-07312-44505')
except ValueError:
    pass

page_numbers = math.ceil(len(potential_purchase_orders) / 100)

# orderresponse = {'Ack': 'Success',
#  'Build': 'E1215_CORE_APIXO_19220561_R1',
#  'HasMoreOrders': 'false',
#  'OrderArray': {'Order': []},
#  'OrdersPerPage': '100',
#  'PageNumber': '1',
#  'PaginationResult': {'TotalNumberOfEntries': '1', 'TotalNumberOfPages': '1'},
#  'ReturnedOrderCountActual': '1',
#  'Timestamp': '2022-01-12T08:08:45.533Z',
#  'Version': '1215'}
#
# for x in potential_purchase_orders:
#
#
#  getordersapi = api.execute('GetOrders', {'OrderIDArray': {'OrderID': [x]}, 'Pagination' : {'EntriesPerPage': 100 ,'PageNumber': 1}})
#  getordersresponse = getordersapi.dict()
#
#  try:
#
#     orderresponse['OrderArray']['Order'].extend(getordersresponse['OrderArray']['Order'])
#
#  except TypeError:
#      continue
#
# getordersresponse = orderresponse

for x in range(page_numbers):

    warning = 'warning'
    page_number = x + 1

    if potential_purchase_orders == []:
        pass

    else:

        getordersapi = api.execute('GetOrders', {'OrderIDArray': {'OrderID': potential_purchase_orders},
                                                 'Pagination': {'PageNumber': 1}})
        getordersresponse = getordersapi.dict()
        # pprint.pprint(getordersresponse)

    try:

        while getordersresponse['Ack'] == 'Warning':

            invalid_orders = getordersresponse['Errors']['LongMessage']
            invalid_orders = invalid_orders.replace('OrderIds ', '')
            invalid_orders = invalid_orders.replace(' are invalid.', '')
            invalid_orders = invalid_orders.split(',')
            for invalid_orders_to_be_popped in invalid_orders:
                potential_purchase_orders.remove(invalid_orders_to_be_popped)

            getordersapi = api.execute('GetOrders', {'OrderIDArray': {'OrderID': potential_purchase_orders},
                                                     'Pagination': {'PageNumber': 1}})
            getordersresponse = getordersapi.dict()
    except Exception as e:
        print(f"DIAGNOSTIC: GetOrders Ack/Warning handling error: {e}")

    try:

        for t in getordersresponse['OrderArray']['Order']:
            ItemTitle = {}
            OrderLineID = {}
            ItemSKU = {}
            ItemID = {}
            Quantity = {}
            purchase_order_from_ebay = t['OrderID']

            # print(t)
            #  print("\n")

            if 'express' in t['ShippingServiceSelected']['ShippingService'].lower():
                ordercancelled = 'y'

            elif 'courier' in t['ShippingServiceSelected']['ShippingService'].lower():
                ordercancelled = 'y'

            elif 'pickup' in t['ShippingServiceSelected']['ShippingService'].lower():
                ordercancelled = 'y'


            # Instead need to try elif 'courier' in t['ShippingServiceSelected']['ShippingService'].lower() AND int(t['ShippingServiceSelected']['ShippingServiceCost']['value']>0)

            elif t['OrderStatus'] == 'CustomCode' or t['OrderStatus'] == 'CancelPending' or t[
                'OrderStatus'] == 'Cancelled':

                ordercancelled = 'y'

            else:
                ordercancelled = 'n'

            totalprice = t['AmountPaid']['value']

            Orderlines = len(t['TransactionArray']['Transaction'])

            for y in range(Orderlines):

                ItemTitle['Name (Item {0})'.format(y + 1)] = t['TransactionArray']['Transaction'][y]['Item']['Title']
                # print(ItemTitle)
                OrderLineID['Orderline (Item {0})'.format(y + 1)] = t['TransactionArray']['Transaction'][y][
                    'OrderLineItemID']
                ItemID['ItemID (Item {0})'.format(y + 1)] = t['TransactionArray']['Transaction'][y]['Item']['ItemID']
                Quantity['Quantity (Item {0})'.format(y + 1)] = t['TransactionArray']['Transaction'][y][
                    'QuantityPurchased']

                try:
                    ItemSKU['SKU (Item {0})'.format(y + 1)] = t['TransactionArray']['Transaction'][y]['Item']['SKU']

                except KeyError:

                    while True:
                        revision_answer = input(
                            '''\n\nSKU not found for ''' + t['TransactionArray']['Transaction'][y]['Item']['Title'] + '''. Do you know what it is? (Enter Number 1-2)
                        1) Yes
                        2) No
                        Enter Response:''')

                        if str(revision_answer) == '1' or revision_answer.lower() == 'y':
                            revision_sku = input('''What is the SKU?
                            Enter Response:''')

                            reviseAPI = api.execute('ReviseItem',
                                                    {"Item": {"ItemID": int(
                                                        t['TransactionArray']['Transaction'][y]['Item']['ItemID']),
                                                        "SKU": str(revision_sku)}})

                            print('Success!')

                            time.sleep(1)

                            ItemSKU['SKU (Item {0})'.format(y + 1)] = revision_sku
                            break
                        elif str(revision_answer) == '2' or revision_answer.lower() == 'n':
                            ItemSKU['SKU (Item {0})'.format(y + 1)] = ''
                            break

                        else:

                            print('Valid answer not found, please try again.')
                            time.sleep(1)
                            continue

            shippingservice = t['ShippingServiceSelected']['ShippingService']
            try:
                SRN = t['ShippingDetails']['SellingManagerSalesRecordNumber']
            except (KeyError, TypeError) as e:
                print(f"DIAGNOSTIC: SRN lookup failed for order {purchase_order_from_ebay}: {e} — using fallback")
                SRN = purchase_order_from_ebay  # Fallback if SellingManagerSalesRecordNumber is missing

            for h in range(len(eBayJSON)):
                if eBayJSON[h]['PurchaseOrderNumber'] == str(purchase_order_from_ebay):
                    eBayJSON[h].update({'ItemTitle': ItemTitle})
                    eBayJSON[h].update({'ItemSKU': ItemSKU})
                    eBayJSON[h].update({'OrderLineID': OrderLineID})
                    eBayJSON[h].update({'ItemID': ItemID})
                    eBayJSON[h].update({'Quantity': Quantity})
                    eBayJSON[h].update({'TotalPrice': totalprice})
                    eBayJSON[h].update({'SalesRecordNumber': SRN})
                    eBayJSON[h].update({'Shipping_Service_eBay': shippingservice})
                    eBayJSON[h].update({'Ordercancelled': ordercancelled})

        for p in getordersresponse['OrderArray']['Order']:
            Order_ID_To_be_popped = p['OrderID']

            try:

                potential_purchase_orders.remove(Order_ID_To_be_popped)

            except ValueError:
                pass
        print('Page ' + str(x + 1) + ' of ' + str(page_numbers) + ' downloaded successfully.')

    except Exception as e:
        print(f"DIAGNOSTIC: GetOrders page {x + 1} processing error: {e}")
        continue
################################ END OF TEST CODE, HOPE ITS FASTER AND MORE RELIABLE!!! ############################

_getmyebayselling_retries = 0
while True:

    try:

        response = api.execute('GetMyeBaySelling', {
            "SoldList": {"Include": True, "IncludeNotes": 'true', "DurationInDays": 10, "OrderStatusFilter": "AwaitingShipment",
                         "Pagination": {"EntriesPerPage": 200}}})
        unformatteddic = response.dict()
        paginationresult = int(unformatteddic['SoldList']['PaginationResult']['TotalNumberOfPages'])
        break

    except Exception as e:
        _getmyebayselling_retries += 1
        print(f"DIAGNOSTIC: GetMyeBaySelling initial call failed (attempt {_getmyebayselling_retries}): {e}")
        if _getmyebayselling_retries >= 5:
            print("DIAGNOSTIC: GetMyeBaySelling initial call failed after 5 attempts — aborting eBay SoldList fetch")
            paginationresult = 0
            break
        continue
# print(paginationresult)
#  pprint.pprint(unformatteddic)

SoldList = {}

# def getSoldList(pagination):
#
#
#     response = api.execute('GetMyeBaySelling', {
#         "SoldList": {"Include": True, "IncludeNotes": True, "DurationInDays": 10, "OrderStatusFilter": "AwaitingShipment" , "Pagination" : {"EntriesPerPage" : 200 , "PageNumber" : int(pagination)}}})
#     unformatteddic = response.dict()
#     #print(unformatteddic)
#     SoldList.update(unformatteddic)
#
# with concurrent.futures.ThreadPoolExecutor() as executor:
#     results = executor.map(getSoldList, range(1, paginationresult+1))

soldlistnumber = 0
for yy in range(int(paginationresult)):

    pass_condition_list = ''
    _soldlist_page_retries = 0

    while True:

        if pass_condition_list == 'true':
            break

        try:
            response = api.execute('GetMyeBaySelling', {
                "SoldList": {"Include": True, "IncludeNotes": 'true', "DurationInDays": 10,
                             "OrderStatusFilter": "AwaitingShipment",
                             "Pagination": {"EntriesPerPage": 200, "PageNumber": int(yy) + 1}}})
            unformatteddic = response.dict()
            # pprint.pprint(unformatteddic['SoldList']['OrderTransactionArray'])

            for t in unformatteddic['SoldList']['OrderTransactionArray']['OrderTransaction']:
                SoldList.update({soldlistnumber: t})
                soldlistnumber = soldlistnumber + 1

            pass_condition_list = 'true'

            break
        except Exception as e:
            _soldlist_page_retries += 1
            print(f"DIAGNOSTIC: GetMyeBaySelling page {int(yy) + 1} failed (attempt {_soldlist_page_retries}): {e}")
            if _soldlist_page_retries >= 5:
                print(f"DIAGNOSTIC: GetMyeBaySelling page {int(yy) + 1} failed after 5 attempts — skipping page")
                break
            continue

# pprint.pprint(SoldList)

# print(SoldList['SoldList']['OrderTransactionArray']['OrderTransaction'][0]['Transaction']['OrderLineItemID'])

for x in eBayJSON:

    if 'Ordercancelled' in x:

        if x['Ordercancelled'] == 'y':
            indexvalue = eBayJSON.index(x)
            eBayJSON.pop(indexvalue)

StickyNotes = {}
preFinalList = []
# pprint.pprint(eBayJSON)
# pprint.pprint(FinalList)
ebayJsonPosition = -1
for keys in eBayJSON:
    ebayJsonPosition = ebayJsonPosition + 1
    Itemlist = []
    try:
        for x in keys['OrderLineID']:
            dicorderline = keys['OrderLineID'][x]

            # print(dicorderline)

            for transactions in SoldList:

                passcondition = 'failed'

                Stickydic = {}

                # print(transactions)
                if 'Transaction' in SoldList[transactions]:
                    orderlinelen = 1

                    ebayorderline = SoldList[transactions]['Transaction'][
                        'OrderLineItemID']  ###  Checking Single Order Lines
                    # print(ebayorderline)

                    if dicorderline == ebayorderline:

                        passcondition = 'passed'

                        length = 0
                        width = 0
                        height = 0
                        weight = 0

                        ItemID = SoldList[transactions]['Transaction']['Item']['ItemID']
                        ItemPrice = SoldList[transactions]['Transaction']['Item']['SellingStatus']['CurrentPrice'][
                            'value']
                        # print('Item ID: ' + ItemID)
                        # print('dictionary orderline = ' + dicorderline)
                        # print('Ebay orderline = ' + ebayorderline)

                        # print('bing')

                        ebayItemName = SoldList[transactions]['Transaction']['Item']['Title']

                        try:
                            ebaySKU = SoldList[transactions]['Transaction']['Item']['SKU']

                            if ebaySKU == 'DD3100NBK':
                                print('bing')
                                print('bong')




                        except KeyError:
                            ebaySKU = ''

                        try:
                            ItemNote = SoldList[transactions]['Transaction']['Item']['PrivateNotes']



                        except KeyError:

                            while True:

                                passcondition = 'fail'

                                ItemNote = ''

                                PostageType = input(
                                    '''\n\n\n\n\n\nPostage type not found for ''' + ebayItemName + ' (SKU : ' + ebaySKU + ''').

                                Which of the following best suits it?

                                1) Minilope
                                2) Devilope
                                3) Popelope
                                4) Satchel / eParcel
                                5) Dropship
                                6) I don't know! (Makes no permanent change. Will just add to Label)
                                7) Skip this order!

                                Enter Response:''')

                                if PostageType == str(1):
                                    ItemNote = 'Mini.'

                                    if float(ItemPrice) > minimum_tracking_amount:
                                        ItemNote = 'Satchel'
                                        height = 21
                                        width = 21
                                        length = 3
                                        weight = 0.25

                                        t = math.modf(weight)  # (0.5678000000000338, 1234.0)

                                        minorvalue = round((t[0]) * 1000)
                                        majorvalue = round(t[1])

                                        try:

                                            dimensionupdate = api.execute('ReviseItem', {'Item': {"ItemID": ItemID,
                                                                                                  'ShippingPackageDetails': {
                                                                                                      'MeasurementUnit': 'Metric',
                                                                                                      'PackageDepth': height,
                                                                                                      'PackageLength': length,
                                                                                                      'PackageWidth': width,
                                                                                                      'WeightMajor': int(
                                                                                                          majorvalue),
                                                                                                      'WeightMinor': int(
                                                                                                          minorvalue)}}})

                                        except ConnectionError:

                                            pass

                                    noteresponse = api.execute('SetUserNotes',
                                                               {"Action": "AddOrUpdate", "ItemID": int(ItemID),
                                                                "NoteText": ItemNote})
                                    unformatteddic = response.dict()

                                    print('Success')
                                    time.sleep(1)

                                    break

                                elif PostageType == str(2):
                                    ItemNote = 'Devil.'

                                    if float(ItemPrice) > minimum_tracking_amount:
                                        ItemNote = 'Satchel'
                                        height = 21
                                        width = 21
                                        length = 3
                                        weight = 0.25

                                        t = math.modf(weight)  # (0.5678000000000338, 1234.0)

                                        minorvalue = round((t[0]) * 1000)
                                        majorvalue = round(t[1])

                                        try:

                                            dimensionupdate = api.execute('ReviseItem', {'Item': {"ItemID": ItemID,
                                                                                                  'ShippingPackageDetails': {
                                                                                                      'MeasurementUnit': 'Metric',
                                                                                                      'PackageDepth': height,
                                                                                                      'PackageLength': length,
                                                                                                      'PackageWidth': width,
                                                                                                      'WeightMajor': int(
                                                                                                          majorvalue),
                                                                                                      'WeightMinor': int(
                                                                                                          minorvalue)}}})

                                        except ConnectionError:

                                            pass

                                    noteresponse = api.execute('SetUserNotes',
                                                               {"Action": "AddOrUpdate", "ItemID": int(ItemID),
                                                                "NoteText": ItemNote})
                                    unformatteddic = response.dict()

                                    print('Success')
                                    time.sleep(1)

                                    break

                                elif PostageType == str(3):
                                    ItemNote = 'Label.'

                                    noteresponse = api.execute('SetUserNotes',
                                                               {"Action": "AddOrUpdate", "ItemID": int(ItemID),
                                                                "NoteText": ItemNote})
                                    unformatteddic = response.dict()

                                    print('Success')
                                    time.sleep(1)

                                    break

                                elif PostageType == str(4):
                                    ItemNote = 'Satchel.'

                                    noteresponse = api.execute('SetUserNotes',
                                                               {"Action": "AddOrUpdate", "ItemID": int(ItemID),
                                                                "NoteText": ItemNote})
                                    unformatteddic = response.dict()

                                    # GetItemAPI = response = api.execute('GetItem', {"ItemID": ItemID})
                                    GetItemAPI = api.execute('GetItem', {"ItemID": ItemID})
                                    GetItemAPIdic = GetItemAPI.dict()

                                    if 'PackageDepth' not in GetItemAPIdic['Item']['ShippingPackageDetails']:

                                        while True:
                                            ShippingDimensions = input(
                                                '''\n\n\n\n\n\n\nShipping dimensions not found for ''' + ebayItemName + ' (SKU : ' + ebaySKU + ''').

                                                                Which of the following best suits it?

                                                                1) Bubble Mailer (<250g weight)
                                                                2) Bubble Mailer (>250g weight)
                                                                3) Other (will need to measure)
                                                                4) I don't know! (will put into label for now)

                                                                Enter Response:''')

                                            if ShippingDimensions == str(1) or ShippingDimensions == str(2):

                                                weight = 0.5
                                                thickness_test_result = 'false'

                                                if ShippingDimensions == str(1):
                                                    weight = 0.25

                                                    while True:

                                                        thickness_test = input('''

                                                    Would this package be less than 3cm thick?

                                                    1) Yes
                                                    2) No
                                                    3) I don't know

                                                    Enter response here:''')

                                                        if thickness_test == str(1):
                                                            thickness_test_result = 'true'
                                                            break

                                                        if thickness_test == str(2) or thickness_test == str(3):
                                                            thickness_test_result = 'false'
                                                            break

                                                        else:
                                                            print('Valid input not found. Try again.')
                                                            time.sleep(1)
                                                            continue

                                                try:
                                                    thickness_test

                                                except:
                                                    thickness_test = '3'

                                                if thickness_test == str(3):
                                                    ItemNote = 'Label.'

                                                    noteresponse = api.execute('SetUserNotes',
                                                                               {"Action": "AddOrUpdate",
                                                                                "ItemID": int(ItemID),
                                                                                "NoteText": ItemNote})

                                                    unformatteddic = response.dict()

                                                    passcondition = 'true'

                                                    break

                                                if thickness_test_result == 'true':

                                                    height = 21
                                                    width = 21
                                                    length = 3

                                                else:

                                                    length = 18
                                                    width = 23
                                                    height = 4

                                                t = math.modf(weight)  # (0.5678000000000338, 1234.0)

                                                minorvalue = round((t[0]) * 1000)
                                                majorvalue = round(t[1])

                                                try:

                                                    dimensionupdate = api.execute('ReviseItem', {
                                                        'Item': {"ItemID": ItemID,
                                                                 'ShippingPackageDetails': {'MeasurementUnit': 'Metric',
                                                                                            'PackageDepth': height,
                                                                                            'PackageLength': length,
                                                                                            'PackageWidth': width,
                                                                                            'WeightMajor': int(
                                                                                                majorvalue),
                                                                                            'WeightMinor': int(
                                                                                                minorvalue)}}})
                                                    passcondition = 'true'
                                                    break

                                                except ConnectionError:
                                                    passcondition = 'true'
                                                    break




                                            elif ShippingDimensions == str(3):

                                                weight = float(input('What is the weight? (kg)'))
                                                length = float(input('What is the length? (cm)'))
                                                width = float(input('What is the width? (cm)'))
                                                height = float(input('What is the height? (cm)'))

                                                t = math.modf(weight)  # (0.5678000000000338, 1234.0)

                                                minorvalue = round((t[0]) * 1000)
                                                majorvalue = round(t[1])

                                                try:

                                                    dimensionupdate = api.execute('ReviseItem',
                                                                                  {'Item': {"ItemID": ItemID,
                                                                                            'ShippingPackageDetails': {
                                                                                                'MeasurementUnit': 'Metric',
                                                                                                'PackageDepth': height,
                                                                                                'PackageLength': length,
                                                                                                'PackageWidth': width,
                                                                                                'WeightMajor': majorvalue,
                                                                                                'WeightMinor': minorvalue}}})
                                                    passcondition = 'true'

                                                    break
                                                except:
                                                    passcondition = 'true'

                                                    break



                                            elif ShippingDimensions == str(4):

                                                ItemNote = 'Label.'

                                                noteresponse = api.execute('SetUserNotes',
                                                                           {"Action": "AddOrUpdate",
                                                                            "ItemID": int(ItemID),
                                                                            "NoteText": ItemNote})

                                                unformatteddic = response.dict()

                                                passcondition = 'true'

                                                break

                                            else:
                                                print('Valid input not found. Try again.')
                                                time.sleep(1)
                                                continue

                                elif PostageType == str(5):
                                    ItemNote = 'Dropship.'

                                    noteresponse = api.execute('SetUserNotes',
                                                               {"Action": "AddOrUpdate", "ItemID": int(ItemID),
                                                                "NoteText": ItemNote})
                                    unformatteddic = response.dict()

                                    print('Success')
                                    time.sleep(1)

                                    break

                                elif PostageType == str(6):
                                    ItemNote = 'Label.'

                                    print('Success')
                                    time.sleep(1)
                                    break

                                elif PostageType == str(7):
                                    ItemNote = 'Skip.'

                                    print('Success')
                                    time.sleep(1)
                                    break

                                if passcondition == 'true':
                                    break

                                else:
                                    print('Valid input not found. Try again.')
                                    time.sleep(1)
                                    continue

                                    ####Need to check if there's any dimensions and add them to JSON file. If not, add dimensions.

                                    ####Do same with multiliners

                                print('Success')
                                time.sleep(1)
                        ###############################################################################################
                        if 'satchel' in ItemNote.lower():

                            GetItemAPI = api.execute('GetItem', {"ItemID": ItemID})
                            GetItemAPIdic = GetItemAPI.dict()
                            #  pprint.pprint(GetItemAPIdic)

                            if 'PackageDepth' not in GetItemAPIdic['Item']['ShippingPackageDetails']:

                                while True:

                                    ShippingDimensions = input(
                                        '''
                                        Shipping dimensions not found for ''' + ebayItemName + ' (SKU : ' + ebaySKU + ''').

                                                                                    Which of the following best suits it?

                                                                                    1) Bubble Mailer (<250g weight)
                                                                                    2) Bubble Mailer (>250g weight)
                                                                                    3) Other (will need to measure)
                                                                                    4) I don't know! (will put into label for now)

                                                                                    Enter Response:''')

                                    if ShippingDimensions == str(1) or ShippingDimensions == str(2):

                                        weight = 0.5

                                        thickness_test_result = 'false'

                                        if ShippingDimensions == str(1):
                                            weight = 0.25

                                            while True:

                                                thickness_test = input('''

                                            Would this package be less than 3cm thick?

                                            1) Yes
                                            2) No
                                            3) I don't know

                                            Enter response here:''')

                                                if thickness_test == str(1):
                                                    thickness_test_result = 'true'
                                                    break

                                                if thickness_test == str(2) or thickness_test == str(3):
                                                    thickness_test_result = 'false'
                                                    weight = 0
                                                    length = 0
                                                    width = 0
                                                    height = 0

                                                    ItemNote = 'Label.'
                                                    break

                                                else:
                                                    print('Valid input not found. Try again.')
                                                    time.sleep(1)
                                                    continue

                                        # if thickness_test == str(3):
                                        #     weight = 0
                                        #     length = 0
                                        #     width = 0
                                        #     height = 0
                                        #
                                        #     ItemNote = 'Label.'
                                        #     break

                                        if thickness_test_result == 'true':

                                            height = 21
                                            width = 21
                                            length = 3

                                        else:

                                            length = 18
                                            width = 23
                                            height = 4

                                        t = math.modf(weight)  # (0.5678000000000338, 1234.0)

                                        minorvalue = round((t[0]) * 1000)
                                        majorvalue = round(t[1])

                                        try:

                                            dimensionupdate = api.execute('ReviseItem', {'Item': {"ItemID": ItemID,
                                                                                                  'ShippingPackageDetails': {
                                                                                                      'MeasurementUnit': 'Metric',
                                                                                                      'PackageDepth': height,
                                                                                                      'PackageLength': length,
                                                                                                      'PackageWidth': width,
                                                                                                      'WeightMajor': int(
                                                                                                          majorvalue),
                                                                                                      'WeightMinor': int(
                                                                                                          minorvalue)}}})

                                            break

                                        except ConnectionError:

                                            break


                                    elif ShippingDimensions == str(3):

                                        weight = float(input('What is the weight? (kg)'))
                                        length = float(input('What is the length? (cm)'))
                                        width = float(input('What is the width? (cm)'))
                                        height = float(input('What is the height? (cm)'))

                                        t = math.modf(weight)  # (0.5678000000000338, 1234.0)

                                        minorvalue = round((t[0]) * 1000)
                                        majorvalue = round(t[1])

                                        try:

                                            dimensionupdate = api.execute('ReviseItem', {'Item': {"ItemID": ItemID,
                                                                                                  'ShippingPackageDetails': {
                                                                                                      'MeasurementUnit': 'Metric',
                                                                                                      'PackageDepth': height,
                                                                                                      'PackageLength': length,
                                                                                                      'PackageWidth': width,
                                                                                                      'WeightMajor': majorvalue,
                                                                                                      'WeightMinor': minorvalue}}})

                                            break
                                        except:

                                            break

                                    elif ShippingDimensions == str(4):

                                        weight = 0
                                        length = 0
                                        width = 0
                                        height = 0

                                        ItemNote = 'Label.'
                                        break



                                    else:
                                        continue



                            else:
                                length = GetItemAPIdic['Item']['ShippingPackageDetails']['PackageLength']['value']
                                width = GetItemAPIdic['Item']['ShippingPackageDetails']['PackageWidth']['value']
                                height = GetItemAPIdic['Item']['ShippingPackageDetails']['PackageDepth']['value']
                                weightmajor = GetItemAPIdic['Item']['ShippingPackageDetails']['WeightMajor']['value']
                                weightminor = GetItemAPIdic['Item']['ShippingPackageDetails']['WeightMinor']['value']
                                weight = str(float(weightmajor) + float(float(weightminor) / 1000))

                                if float(weight) <= 0.5:
                                    sendle_weight_results = item_check(ItemID)

                                    item_thickness_results = item_thickness_check(ItemID)

                                    if sendle_weight_results == 'true':

                                        while True:
                                            sendleweighttest = input(f'''

                                        Hold up there cowboy - any chance the following package would be less than 250g?

                                        {ebayItemName} (SKU: {ebaySKU})

                                        1) Ohhh boy yessir!
                                        2) No
                                        3) I don't know!

                                        Enter response here:''')

                                            if sendleweighttest == str(1):
                                                weight = 0.25

                                                t = math.modf(weight)  # (0.5678000000000338, 1234.0)

                                                minorvalue = round((t[0]) * 1000)
                                                majorvalue = round(t[1])

                                                try:

                                                    dimensionupdate = api.execute('ReviseItem',
                                                                                  {'Item': {"ItemID": ItemID,
                                                                                            'ShippingPackageDetails': {
                                                                                                'MeasurementUnit': 'Metric',
                                                                                                'WeightMajor': majorvalue,
                                                                                                'WeightMinor': minorvalue}}})

                                                    item_insert(ItemID)
                                                    break

                                                except:

                                                    break

                                            elif sendleweighttest == str(2):

                                                item_insert(ItemID)

                                                break

                                            elif sendleweighttest == str(3):

                                                break

                                            else:
                                                print('Valid input not found, try again')
                                                time.sleep(1)
                                                continue

                                    if item_thickness_results == 'true':
                                        while True:
                                            sendleweighttest = console.input(f'''

        ___
     __|___|__
      ('o_o')                             
      _\~-~/_    ______.                  STOP RIGHT THERE. 
     //\__/\ \ ~(_]---'                   
    / )O  O( .\/_)                        I AM A [red]TERRORIST[/].
    \ \    / \_/            
    )/_|  |_\\                             WOULD THE FOLLOWING PACKAGE BE [red]LESS THAN 3cm THICK[/]?
   // /(\/)\ \\     
   /_/      \_\\      
  (_||      ||_)      
    \| |__| |/
     | |  | |                             
     | |  | |                            
     |_|  |_|                             
     /_\  /_\\                            


Item: {ebayItemName}   

SKU: {ebaySKU}           

                                          1) Yes Mr Terrorist.
                                          2) No sir   
                                          3) I don't know!
                                          Enter response here:''')

                                            if sendleweighttest == str(1):
                                                height = 21
                                                width = 21
                                                length = 3
                                                weightvalue = 0.25
                                                t = math.modf(weightvalue)  # (0.5678000000000338, 1234.0)

                                                minorvalue = round((t[0]) * 1000)
                                                majorvalue = round(t[1])

                                                try:

                                                    dimensionupdate = api.execute('ReviseItem',
                                                                                  {'Item': {"ItemID": ItemID,
                                                                                            'ShippingPackageDetails': {
                                                                                                'MeasurementUnit': 'Metric',
                                                                                                'PackageDepth': height,
                                                                                                'PackageLength': length,
                                                                                                'PackageWidth': width, }}})

                                                    item_thickness_insert(ItemID)
                                                    break

                                                except:

                                                    break


                                            elif sendleweighttest == str(2):

                                                item_thickness_insert(ItemID)
                                                break

                                            elif sendleweighttest == str(3):

                                                break

                                            else:
                                                print('Valid input not found, try again')
                                                time.sleep(1)
                                                continue
                        ###############################################################################
                        # for orderlines in range(orderlinelen):
                        #     Stickydic['Sticky Note (Item {0})'.format(orderlines + 1)] = ItemNote
                        #     Stickydic['Length (cm) (Item {0})'.format(orderlines + 1)] = length
                        #     Stickydic['Height (cm) (Item {0})'.format(orderlines + 1)] = height
                        #     Stickydic['Width (cm) (Item {0})'.format(orderlines + 1)] = width
                        #     Stickydic['Weight (kg) (Item {0})'.format(orderlines + 1)] = weight

                        keys.update({'Sticky Note': ItemNote, 'Height': height, 'Width': width, 'Length': length,
                                     'Weight': weight})
                        preFinalList.append(keys)
                        # eBayJSON.append(keys)


                    else:
                        continue

                elif 'Order' in SoldList[transactions]:

                    orderlinelen = len(SoldList[transactions]['Order']['TransactionArray']['Transaction'])

                    for multitransactions in SoldList[transactions]['Order']['TransactionArray'][
                        'Transaction']:  ###  Checking Multi Order Lines

                        ebayorderline = multitransactions['OrderLineItemID']
                        itemprice = multitransactions['Item']['SellingStatus']['CurrentPrice']['value']
                        # print(ebayorderline)
                        #  print('bing')
                        # print('dictionary orderline = ' + dicorderline)
                        # print('Ebay orderline = ' + ebayorderline)

                        if dicorderline == ebayorderline:

                            ItemID = multitransactions['Item']['ItemID']
                            # print('Item ID: ' +ItemID)
                            # print(multitransactions)

                            try:
                                ItemNote = multitransactions['Item']['PrivateNotes']
                            except KeyError:
                                ItemNote = ''
                            # print(ItemNote)
                            # print('bing')

                            Itemlist.append(ItemNote)

                            if len(Itemlist) == orderlinelen:

                                for orderlines in range(len(Itemlist)):
                                    Stickydic['Sticky Note (Item {0})'.format(orderlines + 1)] = Itemlist[orderlines]

                                keys.update({'Sticky Note': Stickydic})
                                preFinalList.append(keys)
                            # eBayJSON.append(keys)



                        else:
                            continue
    except (KeyError, AttributeError, TypeError) as e:
        print(f"DIAGNOSTIC: SoldList matching error for order {keys.get('PurchaseOrderNumber', 'unknown')}: {e}")
        continue
t = 0
# pprint.pprint(preFinalList)

print(f"\n=== DIAGNOSTIC: SoldList matching summary ===")
print(f"  eBayJSON entries entering SoldList match: {numdic}")
print(f"  SoldList entries available:               {len(SoldList)}")
print(f"  Survived into preFinalList:               {len(preFinalList)}")
print(f"  Lost during matching:                     {numdic - len(preFinalList)}")
print(f"===============================================\n")

eBayJSON = preFinalList

if len(eBayJSON) == 0:
    print("DIAGNOSTIC: 0 eBay orders survived into preFinalList — all orders were lost during processing")
    input(
        "\n\nUh oh dog this script did NOT work. Sorry this is one bug I haven't fucking nutted out just yet.\n\n Usually works though if you try again.")
    sys.exit()
elif float(numdic) / float(len(eBayJSON)) < 0.3:
    input(
        "\n\nUh oh dog this script did NOT work. Sorry this is one bug I haven't fucking nutted out just yet.\n\n Usually works though if you try again.")
    sys.exit()

for x in range(len(eBayJSON)):  #### JUST DOING ANOTHER CHECK TO SEE WHICH ORDERS ARE ALREADY SENT

    passfailed = 'false'

    x = x - t

    orderline = len(eBayJSON[x]['ItemTitle'])

    TotalPrice = eBayJSON[x]['TotalPrice']

    if 'Sticky Note' not in eBayJSON[x]:
        eBayJSON.pop(x)
        t = t + 1
        continue

    stickynote = eBayJSON[x]['Sticky Note']

    if orderline == 1:

        if 'ordered' in eBayJSON[x]['Sticky Note'].lower() or 'po' in \
                eBayJSON[x]['Sticky Note'].lower() or 'stock' in \
                eBayJSON[x]['Sticky Note'].lower() or 'awaiting' in \
                eBayJSON[x]['Sticky Note'].lower() or 'no' in \
                eBayJSON[x]['Sticky Note'].lower() or 'messaged' in \
                eBayJSON[x]['Sticky Note'].lower() or 'cust' in \
                eBayJSON[x]['Sticky Note'].lower() or 'list' in \
                eBayJSON[x]['Sticky Note'].lower() or 'cp' in \
                eBayJSON[x]['Sticky Note'].lower() or 'courier' in \
                eBayJSON[x]['Sticky Note'].lower() or 'fastway' in \
                eBayJSON[x]['Sticky Note'].lower() or 'fw' in \
                eBayJSON[x]['Sticky Note'].lower() or 'ordered' in \
                eBayJSON[x]['Sticky Note'].lower() or 'transdirect' in \
                eBayJSON[x]['Sticky Note'].lower() or 'dropship' in \
                eBayJSON[x]['Sticky Note'].lower() or 'sent' in \
                eBayJSON[x]['Sticky Note'].lower():
            eBayJSON.pop(x)
            t = t + 1
            continue

        if 'mini' not in eBayJSON[x]['Sticky Note'].lower():
            if 'devil' not in eBayJSON[x]['Sticky Note'].lower():
                if 'label' not in eBayJSON[x]['Sticky Note'].lower():
                    if 'satchel' not in eBayJSON[x]['Sticky Note'].lower():
                        eBayJSON.pop(x)
                        t = t + 1
                        continue
        # pprint.pprint(eBayJSON)

        if int(eBayJSON[x]['Quantity'][
                   'Quantity (Item 1)']) > 1:  ########## START OF THE 1 LINER, MULTIPLE QUANTITIES ######

            print('''\n\n\n\n\nCustomer has ordered the following items:\n\n''')

            print('(QTY: ' + eBayJSON[x]['Quantity']['Quantity (Item 1)'] + ') ' +
                  eBayJSON[x]['ItemTitle']['Name (Item 1)'] + ' (SKU: ' +
                  eBayJSON[x]['ItemSKU']['SKU (Item 1)'] + ')')

            print('\nORDER TOTAL: $' + TotalPrice)

            while True:
                PostageType = input('''\n\nWhich of the following best suits it?

                                                                        1) Minilope
                                                                        2) Devilope
                                                                        3) Popelope
                                                                        4) Satchel / eParcel
                                                                        5) Dropship
                                                                        6) I don't know! (Makes no permanent change. Will just add to Label)
                                                                        7) Skip this order!

                                                                        Enter Response:''')

                if PostageType == str(1):
                    stickynote = 'Minilope'

                    if float(TotalPrice) > minimum_tracking_amount:
                        stickynote = 'Satchel'
                        height = 21
                        width = 21
                        length = 3
                        weight = 0.25
                        eBayJSON[x].update({'Height': height})
                        eBayJSON[x].update({'Width': width})
                        eBayJSON[x].update({'Length': length})
                        eBayJSON[x].update({'Weight': weight})
                        eBayJSON[x].update({'Final_Postage': stickynote})
                        break
                    else:
                        eBayJSON[x].update({'Final_Postage': stickynote})
                        break




                elif PostageType == str(2):
                    stickynote = 'Devilope'

                    if float(TotalPrice) > minimum_tracking_amount:
                        stickynote = 'Satchel'
                        height = 21
                        width = 21
                        length = 3
                        weight = 0.25
                        eBayJSON[x].update({'Height': height})
                        eBayJSON[x].update({'Width': width})
                        eBayJSON[x].update({'Length': length})
                        eBayJSON[x].update({'Weight': weight})
                        eBayJSON[x].update({'Final_Postage': stickynote})
                        break
                    else:
                        eBayJSON[x].update({'Final_Postage': stickynote})
                        break


                elif PostageType == str(3):
                    stickynote = 'Label'

                    eBayJSON[x].update({'Final_Postage': stickynote})
                    break

                elif PostageType == str(4):
                    while True:

                        stickynote = 'Satchel'

                        ShippingDimensions = input('''
                        Will this fit into a bubble mailer?.

                                                        1) Yes (<250g weight)
                                                        2) Yes (>250g weight)
                                                        3) No (Will put into label until I get better at coding)

                                                        Enter Response:''')

                        if ShippingDimensions == str(1) or ShippingDimensions == str(2):
                            weight = 0.5

                            if ShippingDimensions == str(1):

                                weight = 0.25

                                while True:

                                    thickness_test = input('''

                                    Would this package be less than 3cm thick?

                                    1) Yes
                                    2) No
                                    3) I don't know

                                    Enter response here:''')

                                    if thickness_test == str(1):
                                        thickness_test_result = 'true'
                                        break

                                    if thickness_test == str(2) or thickness_test == str(3):
                                        thickness_test_result = 'false'
                                        break

                                    else:
                                        print('Valid input not found. Try again.')
                                        time.sleep(1)
                                        continue

                            try:

                                if thickness_test == str(3):
                                    stickynote = 'Label'
                                    eBayJSON[x].update({'Final_Postage': stickynote})
                                    passcondition = 'true'

                                    break
                            except:
                                pass

                            try:
                                thickness_test_result
                            except:
                                thickness_test_result = 'false'


                            if thickness_test_result == 'true':

                                height = 21
                                width = 21
                                length = 3

                            else:

                                length = 18
                                width = 23
                                height = 4

                            eBayJSON[x].update({'Height': height})
                            eBayJSON[x].update({'Width': width})
                            eBayJSON[x].update({'Length': length})
                            eBayJSON[x].update({'Weight': weight})
                            eBayJSON[x].update({'Final_Postage': stickynote})

                            passcondition = 'true'

                            break

                        elif ShippingDimensions == str(3):
                            stickynote = 'Label'
                            eBayJSON[x].update({'Final_Postage': stickynote})
                            passcondition = 'true'

                            break

                        else:
                            input('No correct value found. Press Enter to try again.')
                            continue


                elif PostageType == str(5):
                    stickynote = 'Dropship'
                    eBayJSON[x].update({'Final_Postage': stickynote})

                    break

                elif PostageType == str(6):

                    stickynote = 'Label'
                    eBayJSON[x].update({'Final_Postage': stickynote})
                    break

                elif PostageType == str(7):

                    stickynote = 'Skip'
                    eBayJSON[x].update({'Final_Postage': stickynote})
                    break

                if passcondition == 'true':
                    break

                else:
                    print('\nValid number not found. Try again. \n')
                    time.sleep(1)
                    continue  ########## END OF THE 1 LINER, MULTIPLE QUANTITIES ######

        if 'mini' in stickynote.lower() or 'devil' in stickynote.lower():

            if float(TotalPrice) > minimum_tracking_amount:

                stickynote = 'Satchel.'

                height = 21
                width = 21
                length = 3
                weight = 0.25

                eBayJSON[x].update({'Final_Postage': stickynote})
                eBayJSON[x].update({'Height': height})
                eBayJSON[x].update({'Width': width})
                eBayJSON[x].update({'Length': length})
                eBayJSON[x].update({'Weight': weight})


            else:
                eBayJSON[x].update({'Final_Postage': stickynote})


        else:
            eBayJSON[x].update({'Final_Postage': stickynote})

pprint.pprint(eBayJSON)
#############################JAN 2022 UPDATE START ########################################

### Merges eBay orders together, appending all necessary categories + makes strings versions of orderID, SRN, and purchase order numbers

for order in eBayJSON:  # Makes all the OrderIDs lists
    OrderID = order['OrderID']
    OrderIDList = {'OrderID': [OrderID]}
    order.update(OrderIDList)

    SRN = order['SalesRecordNumber']
    SRNList = {'SalesRecordNumber': [SRN]}
    order.update(SRNList)

    PurchaseOrderNumber = order['PurchaseOrderNumber']
    PurchaseOrderNumberList = {'PurchaseOrderNumber': [PurchaseOrderNumber]}
    order.update(PurchaseOrderNumberList)

for order in range(
        len(eBayJSON)):  # Itereates throgh each order and 'merges' orders with the same address + customer name. Deletes Duplicate.
    # Going through each order in eBayJSON
    suborderminus = 0
    try:
        order_id = eBayJSON[order]['OrderID']  # grabs the 'primary' orders order id (eg Order 0)

        for sub_order in range(len(eBayJSON)):  # Begins a loop looking for duplicate orders
            sub_order -= suborderminus
            if eBayJSON[sub_order]['OrderID'] == order_id:  # Passes if the secondary loop finds the primary order
                pass

            else:
                if eBayJSON[order]['BuyerName'] == eBayJSON[sub_order]['BuyerName'] and eBayJSON[order]['Phone'] == \
                        eBayJSON[sub_order]['Phone'] and eBayJSON[order]['Postcode'] == eBayJSON[sub_order][
                    'Postcode'] and eBayJSON[order]['AddressLine1'] == eBayJSON[sub_order]['AddressLine1']:
                    ###Found two orders going to the same name and address

                    orderlinelen = len(eBayJSON[order]['Quantity'])
                    suborderlinelen = len(eBayJSON[sub_order]['Quantity'])

                    ####################NEED TO EXTEND THE APROPRIATE LISTS, BUT, ALSO NEED AN ALGORITHM TO UPDATE THE (ITEM 1, ITEM 2), ELSE THERE WILL
                    #################### BE OVERLAPPING KEYS OF THE SAME VALUE

                    # Lists to extend = ItemID, ItemSKU, ItemTitle, OrderLineID, PurchaseOrderNumber, Quantity, SKU, OrderID, SalesRecordNumber
                    # PurchaseOrderNumber, OrderID and SalesRecordNumber wont need algorithm

                    eBayJSON[order]['PurchaseOrderNumber'].extend(eBayJSON[sub_order]['PurchaseOrderNumber'])
                    eBayJSON[order]['OrderID'].extend(eBayJSON[sub_order]['OrderID'])
                    eBayJSON[order]['SalesRecordNumber'].extend(eBayJSON[sub_order]['SalesRecordNumber'])

                    ItemIDmulti = {}  #
                    ItemSKUmulti = {}
                    ItemTitlemulti = {}
                    OrderLineIDmulti = {}
                    Quantitymulti = {}
                    SKUmulti = {}

                    for y in range(suborderlinelen):
                        ItemIDmulti['ItemID (Item {0})'.format(y + 1 + orderlinelen)] = eBayJSON[sub_order]['ItemID'][
                            'ItemID (Item {0})'.format(y + 1)]
                        ItemSKUmulti['SKU (Item {0})'.format(y + 1 + orderlinelen)] = eBayJSON[sub_order]['ItemSKU'][
                            'SKU (Item {0})'.format(y + 1)]
                        ItemTitlemulti['Name (Item {0})'.format(y + 1 + orderlinelen)] = \
                            eBayJSON[sub_order]['ItemTitle']['Name (Item {0})'.format(y + 1)]
                        OrderLineIDmulti['Orderline (Item {0})'.format(y + 1 + orderlinelen)] = \
                            eBayJSON[sub_order]['OrderLineID']['Orderline (Item {0})'.format(y + 1)]
                        Quantitymulti['Quantity (Item {0})'.format(y + 1 + orderlinelen)] = \
                            eBayJSON[sub_order]['Quantity']['Quantity (Item {0})'.format(y + 1)]
                        SKUmulti['SKU (Item {0})'.format(y + 1 + orderlinelen)] = eBayJSON[sub_order]['SKU'][
                            'SKU (Item {0})'.format(y + 1)]

                    eBayJSON[order]['ItemID'].update(ItemIDmulti)
                    eBayJSON[order]['ItemSKU'].update(ItemSKUmulti)
                    eBayJSON[order]['ItemTitle'].update(ItemTitlemulti)
                    eBayJSON[order]['Quantity'].update(Quantitymulti)
                    eBayJSON[order]['SKU'].update(SKUmulti)
                    eBayJSON[order]['OrderLineID'].update(OrderLineIDmulti)

                    del eBayJSON[sub_order]
                    suborderminus += 1

    except IndexError:
        pass

for orders in eBayJSON:
    pre_string_order_id = ''

    for orderid in orders['OrderID']:
        pre_string_order_id = pre_string_order_id + ' + ' + orderid

    pre_string_order_id = pre_string_order_id[3:]
    string_order_id = {'String_OrderID': pre_string_order_id}

    orders.update(string_order_id)

    pre_string_order_id = ''

    for purchaseordernumber in orders['PurchaseOrderNumber']:
        pre_string_order_id = pre_string_order_id + ' + ' + purchaseordernumber

    pre_string_order_id = pre_string_order_id[3:]
    string_purchaseordernumber = {'String_PurchaseOrderNumber': pre_string_order_id}

    orders.update(string_purchaseordernumber)

    pre_string_order_id = ''

    for srn in orders['SalesRecordNumber']:
        pre_string_order_id = pre_string_order_id + ' + ' + srn

    pre_string_order_id = pre_string_order_id[3:]
    string_srn = {'String_SalesRecordNumber': pre_string_order_id}

    orders.update(string_srn)
t = 0
#############################JAN 2022 UPDATE END ########################################
for x in range(len(eBayJSON)):
    passfailed = 'false'

    x = x - t

    orderline = len(eBayJSON[x]['ItemTitle'])

    for y in range(orderline):

        try:

            if 'ordered' in eBayJSON[x]['Sticky Note']['Sticky Note (Item ' + str(y + 1) + ')'].lower() or 'po' in \
                    eBayJSON[x]['Sticky Note']['Sticky Note (Item ' + str(y + 1) + ')'].lower() or 'stock' in \
                    eBayJSON[x]['Sticky Note']['Sticky Note (Item ' + str(y + 1) + ')'].lower() or 'awaiting' in \
                    eBayJSON[x]['Sticky Note']['Sticky Note (Item ' + str(y + 1) + ')'].lower() or 'no' in \
                    eBayJSON[x]['Sticky Note']['Sticky Note (Item ' + str(y + 1) + ')'].lower() or 'messaged' in \
                    eBayJSON[x]['Sticky Note']['Sticky Note (Item ' + str(y + 1) + ')'].lower() or 'cust' in \
                    eBayJSON[x]['Sticky Note']['Sticky Note (Item ' + str(y + 1) + ')'].lower():
                eBayJSON.pop(x)
                t = t + 1
                passfailed = 'true'

                break
        except:
            pass

t = 0

for x in range(len(eBayJSON)):

    passfailed = 'false'

    x = x - t

    orderline = len(eBayJSON[x]['ItemTitle'])

    TotalPrice = eBayJSON[x]['TotalPrice']

    if orderline > 1:

        # pprint.pprint(eBayJSON[x])

        namemulti = {}
        SKUmulti = {}
        quantitymulti = {}

        g = eBayJSON[x]['Quantity']

        if passfailed == 'true':
            continue

        while True:

            passcondition = 'failed'

            #    pprint.pprint(eBayJSON)

            print('''\n\n\n\n\nCustomer has ordered the following items:\n\n''')

            for y in range(orderline):
                print('(QTY: ' + eBayJSON[x]['Quantity']['Quantity (Item ' + str(y + 1) + ')'] + ') ' +
                      eBayJSON[x]['ItemTitle']['Name (Item ' + str(y + 1) + ')'] + ' (SKU: ' +
                      eBayJSON[x]['ItemSKU']['SKU (Item ' + str(y + 1) + ')'] + ')')

                namemulti['Name (Item {0})'.format(y + 1)] = eBayJSON[x]['ItemTitle']['Name (Item ' + str(y + 1) + ')']
                SKUmulti['SKU (Item {0})'.format(y + 1)] = eBayJSON[x]['ItemSKU']['SKU (Item ' + str(y + 1) + ')']
                quantitymulti['Quantity (Item {0})'.format(y + 1)] = eBayJSON[x]['Quantity'][
                    'Quantity (Item ' + str(y + 1) + ')']

            PostageType = input('''\n\nWhich of the following best suits it?

                                                        1) Minilope
                                                        2) Devilope
                                                        3) Popelope
                                                        4) Satchel / eParcel
                                                        5) Dropship
                                                        6) I don't know! (Makes no permanent change. Will just add to Label)
                                                        7) Skip this order!

                                                        Enter Response:''')

            if PostageType == str(1):
                stickynote = 'Minilope'

                if float(TotalPrice) > minimum_tracking_amount:
                    stickynote = 'Satchel'
                    height = 21
                    width = 21
                    length = 3
                    weight = 0.25
                    eBayJSON[x].update({'Height': height})
                    eBayJSON[x].update({'Width': width})
                    eBayJSON[x].update({'Length': length})
                    eBayJSON[x].update({'Weight': weight})
                    eBayJSON[x].update({'Final_Postage': stickynote})
                    break
                else:
                    eBayJSON[x].update({'Final_Postage': stickynote})
                    break




            elif PostageType == str(2):
                stickynote = 'Devilope'

                if float(TotalPrice) > minimum_tracking_amount:
                    stickynote = 'Satchel'
                    height = 21
                    width = 21
                    length = 3
                    weight = 0.25
                    eBayJSON[x].update({'Height': height})
                    eBayJSON[x].update({'Width': width})
                    eBayJSON[x].update({'Length': length})
                    eBayJSON[x].update({'Weight': weight})
                    eBayJSON[x].update({'Final_Postage': stickynote})
                    break
                else:
                    eBayJSON[x].update({'Final_Postage': stickynote})
                    break


            elif PostageType == str(3):
                stickynote = 'Label'

                eBayJSON[x].update({'Final_Postage': stickynote})
                break

            elif PostageType == str(4):
                while True:

                    stickynote = 'Satchel'

                    ShippingDimensions = input('''
                    Will this fit into a bubble mailer?.

                                        1) Yes (<250g weight)
                                        1) Yes (>250g weight)
                                        2) No (Will put into label until I get better at coding)

                                        Enter Response:''')

                    if ShippingDimensions == str(1) or ShippingDimensions == str(2):

                        weight = 0.5

                        if ShippingDimensions == str(1):
                            weight = 0.25

                            while True:

                                thickness_test = input('''

                                                               Would this package be less than 3cm thick?

                                                               1) Yes
                                                               2) No
                                                               3) I don't know

                                                               Enter response here:''')

                                if thickness_test == str(1):
                                    thickness_test_result = 'true'
                                    break

                                if thickness_test == str(2) or thickness_test == str(3):
                                    thickness_test_result = 'false'
                                    break

                                else:
                                    print('Valid input not found. Try again.')
                                    time.sleep(1)
                                    continue

                            if thickness_test == str(3):
                                stickynote = 'Label'
                                eBayJSON[x].update({'Final_Postage': stickynote})
                                passcondition = 'true'

                                break

                            if thickness_test_result == 'true':

                                height = 21
                                width = 21
                                length = 3

                            else:

                                length = 18
                                width = 23
                                height = 4

                        eBayJSON[x].update({'Height': height})
                        eBayJSON[x].update({'Width': width})
                        eBayJSON[x].update({'Length': length})
                        eBayJSON[x].update({'Weight': weight})
                        eBayJSON[x].update({'Final_Postage': stickynote})

                        passcondition = 'true'

                        break

                    elif ShippingDimensions == str(3):
                        stickynote = 'Label'
                        eBayJSON[x].update({'Final_Postage': stickynote})
                        passcondition = 'true'

                        break

                    else:
                        input('No correct value found. Press Enter to try again.')
                        continue


            elif PostageType == str(5):
                stickynote = 'Dropship'
                eBayJSON[x].update({'Final_Postage': stickynote})

                break

            elif PostageType == str(6):

                stickynote = 'Label'
                eBayJSON[x].update({'Final_Postage': stickynote})
                break

            elif PostageType == str(7):

                stickynote = 'Skip'
                eBayJSON[x].update({'Final_Postage': stickynote})
                break

            if passcondition == 'true':
                break

            else:
                print('\nValid number not found. Try again. \n')
                time.sleep(1)
                continue

        # eBayJSON[x].update({'Final_Postage': stickynote})
        # eBayJSON[x].update({'Width': width})
        # eBayJSON[x].update({'Length': length})
        # eBayJSON[x].update({'Weight': weight})

# pprint.pprint(eBayJSON)

############# THIS IS THE END OF THE EBAY ORDERS (eBayJSON)
# if eBayJSON == []:
#     input(
#         "\n\nUh oh dog this script did NOT work. Sorry this is one bug I haven't fucking nutted out just yet.\n\n Usually works though if you try again.")
#     sys.exit()
netoresults = OrdersJSON
ebayresults = eBayJSON

for x in netoresults['Satchel']:

    if x['height'] == 0 or x['length'] == 0 or x['width'] == 0 or x['weight'] == 0:
        indexvalue = netoresults['Satchel'].index(x)
        x['PostageType'] = 'Label'
        netoresults['Label'].append(x)
        netoresults['Satchel'].pop(indexvalue)

for x in netoresults:

    for lopes in netoresults[x]:  #####Adding items to master pick sheet

        try:

            if lopes['Orderlines'] == 1 and lopes['Quantity'] == '1':
                maxrow = wsMasterPickSheet.max_row
                maxrow = maxrow + 1
                wsMasterPickSheet['A' + str(maxrow)].value = lopes['ItemName']
                wsMasterPickSheet['B' + str(maxrow)].value = lopes['SKU']
                wsMasterPickSheet['C' + str(maxrow)].value = lopes['Quantity']

            else:
                for t in range(len(lopes['ItemName'])):
                    maxrow = wsMasterPickSheet.max_row
                    maxrow = maxrow + 1
                    wsMasterPickSheet['A' + str(maxrow)].value = lopes['ItemName']['Name (Item ' + str(t + 1) + ')']
                    wsMasterPickSheet['B' + str(maxrow)].value = lopes['SKU']['SKU (Item ' + str(t + 1) + ')']
                    wsMasterPickSheet['C' + str(maxrow)].value = lopes['Quantity']['Quantity (Item ' + str(t + 1) + ')']
        except TypeError:
            continue

    if 'mini' in x.lower():  ######Adding items to the envelope sheets

        for y in netoresults[x]:  # iterating through each item in 'Minilope

            try:
                y['ItemName'] = y['ItemName'].replace("'", "")
                y['BuyerName'] = y['BuyerName'].replace("'", "")
                y['Company'] = y['Company'].replace("'", "")
                y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                y['City'] = y['City'].replace("'", "")

                try:
                    cursor = connection.cursor()
                except:
                    connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                    cursor = connection.cursor()

                try:
                    cursor.execute(
                        fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][0:50]}', '', '{y['Quantity'][0:50]}', '{y['ItemName'][0:50]}', '{y['SKU'][0:50]}', '{y['OrderID'][0]}', 'Minilope','' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                except:
                    pass
            except:

                y['BuyerName'] = y['BuyerName'].replace("'", "")
                y['Company'] = y['Company'].replace("'", "")
                y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                y['City'] = y['City'].replace("'", "")

                for lines in range(len(y['SKU'])):
                    y['ItemName'][f'Name (Item {str(lines + 1)})'] = y['ItemName'][
                        f'Name (Item {str(lines + 1)})'].replace("'", "")
                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()

                    try:
                        cursor.execute(
                            fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '', '{y['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{y['ItemName'][f'Name (Item {str(lines + 1)})'][0:200]}', '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '{y['OrderID'][0][0:50]}', 'Minilope','' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                    except:
                        pass
            y['String_OrderID'] = str(y['item_count_string'])
            y['String_SalesRecordNumber'] = str(y['item_count_string'])
            mergeMinilopes.append(y)

            maxrow = wsMini.max_row
            maxrow = maxrow + 1
            wsMini['A' + str(maxrow)].value = y['String_OrderID']
            wsMini['B' + str(maxrow)].value = y['BuyerName']
            wsMini['C' + str(maxrow)].value = y['Email']
            wsMini['D' + str(maxrow)].value = y['AddressLine1']
            wsMini['E' + str(maxrow)].value = y['AddressLine2']
            wsMini['F' + str(maxrow)].value = y['City']
            wsMini['G' + str(maxrow)].value = y['State']
            wsMini['H' + str(maxrow)].value = y['Postcode']
            # wsMini['I' + str(maxrow)].value = y['ItemName']
            # wsMini['J' + str(maxrow)].value = y['SKU']
            # wsMini['K' + str(maxrow)].value = y['Quantity']
            wsMini['L' + str(maxrow)].value = y['TotalPrice']
            wsMini['M' + str(maxrow)].value = y['PostageType']  # THINK THIS WAS SUPPOSED TO BE EXPRESS?
            wsMini['N' + str(maxrow)].value = y['PostageType']
            # wsMini['P' + str(maxrow)].value = str(y['height'])
            # wsMini['Q' + str(maxrow)].value = str(y['length'])
            # wsMini['R' + str(maxrow)].value = str(y['width'])
            # wsMini['S' + str(maxrow)].value = str(y['weight'])
            wsMini['T' + str(maxrow)].value = str(y['Phone'])
            wsMini['U' + str(maxrow)].value = str(y['Company'])


    elif 'devil' in x.lower():

        for y in netoresults[x]:  # iterating through each item in 'Minilope

            try:
                y['ItemName'] = y['ItemName'].replace("'", "")
                y['BuyerName'] = y['BuyerName'].replace("'", "")
                y['Company'] = y['Company'].replace("'", "")
                y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                y['City'] = y['City'].replace("'", "")

                try:
                    cursor = connection.cursor()
                except:
                    connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                    cursor = connection.cursor()

                try:
                    cursor.execute(
                        fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed']}','' , '{y['SKU'][0:50]}', '', '{y['Quantity'][0:50]}', '{y['ItemName'][0:50]}', '{y['SKU'][0:50]}', '{y['OrderID'][0]}', 'Devilope','' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                except:
                    pass
            except:

                y['BuyerName'] = y['BuyerName'].replace("'", "")
                y['Company'] = y['Company'].replace("'", "")
                y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                y['City'] = y['City'].replace("'", "")

                for lines in range(len(y['SKU'])):
                    y['ItemName'][f'Name (Item {str(lines + 1)})'] = y['ItemName'][
                        f'Name (Item {str(lines + 1)})'].replace("'", "")
                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()
                    try:
                        cursor.execute(
                            fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '', '{y['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{y['ItemName'][f'Name (Item {str(lines + 1)})'][0:200]}', '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '{y['OrderID'][0][0:50]}', 'Devilope','' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                    except:
                        pass
            y['String_OrderID'] = str(y['item_count_string'])
            y['String_SalesRecordNumber'] = str(y['item_count_string'])

            mergeDevilopes.append(y)

            maxrow = wsDevil.max_row

            maxrow = maxrow + 1
            wsDevil['A' + str(maxrow)].value = y['String_OrderID']
            wsDevil['B' + str(maxrow)].value = y['BuyerName']
            wsDevil['C' + str(maxrow)].value = y['Email']
            wsDevil['D' + str(maxrow)].value = y['AddressLine1']
            wsDevil['E' + str(maxrow)].value = y['AddressLine2']
            wsDevil['F' + str(maxrow)].value = y['City']
            wsDevil['G' + str(maxrow)].value = y['State']
            wsDevil['H' + str(maxrow)].value = y['Postcode']
            # wsDevil['I' + str(maxrow)].value = y['ItemName']
            # wsDevil['J' + str(maxrow)].value = y['SKU']
            # wsDevil['K' + str(maxrow)].value = y['Quantity']
            wsDevil['L' + str(maxrow)].value = y['TotalPrice']
            wsDevil['M' + str(maxrow)].value = y['PostageType']  # THINK THIS WAS SUPPOSED TO BE EXPRESS?
            wsDevil['N' + str(maxrow)].value = y['PostageType']
            # wsDevil['P' + str(maxrow)].value = str(y['height'])
            # wsDevil['Q' + str(maxrow)].value = str(y['length'])
            # wsDevil['R' + str(maxrow)].value = str(y['width'])
            # wsDevil['S' + str(maxrow)].value = str(y['weight'])
            wsDevil['T' + str(maxrow)].value = str(y['Phone'])
            wsDevil['U' + str(maxrow)].value = str(y['Company'])




    elif 'label' in x.lower():

        for y in netoresults[x]:  # iterating through each item in 'Minilope

            try:
                y['ItemName'] = y['ItemName'].replace("'", "")
                y['BuyerName'] = y['BuyerName'].replace("'", "")
                y['Company'] = y['Company'].replace("'", "")
                y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                y['City'] = y['City'].replace("'", "")

                try:
                    cursor = connection.cursor()
                except:
                    connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                    cursor = connection.cursor()

                try:
                    cursor.execute(
                        fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][0:50]}', '', '{y['Quantity'][0:50]}', '{y['ItemName'][0:50]}', '{y['SKU'][0:50]}', '{y['OrderID'][0][0:50]}', 'Label','' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                except:
                    pass
            except:

                y['BuyerName'] = y['BuyerName'].replace("'", "")
                y['Company'] = y['Company'].replace("'", "")
                y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                y['City'] = y['City'].replace("'", "")

                for lines in range(len(y['SKU'])):
                    y['ItemName'][f'Name (Item {str(lines + 1)})'] = y['ItemName'][
                        f'Name (Item {str(lines + 1)})'].replace("'", "")
                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()

                    try:
                        cursor.execute(
                            fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '', '{y['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{y['ItemName'][f'Name (Item {str(lines + 1)})'][0:200]}', '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '{y['OrderID'][0]}', 'Label','' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                    except:
                        pass
            y['String_SalesRecordNumber'] = str(y['item_count_string'])
            y['String_OrderID'] = str(y['item_count_string'])
            mergeLabel.append(y)

            maxrow = wsLabel.max_row
            maxrow = maxrow + 1
            wsLabel['A' + str(maxrow)].value = y['String_OrderID']
            wsLabel['B' + str(maxrow)].value = y['BuyerName']
            wsLabel['C' + str(maxrow)].value = y['Email']
            wsLabel['D' + str(maxrow)].value = y['AddressLine1']
            wsLabel['E' + str(maxrow)].value = y['AddressLine2']
            wsLabel['F' + str(maxrow)].value = y['City']
            wsLabel['G' + str(maxrow)].value = y['State']
            wsLabel['H' + str(maxrow)].value = y['Postcode']
            # wsLabel['I' + str(maxrow)].value = y['ItemName']
            # wsLabel['J' + str(maxrow)].value = y['SKU']
            # wsLabel['K' + str(maxrow)].value = y['Quantity']
            wsLabel['L' + str(maxrow)].value = y['TotalPrice']
            wsLabel['M' + str(maxrow)].value = y['PostageType']  # THINK THIS WAS SUPPOSED TO BE EXPRESS?
            wsLabel['N' + str(maxrow)].value = y['PostageType']
            # wsLabel['P' + str(maxrow)].value = str(y['height'])
            # wsLabel['Q' + str(maxrow)].value = str(y['length'])
            # wsLabel['R' + str(maxrow)].value = str(y['width'])
            # wsLabel['S' + str(maxrow)].value = str(y['weight'])
            wsLabel['T' + str(maxrow)].value = str(y['Phone'])
            wsLabel['U' + str(maxrow)].value = str(y['Company'])

    else:
        continue


def SatchelQuote(OrderNumber):  ## Concurent futures for getting quotes from couriers

    global lowesttranscourier, next_cp_day, cp_auth, cp_auth_encoded, cp_headers, cp_validate_body, cp_body, cp_item_array, final_freightster_weight, allied_failed, allied_client, allied_account

    try:
        allied_failed

    except:
        allied_failed = ''

    dai_failed= ''
    po_box = ''
    auspostfailed = ''
    sendlefailed = ''
    transdirectfailed = ''
    fastwayfailed = ''
    fastway_data = ''
    couriers_failed = ''
    freightsterfailed = ''
    cbdexpressfailed = ''
    toll_failed = ''
    satchel = 'false'
    srn = netoresults['Satchel'][OrderNumber]['String_OrderID']
    name = netoresults['Satchel'][OrderNumber]['BuyerName']
    email = netoresults['Satchel'][OrderNumber]['Email']
    address1 = netoresults['Satchel'][OrderNumber]['AddressLine1']
    address2 = netoresults['Satchel'][OrderNumber]['AddressLine2']
    city = netoresults['Satchel'][OrderNumber]['City']
    state = netoresults['Satchel'][OrderNumber]['State']
    postcode = netoresults['Satchel'][OrderNumber]['Postcode']
    # itemname = netoresults['Satchel'][OrderNumber]['ItemName']
    # sku = netoresults['Satchel'][OrderNumber]['SKU']
    # quantity = netoresults['Satchel'][OrderNumber]['Quantity']
    totalprice = netoresults['Satchel'][OrderNumber]['TotalPrice']
    postagetype = netoresults['Satchel'][OrderNumber]['PostageType']
    height = str(netoresults['Satchel'][OrderNumber]['height'])
    length = str(netoresults['Satchel'][OrderNumber]['length'])
    width = str(netoresults['Satchel'][OrderNumber]['width'])
    weightvalue = str(netoresults['Satchel'][OrderNumber]['weight'])
    phone = str(netoresults['Satchel'][OrderNumber]['Phone'])
    company = str(netoresults['Satchel'][OrderNumber]['Company'])
    item_string_counter = netoresults['Satchel'][OrderNumber]['item_count_string']

    try:

        split_name = name.split(' ')

        first_name = split_name[0]

        last_name = ''
        for x in range(1, len(split_name)):
            last_name = f'{last_name} {split_name[x]}'

        last_name = last_name.strip()
    except:
        first_name = name
        last_name = ''

    ########################ALLIED EXPRESS AUTHENTICATION

    # today = datetime.date.today()
    # if today.isoweekday() in set((6, 7)):
    #     today += datetime.timedelta(days=today.isoweekday() % 5)
    # next_day_allied = str(today.day) + '/' + str(today.month) + '/' + str(today.year) + " 10:00:00"
    #
    # history = HistoryPlugin()
    # session = Session()
    # transport = Transport(session=session)
    # wsdl = 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS?wsdl'
    #
    # try:
    #     allied_client = zeep.Client(wsdl=wsdl, transport=transport, plugins=[history])
    #
    #     allied_client.transport.session.proxies = {
    #         # Utilize for all http/https connections
    #         'http': 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS', }
    #     allied_account = allied_client.service.getAccountDefaults('755cf13abb3934695f03bd4a75cfbca7', "SCAMUS", "VIC",
    #                                                               "AOE")
    #
    # except:
    #     allied_failed = 'true'

    #####GET GOOGLE ADDRESS BELOW, COLLAPSED BECAUSE CODE  IS JUST REPEATED FUNCTION

    try:
        if 'ebay' in address1:
            address = address2 + " " + city + " " + state + " " + postcode

            # get_google_address()

            dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
            qstr = urllib.parse.urlencode(dict1)
            URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
            URL = URL + qstr
            response = urllib.request.urlopen(URL)
            data = json.load(response)
            # pprint.pprint(data['candidates'])
            placeid = data['candidates'][0]['place_id']
            payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
            r = requests.get(
                'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
            # pprint.pprint(r.json())
            data = r.json()
            #  pprint.pprint(data['result']['address_components'])
            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'locality':
                    suburb = data['result']['address_components'][t]['long_name']
                    break

                else:
                    continue

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'postal_code':
                    postal_code = data['result']['address_components'][t]['long_name']
                    break

                else:
                    continue

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                    google_state = data['result']['address_components'][t]['short_name']
                    break

                else:
                    continue

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'subpremise':
                    google_street0 = data['result']['address_components'][t]['short_name']
                    break
                else:
                    google_street0 = ''

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'street_number':
                    google_street1 = data['result']['address_components'][t]['short_name']
                    break

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'route':
                    google_street2 = data['result']['address_components'][t]['long_name']
                    break

            if google_street0 == '':

                backup_google_street_address = google_street1 + ' ' + google_street2

            else:

                backup_google_street_address = google_street0 + ' ' + google_street1 + ' ' + google_street2

            postcode = postal_code
            city = suburb
            state = google_state

        else:
            address = address1 + " " + address2 + " " + city + " " + state + " " + postcode

            # get_google_address()

            dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
            qstr = urllib.parse.urlencode(dict1)
            URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
            URL = URL + qstr
            response = urllib.request.urlopen(URL)
            data = json.load(response)
            # pprint.pprint(data['candidates'])
            placeid = data['candidates'][0]['place_id']
            payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
            r = requests.get(
                'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
            #    pprint.pprint(r.json())
            data = r.json()
            #  pprint.pprint(data['result']['address_components'])
            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'locality':
                    suburb = data['result']['address_components'][t]['long_name']
                    break

                else:
                    continue

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'postal_code':
                    postal_code = data['result']['address_components'][t]['long_name']
                    break

                else:
                    continue

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                    google_state = data['result']['address_components'][t]['short_name']
                    break

                else:
                    continue

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'subpremise':
                    google_street0 = data['result']['address_components'][t]['short_name']
                    break
                else:
                    google_street0 = ''

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'street_number':
                    google_street1 = data['result']['address_components'][t]['short_name']
                    break

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'route':
                    google_street2 = data['result']['address_components'][t]['long_name']
                    break

            if google_street0 == '':

                backup_google_street_address = google_street1 + ' ' + google_street2

            else:

                backup_google_street_address = google_street0 + ' ' + google_street1 + ' ' + google_street2

            postcode = postal_code
            city = suburb
            state = google_state

            postcode = postal_code
            city = suburb
            state = google_state



    except (IndexError, NameError):
        suburb = city
        pass

    if 'po box' in address1.lower() or 'po box' in address2.lower() or 'parcel locker' in address1.lower() or 'parcel locker' in address2.lower() or 'p.o' in address1.lower() or 'p.o' in address2.lower() or 'parcel collect' in address1.lower() or 'parcel collect' in address2.lower() or 'pobox' in address1.lower() or 'pobox' in address2.lower() or 'locker' in address1.lower() or 'locker' in address2.lower() or 'collect' in address1.lower() or 'collect' in address2.lower() or 'parcel' in address1.lower() or 'parcel' in address2.lower() or 'pmb' in address1.lower() or 'pmb' in address2.lower() or 'p/o' in address1.lower() or 'p/o' in address2.lower() or 'post office box' in address1.lower() or 'post office box' in address2.lower() or 'lpo' in address1.lower() or 'lpo' in address2.lower() or 'post box' in address1.lower() or 'post box' in address2.lower():
        sendlefailed = 'true'
        transdirectfailed = 'true'
        fastwayfailed = 'true'
        couriers_failed = 'true'
        freightsterfailed = 'true'
        cbdexpressfailed = 'true'
        allied_failed = 'true'
        toll_failed = 'true'
        dai_failed = 'true'
        po_box = 'true'

    volumevalue = (float(length) * float(width) * float(height)) * 0.000001

    data = {
        "length": float(length),
        "width": float(width),
        "height": float(height)
    }

    r = requests.post(base_url + '/api/utils/calc-cubic', headers=Fastway_Headers, json=data)

    try:

        response = r.text
        response = json.loads(response)
        if 'errors' in response:
            fastwayfailed = 'true'

            cubicweight = round(((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250, 2)


        else:

            cubicweight = response['data']['cubicWeight']

    except:
        cubicweight = round(((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250, 2)

    if ((float(length) * float(width) * float(height)) / 4000) > 25 or float(
            weightvalue) > 25 or float(length) > 120 or float(width) > 120 or float(height) > 120:
        cbdexpressfailed = 'true'

    if float(cubicweight) > 40 or float(weightvalue) > 25:
        fastwayfailed = 'true'

    if float(length) > 105 or float(width) > 105 or float(height) > 105 or float(weightvalue) > 22:
        auspostfailed = 'true'
        dai_failed = 'true'

    if float(length) > 105 or float(width) > 105 or float(height) > 105 or float(weightvalue) > 20:
        freightsterfailed = 'true'

    if (float(length) * float(width) * float(height)) / 1000 > 100 or float(weightvalue) > 25 or float(
            length) > 120 or float(width) > 120 or float(height) > 120:
        sendlefailed = 'true'

    if float(cubicweight) > 40 or float(length) > 180 or float(width) > 180 or float(height) > 180:
        couriers_failed = 'true'

    try:

        if auspostfailed == 'true':
            auspostprice = 10000


        else:

            headers = {'Content-Type': 'application/json', 'Accept': 'application/json',
                       'Account-Number': accountnumber,
                       "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
                       "Accept-Encoding": "*",
                       "Connection": "keep-alive"}

            payload = {"shipments": [{
                "from": {
                    "suburb": "FOOTSCRAY",
                    "state": "VIC",
                    "postcode": "3011"
                },
                "to": {
                    "suburb": str(city),
                    "state": str(state),
                    "postcode": str(postcode)
                },
                "items": [
                    {
                        "product_id": "3D55",
                        "length": str(length),
                        "height": str(height),
                        "width": str(width),
                        "weight": str(weightvalue)}]}]}

            try:
                r = requests.post('https://digitalapi.auspost.com.au/shipping/v1/prices/shipments', headers=headers,
                                  auth=HTTPBasicAuth(username, secret), json=payload)
                response = r.text
                response = json.loads(response)
                #  pprint.pprint(response)
                auspostprice = response['shipments'][0]['shipment_summary']['total_cost']

                # if str(state) == 'WA':
                #     auspostprice = str(float(auspostprice) *1.4)
            except json.decoder.JSONDecodeError:
                auspostprice = 1000

    except KeyError:
        auspostprice = 1000

    #    print('Aus Post price: ' + str(auspostprice))

    sendlefailed = 'true'

    if sendlefailed == 'true' or float(totalprice) > 200:
        sendleprice = 1000
        pass

    else:

        data = {
            "pickup_suburb": "Maidstone",
            "pickup_postcode": "3012",
            "delivery_suburb": city,
            "delivery_postcode": str(postcode),
            "weight_value": str(weightvalue),
            'weight_units': 'kg',
            'volume_value': str(volumevalue),
            'volume_units': 'm3'}

        r = requests.get('https://api.sendle.com/api/quote',
                         auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'), json=data)
        sendleresponse = json.loads(r.text)
        #     pprint.pprint(sendleresponse)

        try:
            sendleprice = sendleresponse[0]['quote']['gross']['amount']

        except (KeyError, ValueError):
            sendleprice = 1000

    if dai_failed == 'true' or float(totalprice) > 200:
        dai_price = 1000
        pass

    else:
        dai_volume_value = ((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250
        # if float(dai_volume_value) > float(weightvalue):
        #     final_dai_weight = dai_volume_value
        # elif float(weightvalue) > dai_volume_value:
        #     final_dai_weight = weightvalue
        # else:
        #     final_dai_weight = weightvalue
        final_dai_weight = weightvalue
        final_dai_weight = round(float(final_dai_weight), 2)

        dai_postcode_wb = openpyxl.load_workbook(rf"\\SERVER\Project Folder\Python\Courier Info\Dai_Postcodes.xlsx")

        dai_postcode_sheet = dai_postcode_wb['Sheet']

        dai_zone_sheet1 = ''

        dai_rates_wb = openpyxl.load_workbook(rf"\\SERVER\Project Folder\Python\Courier Info\Dai_Rates.xlsx")

        dai_rates_sheet = dai_rates_wb['Sheet']

        for x in range(2,
                       dai_postcode_sheet.max_row + 1):  ###ITERATING THROUGH THE DAIPOST POSTCODE SHEET TO FIND REGION

            sheet_postcode = str(dai_postcode_sheet['A' + str(x)].value)

            if str(len(sheet_postcode)) == '3':
                sheet_postcode = f'0{sheet_postcode}'

            if postcode == sheet_postcode:
                dai_zone_sheet1 = str(dai_postcode_sheet['B' + str(x)].value)

                for xx in range(2, dai_rates_sheet.max_row + 1):

                    dai_zone_sheet2 = str(dai_rates_sheet['A' + str(xx)].value)

                    if dai_zone_sheet1.lower() == dai_zone_sheet2.lower():  ###Two zones match

                        if final_dai_weight <= 0.5:
                            dai_price = float(dai_rates_sheet['B' + str(xx)].value) * 1.1

                        elif 0.501 <= final_dai_weight <= 1.00:
                            dai_price = float(dai_rates_sheet['C' + str(xx)].value) * 1.1

                        elif 1.01 <= final_dai_weight <= 1.5:
                            dai_price = float(dai_rates_sheet['D' + str(xx)].value) * 1.1

                        elif 1.501 <= final_dai_weight <= 2.0:
                            dai_price = float(dai_rates_sheet['D' + str(xx)].value) * 1.1

                        elif 2.01 <= final_dai_weight <= 2.5:
                            dai_price = float(dai_rates_sheet['E' + str(xx)].value) * 1.1

                        elif 2.501 <= final_dai_weight <= 3.0:
                            dai_price = float(dai_rates_sheet['E' + str(xx)].value) * 1.1

                        elif 3.01 <= final_dai_weight <= 3.50:
                            dai_price = float(dai_rates_sheet['F' + str(xx)].value) * 1.1

                        elif 3.501 <= final_dai_weight <= 4.0:
                            dai_price = float(dai_rates_sheet['F' + str(xx)].value) * 1.1

                        elif 4.01 <= final_dai_weight <= 4.50:
                            dai_price = float(dai_rates_sheet['G' + str(xx)].value) * 1.1

                        elif 4.501 <= final_dai_weight <= 5.00:
                            dai_price = float(dai_rates_sheet['G' + str(xx)].value) * 1.1

                        elif 5.01 <= final_dai_weight <= 7.0:
                            dai_price = float(dai_rates_sheet['H' + str(xx)].value) * 1.1

                        elif 7.01 <= final_dai_weight <= 10.0:
                            dai_price = float(dai_rates_sheet['I' + str(xx)].value) * 1.1

                        elif 10.01 <= final_dai_weight <= 15.00:
                            dai_price = float(dai_rates_sheet['J' + str(xx)].value) * 1.1

                        elif 15.01 <= final_dai_weight <= 22.00:
                            dai_price = float(dai_rates_sheet['K' + str(xx)].value) * 1.1

                        else:
                            dai_price = 1000

                        if po_box == 'true':
                            dai_price += 0.8

                        dai_price = round(dai_price, 2)

    if freightsterfailed == 'true' or float(totalprice) > 200:
        freightster_price = 1000
        pass

    else:

        try:

            for x in range(2, sheet.max_row + 1):
                if sheet['A' + str(x)].value.lower() == city.lower():
                    if int(sheet['B' + str(x)].value) == int(postcode):
                        freightster_pricing_zone = sheet['D' + str(x)].value

            freightster_volume_value = ((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250

            # print(f'l : {length}')
            # print(f'w : {width}')
            # print(f'l : {height}')
            if float(freightster_volume_value) > float(weightvalue):
                final_freightster_weight = freightster_volume_value
            elif float(weightvalue) > freightster_volume_value:
                final_freightster_weight = weightvalue

            if float(length) <= 4.0 and float(width) <= 23.0 and float(height) <= 23.0 and float(weightvalue) <= 0.25:
                final_freightster_weight = 0.250

            elif float(length) <= 18.0 and float(width) <= 23.0 and float(height) <= 4.0 and float(weightvalue) <= 0.25:
                final_freightster_weight = 0.250

            elif float(weightvalue) <= 0.25 and freightster_volume_value < 0.331:
                final_freightster_weight = 0.250

            elif float(length) <= 18.0 and float(width) <= 23.0 and float(height) <= 4.0 and float(weightvalue) <= 0.5:
                final_freightster_weight = 0.500

            if freightster_pricing_zone == 'ACT' or freightster_pricing_zone == 'CBR' or freightster_pricing_zone == 'NSW':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.60

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 5.5

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 6.5

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 7.5

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 8.5

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 22.5

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 28.5

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'ADL':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.60

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 5

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 5.5

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 6.4

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 7.5

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 19

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 28

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'BNE':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.60

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 5

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 5

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 6.25

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 7.3

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 18

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 27

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'MEL':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 3.75

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 4.2

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 4.5

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 4.85

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 5.5

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 7

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 7.5

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'PER':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.6

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 5.6

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 6.45

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 8.15

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 8.35

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 22

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 30.5

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'QLD':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.6

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 5.5

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 7

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 8

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 9

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 22.5

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 28.5

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'SA':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.6

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 5.5

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 5

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 8

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 9

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 22.5

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 28.5

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'SYD':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.6

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 4.6

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 5

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 5.95

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 7.15

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 17

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 18

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'VIC':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.6

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 4.75

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 4.95

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 5.85

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 8.5

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 18.5

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 20.9

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'WA':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.6

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 5.5

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 7.12

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 8.26

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 10.26

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 25.5

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 31.5

                else:
                    freightster_price = 1000

            else:
                freightster_price = 1000

            payload = {"order": {"serviceCode": 12,
                                 "consignee": {"company": company,
                                               "name": name,
                                               "address1": address1,
                                               "address2": address2,
                                               "city": city,
                                               "postcode": postcode,
                                               "state": state,
                                               "phone": phone,
                                               "email": email},
                                 "sender": {"name": "Kyal Scarlett",
                                            "address1": "286-288 Ballarat Rd",
                                            "address2": "",
                                            "city": "Footscray",
                                            "postcode": "3011",
                                            "state": "VIC",
                                            "phone": "0382563460",
                                            "email": "kyal@scarlettmusic.com.au"},
                                 "shipment": {"reference": OrderID,
                                              "description": OrderID,
                                              "weight": str(final_freightster_weight)}}}

            while True:

                r = requests.post('https://freightster.com.au/api/v1/shippingAPI/create', json=payload,
                                  headers=freightster_headers)
                if r.status_code == 429:
                    continue
                else:
                    break
            freightster_response = json.loads(r.text)
            freightster_price = freightster_price * 1.1  #####Adding GST
            if 'NEX' in freightster_response['response_data']['tracking_number']:
                freightster_price = 1000

            if freightster_response['status'] is False:
                freightster_price = 1000

        except:
            freightster_price = 1000

    #  print('Sendle price: ' + str(sendleprice))

    if toll_failed == 'true':
        tollprice = 1000

    else:
        toll_item_array = []
        additional_costs = 0

        # for _ in range(int(all_items[6])):

        toll_item_array.append({"Commodity": {"CommodityCode": "Z",
                                              "CommodityDescription": "ALL FREIGHT"},
                                "ShipmentItemTotals": {
                                    "ShipmentItemCount": '1'},
                                "Dimensions": {
                                    "Width": str(math.ceil(float(width))),
                                    "Length": str(math.ceil(float(length))),
                                    "Height": str(math.ceil(float(height))),
                                    "Volume": str(volumevalue),
                                    "Weight": str(weightvalue)
                                }})

        package = [math.ceil(float(weight)), math.ceil(float(length)),
                   math.ceil(float(height))]
        sorted_package = sorted(package)

        if float(weightvalue) > 35:
            additional_costs += 77.55

        if float(weightvalue) > 35 or sorted_package[0] > 180 or sorted_package[1] > 180 or sorted_package[
            2] > 180 or float(volumevalue) > 0.7:
            additional_costs += 50

        elif float(weightvalue) > 30 or sorted_package[0] > 60 or sorted_package[1] > 80 or \
                sorted_package[2] > 120 or float(volumevalue) > 0.7:
            additional_costs += 12

        if float(totalprice) < 500:
            additional_costs += 6.95
        else:
            additional_costs += float(totalprice) * 0.02

        message_identifier = str(uuid.uuid4())
        now = datetime.datetime.now()
        current_time = f"{now.strftime('%Y-%m-%d')}T{now.strftime('%H:%M:%S')}+10:00"

        next_day = next_business_day()
        next_toll_day = f"{next_day.strftime('%Y-%m-%d')}T09:00:00+10:00"

        environment = 'PRD'  #######################Will have to change this in prod

        headers = {'Content-Type': 'application/json', 'Accept': 'application/json'}

        test_rate_enquiry_headers = {
            "MessageVersion": "3.1",
            "MessageIdentifier": message_identifier,
            "CreateTimestamp": current_time,
            "DocumentType": "RateEnquiry",
            "Environment": environment,
            "MessageSender": "SCARLETTMUSIC",
            "MessageReceiver": "TOLL",
            'SourceSystemCode': 'SCAR'
        }
        # rate_enquiry_test_url = "https://api-uat.teamglobalexp.com:6930/gateway/TollMessageRateEnquiryRestService/1.0/tom/rateEnquiry"
        rate_enquiry_prod_url = "https://api.teamglobalexp.com:6930/gateway/TollMessageRateEnquiryRestService/1.0/tom/rateEnquiry"
        price = totalprice
        payload = {
            "Request": {
                "BusinessID": "IPEC",
                "SystemFields": {"PickupDateTime": next_toll_day},
                "ShipmentService": {
                    "ServiceCode": "X",
                    "ServiceDescription": "Road Express"
                },
                "ShipmentFlags": {
                    "ExtraServiceFlag": "true"},
                "ShipmentFinancials": {
                    "ExtraServicesAmount": {
                        "Currency": "AUD",
                        "Value": str(round(float(totalprice)))}},
                "BillToParty": {
                    "AccountCode": "80119621"
                },
                "ConsignorParty": {
                    "PhysicalAddress": {
                        "Suburb": "FOOTSCRAY",
                        "StateCode": "VIC",
                        "PostalCode": "3011",
                        "CountryCode": "AU"
                    }
                },
                "ConsigneeParty": {
                    "PhysicalAddress": {
                        "Suburb": city,
                        "StateCode": state,
                        "PostalCode": str(postcode),
                        "CountryCode": "AU"
                    }
                },
                "ShipmentItems": {
                    "ShipmentItem": toll_item_array
                }
            }
        }

        toll_message = {"@version": "3.1",
                        "@encoding": "utf-8",
                        "TollMessage": {"Header": test_rate_enquiry_headers,
                                        "RateEnquiry": payload}}

        attempts = 0

        while attempts < 6:

            r = requests.post(url=rate_enquiry_prod_url, auth=('accounts@scarlettmusic.com.au', 't2TrAPsNTB'),
                              json=toll_message, headers=headers, timeout=100)

            response = r.text

            try:

                response = json.loads(response)

                if 'timeout' in response['TollMessage']['ErrorMessages']['ErrorMessage'][0]['ErrorMessage'].lower():
                    message_identifier = str(uuid.uuid4())
                    attempts += 1
                    continue

                else:
                    courierprice = 1000
                    break

            except:
                break

        try:
            tollprice = float(
                response['TollMessage']['RateEnquiry']['Response']['TotalChargeAmount']['Value']) + additional_costs

            # if 'wa' in state.lower():
            #     tollprice = tollprice * 1.15  # Remove once disaster levy has been lifted

            # print('Toll price: ' + str(tollprice))

        except:
            tollprice = 1000

    #allied_failed = 'true'

    if allied_failed == 'true':
        alliedprice = 1000
        pass

    else:
        pickup_address = {'address1': "286-288 Ballarat Rd",
                          'address2': "",
                          'country': "Australia",
                          'postCode': '3011',
                          'state': 'VIC',
                          'suburb': 'Footscray'}

        reciever_address = {'address1': address1,
                            'address2': address2,
                            'country': "Australia",
                            'postCode': postcode,
                            'state': state,
                            'suburb': city}

        Jobstop_pickupstop = {'companyName': 'Scarlett Music',
                              'contact': 'Kyal Scarlett',
                              'emailAddress': 'info@scarlettmusic.com.au',
                              'geographicAddress': pickup_address,
                              'phoneNumber': '03 9318 5751',
                              'stopNumber': 1,
                              'stopType': 'P'}

        Jobstop_deliverystop = {'companyName': company,
                                'contact': name,
                                'emailAddress': email,
                                'geographicAddress': reciever_address,
                                'phoneNumber': phone,
                                'stopNumber': 2,
                                'stopType': 'D'}

        Jobstop_final = [Jobstop_pickupstop, Jobstop_deliverystop]

        allied_item_array = []
        allied_total_volume = 0
        allied_total_weight = 0
        allied_total_items = 0
        #   multi_items_to_ship.append((weightvalue, length, height, width,volumevalue, cubicweight, quantity, satchel))

        allied_item_array.append({'dangerous': 'false',
                                  'height': str(height),
                                  'itemCount': str(1),
                                  'length': str(length),
                                  'volume': str(volumevalue),
                                  'weight': str(weightvalue),
                                  'width': str(width)})

        allied_total_volume = str(volumevalue)
        allied_total_weight = str(weightvalue)
        allied_total_items = '1'

        # item = {'dangerous': 'false',
        #         'height': float(height),
        #         'itemCount': 1,
        #         'length': float(length),
        #         'volume': float(volumevalue),
        #         'weight': float(weightvalue),
        #         'width': float(width)}

        ##IF MULTIPLE ITEMS, easily put into dic, item = [{item1, item2}]

        Service_Level = 'R'  # Overnight Express

        pickup_instructions = "The music shop, open 9am-6pm. Best parking is at The Palms across the road."
        reference_number = item_string_counter
        booked_by = 'Kyal Scarlett'

        job_number = int(re.sub("[^0-9]", "", reference_number))

        Job = {'account': allied_account,
               'cubicWeight': cubicweight,
               'Docket': "SCM",
               'instructions': pickup_instructions,
               'cubedItems': allied_item_array,
               'itemCount': allied_total_items,
               'weight': allied_total_weight,
               'volume': allied_total_volume,
               'items': allied_item_array,
               'jobStops': Jobstop_final,
               'serviceLevel': Service_Level,
               'referenceNumbers': reference_number,
               'bookedBy': booked_by,
               'readyDate': today,
               'jobNumber': job_number,
               'vehicle': {'vehicleID': 1}}

        JobIDs = {'jobIds': job_number}

        try:

            Job = allied_client.service.validateBooking('755cf13abb3934695f03bd4a75cfbca7', Job)

            job_price = allied_client.service.calculatePrice('755cf13abb3934695f03bd4a75cfbca7', Job)
            alliedprice = job_price['totalCharge']
            alliedprice = round((float(alliedprice) * 1.269) * 1.1, 2)
            if alliedprice == 0.0:
                alliedprice = 1000
            # print('Allied price: $' + str(alliedprice))

        except zeep.exceptions.Fault:
            alliedprice = 1000

    if 'vic' not in state.lower():
        cbdexpressfailed = 'true'
    cbdexpressfailed = 'true'

    if cbdexpressfailed == 'true':
        cbdprice = 1000

    else:

        cbdprice = 1000

        cbd_workbook = openpyxl.load_workbook(
            rf'\\SERVER\Project Folder\Python\Courier Info\CBD_Express_Areas.xlsx')

        cbd_sheet = cbd_workbook['Sheet']
        courierprice = 1000
        for xx in range(2, cbd_sheet.max_row + 1):
            cbd_suburb_check = cbd_sheet['A' + str(xx)].value.strip().lower()
            if cbd_suburb_check == city.lower().strip():
                cbdprice = float(cbd_sheet['C' + str(xx)].value)

    if float(weightvalue) > 25:
        tailgate = 'true'
    else:
        tailgate = 'false'

    if transdirectfailed == 'true':
        transdirectprice = 1000
        pass

    else:

        if email == '':
            email = 'info@scarlettmusic.com.au'

        if phone is None:
            phone = '0417557472'

        phone = phone.replace(' ', '')

        try:
            int(phone)

        except ValueError:

            phone = '0417557472'

        headers = {'Api-key': '74a242a1f4b8ba6ea5e011a80bdd2732', 'Content-Type': 'application/json',
                   "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
                   "Accept-Encoding": "*",
                   "Connection": "keep-alive"}

        try:

            payload = {
                "declared_value": "0",
                "referrer": "API",
                "requesting_site": "www.scarlettmusic.com.au",
                "tailgate_pickup": "false",
                "tailgate_delivery": str(tailgate),
                "items": [
                    {
                        "weight": str(weightvalue),
                        "height": str(height),
                        "width": str(width),
                        "length": str(length),
                        "quantity": 1,
                        "description": "carton"
                    }
                ],
                "sender": {
                    "address": "288 Ballarat Rd",
                    "company_name": "Scarlett Music",
                    "email": "info@scarlettmusic.com.au",
                    "name": "Lorelle Scarlett",
                    "postcode": "3011",
                    "phone": "0417557472",
                    "state": "VIC",
                    "suburb": "FOOTSCRAY",
                    "type": "business",
                    "country": "AU"
                },
                "receiver": {
                    "address": str(address1) + ', ' + str(address2),
                    "company_name": str(company),
                    "email": str(email),
                    "name": str(name),
                    "postcode": str(postcode),
                    "phone": int(phone),
                    "state": str(state),
                    "suburb": str(city),
                    "type": "residential",
                    "country": "AU"
                }
            }
            try:
                r = requests.post('https://www.transdirect.com.au/api/bookings/v4',
                                  auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'), headers=headers,
                                  json=payload)

                ##Mock API https://private-anon-2b516758f2-transdirectapiv4.apiary-mock.com/api/bookings/v4
                ###Legit API https://www.transdirect.com.au/api/bookings/v4

                response = r.text
                response = json.loads(response)

                if 'errors' in response:
                    try:

                        if 'ebay' in address1:
                            address = address2 + " " + city + " " + state + " " + postcode

                            # get_google_address()

                            dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                            qstr = urllib.parse.urlencode(dict1)
                            URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                            URL = URL + qstr
                            response = urllib.request.urlopen(URL)
                            data = json.load(response)
                            # pprint.pprint(data['candidates'])
                            placeid = data['candidates'][0]['place_id']
                            payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                            r = requests.get(
                                'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                            #  pprint.pprint(r.json())
                            data = r.json()
                            #   pprint.pprint(data['result']['address_components'])
                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'locality':
                                    suburb = data['result']['address_components'][t]['long_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'postal_code':
                                    postal_code = data['result']['address_components'][t]['long_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                                    google_state = data['result']['address_components'][t]['short_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'subpremise':
                                    google_street0 = data['result']['address_components'][t]['short_name']
                                    break
                                else:
                                    google_street0 = ''

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'street_number':
                                    google_street1 = data['result']['address_components'][t]['short_name']
                                    break

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'route':
                                    google_street2 = data['result']['address_components'][t]['long_name']
                                    break

                            if google_street0 == '':

                                backup_google_street_address = google_street1 + ' ' + google_street2

                            else:

                                backup_google_street_address = google_street0 + ' ' + google_street1 + ' ' + google_street2

                            postcode = postal_code
                            city = suburb
                            state = google_state

                        else:
                            address = address1 + " " + address2 + " " + city + " " + state + " " + postcode

                            # get_google_address()

                            dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                            qstr = urllib.parse.urlencode(dict1)
                            URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                            URL = URL + qstr
                            response = urllib.request.urlopen(URL)
                            data = json.load(response)
                            # pprint.pprint(data['candidates'])
                            placeid = data['candidates'][0]['place_id']
                            payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                            r = requests.get(
                                'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                            #    pprint.pprint(r.json())
                            data = r.json()
                            #  pprint.pprint(data['result']['address_components'])
                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'locality':
                                    suburb = data['result']['address_components'][t]['long_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'postal_code':
                                    postal_code = data['result']['address_components'][t]['long_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                                    google_state = data['result']['address_components'][t]['short_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'subpremise':
                                    google_street0 = data['result']['address_components'][t]['short_name']
                                    break
                                else:
                                    google_street0 = ''

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'street_number':
                                    google_street1 = data['result']['address_components'][t]['short_name']
                                    break

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'route':
                                    google_street2 = data['result']['address_components'][t]['long_name']
                                    break

                            if google_street0 == '':

                                backup_google_street_address = google_street1 + ' ' + google_street2

                            else:

                                backup_google_street_address = google_street0 + ' ' + google_street1 + ' ' + google_street2

                            postcode = postal_code
                            city = suburb
                            state = google_state

                        payload = {
                            "declared_value": "0",
                            "referrer": "API",
                            "requesting_site": "www.scarlettmusic.com.au",
                            "tailgate_pickup": "false",
                            "tailgate_delivery": str(tailgate),
                            "items": [
                                {
                                    "weight": str(weightvalue),
                                    "height": str(height),
                                    "width": str(width),
                                    "length": str(length),
                                    "quantity": 1,
                                    "description": "carton"
                                }
                            ],
                            "sender": {
                                "address": "288 Ballarat Rd",
                                "company_name": "Scarlett Music",
                                "email": "info@scarlettmusic.com.au",
                                "name": "Lorelle Scarlett",
                                "postcode": "3011",
                                "phone": "0417557472",
                                "state": "VIC",
                                "suburb": "FOOTSCRAY",
                                "type": "business",
                                "country": "AU"
                            },
                            "receiver": {
                                "address": str(backup_google_street_address),
                                "company_name": str(company),
                                "email": str(email),
                                "name": str(name),
                                "postcode": str(postcode),
                                "phone": int(phone),
                                "state": str(state),
                                "suburb": str(city),
                                "type": "residential",
                                "country": "AU"
                            }
                        }

                        try:
                            r = requests.post('https://www.transdirect.com.au/api/bookings/v4',
                                              auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'),
                                              headers=headers,
                                              json=payload)

                            ##Mock API https://private-anon-2b516758f2-transdirectapiv4.apiary-mock.com/api/bookings/v4
                            ###Legit API https://www.transdirect.com.au/api/bookings/v4

                            response = r.text
                            response = json.loads(response)

                            if 'id' in response:
                                google = 'true'
                        # pprint.pprint(response)

                        except json.decoder.JSONDecodeError:
                            transdirectprice = 1000
                    except:
                        transdirectprice = 1000
            # pprint.pprint(response)

            except json.decoder.JSONDecodeError:

                transdirectprice = 1000

            id = response['id']

            quotelen = len(response['quotes'])
            # print('Amount of quotes: ' + str(quotelen))
            namesofquotes = list(response['quotes'])
            # print(namesofquotes)

            quotes = {}

            for x in range(quotelen):
                quotes[str(namesofquotes[x])] = str(response['quotes'][namesofquotes[x]]['total'])

            # print(quotes)

            quotes['couriers_please_multi_21'] = \
                response['quotes']['couriers_please_domestic_proirity_authority']['tiers'][1]['total']
            quotes['fastway_multi_7'] = response['quotes']['fastway']['tiers'][1]['total']

            if float(totalprice) > 250:
                quotes['couriers_please_domestic_proirity_authority'] = '1000'

            intquotes = dict(
                (k, float(v)) for k, v in quotes.items())  ### Converting all values into FLOAT and then into INT
            lowesttranscourier = min(intquotes, key=intquotes.get)

            transdirectprice = quotes[lowesttranscourier]



        except (KeyError, TypeError, IndexError):

            try:

                if 'ebay' in address1:
                    address = address2 + " " + city + " " + state + " " + postcode

                    # get_google_address()

                    dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    qstr = urllib.parse.urlencode(dict1)
                    URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                    URL = URL + qstr
                    response = urllib.request.urlopen(URL)
                    data = json.load(response)
                    # pprint.pprint(data['candidates'])
                    placeid = data['candidates'][0]['place_id']
                    payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    r = requests.get(
                        'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                    #  pprint.pprint(r.json())
                    data = r.json()
                    #   pprint.pprint(data['result']['address_components'])
                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'locality':
                            suburb = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'postal_code':
                            postal_code = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                            google_state = data['result']['address_components'][t]['short_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'subpremise':
                            google_street0 = data['result']['address_components'][t]['short_name']
                            break
                        else:
                            google_street0 = ''

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'street_number':
                            google_street1 = data['result']['address_components'][t]['short_name']
                            break

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'route':
                            google_street2 = data['result']['address_components'][t]['long_name']
                            break

                    if google_street0 == '':

                        backup_google_street_address = google_street1 + ' ' + google_street2

                    else:

                        backup_google_street_address = google_street0 + ' ' + google_street1 + ' ' + google_street2

                    postcode = postal_code
                    city = suburb
                    state = google_state

                else:
                    address = address1 + " " + address2 + " " + city + " " + state + " " + postcode

                    # get_google_address()

                    dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    qstr = urllib.parse.urlencode(dict1)
                    URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                    URL = URL + qstr
                    response = urllib.request.urlopen(URL)
                    data = json.load(response)
                    # pprint.pprint(data['candidates'])
                    placeid = data['candidates'][0]['place_id']
                    payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    r = requests.get(
                        'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                    #    pprint.pprint(r.json())
                    data = r.json()
                    #  pprint.pprint(data['result']['address_components'])
                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'locality':
                            suburb = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'postal_code':
                            postal_code = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                            google_state = data['result']['address_components'][t]['short_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'subpremise':
                            google_street0 = data['result']['address_components'][t]['short_name']
                            break
                        else:
                            google_street0 = ''

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'street_number':
                            google_street1 = data['result']['address_components'][t]['short_name']
                            break

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'route':
                            google_street2 = data['result']['address_components'][t]['long_name']
                            break

                    if google_street0 == '':

                        backup_google_street_address = google_street1 + ' ' + google_street2

                    else:

                        backup_google_street_address = google_street0 + ' ' + google_street1 + ' ' + google_street2

                    postcode = postal_code
                    city = suburb
                    state = google_state

                payload = {
                    "declared_value": "0",
                    "referrer": "API",
                    "requesting_site": "www.scarlettmusic.com.au",
                    "tailgate_pickup": "false",
                    "tailgate_delivery": str(tailgate),
                    "items": [
                        {
                            "weight": str(weightvalue),
                            "height": str(height),
                            "width": str(width),
                            "length": str(length),
                            "quantity": 1,
                            "description": "carton"
                        }
                    ],
                    "sender": {
                        "address": "288 Ballarat Rd",
                        "company_name": "Scarlett Music",
                        "email": "info@scarlettmusic.com.au",
                        "name": "Lorelle Scarlett",
                        "postcode": "3011",
                        "phone": "0417557472",
                        "state": "VIC",
                        "suburb": "FOOTSCRAY",
                        "type": "business",
                        "country": "AU"
                    },
                    "receiver": {
                        "address": str(backup_google_street_address),
                        "company_name": str(company),
                        "email": str(email),
                        "name": str(name),
                        "postcode": str(postcode),
                        "phone": int(phone),
                        "state": str(state),
                        "suburb": str(city),
                        "type": "residential",
                        "country": "AU"
                    }
                }
                try:
                    r = requests.post('https://www.transdirect.com.au/api/bookings/v4',
                                      auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'), headers=headers,
                                      json=payload)

                    ##Mock API https://private-anon-2b516758f2-transdirectapiv4.apiary-mock.com/api/bookings/v4
                    ###Legit API https://www.transdirect.com.au/api/bookings/v4

                    response = r.text
                    response = json.loads(response)
                # pprint.pprint(response)

                except json.decoder.JSONDecodeError:

                    transdirectprice = 1000

                id = response['id']

                quotelen = len(response['quotes'])
                # print('Amount of quotes: ' + str(quotelen))
                namesofquotes = list(response['quotes'])
                # print(namesofquotes)

                quotes = {}

                for x in range(quotelen):
                    quotes[str(namesofquotes[x])] = str(response['quotes'][namesofquotes[x]]['total'])

                # print(quotes)
                quotes['couriers_please_multi_21'] = \
                    response['quotes']['couriers_please_domestic_proirity_authority']['tiers'][1]['total']
                quotes['fastway_multi_7'] = response['quotes']['fastway']['tiers'][1]['total']

                if float(totalprice) > 250:
                    quotes['couriers_please_domestic_proirity_authority'] = '1000'

                intquotes = dict(
                    (k, float(v)) for k, v in quotes.items())  ### Converting all values into FLOAT and then into INT
                lowesttranscourier = min(intquotes, key=intquotes.get)

                transdirectprice = quotes[lowesttranscourier]

            except (KeyError, TypeError, IndexError):

                transdirectprice = 1000

    #       print('Transdirect price: ' + str(transdirectprice))

    if fastwayfailed == 'true':
        fastwayprice = 1000

    else:

        if email == '':
            email = 'info@scarlettmusic.com.au'

        if phone is None:
            phone = '0417557472'

        phone = phone.replace(' ', '')

        try:
            int(phone)

        except ValueError:

            phone = '0417557472'

        ####### INITIAL FASTWAY VALIDATION ######

        if 'ebay' in address1:

            street_address = str(address2)

            data = {
                "streetAddress": street_address,
                "additionalDetails": "",
                "locality": str(city),
                "stateOrProvince": str(state),
                "postalCode": str(postcode),
                "country": "AU",
                "lat": 0,
                "lng": 0,
                "userCreated": False,
            }


        else:

            street_address = str(address1) + ', ' + str(address2)

            data = {
                "streetAddress": street_address,
                "additionalDetails": "",
                "locality": str(city),
                "stateOrProvince": str(state),
                "postalCode": str(postcode),
                "country": "AU",
                "lat": 0,
                "lng": 0,
                "userCreated": False,
            }

            try:

                r = requests.post(base_url + '/api/addresses/validate', headers=Fastway_Headers, json=data)
                response = r.text
                response = json.loads(response)
                # pprint.pprint(response)
            except:
                fastwayprice = 1000

        try:
            country = response['data']['country']
            addressid = response['data']['addressId']
            latitude = response['data']['lat']
            longitude = response['data']['lng']
            suburb = response['data']['locality']
            placeId = response['data']['placeId']
            postcode = response['data']['postalCode']
            state = response['data']['stateOrProvince']
            street_address = response['data']['streetAddress']
            hash = response['data']['hash']

            # package = [float(length), float(width), float(height)]
            # sorted_package = sorted(package)

            package = [float(length), float(width), float(height)]
            sorted_package = sorted(package)

            if float(weightvalue) <= 0.3 and float(sorted_package[0]) <= 3 and float(sorted_package[1]) <= 21 and float(
                    sorted_package[2]) <= 21:
                satchel = '300gm'

            elif float(cubicweight) <= 0.5 and float(weightvalue) <= 0.5:
                satchel = 'A5'

            elif float(cubicweight) <= 1 and float(weightvalue) <= 1:
                satchel = 'A4'

            elif float(cubicweight) <= 3 and float(weightvalue) <= 3:
                satchel = 'A3'

            elif float(cubicweight) <= 5 and float(weightvalue) <= 5:
                satchel = 'A2'

            if satchel == 'false':

                fastway_data = {
                    "To": {
                        "ContactName": name,
                        "BusinessName": company,
                        "PhoneNumber": phone,
                        "Email": email,
                        "Address": {
                            "StreetAddress": street_address,
                            "Locality": suburb,
                            "StateOrProvince": state,
                            "PostalCode": postcode,
                            "Country": country
                        }
                    },
                    "Items": [
                        {
                            "Quantity": 1,
                            "Reference": item_string_counter,
                            "PackageType": "P",
                            "WeightDead": weightvalue,
                            "WeightCubic": cubicweight,
                            "Length": length,
                            "Width": width,
                            "Height": height
                        }
                    ],
                    "ExternalRef1": item_string_counter,
                    "ExternalRef2": item_string_counter,
                }

            else:

                fastway_data = {
                    "To": {
                        "ContactName": name,
                        "BusinessName": company,
                        "PhoneNumber": phone,
                        "Email": email,
                        "Address": {
                            "StreetAddress": street_address,
                            "Locality": suburb,
                            "StateOrProvince": state,
                            "PostalCode": postcode,
                            "Country": country
                        }
                    },
                    "Items": [
                        {
                            "Quantity": 1,
                            "PackageType": "S",
                            "SatchelSize": satchel,
                            "Reference": item_string_counter
                        }
                    ],
                    "ExternalRef1": item_string_counter,
                    "ExternalRef2": item_string_counter,
                }

            r = requests.post(base_url + '/api/consignments/quote', headers=Fastway_Headers, json=fastway_data)
            response = r.text
            response = json.loads(response)

            fastwayprice = response['data']['total']

            # print('Fastway price: ' + str(fastwayprice))

        except:

            try:

                if 'ebay' in address1:
                    address = address2 + " " + city + " " + state + " " + postcode

                    # get_google_address()

                    dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    qstr = urllib.parse.urlencode(dict1)
                    URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                    URL = URL + qstr
                    response = urllib.request.urlopen(URL)
                    data = json.load(response)
                    # pprint.pprint(data['candidates'])
                    placeid = data['candidates'][0]['place_id']
                    payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    r = requests.get(
                        'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                    #  pprint.pprint(r.json())
                    data = r.json()
                    #   pprint.pprint(data['result']['address_components'])
                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'locality':
                            suburb = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'postal_code':
                            postal_code = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                            google_state = data['result']['address_components'][t]['short_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'subpremise':
                            google_street0 = data['result']['address_components'][t]['short_name']
                            break
                        else:
                            google_street0 = ''

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'street_number':
                            google_street1 = data['result']['address_components'][t]['short_name']
                            break

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'route':
                            google_street2 = data['result']['address_components'][t]['long_name']
                            break

                    if google_street0 == '':

                        backup_google_street_address = google_street1 + ' ' + google_street2

                    else:

                        backup_google_street_address = google_street0 + '/' + google_street1 + ' ' + google_street2

                    postcode = postal_code
                    city = suburb
                    state = google_state

                else:
                    address = address1 + " " + address2 + " " + city + " " + state + " " + postcode

                    # get_google_address()

                    dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    qstr = urllib.parse.urlencode(dict1)
                    URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                    URL = URL + qstr
                    response = urllib.request.urlopen(URL)
                    data = json.load(response)
                    # pprint.pprint(data['candidates'])
                    placeid = data['candidates'][0]['place_id']
                    payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    r = requests.get(
                        'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                    #    pprint.pprint(r.json())
                    data = r.json()
                    #  pprint.pprint(data['result']['address_components'])
                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'locality':
                            suburb = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'postal_code':
                            postal_code = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                            google_state = data['result']['address_components'][t]['short_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'subpremise':
                            google_street0 = data['result']['address_components'][t]['short_name']
                            break
                        else:
                            google_street0 = ''

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'street_number':
                            google_street1 = data['result']['address_components'][t]['short_name']
                            break

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'route':
                            google_street2 = data['result']['address_components'][t]['long_name']
                            break

                    if google_street0 == '':

                        backup_google_street_address = google_street1 + ' ' + google_street2

                    else:

                        backup_google_street_address = google_street0 + ' ' + google_street1 + ' ' + google_street2

                    postcode = postal_code
                    city = suburb
                    state = google_state

                data = {
                    "streetAddress": str(backup_google_street_address),
                    "additionalDetails": "",
                    "locality": str(city),
                    "stateOrProvince": str(state),
                    "postalCode": str(postcode),
                    "country": "AU",
                    "lat": 0,
                    "lng": 0,
                    "userCreated": False,
                }

                try:
                    r = requests.post(base_url + '/api/addresses/validate', headers=Fastway_Headers, json=data)
                    response = r.text
                    response = json.loads(response)
                    # pprint.pprint(response)

                    try:
                        country = response['data']['country']
                        addressid = response['data']['addressId']
                        latitude = response['data']['lat']
                        longitude = response['data']['lng']
                        suburb = response['data']['locality']
                        placeId = response['data']['placeId']
                        Postcode = response['data']['postalCode']
                        State = response['data']['stateOrProvince']
                        AddressLine1 = response['data']['streetAddress']
                        hash = response['data']['hash']

                        if satchel == 'false':

                            fastway_data = {
                                "To": {
                                    "ContactName": name,
                                    "BusinessName": company,
                                    "PhoneNumber": phone,
                                    "Email": email,
                                    "Address": {
                                        "StreetAddress": street_address,
                                        "Locality": suburb,
                                        "StateOrProvince": state,
                                        "PostalCode": postcode,
                                        "Country": country
                                    }
                                },
                                "Items": [
                                    {
                                        "Quantity": 1,
                                        "Reference": item_string_counter,
                                        "PackageType": "P",
                                        "WeightDead": weightvalue,
                                        "WeightCubic": cubicweight,
                                        "Length": length,
                                        "Width": width,
                                        "Height": height
                                    }
                                ],
                                "ExternalRef1": item_string_counter,
                                "ExternalRef2": item_string_counter,
                            }

                        else:

                            fastway_data = {
                                "To": {
                                    "ContactName": name,
                                    "BusinessName": company,
                                    "PhoneNumber": phone,
                                    "Email": email,
                                    "Address": {
                                        "StreetAddress": street_address,
                                        "Locality": suburb,
                                        "StateOrProvince": state,
                                        "PostalCode": postcode,
                                        "Country": country
                                    }
                                },
                                "Items": [
                                    {
                                        "Quantity": 1,
                                        "PackageType": "S",
                                        "SatchelSize": satchel
                                    }
                                ],
                                "ExternalRef1": item_string_counter,
                                "ExternalRef2": item_string_counter,
                            }

                        r = requests.post(base_url + '/api/consignments/quote', headers=Fastway_Headers,
                                          json=fastway_data)
                        response = r.text
                        response = json.loads(response)

                        fastwayprice = response['data']['total']

                        # print('Fastway price: ' + str(fastwayprice))

                    except:

                        fastwayprice = 1000

                except:

                    fastwayprice = 1000

            except:
                fastwayprice = 1000
    couriers_failed = 'true'
    if couriers_failed == 'true':
        courierprice = 1000

    else:

        if email == '':
            email = 'info@scarlettmusic.com.au'

        if phone is None:
            phone = '0417557472'

        phone = phone.replace(' ', '')

        try:
            int(phone)

        except ValueError:

            phone = '0417557472'

        cp_body = {
            "fromSuburb": "Footscray",
            "fromPostcode": 3011,
            "toSuburb": suburb,
            "toPostcode": postcode,
            "items": [{
                "length": math.ceil(float(length)),
                "height": math.ceil(float(height)),
                "width": math.ceil(float(width)),
                "physicalWeight": float(weightvalue),
                "quantity": 1
            }]
        }
        additional_costs = 0
        package = [math.ceil(float(length)), math.ceil(float(height)),
                   math.ceil(float(width))]
        sorted_package = sorted(package)
        if sorted_package[0] > 105 or sorted_package[1] > 105 or sorted_package[
            2] > 105:
            additional_costs += 15
        cp_url = 'https://api.couriersplease.com.au/v2/domestic/quote'

        try:

            r = requests.post(cp_url, headers=cp_headers, json=cp_body)

        except:
            courierprice = 1000

        try:
            response = r.text
            response = json.loads(response)
            # pprint.pprint(response)

            quotelen = len(response['data'])
            # print('Amount of quotes: ' + str(quotelen))
            # namesofquotes = list(response['quotes'])
            # print(namesofquotes)

            quotes = {}

            for x in range(quotelen):
                quotes[str(response['data'][x]['RateCardCode'])] = str((float(
                    response['data'][x]['CalculatedFreightCharge']) + float(
                    response['data'][x]['CalculatedFuelCharge'])) * 1.1)

            # pprint.pprint(quotes)

            intquotes = dict(
                (k, float(v)) for k, v in quotes.items())  ### Converting all values into FLOAT and then into INT
            lowest_cp_service = min(intquotes, key=intquotes.get)

            courierprice = float(quotes[lowest_cp_service])*1.025

            courierprice+= additional_costs

            cp_url = 'https://api.couriersplease.com.au/v1/domestic/shipment/validate'
            # print(lowest_cp_service)

            cp_validate_body = {
                "pickupFirstName": "Kyal",
                "pickupLastName": "Scarlett",
                "pickupCompanyName": "Scarlett Music",
                "pickupEmail": "kyal@scarlettmusic.com.au",
                "pickupAddress1": "286-288 Ballarat Rd",
                "pickupAddress2": "",
                "pickupSuburb": "Footscray",
                "pickupState": "VIC",
                "pickupPostcode": "3011",
                "pickupPhone": "0393185751",
                "pickupIsBusiness": "true",
                "destinationFirstName": first_name,
                "destinationLastName": last_name,
                "destinationCompanyName": company,
                "destinationEmail": email,
                "destinationAddress1": address1,
                "destinationAddress2": address2,
                "destinationSuburb": city,
                "destinationState": state,
                "destinationPostcode": postcode,
                "destinationPhone": phone,
                "destinationIsBusiness": "false",
                "contactFirstName": "Kyal",
                "contactLastName": "Scarlett",
                "contactCompanyName": "Scarlett Music",
                "contactEmail": "kyal@scarlettmusic.com.au",
                "contactAddress1": "286-288 Ballarat Rd",
                "contactAddress2": "",
                "contactSuburb": "Footscray",
                "contactState": "VIC",
                "contactPostcode": "3011",
                "contactPhone": "0393185751",
                "contactIsBusiness": "true",
                "referenceNumber": item_string_counter,
                "termsAccepted": "true",
                "dangerousGoods": "false",
                "rateCardId": lowest_cp_service,
                "specialInstruction": "",
                "isATL": "false",
                "readyDateTime": next_cp_day,
                "items": [{
                    "length": math.ceil(float(length)),
                    "height": math.ceil(float(height)),
                    "width": math.ceil(float(width)),
                    "physicalWeight": float(weightvalue),
                    "quantity": 1
                }]}

            r = requests.post(cp_url, headers=cp_headers, json=cp_validate_body)
            response = r.text
            response = json.loads(response)
            # pprint.pprint(response)

            if response['responseCode'] != 'SUCCESS':
                ##create shipment / print label
                cp_price = 1000

        #  print(f"Courier's Please: {courierprice}")
        except:

            try:

                if 'ebay' in address1:
                    address = address2 + " " + city + " " + state + " " + postcode

                    # get_google_address()

                    dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    qstr = urllib.parse.urlencode(dict1)
                    URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                    URL = URL + qstr
                    response = urllib.request.urlopen(URL)
                    data = json.load(response)
                    # pprint.pprint(data['candidates'])
                    placeid = data['candidates'][0]['place_id']
                    payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    r = requests.get(
                        'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                    #  pprint.pprint(r.json())
                    data = r.json()
                    #   pprint.pprint(data['result']['address_components'])
                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'locality':
                            suburb = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'postal_code':
                            postal_code = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][
                            0] == 'administrative_area_level_1':
                            google_state = data['result']['address_components'][t]['short_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'subpremise':
                            google_street0 = data['result']['address_components'][t]['short_name']
                            break
                        else:
                            google_street0 = ''

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'street_number':
                            google_street1 = data['result']['address_components'][t]['short_name']
                            break

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'route':
                            google_street2 = data['result']['address_components'][t]['long_name']
                            break

                    if google_street0 == '':

                        backup_google_street_address = google_street1 + ' ' + google_street2

                    else:

                        backup_google_street_address = google_street0 + '/' + google_street1 + ' ' + google_street2

                    postcode = postal_code
                    city = suburb
                    state = google_state

                else:
                    address = address1 + " " + address2 + " " + city + " " + state + " " + postcode

                    # get_google_address()

                    dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    qstr = urllib.parse.urlencode(dict1)
                    URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                    URL = URL + qstr
                    response = urllib.request.urlopen(URL)
                    data = json.load(response)
                    # pprint.pprint(data['candidates'])
                    placeid = data['candidates'][0]['place_id']
                    payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    r = requests.get(
                        'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                    #    pprint.pprint(r.json())
                    data = r.json()
                    #  pprint.pprint(data['result']['address_components'])
                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'locality':
                            suburb = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'postal_code':
                            postal_code = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][
                            0] == 'administrative_area_level_1':
                            google_state = data['result']['address_components'][t]['short_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'subpremise':
                            google_street0 = data['result']['address_components'][t]['short_name']
                            break
                        else:
                            google_street0 = ''

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'street_number':
                            google_street1 = data['result']['address_components'][t]['short_name']
                            break

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'route':
                            google_street2 = data['result']['address_components'][t]['long_name']
                            break

                    if google_street0 == '':

                        backup_google_street_address = google_street1 + ' ' + google_street2

                    else:

                        backup_google_street_address = google_street0 + '/' + google_street1 + ' ' + google_street2

                    city = suburb
                    state = google_state

                cp_body = {
                    "fromSuburb": "Footscray",
                    "fromPostcode": 3011,
                    "toSuburb": suburb,
                    "toPostcode": postcode,
                    "items": [{
                        "length": math.ceil(float(length)),
                        "height": math.ceil(float(height)),
                        "width": math.ceil(float(width)),
                        "physicalWeight": float(weightvalue),
                        "quantity": 1
                    }]
                }

                lowest_cp_service = ''

                cp_validate_body = {
                    "pickupFirstName": "Kyal",
                    "pickupLastName": "Scarlett",
                    "pickupCompanyName": "Scarlett Music",
                    "pickupEmail": "kyal@scarlettmusic.com.au",
                    "pickupAddress1": "286-288 Ballarat Rd",
                    "pickupAddress2": "",
                    "pickupSuburb": "Footscray",
                    "pickupState": "VIC",
                    "pickupPostcode": "3011",
                    "pickupPhone": "0393185751",
                    "pickupIsBusiness": "true",
                    "destinationFirstName": first_name,
                    "destinationLastName": last_name,
                    "destinationCompanyName": company,
                    "destinationEmail": email,
                    "destinationAddress1": address1,
                    "destinationAddress2": address2,
                    "destinationSuburb": city,
                    "destinationState": state,
                    "destinationPostcode": postcode,
                    "destinationPhone": phone,
                    "destinationIsBusiness": "false",
                    "contactFirstName": "Kyal",
                    "contactLastName": "Scarlett",
                    "contactCompanyName": "Scarlett Music",
                    "contactEmail": "kyal@scarlettmusic.com.au",
                    "contactAddress1": "286-288 Ballarat Rd",
                    "contactAddress2": "",
                    "contactSuburb": "Footscray",
                    "contactState": "VIC",
                    "contactPostcode": "3011",
                    "contactPhone": "0393185751",
                    "contactIsBusiness": "true",
                    "referenceNumber": item_string_counter,
                    "termsAccepted": "true",
                    "dangerousGoods": "false",
                    "rateCardId": lowest_cp_service,
                    "specialInstruction": "",
                    "isATL": "false",
                    "readyDateTime": next_cp_day,
                    "items": [{
                        "length": math.ceil(float(length)),
                        "height": math.ceil(float(height)),
                        "width": math.ceil(float(width)),
                        "physicalWeight": float(weightvalue),
                        "quantity": 1
                    }]}

                additional_costs = 0
                package = [math.ceil(float(length)), math.ceil(float(height)),
                           math.ceil(float(width))]
                sorted_package = sorted(package)
                if sorted_package[0] > 105 or sorted_package[1] > 105 or sorted_package[
                    2] > 105:
                    additional_costs += 15

                cp_url = 'https://api.couriersplease.com.au/v2/domestic/quote'

                try:
                    r = requests.post(cp_url, headers=cp_headers, json=cp_body)

                    response = r.text
                    response = json.loads(response)

                except:
                    courierprice = 1000
                # pprint.pprint(response)

                quotelen = len(response['data'])
                # print('Amount of quotes: ' + str(quotelen))
                # namesofquotes = list(response['quotes'])
                # print(namesofquotes)

                quotes = {}

                for x in range(quotelen):
                    quotes[str(response['data'][x]['RateCardCode'])] = str((float(
                        response['data'][x]['CalculatedFreightCharge']) + float(
                        response['data'][x]['CalculatedFuelCharge'])) * 1.1)

                # pprint.pprint(quotes)

                intquotes = dict(
                    (k, float(v)) for k, v in quotes.items())  ### Converting all values into FLOAT and then into INT
                lowest_cp_service = min(intquotes, key=intquotes.get)

                courierprice = float(quotes[lowest_cp_service])*1.025

                if 'wa' in state.lower():
                    courierprice = courierprice*1.20

                courierprice += additional_costs

                cp_url = 'https://api.couriersplease.com.au/v1/domestic/shipment/validate'
                # print(lowest_cp_service)

                r = requests.post(cp_url, headers=cp_headers, json=cp_validate_body)
                response = r.text
                response = json.loads(response)
                # pprint.pprint(response)

                if response['responseCode'] != 'SUCCESS':
                    ##create shipment / print label
                    cp_price = 1000


            #      print(f"Courier's Please: {courierprice}")
            except:
                courierprice = 1000

    if weightvalue == 0:
        transdirectprice = 1000

    try:
        dai_price
    except:
        dai_price = 1000

    lowestcourier = {'auspostprice': float(auspostprice), 'sendleprice': float(sendleprice),
                     'transdirectprice': float(transdirectprice), 'fastwayprice': float(fastwayprice),
                     'courierprice': float(courierprice), 'freightsterprice': float(freightster_price),
                     'cbdprice': float(cbdprice), 'alliedprice': float(alliedprice), 'tollprice': float(tollprice),
                     'daiprice': float(dai_price)}
    lowestcourier = min(lowestcourier, key=lowestcourier.get)
    print(lowestcourier + ' booked!')

    if lowestcourier == 'sendleprice':
        finalcourier = 'Sendle'

    if lowestcourier == 'auspostprice':
        finalcourier = 'Australia Post'

    if lowestcourier == 'fastwayprice':
        finalcourier = 'Fastway'

    if lowestcourier == 'courierprice':
        finalcourier = 'Couriers Please'

    if lowestcourier == 'freightsterprice':
        finalcourier = 'Freightster'

    if lowestcourier == 'alliedprice':
        finalcourier = 'Allied Express'

    if lowestcourier == 'cbdprice':
        finalcourier = 'CBDExpress'

    if lowestcourier == 'tollprice':
        finalcourier = 'Toll'

    if lowestcourier == 'daiprice':
        finalcourier = 'Dai Post'

    if lowestcourier == 'transdirectprice':
        finalcourier = 'Transdirect'

        lowesttranscourier = lowesttranscourier.replace('_', ' ')
        lowesttranscourier = lowesttranscourier.title()
        lowesttranscourier = lowesttranscourier.replace('Tnt', 'TNT')

    try:
        id
    except NameError:
        id = ''

    try:
        fastway_data
    except NameError:
        fastway_data = 'NA'

    try:
        lowesttranscourier
    except NameError:
        lowesttranscourier = ''

    try:
        Job
    except:
        Job = ''

    try:
        JobIDs
    except:
        JobIDs = ''

    try:
        cp_validate_body
    except NameError:
        cp_validate_body = ''

    return finalcourier, lowesttranscourier, id, fastway_data, cp_validate_body, Job, JobIDs


satchelrange = len(netoresults['Satchel'])

with concurrent.futures.ThreadPoolExecutor(max_workers=None) as executor:
    results = executor.map(SatchelQuote, range(0, satchelrange))

    satchelposition = 0

    for g in results:
        netoresults['Satchel'][satchelposition].update({'Final_Courier': g[0]})

        if g[0] == 'Transdirect':
            netoresults['Satchel'][satchelposition].update({'Final_Transdirect_Courier': g[1]})
            netoresults['Satchel'][satchelposition].update({'Transdirect_ID': g[2]})

        if g[0] == 'Fastway':
            netoresults['Satchel'][satchelposition].update({'fastway_data': g[3]})

        if g[0] == 'Couriers Please':
            netoresults['Satchel'][satchelposition].update({'cp_validate_body': g[4]})

        if g[0] == 'Allied Express':
            netoresults['Satchel'][satchelposition].update({'Job': g[5]})
            netoresults['Satchel'][satchelposition].update({'JobIDs': g[6]})

        satchelposition = satchelposition + 1

# pprint.pprint(netoresults)
Label_Location = []

couriers_please_weight = 0
couriers_please_item_count = 0

for x in netoresults:

    if 'satchel' in x.lower():

        for y in netoresults[x]:  # iterating through each item in 'Minilope

            try:

                if y['Final_Courier'] == 'Fastway':
                    y['Phone'] = y['Phone'].replace('+61', '0')
                    y['Phone'] = y['Phone'].replace(' ', '')

                    fastway_data = y['fastway_data']

                    fastway_attempt = 0

                    while True:

                        try:

                            r = requests.post(base_url + '/api/consignments', headers=Fastway_Headers,
                                              json=fastway_data, timeout=120)
                            response = r.text
                            response = json.loads(response)

                        except:
                            fastway_attempts += 1
                            if fastway_attempts > 5:
                                break

                            continue

                        if 'errors' in response:
                            break

                        else:
                            break
                    # pprint.pprint(response)

                    id = response['data']['conId']
                    tracking_number = response['data']['items'][0]['label']
                    r = requests.get(base_url + '/api/consignments/' + str(id) + '/labels?pageSize=4x6',
                                     headers=Fastway_Headers)
                    response = r.content

                    label_name = label_location + "\\" + str(y['String_OrderID']) + ".pdf"

                    Label_Location.append(label_name)

                    with open(label_name, 'wb') as f:
                        f.write(response)

                    try:
                        y['ItemName'] = y['ItemName'].replace("'", "")
                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        try:
                            cursor = connection.cursor()
                        except:
                            connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                           "5432")
                            cursor = connection.cursor()

                        try:
                            cursor.execute(
                                fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed']}','' , '{y['SKU']}', '', '{y['Quantity']}', '{y['ItemName'][0:50]}', '{y['SKU']}', '{y['OrderID'][0]}', '{y['Final_Courier']}','{tracking_number}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                        except:
                            pass
                    except:

                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        for lines in range(len(y['SKU'])):
                            y['ItemName'][f'Name (Item {str(lines + 1)})'] = y['ItemName'][
                                f'Name (Item {str(lines + 1)})'].replace("'", "")
                            try:
                                cursor = connection.cursor()
                            except:
                                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                               "5432")
                                cursor = connection.cursor()

                            try:
                                cursor.execute(
                                    fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '', '{y['Quantity'][f'Quantity (Item {str(lines + 1)})']}', '{y['ItemName'][f'Name (Item {str(lines + 1)})'][0:200]}', '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '{y['OrderID'][0][0:50]}', '{y['Final_Courier']}','{tracking_number}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")

                            except:
                                pass

                if y['Final_Courier'] == 'Toll':

                    y['Phone'] = y['Phone'].replace('+61', '0')
                    y['Phone'] = y['Phone'].replace(' ', '')
                    y['BuyerName'] = y['BuyerName'].replace("'", "")
                    y['BuyerName'] = y['BuyerName'].replace("’", "")
                    y['BuyerName'] = y['BuyerName'].replace("’", "")


                    volumevalue = (float(y['length']) * float(y['width']) * float(y['height'])) * 0.000001

                    message_identifier = str(uuid.uuid4())  # getting unique string for toll order
                    ###Getting current date / time
                    now = datetime.datetime.now()
                    current_time = f"{now.strftime('%Y-%m-%d')}T{now.strftime('%H:%M:%S')}+10:00"

                    next_day = next_business_day()
                    next_toll_day = f"{next_day.strftime('%Y-%m-%d')}T09:00:00+10:00"

                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()
                    cursor.execute("SELECT MAX(ShipmentID), MAX(SSCC) FROM Toll;")
                    results = cursor.fetchall()
                    pprint.pprint(results)
                    max_shipment_id = int(results[0][0])
                    max_sscc = int(results[0][1])
                    environment = "PRD"

                    manifest_test_url = "https://api-uat.teamglobalexp.com:6930/gateway/TollMessageManifestRestService/1.0/tom/receiveManifest"
                    manifest_prod_url = "https://api.teamglobalexp.com:6930/gateway/TollMessageManifestRestService/1.0/tom/receiveManifest"

                    headers = {'Content-Type': 'application/json', 'Accept': 'application/json'}


                    def round_up_to_nearest_10(num):
                        return math.ceil(num / 10) * 10


                    toll_item_array = []
                    toll_sql_insert = []

                    if y['AddressLine2'] == '':
                        y['AddressLine2'] = y['AddressLine1']

                    total_toll_information = {
                        "ShipmentFinancials": {
                            "ExtraServicesAmount": {
                                "Currency": "AUD",
                                "Value": str(round(float(y['TotalPrice'])))}},
                        "ShipmentID": str(max_shipment_id),
                        "CreateDateTime": current_time,
                        "ConsigneeParty": {
                            "PartyName": y['BuyerName'],
                            "PhysicalAddress": {
                                "AddressLine1": y['AddressLine1'],
                                "AddressLine2": y['AddressLine2'],

                                "Suburb": y['City'],
                                "PostalCode": y['Postcode'],
                                "StateCode": y['State'],
                                "CountryName": "Australia",
                                "CountryCode": "AU"
                            },
                            "Contact": {
                                "Name": y['BuyerName'],
                                "Phone": {
                                    "Number": y['Phone']
                                }
                            }
                        },

                        "BillToParty": {
                            "AccountCode": "80119621",
                            "PartyName": "SCARLETTMUSIC",
                            "PhysicalAddress": {
                                "AddressLine1": "288 Ballarat Rd",

                                "Suburb": "Footscray",
                                "PostalCode": "3011",
                                "StateCode": "VIC",
                                "CountryName": "Australia",
                                "CountryCode": "AU"
                            },
                            "Contact": {
                                "Name": "Kyal Scarlett",
                                "Phone": {
                                    "Number": "0382563460"
                                }
                            }
                        },
                        "ShipmentItemCollection": {
                            "ShipmentItem": []
                        }
                    }

                    toll_item_array = {
                        "ShipmentFinancials": {
                            "ExtraServicesAmount": {
                                "Currency": "AUD",
                                "Value": str(round(float(y['TotalPrice'])))}},
                        "ShipmentID": str(max_shipment_id),
                        "CreateDateTime": current_time,
                        "ConsigneeParty": {
                            "PartyName": y['BuyerName'],
                            "PhysicalAddress": {
                                "AddressLine1": y['AddressLine1'],
                                "AddressLine2": y['AddressLine2'],

                                "Suburb": y['City'],
                                "PostalCode": y['Postcode'],
                                "StateCode": y['State'],
                                "CountryName": "Australia",
                                "CountryCode": "AU"
                            },
                            "Contact": {
                                "Name": y['BuyerName'],
                                "Phone": {
                                    "Number": y['Phone']
                                }
                            }
                        },

                        "BillToParty": {
                            "AccountCode": "80119621",
                            "PartyName": "SCARLETTMUSIC",
                            "PhysicalAddress": {
                                "AddressLine1": "288 Ballarat Rd",

                                "Suburb": "Footscray",
                                "PostalCode": "3011",
                                "StateCode": "VIC",
                                "CountryName": "Australia",
                                "CountryCode": "AU"
                            },
                            "Contact": {
                                "Name": "Kyal Scarlett",
                                "Phone": {
                                    "Number": "0382563460"
                                }
                            }
                        },
                        "ShipmentItemCollection": {
                            "ShipmentItem": []
                        }
                    }

                    sscc_array = []

                    max_shipment_id += 1
                    max_sscc += 1

                    sscc = str(max_sscc)
                    sscc_equation_1 = (int(sscc[16]) + int(sscc[14]) + int(sscc[12]) + int(sscc[10]) + int(
                        sscc[8]) + int(
                        sscc[6]) + int(sscc[4]) + int(sscc[2]) + int(sscc[0])) * 3
                    sscc_equation_2 = int(sscc[15]) + int(sscc[13]) + int(sscc[11]) + int(sscc[9]) + int(
                        sscc[7]) + int(
                        sscc[5]) + int(sscc[3]) + int(sscc[1])

                    sscc_equation_3 = sscc_equation_1 + sscc_equation_2

                    sscc_equation_4 = round_up_to_nearest_10(sscc_equation_3)

                    sscc_check_digit = sscc_equation_4 - sscc_equation_3

                    sscc = '00' + sscc + str(sscc_check_digit)

                    sscc_array.append({
                        "Value": sscc,
                        "SchemeName": "SSCC"
                    })

                    toll_item_array['ShipmentItemCollection']['ShipmentItem'].append(
                        {
                            "IDs": {
                                "ID": sscc_array
                            },

                            "ShipmentItemTotals": {
                                "ShipmentItemCount": '1'
                            },
                            "ShipmentService": {
                                "ServiceCode": "X",
                                "ServiceDescription": "ROAD EXPRESS",
                                "ShipmentProductCode": "1"
                            },
                            "Description": "Carton",
                            "Dimensions": {
                                "Volume": str(volumevalue),
                                "Weight": str(round(float(y['weight'])))
                            },
                            "References": {
                                "Reference": [
                                    {
                                        "ReferenceType": "ConsignorItemReference",
                                        "ReferenceValue": y['String_OrderID']
                                    },
                                    {
                                        "ReferenceType": "ConsigneeItemReference",
                                        "ReferenceValue": y['String_OrderID']
                                    }
                                ]
                            }
                        })
                    ########################## Need more information in print API, so making seperate list
                    total_toll_information['ShipmentItemCollection']['ShipmentItem'].append({
                        "IDs": {
                            "ID": sscc_array
                        },

                        "ShipmentItemTotals": {
                            "ShipmentItemCount": '1'
                        },
                        "ShipmentService": {
                            "ServiceCode": "X",
                            "ServiceDescription": "ROAD EXPRESS",
                            "ShipmentProductCode": "1"
                        },
                        "Description": "Carton",
                        "Dimensions": {
                            "Volume": str(volumevalue),
                            "Weight": str(round(float(y['weight']), 1)),
                            "Length": str(math.ceil(float(y['length']))),
                            "Width": str(math.ceil(float(y['width']))),
                            "Height": str(math.ceil(float(y['height'])))
                        },
                        "References": {
                            "Reference": [
                                {
                                    "ReferenceType": "ConsignorItemReference",
                                    "ReferenceValue": y['String_OrderID']
                                },
                                {
                                    "ReferenceType": "ConsigneeItemReference",
                                    "ReferenceValue": y['String_OrderID']
                                }
                            ]
                        }
                    })

                    toll_sql_insert.append(
                        f"INSERT INTO Toll(ShipmentID, SSCC, Name, Address1, Address2, Suburb, Postcode, State, Reference, phone_number, item_count, current_sscc, length, width, weight, height, volume) VALUES ('{max_shipment_id}', '{max_sscc}', '{y['BuyerName']}', '{y['AddressLine1']}', '{y['AddressLine2']}', '{y['City']}', '{y['Postcode']}', '{y['State']}', '{y['String_OrderID']}', '{y['Phone']}', '1', '{sscc}', '{str(y['length'])}', '{str(y['width'])}', '{str(y['width'])}', '{str(y['height'])}', '{str(volumevalue)}'); COMMIT;")

                    y['City'] = y['City'].replace("'", "")
                    y['City'] = y['City'].replace("’", "")
                    y['City'] = y['City'].replace("’", "")
                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()

                    try:

                        cursor.execute(
                            f"INSERT INTO Toll(ShipmentID, SSCC, Name, Address1, Address2, Suburb, Postcode, State, Reference, phone_number, item_count, current_sscc, length, width, weight, height, volume) VALUES ('{max_shipment_id}', '{max_sscc}', '{y['BuyerName']}', '{y['AddressLine1']}', '{y['AddressLine2']}', '{y['City']}', '{y['Postcode']}', '{y['State']}', '{y['String_OrderID']}', '{y['Phone']}', '1', '{sscc}', '{str(y['length'])}', '{str(y['width'])}', '{str(y['width'])}', '{str(y['height'])}', '{str(volumevalue)}'); COMMIT;")
                    except:
                        pass
                    message_identifier = str(uuid.uuid4())

                    test_rate_enquiry_headers = {
                        "MessageVersion": "3.1",
                        "MessageIdentifier": message_identifier,
                        "CreateTimestamp": current_time,
                        "DocumentType": "Manifest",
                        "Environment": environment,
                        "MessageSender": "SCARLETTMUSIC",
                        "MessageReceiver": "TOLL",
                        'SourceSystemCode': 'XH56'
                    }

                    toll_message = {"@version": "3.1",
                                    "@encoding": "utf-8",
                                    "TollMessage": {"Header": test_rate_enquiry_headers,
                                                    "Manifest": {
                                                        "BusinessID": "IPEC",
                                                        "CreateDateTime": current_time,
                                                        "DatePeriodCollection": {
                                                            "DatePeriod": [{
                                                                "DateType": "DespatchDate",
                                                                "DateTime": next_toll_day
                                                            }]},

                                                        "ConsignorParty": {
                                                            "PartyName": "Scarlett Music",
                                                            "PhysicalAddress": {
                                                                "AddressLine1": "288 Ballarat Rd",

                                                                "Suburb": "FOOTSCRAY",
                                                                "PostalCode": "3011",
                                                                "StateCode": "VIC",
                                                                "CountryName": "Australia",
                                                                "CountryCode": "AU"
                                                            },
                                                            "Contact": {
                                                                "Name": "Michael Demetre",
                                                                "Phone": {
                                                                    "Number": "1800688586"
                                                                }
                                                            }
                                                        },
                                                        "ShipmentCollection": {
                                                            "Shipment": [toll_item_array]}}}}

                    message_identifier = str(uuid.uuid4())

                    r = requests.post(url=manifest_prod_url, auth=('accounts@scarlettmusic.com.au', 't2TrAPsNTB'),
                                      json=toll_message, headers=headers)

                    response = r.text

                    response = json.loads(response)

                    try:

                        if str(response['TollMessage']['ResponseMessages']['ResponseMessage'][0]['ResponseID'][
                                   'Value']) == '200':
                            # for shipment_lines in payload[]
                            pass

                        else:
                            print(response['TollMessage']['ErrorMessages']['ErrorMessage'][0]['ErrorMessage'])
                            input('see above error and press enter')
                    except:
                        print(response)

                    ####Will need to change environment in prod

                    ############## ABOVE IS THE CREATE SHIPMENT API, BELOW IS THE PRINT ORDER API ############

                    # message_identifier = str(uuid.uuid4())  # getting unique string for toll order
                    ###Getting current date / time
                    now = datetime.datetime.now()
                    current_time = now.strftime("%Y-%m-%dT%H:%M:%SZ")

                    next_day = next_business_day()
                    next_toll_day = next_day.strftime("%Y-%m-%dT09:00:00Z")

                    environment = "PRD"

                    headers = {"Content-Type": "application/json",
                               "Accept": "application/json",
                               "x-api-key": "0O87zfzRVu4yf84mxgRiQ2uw9Q86Xb1Z8JrcnJfL",
                               "callId": "SCARLETTMUSIC",
                               "x-mytoll-identity": "727aba8b-0807-4a90-93be-0231d61d4806",
                               "Channel": "CES",
                               "x-mytoll-token": "eyJhbGciOiJSUzUxMiJ9.eyJ1c2VySWQiOiI3MjdhYmE4Yi0wODA3LTRhOTAtOTNiZS0wMjMxZDYxZDQ4MDYiLCJUJlQiOnRydWUsImdlbmVyYXRlRGF0ZSI6MTY4NjcxNjAxMTM3MSwiY3VzdG9tTmFtZSI6IjE0LTA2LTIzX015VGVhbUdFVG9rZW4iLCJjZXJ0aWZpY2F0ZU5hbWUiOiJsb2NhbGhjbCIsIkMmQyI6dHJ1ZSwidW5pcXVlSWQiOiJjNDcxNzgyOTZkMGQzM2RjY2NlMWY4MjQwODFhYjQzYTk5MzY0NGMxMDY1OGFkZWE2YjhlNGI5OGFkNGEzMGZmIiwiZXhwIjoxNzQ5ODc0NDExfQ.A-gOQU6Pc1_yuFkHTqQ219So4lkeoRI0CxtQrZlAsF9VBgqt085lffV_QRGDBPeogjLL5bae-XloKfPO-Ah23HErGHh_oXw_9CkRg8mcG7tkBZsf8StPPN-6HD1i-9iFioJvRE6d9njkVdePapet1FkBuWVg9WOKp8ft_516XR_pok1JmG_fnA55nDBADMDvUHFPW_YUqaoNbJmLpjf7CV0RGiT4pASilzQ4Ut4cuZ0NxQ3d-bQXBQetL5BxQzYNfANsxRD25icSGmi06alngfIFFxoCqBnuxYs_QCT1BvJHJw5e9LUMnXEGzuNAwx_6baRta7Fjq6UsuQPZ8zU-VA"}
                    # Need to change x-api-key to 0O87zfzRVu4yf84mxgRiQ2uw9Q86Xb1Z8JrcnJfL in prod
                    test_rate_enquiry_headers = {
                        "MessageVersion": "1.0",
                        "MessageIdentifier": message_identifier,
                        "CreateTimestamp": current_time,
                        "Environment": environment,
                        "MessageSender": "SCARLETTMUSIC",
                        "MessageReceiver": "Toll",
                        "SourceSystemCode": "XH56"
                    }
                    manifest_test_url = "https://au-print-sit-apigw.internal.mytoll.com/printDocument"
                    manifest_prod_url = "https://au-print-prod-apigw.internal.myteamge.com/printDocument"

                    toll_item_count = 0
                    total_volume_count = 0
                    total_weight_count = 0

                    shipment_id = total_toll_information['ShipmentID']
                    name = total_toll_information['ConsigneeParty']['PartyName']
                    address1 = total_toll_information['ConsigneeParty']['PhysicalAddress']['AddressLine1']
                    address2 = total_toll_information['ConsigneeParty']['PhysicalAddress']['AddressLine2']
                    suburb = total_toll_information['ConsigneeParty']['PhysicalAddress']['Suburb']
                    postcode = total_toll_information['ConsigneeParty']['PhysicalAddress']['PostalCode']
                    state = total_toll_information['ConsigneeParty']['PhysicalAddress']['StateCode']
                    phone = total_toll_information['ConsigneeParty']['Contact']['Phone']['Number']

                    address_type = 'Residential'

                    payload = {
                        "BusinessID": "IPEC",
                        "PrintDocumentType": "Label",
                        "PrintSettings": {
                            "IsLabelThermal": "false",
                            "IsZPLRawResponseRequired": "false",
                            "PDF": {
                                "IsPDFA4": "false",
                                "PDFSettings": {
                                    "StartQuadrant": "1"
                                }
                            }
                        },
                        "ConsignorParty": {
                            "Contact": {
                                "Name": "Kyal",
                                "Phone": {
                                    "Number": "61422747033"
                                }
                            },
                            "PartyName": "Scarlett Music",
                            "PhysicalAddress": {
                                "AddressLine1": "288 Ballarat Rd",
                                "AddressType": "Business",
                                "CountryCode": "AU",
                                "PostalCode": "3011",
                                "StateCode": "VIC",
                                "Suburb": "FOOTSCRAY"
                            }
                        },
                        "CreateDateTime": current_time,
                        "ShipmentCollection": {
                            "Shipment": [
                                {
                                    "BillToParty": {
                                        "AccountCode": "80119621"
                                    },
                                    "ConsigneeParty": {

                                        "Contact": {
                                            "Name": "Kyal",
                                            "Phone": {
                                                "Number": "0382563460"
                                            }
                                        },
                                        "PartyName": name,
                                        "PhysicalAddress": {
                                            "AddressType": address_type,  ##Business / residential
                                            "AddressLine1": address1,
                                            "AddressLine2": address2,

                                            "Suburb": suburb,
                                            "PostalCode": postcode,
                                            "StateCode": state,
                                            "CountryName": "Australia",
                                            "CountryCode": "AU"
                                        }
                                    },
                                    "CreateDateTime": current_time,
                                    "DatePeriodCollection": {
                                        "DatePeriod": [
                                            {
                                                "DateTime": next_toll_day,
                                                "DateType": "DespatchDate"
                                            }
                                        ]
                                    },
                                    "Orders": {
                                        "Order": [
                                            {}
                                        ]
                                    },
                                    "References": {
                                        "Reference": [
                                            {
                                                "ReferenceType": "ShipmentReference1",
                                                "ReferenceValue": "ShipmentReference1"
                                            }
                                        ]
                                    },
                                    "ShipmentID": shipment_id,
                                    "ShipmentItemCollection": {
                                        "ShipmentItem": []
                                    },
                                    "ShipmentTotals": {
                                        "MiscellaneousItemCount": 0,
                                        "Volume": {
                                            "UOM": "m3",
                                            "Value": 1
                                        },
                                        "Weight": {
                                            "UOM": "kg",
                                            "Value": 1
                                        }
                                    },

                                }
                            ]
                        }
                    }

                    for t in total_toll_information['ShipmentItemCollection']["ShipmentItem"]:
                        message_identifier = str(uuid.uuid4())

                        sscc = t['IDs']['ID']
                        item_count = t['ShipmentItemTotals']['ShipmentItemCount']
                        volume = t['Dimensions']['Volume']
                        weight = round(float(t['Dimensions']['Weight']))
                        length = t['Dimensions']['Length']
                        width = t['Dimensions']['Width']
                        height = t['Dimensions']['Height']

                        total_volume = float(volume) * 1
                        total_weight = float(weight) * 1

                        toll_item_count += int(item_count)
                        total_volume_count += float(total_volume)
                        total_weight_count += float(total_weight)

                        payload['ShipmentCollection']['Shipment'][0]['ShipmentItemCollection']['ShipmentItem'].append({
                            "Commodity": {
                                "CommodityCode": "Z",
                                "CommodityDescription": "ALL FREIGHT"
                            },
                            "Description": "Item- Carton",
                            "Dimensions": {
                                "Height": math.ceil(float(height)),
                                "HeightUOM": "cm3",
                                "Length": math.ceil(float(length)),
                                "LengthUOM": "cm3",
                                "Volume": round(float(volume), 4),
                                "VolumeUOM": "m3",
                                "Weight": round(float(weight)),
                                "WeightUOM": "kg",
                                "Width": math.ceil(float(width)),
                                "WidthUOM": "cm3"
                            },
                            "IDs": {
                                "ID": sscc
                            },
                            "References": {
                                "Reference": [
                                    {
                                        "ReferenceType": "ConsignorItemReference",
                                        "ReferenceValue": y['String_OrderID']
                                    }
                                ]
                            },
                            "ShipmentItemTotals": {
                                "MiscellaneousItemQuantity": 0,
                                "ShipmentItemCount": str(item_count)
                            },
                            "ShipmentService": {
                                "ServiceCode": "X",
                                "ShipmentProductCode": "1"
                            }
                        })

                    payload['ShipmentCollection']['Shipment'][0]['ShipmentTotals']['Volume'][
                        'Value'] = total_volume_count

                    payload['ShipmentCollection']['Shipment'][0]['ShipmentTotals']['Weight']['Value'] = str(
                        round(float(total_weight_count), 1))
                    toll_message = {"TollMessage": {
                        "Header": test_rate_enquiry_headers,
                        "Print": payload}}

                    # print(toll_message)
                    cert_file_path = r"\\SERVER\Python\TOLL\my_client.cert"
                    key_file_path = r"\\SERVER\Python\TOLL\my_client.key"
                    cert = (cert_file_path, key_file_path)

                    r = requests.post(url=manifest_prod_url, auth=("accounts@scarlettmusic.com.au", "t2TrAPsNTB"),
                                      json=toll_message, headers=headers, cert=cert)
                    # r = requests.post(url=manifest_test_url, json=toll_message, headers = headers)
                    # print(curlify.to_curl(r.request))
                    response = r.text

                    response = json.loads(response)
                    pdf = response['TollMessage']['ResponseMessages']['ResponseMessage'][0]['ResponseMessage']
                    # print(pdf)
                    pdf = base64.b64decode(pdf)

                    label_name = label_location + "\\" + str(y['String_OrderID']) + ".pdf"

                    Label_Location.append(label_name)

                    with open(label_name, 'wb') as f:
                        f.write(pdf)
                    ######^^^^^^^ WILL ONLY SPIT OUT PDF IN QUADRANT 1 ^^^^^^^^

                    ######vvvvvvvv CROPPING PDF TO FULL SIZE vvvvvvvvvvvvvv

                    images = convert_from_path(label_name)

                    for page in images:
                        page.save(imagelocations + '\python2.jpg', 'JPEG')

                        im = Image.open(imagelocations + '\python2.jpg')

                        # Size of the image in pixels (size of original image)
                        # (This is not mandatory)
                        # width, height = im.size

                        # Setting the points for cropped image

                        # Cropped image of above dimension
                        # (It will not change original image)
                        im1 = im.crop((0, 0, 900, 1200))
                        im1.save(label_name, 'PDF')

                    #### TOLL PRINTING HAS FINISHED

                    ### BELOW IS FOR BOOKING A TOLL PICKUP

                    message_identifier = str(uuid.uuid4())

                    now = datetime.datetime.now()
                    current_time = now.strftime("%Y-%m-%dT%H:%M:%SZ")

                    next_day = next_business_day()
                    next_toll_day = next_day.strftime("%Y-%m-%dT09:00:00Z")
                    business_close_time = next_day.strftime("%Y-%m-%dT16:00:00Z")

                    test_booking_url = 'https://api-uat.teamglobalexp.com:6930/gateway/TollMessageBookingRestService/1.0/tom/bookingRequest'
                    production_booking_url = 'https://api.teamglobalexp.com:6930/gateway/TollMessageBookingRestService/1.0/tom/bookingRequest'

                    environment = "PRD"

                    address_type = 'Residential'

                    for t in total_toll_information['ShipmentItemCollection']["ShipmentItem"]:
                        message_identifier = str(uuid.uuid4())

                        sscc = t['IDs']['ID']
                        item_count = t['ShipmentItemTotals']['ShipmentItemCount']
                        volume = t['Dimensions']['Volume']
                        weight = round(float(t['Dimensions']['Weight']))
                        length = t['Dimensions']['Length']
                        width = t['Dimensions']['Width']
                        height = t['Dimensions']['Height']

                        total_volume = float(volume) * 1
                        total_weight = float(weight) * 1

                        toll_item_count += int(item_count)
                        total_volume_count += float(total_volume)
                        total_weight_count += float(total_weight)

                        payload = {
                            "@version": "1.0",
                            "@encoding": "utf-8",
                            "TollMessage": {
                                "Header": {
                                    "MessageVersion": "1.0",
                                    "MessageIdentifier": message_identifier,
                                    "CreateTimestamp": current_time,
                                    "DocumentType": "Booking",
                                    "Environment": environment,
                                    "SourceSystemCode": "XH56",
                                    "MessageSender": "SCARLETTMUSIC",
                                    "MessageReceiver": "TOLL"
                                },
                                "Bookings": {
                                    "Booking": [
                                        {
                                            "BusinessID": "IPEC",
                                            "BookingParty": {
                                                "PartyName": "Scarlett Music",
                                                "AccountCode": "80119621",
                                                "PhysicalAddress": {
                                                    "AddressType": "Business",
                                                    "AddressLine1": "286-288 Ballarat Rd",
                                                    "AddressLine2": "Scarlett Music",
                                                    "Suburb": "Footscray",
                                                    "PostalCode": "3011",
                                                    "StateCode": "VIC",
                                                    "CountryCode": "AU"
                                                },
                                                "Contact": {
                                                    "Name": "Kyal",
                                                    "Phone": {
                                                        "CountryCode": "+61",
                                                        "AreaCode": "03",
                                                        "Number": "82563460"
                                                    },
                                                    "EMail": "kyal@scarlettmusic.com.au"
                                                }
                                            },
                                            "SystemFields": {
                                                "RecordCreateDateTime": current_time,
                                                "PickupDateTime": next_toll_day,
                                                "OpenDateTime": next_toll_day,
                                                "CloseDateTime": business_close_time
                                            },
                                            "References": {
                                                "Reference": [
                                                    {
                                                        "ReferenceType": "PickupLocation",
                                                        "ReferenceValue": "Footscray"
                                                    }
                                                ]
                                            },
                                            "Contact": {
                                                "Name": "Kyal Scarlett",
                                                "Phone": {
                                                    "CountryCode": "+61",
                                                    "AreaCode": "03",
                                                    "Number": "82563460"
                                                },
                                                "EMail": "kyal@scarlettmusic.com.au",
                                                "Note": ""
                                            },
                                            "BookingFlags": {
                                                "BookingInstruction": "The music shop, open 9am-6pm. Best parking is at The Palms across the road.",
                                                "SameLocationFlag": "true",
                                            },
                                            "BookingItems": {
                                                "BookingItem": [
                                                    {
                                                        "ShipmentService": {
                                                            "ServiceCode": "X",
                                                            "ServiceDescription": "Road Express"
                                                        },
                                                        "ItemQuantity": str(item_count),
                                                        #####Plug item details in here
                                                        "Description": "Parcel",
                                                        "ItemTypeCode": "BXCA",
                                                        "ItemTypeName": "Box/Carton",
                                                        "Dimensions": {
                                                            "Length": str(math.ceil(float(length))),
                                                            "LengthUOM": "CMT",
                                                            "Width": str(math.ceil(float(width))),
                                                            "WidthUOM": "CMT",
                                                            "Height": str(math.ceil(float(height))),
                                                            "HeightUOM": "CMT",
                                                            "Volume": str(round(float(volume), 3)),
                                                            "VolumeUOM": "MTQ",
                                                            "Weight": str(round(float(weight))),
                                                            "WeightUOM": "KGM"
                                                        },
                                                        "BookingFinancials": {
                                                            "ChargeCode": "Sender",
                                                            "AccountCode": "80119621"
                                                        },
                                                        "DeliveryParty": {
                                                            "PhysicalAddress": {
                                                                "AddressType": address_type,

                                                                "AddressLine1": address1,
                                                                #####Plug delivery address details in here
                                                                "Suburb": suburb,
                                                                "PostalCode": str(postcode),
                                                                "StateCode": state,
                                                                "CountryCode": "AU"
                                                            }
                                                        },
                                                        "DatePeriodCollection": {
                                                            "DatePeriod": [
                                                                {
                                                                    "DateType": "OrderDate",
                                                                    "DateTime": current_time
                                                                }
                                                            ]
                                                        },
                                                        "BookingItemFlags": {
                                                            "DangerousGoodsFlag": "false"
                                                        }
                                                    }
                                                ]
                                            }
                                        }
                                    ]
                                }
                            }
                        }

                        r = requests.post(url=production_booking_url,
                                          auth=("accounts@scarlettmusic.com.au", "t2TrAPsNTB"),
                                          json=payload)

                        response = r.text

                        response = json.loads(response)

                    try:
                        y['ItemName'] = y['ItemName'].replace("'", "")
                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        try:
                            cursor = connection.cursor()
                        except:
                            connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                           "5432")
                            cursor = connection.cursor()

                        try:
                            cursor.execute(
                                fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][0:50]}', '', '{y['Quantity'][0:50]}', '{y['ItemName'][0:50]}', '{y['SKU']}', '{y['OrderID'][0][0:50]}', '{y['Final_Courier'][0:50]}','{cp_tracking_number}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                        except:
                            pass
                    except:

                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        for lines in range(len(y['SKU'])):
                            y['ItemName'][f'Name (Item {str(lines + 1)})'] = y['ItemName'][
                                f'Name (Item {str(lines + 1)})'].replace("'", "")
                            try:
                                cursor = connection.cursor()
                            except:
                                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                               "5432")
                                cursor = connection.cursor()

                            try:

                                cursor.execute(
                                    fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '', '{y['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{y['ItemName'][f'Name (Item {str(lines + 1)})'][0:200]}', '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '{y['OrderID'][0][0:50]}', '{y['Final_Courier']}','{cp_tracking_number}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")

                            except:
                                pass
                if y['Final_Courier'] == 'Transdirect':

                    id = y['Transdirect_ID']
                    Final_Transdirect_Courier = y['Final_Transdirect_Courier']

                    if Final_Transdirect_Courier == 'Couriers Please Domestic Proirity Authority' or Final_Transdirect_Courier == 'Fastway' or Final_Transdirect_Courier == 'Couriers Please Multi 21' or Final_Transdirect_Courier == 'Fastway Multi 7':
                        tomorrow = datetime.date.today() + datetime.timedelta(days=1)

                        if Final_Transdirect_Courier == 'Couriers Please Domestic Proirity Authority' or Final_Transdirect_Courier == 'Couriers Please Multi 21':
                            courier = 'couriers_please_domestic_proirity_authority'
                            tier = 2
                        elif Final_Transdirect_Courier == 'fastway' or Final_Transdirect_Courier == 'Fastway Multi 7':
                            courier = 'fastway'
                            tier = 1

                        # print(tomorrow)
                        payload = {
                            "courier": courier,
                            "pickup-date": str(next_day),
                            "tier": tier
                        }
                        headers = {'Api-key': '74a242a1f4b8ba6ea5e011a80bdd2732', 'Content-Type': 'application/json',
                                   "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
                                   "Accept-Encoding": "*",
                                   "Connection": "keep-alive"}

                        r = requests.post('https://www.transdirect.com.au/api/bookings/v4/' + str(id) + '/confirm',
                                          auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'),
                                          headers=headers,
                                          json=payload)
                        response = r.text
                        # print(response)
                        if response == '':
                            pass
                        elif response == '{"errors":["Pickup dates invalid."]}':
                            tomorrow = next_business_day()
                            while response == '{"errors":["Pickup dates invalid."]}':
                                tomorrow = tomorrow + datetime.timedelta(days=1)
                                #  print(tomorrow)

                                payload = {"courier": str(courier), "pickup-date": str(tomorrow), "tier": tier}
                                r = requests.post(
                                    'https://www.transdirect.com.au/api/bookings/v4/' + str(id) + '/confirm',
                                    auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'),
                                    headers=headers,
                                    json=payload)
                                response = r.text
                                # print(response)
                                if response == '':
                                    # print('Success!')
                                    break
                                continue

                    else:
                        y['Phone'] = y['Phone'].replace('+61', '0')
                        y['Phone'] = y['Phone'].replace(' ', '')

                        maxrow = wsTransdirect.max_row
                        maxrow = maxrow + 1
                        wsTransdirect['A' + str(maxrow)].value = y['String_OrderID']
                        wsTransdirect['B' + str(maxrow)].value = 'Scarlett Music Order ' + y['String_OrderID']
                        wsTransdirect['C' + str(maxrow)].value = y['Final_Transdirect_Courier']
                        wsTransdirect['D' + str(maxrow)].value = next_day
                        wsTransdirect['E' + str(maxrow)].value = 'Private'
                        wsTransdirect['F' + str(maxrow)].value = y['BuyerName']
                        wsTransdirect['G' + str(maxrow)].value = y['Email']
                        wsTransdirect['H' + str(maxrow)].value = y['Phone']
                        wsTransdirect['I' + str(maxrow)].value = y['Company']

                        if 'ebay' in y['AddressLine1']:
                            wsTransdirect['J' + str(maxrow)].value = y['AddressLine2'].strip()
                        else:
                            wsTransdirect['J' + str(maxrow)].value = y['AddressLine1'].strip() + ' ' + y[
                                'AddressLine2'].strip()

                        wsTransdirect['K' + str(maxrow)].value = y['City']
                        wsTransdirect['L' + str(maxrow)].value = y['Postcode']
                        wsTransdirect['M' + str(maxrow)].value = y['State']  # 'State'
                        wsTransdirect['N' + str(maxrow)].value = 'AU'  # 'Country' # 'AU' always
                        wsTransdirect['O' + str(maxrow)].value = y['weight']  # 'Weight' #kg
                        wsTransdirect['P' + str(maxrow)].value = y['length']  # 'Length' #cm
                        wsTransdirect['Q' + str(maxrow)].value = y['width']  # 'Width' #cm
                        wsTransdirect['R' + str(maxrow)].value = y['height']  # 'Height' #cm
                        wsTransdirect['S' + str(maxrow)].value = '1'  # 'Quantity' #Always 1 i think?
                        wsTransdirect['T' + str(maxrow)].value = 'Carton'  # 'Packaging' #Always 'Carton'

                if y['Final_Courier'] == 'Couriers Please':
                    y['Phone'] = y['Phone'].replace('+61', '0')
                    y['Phone'] = y['Phone'].replace(' ', '')

                    cp_url = 'https://api.couriersplease.com.au/v2/domestic/shipment/create'

                    r = requests.post(cp_url, headers=cp_headers, json=y['cp_validate_body'])
                    response = r.text
                    response = json.loads(response)
                    # pprint.pprint(response)

                    cp_tracking_number = response['data']['consignmentCode']

                    ###########BELOW IS CREATING LABEL FOR ORDER########

                    cp_url = f'https://api.couriersplease.com.au/v1/domestic/shipment/label?consignmentNumber={cp_tracking_number}'
                    r = requests.get(cp_url, headers=cp_headers, json=cp_validate_body)
                    response = r.text

                    try:

                        response = json.loads(response)
                    except:
                        continue
                    # pprint.pprint(response)

                    base64_label = response['data']['label']
                    pdf = base64.b64decode(base64_label)

                    label_name = label_location + "\\" + str(y['String_OrderID']) + ".pdf"

                    Label_Location.append(label_name)

                    with open(label_name, 'wb') as f:
                        f.write(pdf)

                    couriers_please_item_count += 1
                    couriers_please_weight += float(y['weight'])

                    try:
                        y['ItemName'] = y['ItemName'].replace("'", "")
                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        try:
                            cursor = connection.cursor()
                        except:
                            connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                           "5432")
                            cursor = connection.cursor()

                        try:
                            cursor.execute(
                                fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][0:50]}', '', '{y['Quantity'][0:50]}', '{y['ItemName'][0:50]}', '{y['SKU']}', '{y['OrderID'][0][0:50]}', '{y['Final_Courier'][0:50]}','{cp_tracking_number}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                        except:
                            pass
                    except:

                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        for lines in range(len(y['SKU'])):
                            y['ItemName'][f'Name (Item {str(lines + 1)})'] = y['ItemName'][
                                f'Name (Item {str(lines + 1)})'].replace("'", "")
                            try:
                                cursor = connection.cursor()
                            except:
                                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                               "5432")
                                cursor = connection.cursor()

                            try:

                                cursor.execute(
                                    fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '', '{y['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{y['ItemName'][f'Name (Item {str(lines + 1)})'][0:200]}', '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '{y['OrderID'][0][0:50]}', '{y['Final_Courier']}','{cp_tracking_number}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")

                            except:
                                pass

                if y['Final_Courier'] == 'Allied Express':

                    today = datetime.date.today()
                    if today.isoweekday() in set((6, 7)):
                        today += datetime.timedelta(days=today.isoweekday() % 5)
                    next_day_allied = str(today.day) + '/' + str(today.month) + '/' + str(today.year) + " 10:00:00"

                    history = HistoryPlugin()
                    session = Session()
                    transport = Transport(session=session)
                    wsdl = 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS?wsdl'

                    try:
                        allied_client = zeep.Client(wsdl=wsdl, transport=transport, plugins=[history])

                        allied_client.transport.session.proxies = {
                            # Utilize for all http/https connections
                            'http': 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS', }
                        allied_account = allied_client.service.getAccountDefaults('755cf13abb3934695f03bd4a75cfbca7',
                                                                                  "SCAMUS", "VIC",
                                                                                  "AOE")

                    except:
                        allied_failed = 'true'

                    ONE_DAY = datetime.timedelta(days=1)  ###### Getting next business day for pickups
                    HOLIDAYS_AU = holidays.AU(prov='VIC')

                    pickup_address = {'address1': "286-288 Ballarat Rd",
                                      'address2': "",
                                      'country': "Australia",
                                      'postCode': '3011',
                                      'state': 'VIC',
                                      'suburb': 'Footscray'}

                    reciever_address = {'address1': y['AddressLine1'],
                                        'address2': y['AddressLine2'],
                                        'country': "Australia",
                                        'postCode': y['Postcode'],
                                        'state': y['State'],
                                        'suburb': y['City']}

                    Jobstop_pickupstop = {'companyName': 'Scarlett Music',
                                          'contact': 'Kyal Scarlett',
                                          'emailAddress': 'info@scarlettmusic.com.au',
                                          'geographicAddress': pickup_address,
                                          'phoneNumber': '03 9318 5751',
                                          'stopNumber': 1,
                                          'stopType': 'P'}

                    Jobstop_deliverystop = {'companyName': y['BuyerName'],
                                            'contact': y['Company'],
                                            'emailAddress': y['Email'],
                                            'geographicAddress': reciever_address,
                                            'phoneNumber': y['Phone'],
                                            'stopNumber': 2,
                                            'stopType': 'D'}

                    Jobstop_final = [Jobstop_pickupstop, Jobstop_deliverystop]

                    allied_item_array = []
                    allied_total_volume = 0
                    allied_total_weight = 0
                    allied_total_items = 0
                    #   multi_items_to_ship.append((weightvalue, length, height, width,volumevalue, cubicweight, quantity, satchel))

                    volumevalue = (float(y['length']) * float(y['width']) * float(y['height'])) * 0.000001

                    cubicweight = round(((float(y['length']) / 100) * (float(y['width']) / 100) * (float(y['height']) / 100)) * 250, 2)

                    allied_item_array.append({'dangerous': 'false',
                                              'height': str(y['height']),
                                              'itemCount': str(1),
                                              'length': str(y['length']),
                                              'volume': str(volumevalue),
                                              'weight': str(y['weight']),
                                              'width': str(y['width'])})

                    allied_total_volume = str(volumevalue)
                    allied_total_weight = str(y['weight'])
                    allied_total_items = '1'

                    # item = {'dangerous': 'false',
                    #         'height': float(height),
                    #         'itemCount': 1,
                    #         'length': float(length),
                    #         'volume': float(volumevalue),
                    #         'weight': float(weightvalue),
                    #         'width': float(width)}

                    ##IF MULTIPLE ITEMS, easily put into dic, item = [{item1, item2}]

                    Service_Level = 'R'  # Overnight Express

                    pickup_instructions = "The music shop, open 9am-6pm. Best parking is at The Palms across the road."
                    reference_number = x['String_SalesRecordNumber']
                    booked_by = 'Kyal Scarlett'

                    job_number = int(re.sub("[^0-9]", "", reference_number))



                    Job = {'account': allied_account,
                           'cubicWeight': cubicweight,
                           'Docket': "SCM",
                           'instructions': pickup_instructions,
                           'cubedItems': allied_item_array,
                           'itemCount': allied_total_items,
                           'weight': allied_total_weight,
                           'volume': allied_total_volume,
                           'items': allied_item_array,
                           'jobStops': Jobstop_final,
                           'serviceLevel': Service_Level,
                           'referenceNumbers': reference_number,
                           'bookedBy': booked_by,
                           'readyDate': today,
                           'jobNumber': job_number,
                           'vehicle': {'vehicleID': 1}}

                    JobIDs = {'jobIds': job_number}

                    try:

                        Job = allied_client.service.validateBooking('755cf13abb3934695f03bd4a75cfbca7', Job)

                        job_price = allied_client.service.calculatePrice('755cf13abb3934695f03bd4a75cfbca7', Job)
                        alliedprice = job_price['totalCharge']
                        alliedprice = round((float(alliedprice) * 1.269) * 1.1, 2)
                        if alliedprice == 0.0:
                            alliedprice = 1000
                        # print('Allied price: $' + str(alliedprice))

                    except zeep.exceptions.Fault:
                        alliedprice = 1000

                    ########################ALLIED EXPRESS AUTHENTICATION

                    today = datetime.date.today()
                    if today.isoweekday() in set((6, 7)):
                        today += datetime.timedelta(days=today.isoweekday() % 5)
                    next_day_allied = str(today.day) + '/' + str(today.month) + '/' + str(today.year) + " 10:00:00"

                    history = HistoryPlugin()
                    session = Session()
                    transport = Transport(session=session)
                    wsdl = 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS?wsdl'

                    try:
                        allied_client = zeep.Client(wsdl=wsdl, transport=transport, plugins=[history])

                        allied_client.transport.session.proxies = {
                            # Utilize for all http/https connections
                            'http': 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS', }
                        allied_account = allied_client.service.getAccountDefaults('755cf13abb3934695f03bd4a75cfbca7',
                                                                                  "SCAMUS", "VIC",
                                                                                  "AOE")

                    except:
                        continue

                    with allied_client.settings(strict=False):
                        allied_client.service.savePendingJob('755cf13abb3934695f03bd4a75cfbca7', Job)
                        dispatch_jobs = allied_client.service.dispatchPendingJobs('755cf13abb3934695f03bd4a75cfbca7',
                                                                                  JobIDs)
                        xml = etree.tostring(history.last_received["envelope"], encoding="unicode")
                        xml = xmltodict.parse(xml)

                    connote_number = \
                        xml['soapenv:Envelope']['soapenv:Body']['ns1:dispatchPendingJobsResponse']['result']['item'][
                            'docketNumber']
                    reference = \
                        xml['soapenv:Envelope']['soapenv:Body']['ns1:dispatchPendingJobsResponse']['result']['item'][
                            'referenceNumbers']

                    pdf = allied_client.service.getLabel('755cf13abb3934695f03bd4a75cfbca7', "AOE", connote_number,
                                                         reference,
                                                         '3011', 1)
                    xml = etree.tostring(history.last_received["envelope"], encoding="unicode")
                    xml = xmltodict.parse(xml)

                    pdf = xml['soapenv:Envelope']['soapenv:Body']['ns1:getLabelResponse']['result']
                    pdf = base64.b64decode(pdf)

                    label_name = label_location + "\\" + str(y['item_count_string']) + ".pdf"

                    # new_image.save(label_name, 'PDF')

                    with open(label_name, 'wb') as f:
                        f.write(pdf)

                    Label_Location.append(label_name)

                    try:
                        y['ItemName'] = y['ItemName'].replace("'", "")
                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        try:
                            cursor = connection.cursor()
                        except:
                            connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                           "5432")
                            cursor = connection.cursor()

                        try:

                            cursor.execute(
                                fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][0:50]}', '', '{y['Quantity'][0:50]}', '{y['ItemName'][0:50]}', '{y['SKU'][0:50]}', '{y['OrderID'][0][0:50]}', '{y['Final_Courier'][0:50]}','{connote_number}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                        except:
                            pass
                    except:

                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        for lines in range(len(y['SKU'])):
                            y['ItemName'][f'Name (Item {str(lines + 1)})'] = y['ItemName'][
                                f'Name (Item {str(lines + 1)})'].replace("'", "")
                            try:
                                cursor = connection.cursor()
                            except:
                                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                               "5432")
                                cursor = connection.cursor()

                            try:

                                cursor.execute(
                                    fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '', '{y['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{y['ItemName'][f'Name (Item {str(lines + 1)})'][0:200]}', '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '{y['OrderID'][0][0:50]}', '{y['Final_Courier']}','{connote_number}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                            except:
                                pass

                if y['Final_Courier'] == 'Dai Post':
                    weightvalue = y['weight']

                    y['Phone'] = y['Phone'].replace('+61', '0')
                    y['Phone'] = y['Phone'].replace(' ', '')

                    dai_volume_value = ((float(y['length']) / 100) * (float(y['width']) / 100) * (float(y['height']) / 100)) * 250
                    # if float(dai_volume_value) > float(weightvalue):
                    #     final_dai_weight = dai_volume_value
                    # elif float(weightvalue) > dai_volume_value:
                    #     final_dai_weight = weightvalue

                    final_dai_weight = weightvalue
                    try:
                        final_dai_weight = round(float(final_dai_weight), 2)
                    except:
                        final_dai_weight = round(float(weightvalue), 2)

                    ########################################

                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()
                    cursor.execute("SELECT MAX(job_number) FROM dai_post;")
                    results = cursor.fetchall()
                    #pprint.pprint(results)
                    max_job_id = int(results[0][0])

                    name = y['BuyerName']
                    name = name.replace("'", "")
                    name = name.replace("’", "")
                    ###Getting current date / time
                    now = datetime.datetime.now()
                    current_time = f"{now.strftime('%Y-%m-%d')}T{now.strftime('%H:%M:%S')}"

                    signature_required = '0'
                    totalprice = float(y['TotalPrice'])
                    if round(float(totalprice)) > 200:
                        signature_required = '1'

                    payload = {
                        "shipment": {
                            "service": "Parcel Right",
                            "labelformat": "PDF",
                            "account": "SCA",
                            "datetime": current_time,
                            "reference": f"{y['String_SalesRecordNumber']} {y['item_count_string']}",
                            "jobnumber": max_job_id,
                            "signature": signature_required,
                            "value": str(round(float(y['TotalPrice']))),
                            "currency": "AUD",
                            "uom": "kg",
                            "weight": final_dai_weight,
                            "originterminal": "TME",
                            "shipper": {
                                "name": "Scarlett Music",
                                "attention": "Kyal",
                                "addr1": "286-288 Ballarat Rd",
                                "addr2": "",
                                "city": "Footscray",
                                "state": "VIC",
                                "country": "AU",
                                "postal": "3011",
                                "phone": "+61 03 8256 3460",
                                "email": "kyal@scarlettmusic.com.au"
                            },
                            "consignee": {
                                "name": y['BuyerName'],
                                "attention": y['BuyerName'],
                                "addr1": y['AddressLine1'],
                                "addr2": y['AddressLine2'],
                                "city": y['City'],
                                "state": y['State'],
                                "country": "AU",
                                "postal": y['Postcode'],
                                "phone": y['Phone'],
                                "email": y['Email']
                            },
                            "item": [
                                {
                                    "description": f"Scarlett Music Order {y['item_count_string']}",
                                    "qty": "1",
                                    "unit": "pc",
                                    "value": str(round(float(y['TotalPrice'])))
                                }
                            ]
                        }
                    }

                    r = requests.post('https://daiglobaltrack.com/prod/serviceconnect',
                                      auth=HTTPBasicAuth('ScarlettMusic', 'D5es4stu!'), json=payload)

                    response = r.text

                    try:

                        dai_response = json.loads(response)
                    except:
                        continue

                    dai_tracking_number = dai_response['shipmentresponse']['tracknbr']
                    dai_pdf = dai_response['shipmentresponse']['label']

                    dai_pdf = base64.b64decode(dai_pdf)

                    label_name = label_location + "\\" + str(y['String_OrderID']) + ".pdf"

                    Label_Location.append(label_name)

                    with open(label_name, 'wb') as f:
                        f.write(dai_pdf)

                    y['BuyerName'] = y['BuyerName'].replace("'", "")

                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()

                    try:
                        cursor.execute(
                            fr"INSERT INTO dai_post(customer_name, tracking_number, time_created, job_number) VALUES ('{y['BuyerName']}', '{dai_tracking_number}', '{now}', '{max_job_id}'); COMMIT;")
                    except:
                        pass
                    try:
                        y['ItemName'] = y['ItemName'].replace("'", "")
                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        try:
                            cursor = connection.cursor()
                        except:
                            connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                           "5432")
                            cursor = connection.cursor()

                        try:

                            cursor.execute(
                                fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][0:50]}', '', '{y['Quantity'][0:50]}', '{y['ItemName'][0:50]}', '{y['SKU'][0:50]}', '{y['OrderID'][0][0:50]}', '{y['Final_Courier'][0:50]}','{dai_tracking_number}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                        except:
                            pass
                    except:

                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        for lines in range(len(y['SKU'])):
                            y['ItemName'][f'Name (Item {str(lines + 1)})'] = y['ItemName'][
                                f'Name (Item {str(lines + 1)})'].replace("'", "")
                            try:
                                cursor = connection.cursor()
                            except:
                                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                               "5432")
                                cursor = connection.cursor()

                            try:

                                cursor.execute(
                                    fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '', '{y['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{y['ItemName'][f'Name (Item {str(lines + 1)})'][0:200]}', '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '{y['OrderID'][0][0:50]}', '{y['Final_Courier']}','{dai_tracking_number}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                            except:
                                pass


                if y['Final_Courier'] == 'Freightster':

                    weightvalue = y['weight']

                    y['Phone'] = y['Phone'].replace('+61', '0')
                    y['Phone'] = y['Phone'].replace(' ', '')

                    freightster_volume_value = ((float(y['length']) / 100) * (float(y['width']) / 100) * (
                            float(y['height']) / 100)) * 250

                    # print(f'l : {length}')
                    # print(f'w : {width}')
                    # print(f'l : {height}')
                    if float(freightster_volume_value) > float(weightvalue):
                        final_freightster_weight = freightster_volume_value
                    elif float(weightvalue) > freightster_volume_value:
                        final_freightster_weight = weightvalue

                    if float(y['length']) <= 4.0 and float(y['width']) <= 23.0 and float(y['height']) <= 23.0 and float(
                            weightvalue) <= 0.25:
                        final_freightster_weight = 0.250

                    elif float(y['length']) <= 18.0 and float(y['width']) <= 23.0 and float(
                            y['height']) <= 4.0 and float(
                        weightvalue) <= 0.25:
                        final_freightster_weight = 0.250

                    elif float(weightvalue) <= 0.25 and freightster_volume_value < 0.331:
                        final_freightster_weight = 0.250

                    elif float(y['length']) <= 18.0 and float(y['width']) <= 23.0 and float(
                            y['height']) <= 4.0 and float(
                        weightvalue) <= 0.5:
                        final_freightster_weight = 0.500

                    payload = {"order": {"serviceCode": 12,
                                         "consignee": {"company": y['Company'],
                                                       "name": y['BuyerName'],
                                                       "address1": y['AddressLine1'],
                                                       "address2": y['AddressLine2'],
                                                       "city": y['City'],
                                                       "postcode": y['Postcode'],
                                                       "state": y['State'],
                                                       "phone": y['Phone'],
                                                       "email": y['Email']},
                                         "sender": {"name": "Kyal Scarlett",
                                                    "address1": "286-288 Ballarat Rd",
                                                    "address2": "",
                                                    "city": "Footscray",
                                                    "postcode": "3011",
                                                    "state": "VIC",
                                                    "phone": "0382563460",
                                                    "email": "kyal@scarlettmusic.com.au"},
                                         "shipment": {"reference": y['item_count_string'],
                                                      "description": y['item_count_string'],
                                                      "weight": str(final_freightster_weight)}}}

                    while True:
                        r = requests.post('https://freightster.com.au/api/v1/shippingAPI/create', json=payload,
                                          headers=freightster_headers)
                        if r.status_code == 429:
                            continue
                        else:
                            break
                    freightster_response = json.loads(r.text)
                    freightster_orderid = freightster_response['response_data']['order_id']
                    freightster_trackingnumber = freightster_response['response_data']['tracking_number']
                    payload = {"order": {"orderIds": [freightster_orderid]}}

                    while True:

                        r = requests.post('https://freightster.com.au/api/v1/shippingAPI/print', json=payload,
                                          headers=freightster_headers)

                        if r.status_code == 429:
                            continue
                        else:
                            break
                    freightster_response = json.loads(r.text)

                    freightster_pdf = freightster_response['response_data']['labels'][0]['label']
                    freightster_pdf = base64.b64decode(freightster_pdf)
                    buyer_name = y['BuyerName']
                    label_name = label_location + "\\" + str(y['String_OrderID']) + ".pdf"

                    Label_Location.append(label_name)

                    with open(label_name, 'wb') as f:
                        f.write(freightster_pdf)

                    buyer_name = buyer_name.replace("'", "")

                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()
                    try:
                        cursor.execute(
                            f"INSERT INTO Freightster(order_id, tracking_number, name) VALUES ('{freightster_orderid}', '{freightster_trackingnumber}', '{buyer_name}'); COMMIT;")
                    except:
                        pass
                    try:
                        y['ItemName'] = y['ItemName'].replace("'", "")
                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        try:
                            cursor = connection.cursor()
                        except:
                            connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                           "5432")
                            cursor = connection.cursor()

                        try:

                            cursor.execute(
                                fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][0:50]}', '', '{y['Quantity'][0:50]}', '{y['ItemName'][0:50]}', '{y['SKU'][0:50]}', '{y['OrderID'][0][0:50]}', '{y['Final_Courier'][0:50]}','{freightster_trackingnumber}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                        except:
                            pass
                    except:

                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        for lines in range(len(y['SKU'])):
                            y['ItemName'][f'Name (Item {str(lines + 1)})'] = y['ItemName'][
                                f'Name (Item {str(lines + 1)})'].replace("'", "")
                            try:
                                cursor = connection.cursor()
                            except:
                                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                               "5432")
                                cursor = connection.cursor()

                            try:

                                cursor.execute(
                                    fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '', '{y['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{y['ItemName'][f'Name (Item {str(lines + 1)})'][0:200]}', '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '{y['OrderID'][0][0:50]}', '{y['Final_Courier']}','{freightster_trackingnumber}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                            except:
                                pass

                if y['Final_Courier'] == 'CBDExpress':
                    y['Phone'] = y['Phone'].replace('+61', '0')
                    y['Phone'] = y['Phone'].replace(' ', '')

                    volumevalue = str(
                        float(float(y['length']) / 100) * float(float(y['width']) / 100) * float(
                            float(y['height']) / 100))

                    instructions = 'Authority to leave'

                    if float(totalprice) > 150 and volumevalue > 0.0001:
                        instructions = 'Signature on Delivery'

                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()
                    cursor.execute(
                        fr"SELECT MAX(consignment_number) FROM CDBExpress")
                    results = cursor.fetchall()
                    max_consignment = results[0][0]
                    new_consignment = str((int(max_consignment) + 1)).zfill(9)
                    barcode_param = f'SCARLET{new_consignment}'

                    now = datetime.datetime.now()
                    current_time = now.strftime("%Y-%m-%dT%H:%M:%SZ")

                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()

                    try:
                        cursor.execute(
                            f"INSERT INTO CDBExpress(order_id, consignment_number, name, time_created) VALUES ('{y['String_OrderID']}', '{new_consignment}', '{y['BuyerName']}', '{current_time}' ); COMMIT;")
                    except:
                        pass
                    payload = {"barcode": barcode_param,
                               "description": "Scarlett Music Order " + y['String_OrderID'],
                               "weight": {"value": y['weight'], "units": "kg"},
                               "volume": {"value": volumevalue, "units": "m3"},
                               "customer_reference": y['item_count_string'], "sender": {
                            "contact": {
                                "name": "Scarlett Music",
                                "phone": "(03) 9318 5751",
                                "company": "Scarlett Music"
                            },
                            "address": {
                                "address_line1": "286-288 Ballarat Rd",
                                "suburb": "Maidstone",
                                "state_name": "VIC",
                                "postcode": "3012",
                                "country": "Australia"
                            },
                            "instructions": "The music shop, open 9am-6pm. Best parking is at The Palms across the road."
                        }, "receiver": {
                            "contact": {
                                "name": y['BuyerName'],
                                "email": y['Email'],
                                "phone": y['Phone'],
                                "company": y['Company']
                            },
                            "address": {
                                "address_line1": y['AddressLine1'],
                                "address_line2": y['AddressLine2'],
                                "suburb": y['City'],
                                "state_name": y['State'],
                                "postcode": y['Postcode'],
                                "country": "Australia"
                            },
                            "instructions": instructions
                        }
                               }
                    cbdurl = 'https://apis.hubsystems.com.au/booking/'
                    cbdauth = ('code-scarlett', 'syY55DxG41sd8')
                    cbdheaders = {'Content-Type': 'text/scarlett'}
                    data = payload

                    response = requests.post(cbdurl, auth=cbdauth, headers=cbdheaders, json=data)

                    # Create barcode image
                    barcode_image = code128.image(barcode_param, height=100)

                    # Create empty image for barcode + text
                    top_bott_margin = 70
                    l_r_margin = 10
                    new_height = barcode_image.height + (2 * top_bott_margin)
                    new_width = barcode_image.width + (2 * l_r_margin)
                    new_image = Image.new('RGB', (new_width, 700), (255, 255, 255))

                    # put barcode on new image
                    barcode_y = 525
                    new_image.paste(barcode_image, (0, barcode_y))

                    # object to draw text
                    draw = ImageDraw.Draw(new_image)

                    # Define custom text size and font
                    h1_size = 28
                    h2_size = 28
                    h3_size = 16
                    footer_size = 21

                    h1_font = ImageFont.truetype("r'C:\Windows\Fonts\calibri.ttf", h1_size)
                    h2_font = ImageFont.truetype("r'C:\Windows\Fonts\calibri.ttf", h2_size)
                    h3_font = ImageFont.truetype("r'C:\Windows\Fonts\calibri.ttf", h3_size)
                    footer_font = ImageFont.truetype("r'C:\Windows\Fonts\calibri.ttf", footer_size)

                    # Define custom text
                    To = 'To:'

                    product_type = "Courier's By Demand Express"
                    center_product_type = (barcode_image.width / 2) - len(product_type) * 5
                    center_barcode_value = (barcode_image.width / 2) - len(barcode_param) * 8

                    # Draw text on picture
                    draw.text((l_r_margin, 200), To, fill=(0, 0, 0), font=h2_font)
                    draw.text((40, 250), y['BuyerName'], fill=(0, 0, 0), font=h2_font)
                    draw.text((40, 275), y['Company'], fill=(0, 0, 0), font=h2_font)
                    draw.text((40, 300), y['AddressLine1'], fill=(0, 0, 0), font=h2_font)
                    draw.text((40, 325), y['AddressLine2'], fill=(0, 0, 0), font=h2_font)
                    draw.text((40, 350), y['City'], fill=(0, 0, 0), font=h2_font)
                    draw.text((40, 375), y['State'], fill=(0, 0, 0), font=h2_font)
                    draw.text((40, 400), y['Postcode'], fill=(0, 0, 0), font=h2_font)
                    draw.text((40, 425), y['Phone'], fill=(0, 0, 0), font=h2_font)
                    draw.text((40, 490), y['item_count_string'], fill=(0, 0, 0), font=h2_font)

                    draw.text((center_product_type, (490)), product_type, fill=(0, 0, 0), font=footer_font)
                    draw.text((center_barcode_value, (640)), barcode_param, fill=(0, 0, 0), font=h2_font)

                    scarlett_music_logo = Image.open(rf"\\SERVER\Python\website_logo.png")

                    half = 0.2
                    out = scarlett_music_logo.resize([int(half * s) for s in scarlett_music_logo.size])

                    new_image.paste(out, (170, 10))

                    # save in file

                    label_name = label_location + "\\" + str(y['String_OrderID']) + ".pdf"

                    new_image.save(label_name, 'PDF')

                    Label_Location.append(label_name)

                    try:
                        y['ItemName'] = y['ItemName'].replace("'", "")
                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        try:
                            cursor = connection.cursor()
                        except:
                            connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                           "5432")
                            cursor = connection.cursor()

                        try:

                            cursor.execute(
                                fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][0:50]}', '', '{y['Quantity'][0:50]}', '{y['ItemName'][0:50]}', '{y['SKU']}', '{y['OrderID'][0][0:50]}', '{y['Final_Courier']}','{barcode_param}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                        except:
                            pass
                    except:

                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        for lines in range(len(y['SKU'])):
                            y['ItemName'][f'Name (Item {str(lines + 1)})'] = y['ItemName'][
                                f'Name (Item {str(lines + 1)})'].replace("'", "")
                            try:
                                cursor = connection.cursor()
                            except:
                                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                               "5432")
                                cursor = connection.cursor()

                            try:

                                cursor.execute(
                                    fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '', '{y['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{y['ItemName'][f'Name (Item {str(lines + 1)})'][0:200]}', '{y['SKU'][f'SKU (Item {str(lines + 1)})']}', '{y['OrderID'][0]}', '{y['Final_Courier']}','{barcode_param}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                            except:
                                pass

                if y['Final_Courier'] == 'Sendle':
                    y['Phone'] = y['Phone'].replace('+61', '0')
                    y['Phone'] = y['Phone'].replace(' ', '')

                    volumevalue = str(
                        float(float(y['length']) / 100) * float(float(y['width']) / 100) * float(
                            float(y['height']) / 100))

                    payload = {"description": "Scarlett Music Order " + y['String_OrderID'],
                               "weight": {"value": y['weight'], "units": "kg"},
                               "volume": {"value": volumevalue, "units": "m3"},
                               "customer_reference": y['item_count_string'], "sender": {
                            "contact": {
                                "name": "Scarlett Music",
                                "phone": "(03) 9318 5751",
                                "company": "Scarlett Music"
                            },
                            "address": {
                                "address_line1": "286-288 Ballarat Rd",
                                "suburb": "Maidstone",
                                "state_name": "VIC",
                                "postcode": "3012",
                                "country": "Australia"
                            },
                            "instructions": "The music shop, open 9am-6pm. Best parking is at The Palms across the road."
                        }, "receiver": {
                            "contact": {
                                "name": y['BuyerName'],
                                "email": y['Email'],
                                "phone": y['Phone'],
                                "company": y['Company']
                            },
                            "address": {
                                "address_line1": y['AddressLine1'],
                                "address_line2": y['AddressLine2'],
                                "suburb": y['City'],
                                "state_name": y['State'],
                                "postcode": y['Postcode'],
                                "country": "Australia"
                            },
                            "instructions": ''
                        }
                               }

                    r = requests.post('https://api.sendle.com/api/orders/',
                                      auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'),
                                      json=payload)

                    # pprint.pprint(r.text)
                    # pprint.pprint(r.json())
                    response = r.text
                    response = json.loads(response)

                    # pprint.pprint(response)
                    tracking = response['sendle_reference']
                    orderurl = response['order_url']
                    price = response['price']['gross']['amount']
                    croppedpdfurl = response['labels'][1]['url']

                    r = requests.get(croppedpdfurl,
                                     auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'),
                                     stream=True)

                    label_name = label_location + "\\" + str(y['String_OrderID']) + ".pdf"

                    Label_Location.append(label_name)

                    with open(label_name, "wb") as pdf:
                        for chunk in r.iter_content(chunk_size=1024):

                            # writing one chunk at a time to pdf file
                            if chunk:
                                pdf.write(chunk)

                    maxrow = wsSendle.max_row
                    maxrow = maxrow + 1
                    wsSendle['A' + str(maxrow)].value = 'Lorelle Scarlett'  # 'sender_name'
                    wsSendle['B' + str(maxrow)].value = 'Scarlett Music'  # 'sender_company'
                    wsSendle['C' + str(maxrow)].value = '284-288 Ballarat Rd'  # 'sender_address_line1'
                    wsSendle['E' + str(maxrow)].value = 'Footscray'  # 'sender_suburb'
                    wsSendle['F' + str(maxrow)].value = 'VIC'  # 'sender_state_name'
                    wsSendle['G' + str(maxrow)].value = '3011'  # 'sender_postcode'
                    wsSendle['H' + str(maxrow)].value = 'Australia'  # 'sender_country'
                    wsSendle['I' + str(maxrow)].value = '0417557472'  # 'sender_contact_number'
                    wsSendle['J' + str(
                        maxrow)].value = 'At the music shop, open 9am-6pm. Parking is best across the road at The Palms.'  # 'pickup_instructions'  ##At the music shop, open 9am-6pm. Parking is best across the road at The Palms.
                    wsSendle['K' + str(maxrow)].value = y['BuyerName']  # 'receiver_name'
                    wsSendle['L' + str(maxrow)].value = y['Email']  # 'receiver_email'
                    wsSendle['M' + str(maxrow)].value = y['Company']  # 'receiver_company'
                    wsSendle['N' + str(maxrow)].value = y['AddressLine1']  # 'receiver_address_line1'
                    wsSendle['O' + str(maxrow)].value = y['AddressLine2']  # 'receiver_address_line2'
                    wsSendle['P' + str(maxrow)].value = y['City']  # 'receiver_suburb'
                    wsSendle['Q' + str(maxrow)].value = y['State']  # 'receiver_state_name'
                    wsSendle['R' + str(maxrow)].value = y['Postcode']  # 'receiver_postcode'
                    wsSendle['S' + str(maxrow)].value = 'Australia'  # 'receiver_country'
                    wsSendle['T' + str(maxrow)].value = y['Phone']  # 'receiver_contact_number'
                    wsSendle['W' + str(maxrow)].value = y['weight']  # 'kilogram_weight'
                    wsSendle['X' + str(maxrow)].value = str(
                        float(float(y['length']) / 100) * float(float(y['width']) / 100) * float(
                            float(y['height']) / 100))  # 'cubic_metre_volume' ### l x w x h (in meters!)
                    wsSendle['Y' + str(maxrow)].value = 'Scarlett Music Order ' + y['String_OrderID']  # 'description'
                    wsSendle['Z' + str(maxrow)].value = y['String_OrderID']  # 'customer_reference'

                    try:
                        y['ItemName'] = y['ItemName'].replace("'", "")
                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        try:
                            cursor = connection.cursor()
                        except:
                            connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                           "5432")
                            cursor = connection.cursor()

                        try:

                            cursor.execute(
                                fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][0:50]}', '', '{y['Quantity'][0:50]}', '{y['ItemName'][0:50]}', '{y['SKU']}', '{y['OrderID'][0][0:50]}', '{y['Final_Courier']}','{tracking}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                        except:
                            pass
                    except:

                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        for lines in range(len(y['SKU'])):
                            y['ItemName'][f'Name (Item {str(lines + 1)})'] = y['ItemName'][
                                f'Name (Item {str(lines + 1)})'].replace("'", "")
                            try:
                                cursor = connection.cursor()
                            except:
                                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                               "5432")
                                cursor = connection.cursor()

                            try:

                                cursor.execute(
                                    fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed'][0:50]}','' , '{y['SKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '', '{y['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{y['ItemName'][f'Name (Item {str(lines + 1)})'][0:200]}', '{y['SKU'][f'SKU (Item {str(lines + 1)})']}', '{y['OrderID'][0]}', '{y['Final_Courier']}','{tracking}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                            except:
                                pass
                if y['Final_Courier'] == 'Australia Post':
                    y['Phone'] = y['Phone'].replace('+61', '0')
                    y['Phone'] = y['Phone'].replace('+', '')
                    y['Phone'] = y['Phone'].replace(' ', '')

                    headers = {'Content-Type': 'application/json', 'Accept': 'application/json',
                               'Account-Number': accountnumber,
                               "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
                               "Accept-Encoding": "*",
                               "Connection": "keep-alive"}

                    payload = {
                        "shipments": [
                            {
                                "shipment_reference": y['item_count_string'],
                                "customer_reference_1": y['item_count_string'],
                                "customer_reference_2": "SKU-1",
                                "email_tracking_enabled": 'true',
                                "from": {
                                    "name": "Lorelle Scarlett",
                                    "business_name": "Scarlett Music",
                                    "lines": [
                                        "268-288 Ballarat Rd"
                                    ],
                                    "suburb": "FOOTSCRAY",
                                    "state": "VIC",
                                    "postcode": "3011",
                                    "phone": "0422747033",
                                    "email": "lorelle@scarlettmusic.com.au"
                                },
                                "to": {
                                    "name": y['BuyerName'],
                                    "business_name": y['Company'][0:40],
                                    "lines": [
                                        y['AddressLine1'], y['AddressLine2']
                                    ],
                                    "suburb": y['City'],
                                    "state": y['State'],
                                    "postcode": y['Postcode'],
                                    "phone": y['Phone'],
                                    "email": y['Email']
                                },
                                "items": [{
                                    "item_reference": y['String_OrderID'] + '-1',
                                    "product_id": '3D55',
                                    "length": y['length'],
                                    "height": y['height'],
                                    "width": y['width'],
                                    "weight": y['weight'],
                                    "authority_to_leave": 'false',
                                    "allow_partial_delivery": 'true',

                                },

                                ]
                            }
                        ]
                    }

                    r = requests.post('https://digitalapi.auspost.com.au/shipping/v1/shipments', headers=headers,
                                      auth=HTTPBasicAuth(username, secret),
                                      json=payload)

                    response = r.text
                    response = json.loads(response)
                    itemID = response['shipments'][0]['items'][0]['item_id']
                    shipmentID = response['shipments'][0]['shipment_id']
                    totalcost = response['shipments'][0]['shipment_summary']['total_cost']
                    ausposttracking = response['shipments'][0]['items'][0]['tracking_details']['consignment_id']

                    payload = {
                        "wait_for_label_url": 'true',
                        "preferences": [
                            {
                                "type": "PRINT",
                                "format": "PDF",
                                "groups": [
                                    {
                                        "group": "Parcel Post",
                                        "layout": "THERMAL-LABEL-A6-1PP",
                                        "branded": 'true',
                                        "left_offset": 0,
                                        "top_offset": 0
                                    },
                                    {
                                        "group": "Express Post",
                                        "layout": "THERMAL-LABEL-A6-1PP",
                                        "branded": 'false',
                                        "left_offset": 0,
                                        "top_offset": 0
                                    }
                                ]
                            }
                        ],
                        "shipments": [
                            {
                                "shipment_id": shipmentID,
                                "items": [
                                    {
                                        "item_id": itemID
                                    }
                                ]
                            }
                        ]
                    }

                    r = requests.post('https://digitalapi.auspost.com.au/shipping/v1/labels', headers=headers,
                                      auth=HTTPBasicAuth(username, secret),
                                      json=payload)

                    response = r.text
                    response = json.loads(response)
                    # pprint.pprint(response)

                    url = response['labels'][0]['url']

                    r = requests.get(url, stream=True)

                    label_name = label_location + "\\" + str(y['String_OrderID']) + ".pdf"

                    Label_Location.append(label_name)

                    with open(label_name, "wb") as pdf:
                        pdf.write(r.content)

                    maxrow = wsAustraliaPost.max_row
                    maxrow = maxrow + 1
                    wsAustraliaPost['B' + str(maxrow)].value = y['Email']  # 'C_CONSIGNEE_EMAIL' ### Email for reciever
                    wsAustraliaPost['C' + str(
                        maxrow)].value = 'Y'  # 'C_EMAIL_NOTIFICATION' ### Email notifications on? Always make 'Y'
                    wsAustraliaPost['D' + str(maxrow)].value = str((float(y['length']) * float(y['width']) * float(y[
                                                                                                                       'height'])) / 6000)  # 'A_ACTUAL_CUBIC_WEIGHT' ### Formula for this is length (cm) x width (cm) x height (cm) / 6000
                    wsAustraliaPost['E' + str(maxrow)].value = y['length']  # 'A_LENGTH' ### Length (cm)
                    wsAustraliaPost['F' + str(maxrow)].value = y['width']  # 'A_WIDTH' ### Width (cm)
                    wsAustraliaPost['G' + str(maxrow)].value = y['height']  # 'A_HEIGHT' ##Height (cm)
                    wsAustraliaPost['H' + str(maxrow)].value = 'Scarlett Music Order ' + y[
                        'String_OrderID']  # 'G_DESCRIPTION' ### Scarlett music order xxx
                    wsAustraliaPost['I' + str(maxrow)].value = y['weight']  # 'G_WEIGHT' ### Weight in kg
                    wsAustraliaPost['J' + str(maxrow)].value = '3D55'  # 'CHRG_CODE'### Always 7D55 (Standard eParcel)
                    wsAustraliaPost['K' + str(maxrow)].value = y['BuyerName']  # 'CNSGNEE_NAME' ### Customer Name
                    wsAustraliaPost['L' + str(maxrow)].value = y['Company']  # 'CNSGNEE_BUS_NAME'###Company
                    wsAustraliaPost['M' + str(maxrow)].value = y[
                        'AddressLine1']  # 'CNSGNEE_ADDR_LINE1' ### AddressLine1
                    wsAustraliaPost['N' + str(maxrow)].value = y['AddressLine2']  # 'CNSGNEE_ADDR_LINE2' ###AddressLine2
                    wsAustraliaPost['O' + str(maxrow)].value = y['City']  # 'CNSGNEE_SUBURB'###SUBURB
                    wsAustraliaPost['P' + str(maxrow)].value = y['State']  # 'CNSGNEE_STATE_CODE' ##Short State Code
                    wsAustraliaPost['Q' + str(maxrow)].value = 'AU'  # 'CNSGNEE_CNTRY_CODE' ##AU
                    wsAustraliaPost['R' + str(maxrow)].value = y['Phone']  # 'CNSGNEE_PHONE_NBR' ##Phone
                    wsAustraliaPost[
                        'S' + str(maxrow)].value = 'Y'  # 'IS_PHONE_PRNT_REQD' ##Show phone number, always 'Y'
                    wsAustraliaPost['T' + str(maxrow)].value = 'Y'  # 'IS_SIGNTR_REQD' #Y always i guess
                    wsAustraliaPost['U' + str(maxrow)].value = y['String_OrderID']  # 'REF' ##Reference / Order ID
                    wsAustraliaPost['V' + str(maxrow)].value = 'Y'  # 'IS_REF_PRINT_REQD' ##Show reference, always Y
                    wsAustraliaPost['W' + str(maxrow)].value = y['Postcode']  # 'EMAIL_NOTIFICATION' ##always Y

                    try:
                        y['ItemName'] = y['ItemName'].replace("'", "")
                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        try:
                            cursor = connection.cursor()
                        except:
                            connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                           "5432")
                            cursor = connection.cursor()

                        try:

                            cursor.execute(
                                fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed']}','' , '{y['SKU']}', '', '{y['Quantity']}', '{y['ItemName'][0:50]}', '{y['SKU']}', '{y['OrderID'][0]}', '{y['Final_Courier']}','{ausposttracking}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                        except:
                            pass
                    except:

                        y['BuyerName'] = y['BuyerName'].replace("'", "")
                        y['Company'] = y['Company'].replace("'", "")
                        y['AddressLine1'] = y['AddressLine1'].replace("'", "")
                        y['AddressLine2'] = y['AddressLine2'].replace("'", "")
                        y['City'] = y['City'].replace("'", "")

                        for lines in range(len(y['SKU'])):
                            y['ItemName'][f'Name (Item {str(lines + 1)})'] = y['ItemName'][
                                f'Name (Item {str(lines + 1)})'].replace("'", "")
                            try:
                                cursor = connection.cursor()
                            except:
                                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                               "5432")
                                cursor = connection.cursor()

                            try:

                                cursor.execute(
                                    fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{y['String_OrderID'][0:50]}', '{y['String_OrderID'][0:50]}', '{y['SalesChannel'][0:50]}', '{y['BuyerName'][0:50]}', '{y['Company'][0:50]}', '{y['AddressLine1'][0:50]}', '{y['AddressLine2'][0:50]}', '{y['City'][0:50]}', '{y['State'][0:50]}', '{y['Postcode'][0:50]}', 'AU', '{y['Phone'][0:50]}', '{y['Email'][0:50]}', '{y['date_placed']}','' , '{y['SKU'][f'SKU (Item {str(lines + 1)})']}', '', '{y['Quantity'][f'Quantity (Item {str(lines + 1)})']}', '{y['ItemName'][f'Name (Item {str(lines + 1)})'][0:200]}', '{y['SKU'][f'SKU (Item {str(lines + 1)})']}', '{y['OrderID'][0]}', '{y['Final_Courier']}','{ausposttracking}' , '{str(y['item_count'])}', '{str(y['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                            except:
                                pass

            except (KeyError, TypeError):
                pass

##########################START OF EBAY ORDERS, I KNOW IT IS OVERCOMPLICATED SHUT UP##############

for x in ebayresults:
    if 'satchel' in x['Final_Postage'].lower():

        if x['Height'] == 0 or x['Length'] == 0 or x['Width'] == 0 or x['Weight'] == 0:
            x['Final_Postage'] = 'Label'

Satchel_Position = []

for x in range(len(ebayresults)):
    try:

        if 'satchel' in ebayresults[x]['Final_Postage'].lower():
            Satchel_Position.append(x)
    except KeyError:
        continue

for x in ebayresults:

    for t in range(len(x['ItemTitle'])):  #####Adding items to master Picking Sheet

        itemName = x['ItemTitle']['Name (Item ' + str(t + 1) + ')']
        itemSKU = x['ItemSKU']['SKU (Item ' + str(t + 1) + ')']
        itemQuantity = x['Quantity']['Quantity (Item ' + str(t + 1) + ')']
        maxrow = wsMasterPickSheet.max_row
        maxrow = maxrow + 1
        wsMasterPickSheet['A' + str(maxrow)].value = itemName
        wsMasterPickSheet['B' + str(maxrow)].value = itemSKU
        wsMasterPickSheet['C' + str(maxrow)].value = itemQuantity

    if 'mini' in x['Final_Postage'].lower():

        for lines in range(len(x['SKU'])):
            x['ItemTitle'][f'Name (Item {str(lines + 1)})'] = x['ItemTitle'][
                f'Name (Item {str(lines + 1)})'].replace("'", "")

            x['BuyerName'] = x['BuyerName'].replace("'", "")
            x['Company'] = x['Company'].replace("'", "")
            x['AddressLine1'] = x['AddressLine1'].replace("'", "")
            x['AddressLine2'] = x['AddressLine2'].replace("'", "")
            x['City'] = x['City'].replace("'", "")
            try:
                cursor = connection.cursor()
            except:
                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                cursor = connection.cursor()

            try:

                cursor.execute(
                    fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order, Sales_Record_Number ,Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{x['String_OrderID'][0:50]}', '{x['String_PurchaseOrderNumber'][0:50]}', '{x['String_SalesRecordNumber'][0:50]}', 'ebay', '{x['BuyerName'][0:50]}', '{x['Company'][0:50]}', '{x['AddressLine1'][0:50]}', '{x['AddressLine2'][0:50]}', '{x['City'][0:50]}', '{x['State'][0:50]}', '{x['Postcode'][0:50]}', 'AU', '{x['Phone'][0:50]}', '{x['Email'][0:50]}', '{x['date_placed']}','' , '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['OrderLineID'][f'Orderline (Item {str(lines + 1)})']}', '{x['Quantity'][f'Quantity (Item {str(lines + 1)})']}', '{x['ItemTitle'][f'Name (Item {str(lines + 1)})'][0:200]}', '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['String_OrderID'][0:50]}', 'Minilopes','' , '{str(x['item_count'])}', '{str(x['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")

            except:
                pass
        # changing srn to the item_string_count
        x['String_SalesRecordNumber'] = str(x['item_count_string'])
        mergeMinilopes.append(x)

        maxrow = wsMini.max_row
        maxrow = maxrow + 1
        wsMini['A' + str(maxrow)].value = x['String_SalesRecordNumber']
        wsMini['B' + str(maxrow)].value = x['BuyerName']
        wsMini['C' + str(maxrow)].value = x['Email']
        wsMini['D' + str(maxrow)].value = x['AddressLine1']
        wsMini['E' + str(maxrow)].value = x['AddressLine2']
        wsMini['F' + str(maxrow)].value = x['City']
        wsMini['G' + str(maxrow)].value = x['State']
        wsMini['H' + str(maxrow)].value = x['Postcode']
        # wsMini['I' + str(maxrow)].value = y['ItemName']
        # wsMini['J' + str(maxrow)].value = y['SKU']
        # wsMini['K' + str(maxrow)].value = y['Quantity']
        wsMini['L' + str(maxrow)].value = x['TotalPrice']
        wsMini['M' + str(maxrow)].value = x['Final_Postage']  # THINK THIS WAS SUPPOSED TO BE EXPRESS?
        wsMini['N' + str(maxrow)].value = x['Final_Postage']
        # wsMini['P' + str(maxrow)].value = str(y['height'])
        # wsMini['Q' + str(maxrow)].value = str(y['length'])
        # wsMini['R' + str(maxrow)].value = str(y['width'])
        # wsMini['S' + str(maxrow)].value = str(y['weight'])
        wsMini['T' + str(maxrow)].value = str(x['Phone'])
        wsMini['U' + str(maxrow)].value = str(x['Company'])


    elif 'devil' in x['Final_Postage'].lower():

        for lines in range(len(x['SKU'])):
            x['ItemTitle'][f'Name (Item {str(lines + 1)})'] = x['ItemTitle'][
                f'Name (Item {str(lines + 1)})'].replace("'", "")

            x['BuyerName'] = x['BuyerName'].replace("'", "")
            x['Company'] = x['Company'].replace("'", "")
            x['AddressLine1'] = x['AddressLine1'].replace("'", "")
            x['AddressLine2'] = x['AddressLine2'].replace("'", "")
            x['City'] = x['City'].replace("'", "")
            try:
                cursor = connection.cursor()
            except:
                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                cursor = connection.cursor()

            try:

                cursor.execute(
                    fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order,Sales_Record_Number, Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{x['String_OrderID'][0:50]}', '{x['String_PurchaseOrderNumber'][0:50]}', '{x['String_SalesRecordNumber'][0:50]}', 'ebay', '{x['BuyerName'][0:50]}', '{x['Company'][0:50]}', '{x['AddressLine1'][0:50]}', '{x['AddressLine2'][0:50]}', '{x['City'][0:50]}', '{x['State'][0:50]}', '{x['Postcode'][0:50]}', 'AU', '{x['Phone'][0:50]}', '{x['Email'][0:50]}', '{x['date_placed']}','' , '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['OrderLineID'][f'Orderline (Item {str(lines + 1)})']}', '{x['Quantity'][f'Quantity (Item {str(lines + 1)})']}', '{x['ItemTitle'][f'Name (Item {str(lines + 1)})'][0:200]}', '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['String_OrderID'][0:50]}', 'Devilopes','' , '{str(x['item_count'])}', '{str(x['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
            except:
                pass
        maxrow = wsDevil.max_row

        maxrow = maxrow + 1
        wsDevil['A' + str(maxrow)].value = x['String_SalesRecordNumber']
        wsDevil['B' + str(maxrow)].value = x['BuyerName']
        wsDevil['C' + str(maxrow)].value = x['Email']
        wsDevil['D' + str(maxrow)].value = x['AddressLine1']
        wsDevil['E' + str(maxrow)].value = x['AddressLine2']
        wsDevil['F' + str(maxrow)].value = x['City']
        wsDevil['G' + str(maxrow)].value = x['State']
        wsDevil['H' + str(maxrow)].value = x['Postcode']
        # wsDevil['I' + str(maxrow)].value = y['ItemName']
        # wsDevil['J' + str(maxrow)].value = y['SKU']
        # wsDevil['K' + str(maxrow)].value = y['Quantity']
        wsDevil['L' + str(maxrow)].value = x['TotalPrice']
        wsDevil['M' + str(maxrow)].value = x['Final_Postage']  # THINK THIS WAS SUPPOSED TO BE EXPRESS?
        wsDevil['N' + str(maxrow)].value = x['Final_Postage']
        # wsDevil['P' + str(maxrow)].value = str(y['height'])
        # wsDevil['Q' + str(maxrow)].value = str(y['length'])
        # wsDevil['R' + str(maxrow)].value = str(y['width'])
        # wsDevil['S' + str(maxrow)].value = str(y['weight'])
        wsDevil['T' + str(maxrow)].value = str(x['Phone'])
        wsDevil['U' + str(maxrow)].value = str(x['Company'])

        # changing srn to the item_string_count

        x['String_SalesRecordNumber'] = str(x['item_count_string'])
        mergeDevilopes.append(x)






    elif 'label' in x['Final_Postage'].lower():

        for lines in range(len(x['SKU'])):
            x['ItemTitle'][f'Name (Item {str(lines + 1)})'] = x['ItemTitle'][
                f'Name (Item {str(lines + 1)})'].replace("'", "")

            x['BuyerName'] = x['BuyerName'].replace("'", "")
            x['Company'] = x['Company'].replace("'", "")
            x['AddressLine1'] = x['AddressLine1'].replace("'", "")
            x['AddressLine2'] = x['AddressLine2'].replace("'", "")
            x['City'] = x['City'].replace("'", "")
            try:
                cursor = connection.cursor()
            except:
                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                cursor = connection.cursor()

            try:

                cursor.execute(
                    fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order,Sales_Record_Number , Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{x['String_OrderID'][0:50]}', '{x['String_PurchaseOrderNumber'][0:50]}', '{x['String_SalesRecordNumber'][0:50]}', 'ebay', '{x['BuyerName'][0:50]}', '{x['Company'][0:50]}', '{x['AddressLine1'][0:50]}', '{x['AddressLine2'][0:50]}', '{x['City'][0:50]}', '{x['State'][0:50]}', '{x['Postcode'][0:50]}', 'AU', '{x['Phone'][0:50]}', '{x['Email'][0:50]}', '{x['date_placed']}','' , '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['OrderLineID'][f'Orderline (Item {str(lines + 1)})']}', '{x['Quantity'][f'Quantity (Item {str(lines + 1)})']}', '{x['ItemTitle'][f'Name (Item {str(lines + 1)})'][0:200]}', '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['String_OrderID'][0:50]}', 'Label','' , '{str(x['item_count'])}', '{str(x['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
            except:
                pass
        maxrow = wsLabel.max_row
        maxrow = maxrow + 1
        wsLabel['A' + str(maxrow)].value = x['String_SalesRecordNumber']
        wsLabel['B' + str(maxrow)].value = x['BuyerName']
        wsLabel['C' + str(maxrow)].value = x['Email']
        wsLabel['D' + str(maxrow)].value = x['AddressLine1']
        wsLabel['E' + str(maxrow)].value = x['AddressLine2']
        wsLabel['F' + str(maxrow)].value = x['City']
        wsLabel['G' + str(maxrow)].value = x['State']
        wsLabel['H' + str(maxrow)].value = x['Postcode']
        # wsLabel['I' + str(maxrow)].value = y['ItemName']
        # wsLabel['J' + str(maxrow)].value = y['SKU']
        # wsLabel['K' + str(maxrow)].value = y['Quantity']
        wsLabel['L' + str(maxrow)].value = x['TotalPrice']
        wsLabel['M' + str(maxrow)].value = x['Final_Postage']  # THINK THIS WAS SUPPOSED TO BE EXPRESS?
        wsLabel['N' + str(maxrow)].value = x['Final_Postage']
        # wsLabel['P' + str(maxrow)].value = str(y['height'])
        # wsLabel['Q' + str(maxrow)].value = str(y['length'])
        # wsLabel['R' + str(maxrow)].value = str(y['width'])
        # wsLabel['S' + str(maxrow)].value = str(y['weight'])
        wsLabel['T' + str(maxrow)].value = str(x['Phone'])
        wsLabel['U' + str(maxrow)].value = str(x['Company'])

        # changing srn to the item_string_count

        x['String_SalesRecordNumber'] = str(x['item_count_string'])
        mergeLabel.append(x)


    else:
        continue


def SatchelQuote(Satchel_Position):  ## Concurent futures for getting quotes from couriers

    global lowesttranscourier, next_cp_day, cp_auth, cp_auth_encoded, cp_headers, cp_validate_body, cp_body, cp_item_array, freightsterfailed, cbdexpressfailed, allied_failed, allied_client, allied_account

    SatchNumber = Satchel_Position

    try:
        allied_failed

    except:
        allied_failed = ''

    auspostfailed = ''
    sendlefailed = ''
    transdirectfailed = ''
    fastwayfailed = ''
    fastway_data = ''
    couriers_failed = ''
    freightsterfailed = ''
    cbdexpressfailed = ''
    # allied_failed = ''
    toll_failed = ''
    po_box = ''
    dai_failed = ''

    satchel = 'false'
    srn = ebayresults[Satchel_Position]['String_SalesRecordNumber']
    name = ebayresults[Satchel_Position]['BuyerName']
    email = ebayresults[Satchel_Position]['Email']
    address1 = ebayresults[Satchel_Position]['AddressLine1']
    address2 = ebayresults[Satchel_Position]['AddressLine2']
    city = ebayresults[Satchel_Position]['City']
    state = ebayresults[Satchel_Position]['State']
    postcode = ebayresults[Satchel_Position]['Postcode']
    # itemname = ebayresults[Satchel_Position]['ItemName']
    # sku = ebayresults[Satchel_Position]['SKU']
    # quantity = ebayresults[Satchel_Position]['Quantity']
    totalprice = ebayresults[Satchel_Position]['TotalPrice']
    postagetype = ebayresults[Satchel_Position]['Final_Postage']
    height = str(ebayresults[Satchel_Position]['Height'])
    length = str(ebayresults[Satchel_Position]['Length'])
    width = str(ebayresults[Satchel_Position]['Width'])
    weightvalue = str(ebayresults[Satchel_Position]['Weight'])
    phone = str(ebayresults[Satchel_Position]['Phone'])
    company = str(ebayresults[Satchel_Position]['Company'])
    item_string_counter = str(ebayresults[Satchel_Position]['item_count_string'])

    try:
        suburb = city

    except:
        pass

    try:

        split_name = name.split(' ')

        first_name = split_name[0]

        last_name = ''
        for x in range(1, len(split_name)):
            last_name = f'{last_name} {split_name[x]}'

        last_name = last_name.strip()
    except:
        first_name = name
        last_name = ''

    ##Allied express authentication vvvvvvvvvvvvv

    # today = datetime.date.today()
    # if today.isoweekday() in set((6, 7)):
    #     today += datetime.timedelta(days=today.isoweekday() % 5)
    # next_day_allied = str(today.day) + '/' + str(today.month) + '/' + str(today.year) + " 10:00:00"
    #
    # history = HistoryPlugin()
    # session = Session()
    # transport = Transport(session=session)
    #
    # wsdl = 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS?wsdl'
    #
    # try:
    #     allied_client = zeep.Client(wsdl=wsdl, transport=transport, plugins=[history])
    #
    #     allied_client.transport.session.proxies = {
    #         # Utilize for all http/https connections
    #         'http': 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS', }
    #     allied_account = allied_client.service.getAccountDefaults('755cf13abb3934695f03bd4a75cfbca7', "SCAMUS", "VIC",
    #                                                               "AOE")
    #
    # except:
    #     allied_failed = 'true'

    #####GET GOOGLE ADDRESS BELOW, COLLAPSED BECAUSE CODE  IS JUST REPEATED FUNCTION

    try:
        if 'ebay' in address1:
            address = address2 + " " + city + " " + state + " " + postcode

            # get_google_address()

            dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
            qstr = urllib.parse.urlencode(dict1)
            URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
            URL = URL + qstr
            response = urllib.request.urlopen(URL)
            data = json.load(response)
            # pprint.pprint(data['candidates'])
            placeid = data['candidates'][0]['place_id']
            payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
            r = requests.get(
                'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
            #  pprint.pprint(r.json())
            data = r.json()
            #  pprint.pprint(data['result']['address_components'])
            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'locality':
                    suburb = data['result']['address_components'][t]['long_name']
                    break

                else:
                    continue

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'postal_code':
                    postal_code = data['result']['address_components'][t]['long_name']
                    break

                else:
                    continue

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                    google_state = data['result']['address_components'][t]['short_name']
                    break

                else:
                    continue

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'street_number':
                    google_street1 = data['result']['address_components'][t]['short_name']
                    break

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'route':
                    google_street2 = data['result']['address_components'][t]['long_name']
                    break
            backup_google_street_address = google_street1 + ' ' + google_street2

            postcode = postal_code
            city = suburb
            state = google_state

        else:
            address = address1 + " " + address2 + " " + city + " " + state + " " + postcode

            # get_google_address()

            dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
            qstr = urllib.parse.urlencode(dict1)
            URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
            URL = URL + qstr
            response = urllib.request.urlopen(URL)
            data = json.load(response)
            # pprint.pprint(data['candidates'])
            placeid = data['candidates'][0]['place_id']
            payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
            r = requests.get(
                'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
            #   pprint.pprint(r.json())
            data = r.json()
            #   pprint.pprint(data['result']['address_components'])
            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'locality':
                    suburb = data['result']['address_components'][t]['long_name']
                    break

                else:
                    continue

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'postal_code':
                    postal_code = data['result']['address_components'][t]['long_name']
                    break

                else:
                    continue

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                    google_state = data['result']['address_components'][t]['short_name']
                    break

                else:
                    continue

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'street_number':
                    google_street1 = data['result']['address_components'][t]['short_name']
                    break

            for t in range(20):
                if data['result']['address_components'][t]['types'][0] == 'route':
                    google_street2 = data['result']['address_components'][t]['long_name']
                    break
            backup_google_street_address = google_street1 + ' ' + google_street2

            postcode = postal_code
            city = suburb
            state = google_state

            #    print(newSalesRecordNumber + " " + newBuyerCity + " " + address)
        sendlefailed = ''  ### Get google address, collaps

    except (IndexError, NameError):
        pass

    if 'po box' in address1.lower() or 'po box' in address2.lower() or 'parcel locker' in address1.lower() or 'parcel locker' in address2.lower() or 'p.o' in address1.lower() or 'p.o' in address2.lower() or 'parcel collect' in address1.lower() or 'parcel collect' in address2.lower() or 'pobox' in address1.lower() or 'pobox' in address2.lower() or 'locker' in address1.lower() or 'locker' in address2.lower() or 'collect' in address1.lower() or 'collect' in address2.lower() or 'parcel' in address1.lower() or 'parcel' in address2.lower() or 'pmb' in address1.lower() or 'pmb' in address2.lower() or 'p/o' in address1.lower() or 'p/o' in address2.lower() or 'post office box' in address1.lower() or 'post office box' in address2.lower() or 'lpo' in address1.lower() or 'lpo' in address2.lower() or 'post box' in address1.lower() or 'post box' in address2.lower():
        sendlefailed = 'true'
        transdirectfailed = 'true'
        fastwayfailed = 'true'
        couriers_failed = 'true'
        cbdexpressfailed = 'true'
        allied_failed = 'true'
        toll_failed = 'true'
        dai_failed = 'true'
        po_box = 'true'

    volumevalue = (float(length) * float(width) * float(height)) * 0.000001

    data = {
        "length": float(length),
        "width": float(width),
        "height": float(height)
    }

    r = requests.post(base_url + '/api/utils/calc-cubic', headers=Fastway_Headers, json=data)

    try:

        response = r.text
        response = json.loads(response)

        if 'errors' in response:
            fastwayfailed = 'true'

            cubicweight = round(((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250, 2)

        else:

            cubicweight = response['data']['cubicWeight']

    except:
        cubicweight = round(((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250, 2)

    if ((float(length) * float(width) * float(height)) / 4000) > 25 or float(
            weightvalue) > 25 or float(length) > 120 or float(width) > 120 or float(height) > 120:
        cbdexpressfailed = 'true'

    if float(cubicweight) > 40 or float(weightvalue) > 25:
        fastwayfailed = 'true'

    if float(length) > 105 or float(width) > 105 or float(height) > 105 or float(weightvalue) > 22:
        auspostfailed = 'true'
        dai_failed = 'true'

    if (float(length) * float(width) * float(height)) / 1000 > 100 or float(weightvalue) > 25 or float(
            length) > 120 or float(width) > 120 or float(height) > 120:
        sendlefailed = 'true'

    if float(cubicweight) > 40 or float(length) > 180 or float(width) > 180 or float(height) > 180:
        couriers_failed = 'true'

    try:

        if auspostfailed == 'true':
            auspostprice = 10000


        else:

            headers = {'Content-Type': 'application/json', 'Accept': 'application/json',
                       'Account-Number': accountnumber,
                       "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
                       "Accept-Encoding": "*",
                       "Connection": "keep-alive"}

            payload = {"shipments": [{
                "from": {
                    "suburb": "FOOTSCRAY",
                    "state": "VIC",
                    "postcode": "3011"
                },
                "to": {
                    "suburb": str(city),
                    "state": str(state),
                    "postcode": str(postcode)
                },
                "items": [
                    {
                        "product_id": "3D55",
                        "length": str(length),
                        "height": str(height),
                        "width": str(width),
                        "weight": str(weightvalue)}]}]}

            r = requests.post('https://digitalapi.auspost.com.au/shipping/v1/prices/shipments', headers=headers,
                              auth=HTTPBasicAuth(username, secret), json=payload)
            response = r.text
            response = json.loads(response)
            #  pprint.pprint(response)
            auspostprice = response['shipments'][0]['shipment_summary']['total_cost']

            # if str(state) == 'WA':
            #     auspostprice = str(float(auspostprice) * 1.4)

    except KeyError:
        auspostprice = 1000

    # print('Aus Post price: ' + str(auspostprice))

    if dai_failed == 'true' or float(totalprice) > 200:
        dai_price = 1000
        pass

    else:
        dai_volume_value = ((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250
        # if float(dai_volume_value) > float(weightvalue):
        #     final_dai_weight = dai_volume_value
        # elif float(weightvalue) > dai_volume_value:
        #     final_dai_weight = weightvalue
        final_dai_weight = weightvalue
        final_dai_weight = round(float(final_dai_weight), 2)

        dai_postcode_wb = openpyxl.load_workbook(rf"\\SERVER\Project Folder\Python\Courier Info\Dai_Postcodes.xlsx")

        dai_postcode_sheet = dai_postcode_wb['Sheet']

        dai_zone_sheet1 = ''

        dai_rates_wb = openpyxl.load_workbook(rf"\\SERVER\Project Folder\Python\Courier Info\Dai_Rates.xlsx")

        dai_rates_sheet = dai_rates_wb['Sheet']

        for x in range(2,
                       dai_postcode_sheet.max_row + 1):  ###ITERATING THROUGH THE DAIPOST POSTCODE SHEET TO FIND REGION

            sheet_postcode = str(dai_postcode_sheet['A' + str(x)].value)

            if str(len(sheet_postcode)) == '3':
                sheet_postcode = f'0{sheet_postcode}'

            if postcode == sheet_postcode:
                dai_zone_sheet1 = str(dai_postcode_sheet['B' + str(x)].value)

                for xx in range(2, dai_rates_sheet.max_row + 1):

                    dai_zone_sheet2 = str(dai_rates_sheet['A' + str(xx)].value)

                    if dai_zone_sheet1.lower() == dai_zone_sheet2.lower():  ###Two zones match

                        if final_dai_weight <= 0.5:
                            dai_price = float(dai_rates_sheet['B' + str(xx)].value) * 1.1

                        elif 0.501 <= final_dai_weight <= 1.00:
                            dai_price = float(dai_rates_sheet['C' + str(xx)].value) * 1.1

                        elif 1.01 <= final_dai_weight <= 1.5:
                            dai_price = float(dai_rates_sheet['D' + str(xx)].value) * 1.1

                        elif 1.501 <= final_dai_weight <= 2.0:
                            dai_price = float(dai_rates_sheet['D' + str(xx)].value) * 1.1

                        elif 2.01 <= final_dai_weight <= 2.5:
                            dai_price = float(dai_rates_sheet['E' + str(xx)].value) * 1.1

                        elif 2.501 <= final_dai_weight <= 3.0:
                            dai_price = float(dai_rates_sheet['E' + str(xx)].value) * 1.1

                        elif 3.01 <= final_dai_weight <= 3.50:
                            dai_price = float(dai_rates_sheet['F' + str(xx)].value) * 1.1

                        elif 3.501 <= final_dai_weight <= 4.0:
                            dai_price = float(dai_rates_sheet['F' + str(xx)].value) * 1.1

                        elif 4.01 <= final_dai_weight <= 4.50:
                            dai_price = float(dai_rates_sheet['G' + str(xx)].value) * 1.1

                        elif 4.501 <= final_dai_weight <= 5.00:
                            dai_price = float(dai_rates_sheet['G' + str(xx)].value) * 1.1

                        elif 5.01 <= final_dai_weight <= 7.0:
                            dai_price = float(dai_rates_sheet['H' + str(xx)].value) * 1.1

                        elif 7.01 <= final_dai_weight <= 10.0:
                            dai_price = float(dai_rates_sheet['I' + str(xx)].value) * 1.1

                        elif 10.01 <= final_dai_weight <= 15.00:
                            dai_price = float(dai_rates_sheet['J' + str(xx)].value) * 1.1

                        elif 15.01 <= final_dai_weight <= 22.00:
                            dai_price = float(dai_rates_sheet['K' + str(xx)].value) * 1.1

                        else:
                            dai_price = 1000

                        if po_box == 'true':
                            dai_price += 0.8

                        dai_price = round(dai_price, 2)

    sendlefailed = 'true'
    if sendlefailed == 'true' or float(totalprice) > 200:

        sendleprice = 1000
        pass

    else:

        data = {
            "pickup_suburb": "Maidstone",
            "pickup_postcode": "3012",
            "delivery_suburb": city,
            "delivery_postcode": str(postcode),
            "weight_value": str(weightvalue),
            'weight_units': 'kg',
            'volume_value': str(volumevalue),
            'volume_units': 'm3'}

        r = requests.get('https://api.sendle.com/api/quote',
                         auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'), json=data)
        sendleresponse = json.loads(r.text)
        #  pprint.pprint(sendleresponse)

        try:
            sendleprice = sendleresponse[0]['quote']['gross']['amount']

        except KeyError:
            sendleprice = 1000

    #  print('Sendle price: ' + str(sendleprice))
    if freightsterfailed == 'true' or float(totalprice) > 200:
        freightster_price = 1000
        pass

    else:

        try:

            for x in range(2, sheet.max_row + 1):
                if sheet['A' + str(x)].value.lower() == city.lower():
                    if int(sheet['B' + str(x)].value) == int(postcode):
                        freightster_pricing_zone = sheet['D' + str(x)].value

            freightster_volume_value = ((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250

            # print(f'l : {length}')
            # print(f'w : {width}')
            # print(f'l : {height}')
            if float(freightster_volume_value) > float(weightvalue):
                final_freightster_weight = freightster_volume_value
            elif float(weightvalue) > freightster_volume_value:
                final_freightster_weight = weightvalue

            if float(length) <= 4.0 and float(width) <= 23.0 and float(height) <= 23.0 and float(weightvalue) <= 0.25:
                final_freightster_weight = 0.250

            elif float(length) <= 18.0 and float(width) <= 23.0 and float(height) <= 4.0 and float(weightvalue) <= 0.25:
                final_freightster_weight = 0.250

            elif float(weightvalue) <= 0.25 and freightster_volume_value < 0.331:
                final_freightster_weight = 0.250

            elif float(length) <= 18.0 and float(width) <= 23.0 and float(height) <= 4.0 and float(weightvalue) <= 0.5:
                final_freightster_weight = 0.500

            if freightster_pricing_zone == 'ACT' or freightster_pricing_zone == 'CBR' or freightster_pricing_zone == 'NSW':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.60

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 5.5

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 6.5

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 7.5

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 8.5

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 22.5

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 28.5

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'ADL':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.60

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 5

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 5.5

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 6.4

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 7.5

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 19

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 28

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'BNE':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.60

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 5

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 5

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 6.25

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 7.3

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 18

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 27

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'MEL':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 3.75

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 4.2

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 4.5

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 4.85

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 5.5

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 7

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 7.5

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'PER':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.6

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 5.6

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 6.45

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 8.15

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 8.35

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 22

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 30.5

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'QLD':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.6

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 5.5

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 7

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 8

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 9

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 22.5

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 28.5

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'SA':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.6

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 5.5

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 5

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 8

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 9

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 22.5

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 28.5

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'SYD':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.6

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 4.6

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 5

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 5.95

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 7.15

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 17

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 18

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'VIC':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.6

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 4.75

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 4.95

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 5.85

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 8.5

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 18.5

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 20.9

                else:
                    freightster_price = 1000

            elif freightster_pricing_zone == 'WA':
                if float(final_freightster_weight) <= 0.250:
                    freightster_price = 4.6

                elif 0.251 <= float(final_freightster_weight) <= 0.5:
                    freightster_price = 5.5

                elif 0.501 <= float(final_freightster_weight) <= 1.000:
                    freightster_price = 7.12

                elif 1.01 <= float(final_freightster_weight) <= 3.000:
                    freightster_price = 8.26

                elif 3.01 <= float(final_freightster_weight) <= 5.000:
                    freightster_price = 10.26

                elif 5.01 <= float(final_freightster_weight) <= 15:
                    freightster_price = 25.5

                elif 15.01 <= float(final_freightster_weight) <= 20:
                    freightster_price = 31.5

                else:
                    freightster_price = 1000

            else:
                freightster_price = 1000

            payload = {"order": {"serviceCode": 12,
                                 "consignee": {"company": company,
                                               "name": name,
                                               "address1": address1,
                                               "address2": address2,
                                               "city": city,
                                               "postcode": postcode,
                                               "state": state,
                                               "phone": phone,
                                               "email": email},
                                 "sender": {"name": "Kyal Scarlett",
                                            "address1": "286-288 Ballarat Rd",
                                            "address2": "",
                                            "city": "Footscray",
                                            "postcode": "3011",
                                            "state": "VIC",
                                            "phone": "0382563460",
                                            "email": "kyal@scarlettmusic.com.au"},
                                 "shipment": {"reference": item_string_counter,
                                              "description": item_string_counter,
                                              "weight": str(final_freightster_weight)}}}

            while True:

                r = requests.post('https://freightster.com.au/api/v1/shippingAPI/create', json=payload,
                                  headers=freightster_headers)

                if r.status_code == 429:
                    continue
                else:
                    break
            freightster_response = json.loads(r.text)
            freightster_price = freightster_price * 1.1  #####Adding GST
            if 'NEX' in freightster_response['response_data']['tracking_number']:
                freightster_price = 1000
            if freightster_response['status'] is False:
                freightster_price = 1000

        except:
            freightster_price = 1000

    if toll_failed == 'true':
        tollprice = 1000

    else:
        toll_item_array = []
        additional_costs = 0

        # for _ in range(int(all_items[6])):

        toll_item_array.append({"Commodity": {"CommodityCode": "Z",
                                              "CommodityDescription": "ALL FREIGHT"},
                                "ShipmentItemTotals": {
                                    "ShipmentItemCount": '1'},
                                "Dimensions": {
                                    "Width": str(math.ceil(float(width))),
                                    "Length": str(math.ceil(float(length))),
                                    "Height": str(math.ceil(float(height))),
                                    "Volume": str(volumevalue),
                                    "Weight": str(weightvalue)
                                }})

        package = [math.ceil(float(weight)), math.ceil(float(length)),
                   math.ceil(float(height))]
        sorted_package = sorted(package)

        if float(weightvalue) > 35:
            additional_costs += 77.55

        if float(weightvalue) > 35 or sorted_package[0] > 180 or sorted_package[1] > 180 or sorted_package[
            2] > 180 or float(volumevalue) > 0.7:
            additional_costs += 50

        elif float(weightvalue) > 30 or sorted_package[0] > 60 or sorted_package[1] > 80 or \
                sorted_package[2] > 120 or float(volumevalue) > 0.7:
            additional_costs += 12

        if float(totalprice) < 500:
            additional_costs += 6.95
        else:
            additional_costs += float(totalprice) * 0.02

        message_identifier = str(uuid.uuid4())
        now = datetime.datetime.now()
        current_time = f"{now.strftime('%Y-%m-%d')}T{now.strftime('%H:%M:%S')}+10:00"

        next_day = next_business_day()
        next_toll_day = f"{next_day.strftime('%Y-%m-%d')}T09:00:00+10:00"

        environment = 'PRD'  #######################Will have to change this in prod

        headers = {'Content-Type': 'application/json', 'Accept': 'application/json'}

        test_rate_enquiry_headers = {
            "MessageVersion": "3.1",
            "MessageIdentifier": message_identifier,
            "CreateTimestamp": current_time,
            "DocumentType": "RateEnquiry",
            "Environment": environment,
            "MessageSender": "SCARLETTMUSIC",
            "MessageReceiver": "TOLL",
            'SourceSystemCode': 'SCAR'
        }
        # rate_enquiry_test_url = "https://api-uat.teamglobalexp.com:6930/gateway/TollMessageRateEnquiryRestService/1.0/tom/rateEnquiry"
        rate_enquiry_prod_url = "https://api.teamglobalexp.com:6930/gateway/TollMessageRateEnquiryRestService/1.0/tom/rateEnquiry"
        price = totalprice
        payload = {
            "Request": {
                "BusinessID": "IPEC",
                "SystemFields": {"PickupDateTime": next_toll_day},
                "ShipmentService": {
                    "ServiceCode": "X",
                    "ServiceDescription": "Road Express"
                },
                "ShipmentFlags": {
                    "ExtraServiceFlag": "true"},
                "ShipmentFinancials": {
                    "ExtraServicesAmount": {
                        "Currency": "AUD",
                        "Value": str(round(float(totalprice)))}},
                "BillToParty": {
                    "AccountCode": "80119621"
                },
                "ConsignorParty": {
                    "PhysicalAddress": {
                        "Suburb": "FOOTSCRAY",
                        "StateCode": "VIC",
                        "PostalCode": "3011",
                        "CountryCode": "AU"
                    }
                },
                "ConsigneeParty": {
                    "PhysicalAddress": {
                        "Suburb": city,
                        "StateCode": state,
                        "PostalCode": str(postcode),
                        "CountryCode": "AU"
                    }
                },
                "ShipmentItems": {
                    "ShipmentItem": toll_item_array
                }
            }
        }

        toll_message = {"@version": "3.1",
                        "@encoding": "utf-8",
                        "TollMessage": {"Header": test_rate_enquiry_headers,
                                        "RateEnquiry": payload}}

        attempts = 0

        while attempts < 6:

            r = requests.post(url=rate_enquiry_prod_url, auth=('accounts@scarlettmusic.com.au', 't2TrAPsNTB'),
                              json=toll_message, headers=headers, timeout=100)

            response = r.text

            try:

                response = json.loads(response)

                if 'timeout' in response['TollMessage']['ErrorMessages']['ErrorMessage'][0]['ErrorMessage'].lower():
                    message_identifier = str(uuid.uuid4())
                    attempts += 1
                    continue

                else:
                    courierprice = 1000
                    break

            except:
                break

        try:
            tollprice = float(
                response['TollMessage']['RateEnquiry']['Response']['TotalChargeAmount']['Value']) + additional_costs

            # if 'wa' in state.lower():
            #     tollprice = tollprice * 1.15  # Remove once disaster levy has been lifted

            # print('Toll price: ' + str(tollprice))

        except:
            tollprice = 1000

    if 'vic' not in state.lower():
        cbdexpressfailed = 'true'

    cbdexpressfailed = 'true'

    if cbdexpressfailed == 'true':
        cbdprice = 1000

    else:

        cbdprice = 1000

        cbd_workbook = openpyxl.load_workbook(
            rf'\\SERVER\Project Folder\Python\Courier Info\CBD_Express_Areas.xlsx')

        cbd_sheet = cbd_workbook['Sheet']
        courierprice = 1000
        for xx in range(2, cbd_sheet.max_row + 1):
            cbd_suburb_check = cbd_sheet['A' + str(xx)].value.strip().lower()
            if cbd_suburb_check == city.lower().strip():
                cbdprice = float(cbd_sheet['C' + str(xx)].value)

    if allied_failed == 'true':
        alliedprice = 1000
        pass

    else:
        pickup_address = {'address1': "286-288 Ballarat Rd",
                          'address2': "",
                          'country': "Australia",
                          'postCode': '3011',
                          'state': 'VIC',
                          'suburb': 'Footscray'}

        reciever_address = {'address1': address1,
                            'address2': address2,
                            'country': "Australia",
                            'postCode': postcode,
                            'state': state,
                            'suburb': city}

        Jobstop_pickupstop = {'companyName': 'Scarlett Music',
                              'contact': 'Kyal Scarlett',
                              'emailAddress': 'info@scarlettmusic.com.au',
                              'geographicAddress': pickup_address,
                              'phoneNumber': '03 9318 5751',
                              'stopNumber': 1,
                              'stopType': 'P'}

        Jobstop_deliverystop = {'companyName': company,
                                'contact': name,
                                'emailAddress': email,
                                'geographicAddress': reciever_address,
                                'phoneNumber': phone,
                                'stopNumber': 2,
                                'stopType': 'D'}

        Jobstop_final = [Jobstop_pickupstop, Jobstop_deliverystop]

        allied_item_array = []
        allied_total_volume = 0
        allied_total_weight = 0
        allied_total_items = 0
        #   multi_items_to_ship.append((weightvalue, length, height, width,volumevalue, cubicweight, quantity, satchel))

        allied_item_array.append({'dangerous': 'false',
                                  'height': str(height),
                                  'itemCount': str(1),
                                  'length': str(length),
                                  'volume': str(volumevalue),
                                  'weight': str(weightvalue),
                                  'width': str(width)})

        allied_total_volume = str(volumevalue)
        allied_total_weight = str(weightvalue)
        allied_total_items = '1'

        # item = {'dangerous': 'false',
        #         'height': float(height),
        #         'itemCount': 1,
        #         'length': float(length),
        #         'volume': float(volumevalue),
        #         'weight': float(weightvalue),
        #         'width': float(width)}

        ##IF MULTIPLE ITEMS, easily put into dic, item = [{item1, item2}]

        Service_Level = 'R'  # Overnight Express

        pickup_instructions = "The music shop, open 9am-6pm. Best parking is at The Palms across the road."
        reference_number = item_string_counter
        booked_by = 'Kyal Scarlett'

        job_number = int(re.sub("[^0-9]", "", reference_number))

        Job = {'account': allied_account,
               'cubicWeight': cubicweight,
               'Docket': "SCM",
               'instructions': pickup_instructions,
               'cubedItems': allied_item_array,
               'itemCount': allied_total_items,
               'weight': allied_total_weight,
               'volume': allied_total_volume,
               'items': allied_item_array,
               'jobStops': Jobstop_final,
               'serviceLevel': Service_Level,
               'referenceNumbers': reference_number,
               'bookedBy': booked_by,
               'readyDate': today,
               'jobNumber': job_number,
               'vehicle': {'vehicleID': 1}}

        JobIDs = {'jobIds': job_number}

        try:

            Job = allied_client.service.validateBooking('755cf13abb3934695f03bd4a75cfbca7', Job)

            job_price = allied_client.service.calculatePrice('755cf13abb3934695f03bd4a75cfbca7', Job)
            alliedprice = job_price['totalCharge']
            alliedprice = round((float(alliedprice) * 1.269) * 1.1, 2)
            if alliedprice == 0.0:
                alliedprice = 1000
            # print('Allied price: $' + str(alliedprice))

        except zeep.exceptions.Fault:
            alliedprice = 1000

    if float(weightvalue) > 25:
        tailgate = 'true'
    else:
        tailgate = 'false'

    if transdirectfailed == 'true':
        transdirectprice = 1000
        pass

    else:

        if email == '':
            email = 'info@scarlettmusic.com.au'

        headers = {'Api-key': '74a242a1f4b8ba6ea5e011a80bdd2732', 'Content-Type': 'application/json',
                   "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
                   "Accept-Encoding": "*",
                   "Connection": "keep-alive"}

        try:

            try:
                float(phone)

            except ValueError:

                phone = '0417557472'

            payload = {
                "declared_value": "0",
                "referrer": "API",
                "requesting_site": "www.scarlettmusic.com.au",
                "tailgate_pickup": "false",
                "tailgate_delivery": str(tailgate),
                "items": [
                    {
                        "weight": str(weightvalue),
                        "height": str(height),
                        "width": str(width),
                        "length": str(length),
                        "quantity": 1,
                        "description": "carton"
                    }
                ],
                "sender": {
                    "address": "288 Ballarat Rd",
                    "company_name": "Scarlett Music",
                    "email": "info@scarlettmusic.com.au",
                    "name": "Lorelle Scarlett",
                    "postcode": "3011",
                    "phone": "0417557472",
                    "state": "VIC",
                    "suburb": "FOOTSCRAY",
                    "type": "business",
                    "country": "AU"
                },
                "receiver": {
                    "address": str(address1) + ', ' + str(address2),
                    "company_name": str(company),
                    "email": str(email),
                    "name": str(name),
                    "postcode": str(postcode),
                    "phone": int(phone),
                    "state": str(state),
                    "suburb": str(city),
                    "type": "residential",
                    "country": "AU"
                }
            }
            try:
                r = requests.post('https://www.transdirect.com.au/api/bookings/v4',
                                  auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'), headers=headers,
                                  json=payload)

                ##Mock API https://private-anon-2b516758f2-transdirectapiv4.apiary-mock.com/api/bookings/v4
                ###Legit API https://www.transdirect.com.au/api/bookings/v4

                response = r.text
                response = json.loads(response)
                if 'errors' in response:
                    try:

                        if 'ebay' in address1:
                            address = address2 + " " + city + " " + state + " " + postcode

                            # get_google_address()

                            dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                            qstr = urllib.parse.urlencode(dict1)
                            URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                            URL = URL + qstr
                            response = urllib.request.urlopen(URL)
                            data = json.load(response)
                            # pprint.pprint(data['candidates'])
                            placeid = data['candidates'][0]['place_id']
                            payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                            r = requests.get(
                                'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                            #  pprint.pprint(r.json())
                            data = r.json()
                            #   pprint.pprint(data['result']['address_components'])
                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'locality':
                                    suburb = data['result']['address_components'][t]['long_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'postal_code':
                                    postal_code = data['result']['address_components'][t]['long_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                                    google_state = data['result']['address_components'][t]['short_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'street_number':
                                    google_street1 = data['result']['address_components'][t]['short_name']
                                    break

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'route':
                                    google_street2 = data['result']['address_components'][t]['long_name']
                                    break
                            backup_google_street_address = google_street1 + ' ' + google_street2

                            postcode = postal_code
                            city = suburb
                            state = google_state

                        else:
                            address = address1 + " " + address2 + " " + city + " " + state + " " + postcode

                            # get_google_address()

                            dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                            qstr = urllib.parse.urlencode(dict1)
                            URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                            URL = URL + qstr
                            response = urllib.request.urlopen(URL)
                            data = json.load(response)
                            # pprint.pprint(data['candidates'])
                            placeid = data['candidates'][0]['place_id']
                            payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                            r = requests.get(
                                'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                            #    pprint.pprint(r.json())
                            data = r.json()
                            #  pprint.pprint(data['result']['address_components'])
                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'locality':
                                    suburb = data['result']['address_components'][t]['long_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'postal_code':
                                    postal_code = data['result']['address_components'][t]['long_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                                    google_state = data['result']['address_components'][t]['short_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'street_number':
                                    google_street1 = data['result']['address_components'][t]['short_name']
                                    break

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'route':
                                    google_street2 = data['result']['address_components'][t]['long_name']
                                    break
                            backup_google_street_address = google_street1 + ' ' + google_street2

                            postcode = postal_code
                            city = suburb
                            state = google_state

                        payload = {
                            "declared_value": "0",
                            "referrer": "API",
                            "requesting_site": "www.scarlettmusic.com.au",
                            "tailgate_pickup": "false",
                            "tailgate_delivery": str(tailgate),
                            "items": [
                                {
                                    "weight": str(weightvalue),
                                    "height": str(height),
                                    "width": str(width),
                                    "length": str(length),
                                    "quantity": 1,
                                    "description": "carton"
                                }
                            ],
                            "sender": {
                                "address": "288 Ballarat Rd",
                                "company_name": "Scarlett Music",
                                "email": "info@scarlettmusic.com.au",
                                "name": "Lorelle Scarlett",
                                "postcode": "3011",
                                "phone": "0417557472",
                                "state": "VIC",
                                "suburb": "FOOTSCRAY",
                                "type": "business",
                                "country": "AU"
                            },
                            "receiver": {
                                "address": str(backup_google_street_address),
                                "company_name": str(company),
                                "email": str(email),
                                "name": str(name),
                                "postcode": str(postcode),
                                "phone": int(phone),
                                "state": str(state),
                                "suburb": str(city),
                                "type": "residential",
                                "country": "AU"
                            }
                        }

                        try:
                            r = requests.post('https://www.transdirect.com.au/api/bookings/v4',
                                              auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'),
                                              headers=headers,
                                              json=payload)

                            ##Mock API https://private-anon-2b516758f2-transdirectapiv4.apiary-mock.com/api/bookings/v4
                            ###Legit API https://www.transdirect.com.au/api/bookings/v4

                            response = r.text
                            response = json.loads(response)

                            if 'id' in response:
                                google = 'true'
                        # pprint.pprint(response)

                        except json.decoder.JSONDecodeError:
                            transdirectprice = 1000
                    except:
                        transdirectprice = 1000
            # pprint.pprint(response)

            except json.decoder.JSONDecodeError:

                transdirectprice = 1000

            id = response['id']

            quotelen = len(response['quotes'])
            # print('Amount of quotes: ' + str(quotelen))
            namesofquotes = list(response['quotes'])
            # print(namesofquotes)

            quotes = {}

            for x in range(quotelen):
                quotes[str(namesofquotes[x])] = str(response['quotes'][namesofquotes[x]]['total'])

            # print(quotes)

            if 'couriers_please_domestic_proirity_authority' in response['quotes']:

                if 'tiers' in response['quotes']['couriers_please_domestic_proirity_authority']:

                    try:
                        quotes['couriers_please_multi_21'] = \
                            response['quotes']['couriers_please_domestic_proirity_authority']['tiers'][1]['total']

                        couriers_please_tiers = 'true'
                    except:
                        quotes['couriers_please_multi_21'] = '1000'

                else:
                    quotes['couriers_please_multi_21'] = '1000'

            if 'fastway' in response['quotes']:

                if 'tiers' in response['quotes']['fastway']:

                    quotes['fastway_multi_7'] = response['quotes']['fastway']['tiers'][1]['total']

                    fastway_tiers = 'true'

                else:
                    quotes['fastway_multi_7'] = '1000'

            if float(totalprice) > 250:
                quotes['couriers_please_domestic_proirity_authority'] = '1000'
                quotes['couriers_please_multi_21'] = '1000'

            intquotes = dict(
                (k, float(v)) for k, v in quotes.items())  ### Converting all values into FLOAT and then into INT

            try:

                lowesttranscourier = min(intquotes, key=intquotes.get)

                transdirectprice = quotes[lowesttranscourier]
            except ValueError:

                transdirectprice = 1000




        except (KeyError, TypeError):

            try:

                if 'ebay' in address1:
                    address = address2 + " " + city + " " + state + " " + postcode

                    # get_google_address()

                    dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    qstr = urllib.parse.urlencode(dict1)
                    URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                    URL = URL + qstr
                    response = urllib.request.urlopen(URL)
                    data = json.load(response)
                    # pprint.pprint(data['candidates'])
                    placeid = data['candidates'][0]['place_id']
                    payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    r = requests.get(
                        'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                    # pprint.pprint(r.json())
                    data = r.json()
                    #    pprint.pprint(data['result']['address_components'])
                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'locality':
                            suburb = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'postal_code':
                            postal_code = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                            google_state = data['result']['address_components'][t]['short_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'street_number':
                            google_street1 = data['result']['address_components'][t]['short_name']
                            break

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'route':
                            google_street2 = data['result']['address_components'][t]['long_name']
                            break
                    backup_google_street_address = google_street1 + ' ' + google_street2

                    postcode = postal_code
                    city = suburb
                    state = google_state

                else:
                    address = address1 + " " + address2 + " " + city + " " + state + " " + postcode

                    # get_google_address()

                    dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    qstr = urllib.parse.urlencode(dict1)
                    URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                    URL = URL + qstr
                    response = urllib.request.urlopen(URL)
                    data = json.load(response)
                    # pprint.pprint(data['candidates'])
                    placeid = data['candidates'][0]['place_id']
                    payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    r = requests.get(
                        'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                    #   pprint.pprint(r.json())
                    data = r.json()
                    #   pprint.pprint(data['result']['address_components'])
                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'locality':
                            suburb = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'postal_code':
                            postal_code = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                            google_state = data['result']['address_components'][t]['short_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'street_number':
                            google_street1 = data['result']['address_components'][t]['short_name']
                            break

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'route':
                            google_street2 = data['result']['address_components'][t]['long_name']
                            break
                    backup_google_street_address = google_street1 + ' ' + google_street2

                    postcode = postal_code
                    city = suburb
                    state = google_state

                payload = {
                    "declared_value": "0",
                    "referrer": "API",
                    "requesting_site": "www.scarlettmusic.com.au",
                    "tailgate_pickup": "false",
                    "tailgate_delivery": str(tailgate),
                    "items": [
                        {
                            "weight": str(weightvalue),
                            "height": str(height),
                            "width": str(width),
                            "length": str(length),
                            "quantity": 1,
                            "description": "carton"
                        }
                    ],
                    "sender": {
                        "address": "288 Ballarat Rd",
                        "company_name": "Scarlett Music",
                        "email": "info@scarlettmusic.com.au",
                        "name": "Lorelle Scarlett",
                        "postcode": "3011",
                        "phone": "0417557472",
                        "state": "VIC",
                        "suburb": "FOOTSCRAY",
                        "type": "business",
                        "country": "AU"
                    },
                    "receiver": {
                        "address": str(backup_google_street_address),
                        "company_name": str(company),
                        "email": str(email),
                        "name": str(name),
                        "postcode": str(postcode),
                        "phone": int(phone),
                        "state": str(state),
                        "suburb": str(city),
                        "type": "residential",
                        "country": "AU"
                    }
                }
                try:
                    r = requests.post('https://www.transdirect.com.au/api/bookings/v4',
                                      auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'), headers=headers,
                                      json=payload)

                    ##Mock API https://private-anon-2b516758f2-transdirectapiv4.apiary-mock.com/api/bookings/v4
                    ###Legit API https://www.transdirect.com.au/api/bookings/v4

                    response = r.text
                    response = json.loads(response)
                # pprint.pprint(response)

                except json.decoder.JSONDecodeError:

                    transdirectprice = 1000

                id = response['id']

                quotelen = len(response['quotes'])
                # print('Amount of quotes: ' + str(quotelen))
                namesofquotes = list(response['quotes'])
                # print(namesofquotes)

                quotes = {}

                for x in range(quotelen):
                    quotes[str(namesofquotes[x])] = str(response['quotes'][namesofquotes[x]]['total'])

                if 'couriers_please_domestic_proirity_authority' in response['quotes']:

                    if 'tiers' in response['quotes']['couriers_please_domestic_proirity_authority']:

                        try:
                            quotes['couriers_please_multi_21'] = \
                                response['quotes']['couriers_please_domestic_proirity_authority']['tiers'][1]['total']

                            couriers_please_tiers = 'true'
                        except:
                            quotes['couriers_please_multi_21'] = '1000'

                    else:
                        quotes['couriers_please_multi_21'] = '1000'

                if 'fastway' in response['quotes']:

                    if 'tiers' in response['quotes']['fastway']:

                        quotes['fastway_multi_7'] = response['quotes']['fastway']['tiers'][1]['total']

                        fastway_tiers = 'true'

                    else:
                        quotes['fastway_multi_7'] = '1000'

                if float(totalprice) > 250:
                    quotes['couriers_please_domestic_proirity_authority'] = '1000'
                    quotes['couriers_please_multi_21'] = '1000'

                intquotes = dict(
                    (k, float(v)) for k, v in quotes.items())  ### Converting all values into FLOAT and then into INT
                lowesttranscourier = min(intquotes, key=intquotes.get, default='NA')

                transdirectprice = quotes[lowesttranscourier]

            except (KeyError, TypeError, IndexError):

                transdirectprice = 1000

    #   print('Transdirect price: ' + str(transdirectprice))

    if fastwayfailed == 'true':
        fastwayprice = 1000

    else:

        if email == '':
            email = 'info@scarlettmusic.com.au'

        if phone is None:
            phone = '0417557472'

        phone = phone.replace(' ', '')

        ####### INITIAL FASTWAY VALIDATION ######

        if 'ebay' in address1:

            street_address = str(address2)

            data = {
                "streetAddress": street_address,
                "additionalDetails": "",
                "locality": str(city),
                "stateOrProvince": str(state),
                "postalCode": str(postcode),
                "country": "AU",
                "lat": 0,
                "lng": 0,
                "userCreated": False,
            }


        else:

            street_address = str(address1) + ', ' + str(address2)

            data = {
                "streetAddress": street_address,
                "additionalDetails": "",
                "locality": str(city),
                "stateOrProvince": str(state),
                "postalCode": str(postcode),
                "country": "AU",
                "lat": 0,
                "lng": 0,
                "userCreated": False,
            }

        r = requests.post(base_url + '/api/addresses/validate', headers=Fastway_Headers, json=data)
        response = r.text
        response = json.loads(response)
        # pprint.pprint(response)

        try:
            country = response['data']['country']
            addressid = response['data']['addressId']
            latitude = response['data']['lat']
            longitude = response['data']['lng']
            suburb = response['data']['locality']
            placeId = response['data']['placeId']
            postcode = response['data']['postalCode']
            state = response['data']['stateOrProvince']
            street_address = response['data']['streetAddress']
            hash = response['data']['hash']

            # package = [float(length), float(width), float(height)]
            # sorted_package = sorted(package)

            package = [float(length), float(width), float(height)]
            sorted_package = sorted(package)

            if float(weightvalue) <= 0.3 and float(sorted_package[0]) <= 3 and float(sorted_package[1]) <= 21 and float(
                    sorted_package[2]) <= 21:
                satchel = '300gm'

            elif float(cubicweight) <= 0.5 and float(weightvalue) <= 0.5:
                satchel = 'A5'

            elif float(cubicweight) <= 1 and float(weightvalue) <= 1:
                satchel = 'A4'

            elif float(cubicweight) <= 3 and float(weightvalue) <= 3:
                satchel = 'A3'

            elif float(cubicweight) <= 5 and float(weightvalue) <= 5:
                satchel = 'A2'

            if satchel == 'false':

                fastway_data = {
                    "To": {
                        "ContactName": name,
                        "BusinessName": company,
                        "PhoneNumber": phone,
                        "Email": email,
                        "Address": {
                            "StreetAddress": street_address,
                            "Locality": suburb,
                            "StateOrProvince": state,
                            "PostalCode": postcode,
                            "Country": country
                        }
                    },
                    "Items": [
                        {
                            "Quantity": 1,
                            "Reference": item_string_counter,
                            "PackageType": "P",
                            "WeightDead": weightvalue,
                            "WeightCubic": cubicweight,
                            "Length": length,
                            "Width": width,
                            "Height": height
                        }
                    ],
                    "ExternalRef1": item_string_counter,
                    "ExternalRef2": item_string_counter,
                }

            else:

                fastway_data = {
                    "To": {
                        "ContactName": name,
                        "BusinessName": company,
                        "PhoneNumber": phone,
                        "Email": email,
                        "Address": {
                            "StreetAddress": street_address,
                            "Locality": suburb,
                            "StateOrProvince": state,
                            "PostalCode": postcode,
                            "Country": country
                        }
                    },
                    "Items": [
                        {
                            "Quantity": 1,
                            "PackageType": "S",
                            "Reference": item_string_counter,
                            "SatchelSize": satchel
                        }
                    ],
                    "ExternalRef1": item_string_counter,
                    "ExternalRef2": item_string_counter,
                }

            r = requests.post(base_url + '/api/consignments/quote', headers=Fastway_Headers, json=fastway_data)
            response = r.text
            response = json.loads(response)

            fastwayprice = response['data']['total']

            # print('Fastway price: ' + str(fastwayprice))

        except:

            try:

                if 'ebay' in address1:
                    address = address2 + " " + city + " " + state + " " + postcode

                    # get_google_address()

                    dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    qstr = urllib.parse.urlencode(dict1)
                    URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                    URL = URL + qstr
                    response = urllib.request.urlopen(URL)
                    data = json.load(response)
                    # pprint.pprint(data['candidates'])
                    placeid = data['candidates'][0]['place_id']
                    payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    r = requests.get(
                        'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                    #  pprint.pprint(r.json())
                    data = r.json()
                    #   pprint.pprint(data['result']['address_components'])
                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'locality':
                            suburb = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'postal_code':
                            postal_code = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                            google_state = data['result']['address_components'][t]['short_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'subpremise':
                            google_street0 = data['result']['address_components'][t]['short_name']
                            break
                        else:
                            google_street0 = ''

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'street_number':
                            google_street1 = data['result']['address_components'][t]['short_name']
                            break

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'route':
                            google_street2 = data['result']['address_components'][t]['long_name']
                            break

                    if google_street0 == '':

                        backup_google_street_address = google_street1 + ' ' + google_street2

                    else:

                        backup_google_street_address = google_street0 + '/' + google_street1 + ' ' + google_street2

                    postcode = postal_code
                    city = suburb
                    state = google_state

                else:
                    address = address1 + " " + address2 + " " + city + " " + state + " " + postcode

                    # get_google_address()

                    dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    qstr = urllib.parse.urlencode(dict1)
                    URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                    URL = URL + qstr
                    response = urllib.request.urlopen(URL)
                    data = json.load(response)
                    # pprint.pprint(data['candidates'])
                    placeid = data['candidates'][0]['place_id']
                    payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    r = requests.get(
                        'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                    #    pprint.pprint(r.json())
                    data = r.json()
                    #  pprint.pprint(data['result']['address_components'])
                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'locality':
                            suburb = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'postal_code':
                            postal_code = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                            google_state = data['result']['address_components'][t]['short_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'subpremise':
                            google_street0 = data['result']['address_components'][t]['short_name']
                            break
                        else:
                            google_street0 = ''

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'street_number':
                            google_street1 = data['result']['address_components'][t]['short_name']
                            break

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'route':
                            google_street2 = data['result']['address_components'][t]['long_name']
                            break

                    if google_street0 == '':

                        backup_google_street_address = google_street1 + ' ' + google_street2

                    else:

                        backup_google_street_address = google_street0 + ' ' + google_street1 + ' ' + google_street2

                    postcode = postal_code
                    city = suburb
                    state = google_state

                data = {
                    "streetAddress": str(backup_google_street_address),
                    "additionalDetails": "",
                    "locality": str(city),
                    "stateOrProvince": str(state),
                    "postalCode": str(postcode),
                    "country": "AU",
                    "lat": 0,
                    "lng": 0,
                    "userCreated": False,
                }

                try:
                    r = requests.post(base_url + '/api/addresses/validate', headers=Fastway_Headers, json=data)
                    response = r.text
                    response = json.loads(response)
                    # pprint.pprint(response)

                    try:
                        country = response['data']['country']
                        addressid = response['data']['addressId']
                        latitude = response['data']['lat']
                        longitude = response['data']['lng']
                        suburb = response['data']['locality']
                        placeId = response['data']['placeId']
                        Postcode = response['data']['postalCode']
                        State = response['data']['stateOrProvince']
                        AddressLine1 = response['data']['streetAddress']
                        hash = response['data']['hash']

                        if satchel == 'false':

                            fastway_data = {
                                "To": {
                                    "ContactName": name,
                                    "BusinessName": company,
                                    "PhoneNumber": phone,
                                    "Email": email,
                                    "Address": {
                                        "StreetAddress": street_address,
                                        "Locality": suburb,
                                        "StateOrProvince": state,
                                        "PostalCode": postcode,
                                        "Country": country
                                    }
                                },
                                "Items": [
                                    {
                                        "Quantity": 1,
                                        "Reference": item_string_counter,
                                        "PackageType": "P",
                                        "WeightDead": weightvalue,
                                        "WeightCubic": cubicweight,
                                        "Length": length,
                                        "Width": width,
                                        "Height": height
                                    }
                                ],
                                "ExternalRef1": item_string_counter,
                                "ExternalRef2": item_string_counter,
                            }

                        else:

                            fastway_data = {
                                "To": {
                                    "ContactName": name,
                                    "BusinessName": company,
                                    "PhoneNumber": phone,
                                    "Email": email,
                                    "Address": {
                                        "StreetAddress": street_address,
                                        "Locality": suburb,
                                        "StateOrProvince": state,
                                        "PostalCode": postcode,
                                        "Country": country
                                    }
                                },
                                "Items": [
                                    {
                                        "Quantity": 1,
                                        "PackageType": "S",
                                        "SatchelSize": satchel
                                    }
                                ],
                                "ExternalRef1": item_string_counter,
                                "ExternalRef2": item_string_counter,
                            }

                        r = requests.post(base_url + '/api/consignments/quote', headers=Fastway_Headers,
                                          json=fastway_data)
                        response = r.text
                        response = json.loads(response)

                        fastwayprice = response['data']['total']

                        # print('Fastway price: ' + str(fastwayprice))

                    except:

                        fastwayprice = 1000

                except:

                    fastwayprice = 1000
            except:
                fastwayprice = 1000

    couriers_failed = 'true'
    if couriers_failed == 'true':
        courierprice = 1000

    else:

        if email == '':
            email = 'info@scarlettmusic.com.au'

        if phone is None:
            phone = '0417557472'

        phone = phone.replace(' ', '')

        try:
            int(phone)

        except ValueError:

            phone = '0417557472'

        cp_body = {
            "fromSuburb": "Footscray",
            "fromPostcode": 3011,
            "toSuburb": suburb,
            "toPostcode": postcode,
            "items": [{
                "length": math.ceil(float(length)),
                "height": math.ceil(float(height)),
                "width": math.ceil(float(width)),
                "physicalWeight": float(weightvalue),
                "quantity": 1
            }]
        }

        additional_costs = 0
        package = [math.ceil(float(length)), math.ceil(float(height)),
                   math.ceil(float(width))]
        sorted_package = sorted(package)
        if sorted_package[0] > 105 or sorted_package[1] > 105 or sorted_package[
            2] > 105:
            additional_costs += 15

        cp_url = 'https://api.couriersplease.com.au/v2/domestic/quote'

        try:

            r = requests.post(cp_url, headers=cp_headers, json=cp_body)
        except:
            courierprice = 1000

        try:
            response = r.text
            response = json.loads(response)
            # pprint.pprint(response)

            quotelen = len(response['data'])
            # print('Amount of quotes: ' + str(quotelen))
            # namesofquotes = list(response['quotes'])
            # print(namesofquotes)

            quotes = {}

            for x in range(quotelen):
                quotes[str(response['data'][x]['RateCardCode'])] = str((float(
                    response['data'][x]['CalculatedFreightCharge']) + float(
                    response['data'][x]['CalculatedFuelCharge'])) * 1.1)

            # pprint.pprint(quotes)

            intquotes = dict(
                (k, float(v)) for k, v in quotes.items())  ### Converting all values into FLOAT and then into INT
            lowest_cp_service = min(intquotes, key=intquotes.get)

            courierprice = float(quotes[lowest_cp_service])*1.025

            courierprice+= additional_costs

            cp_url = 'https://api.couriersplease.com.au/v1/domestic/shipment/validate'
            # print(lowest_cp_service)

            cp_validate_body = {
                "pickupFirstName": "Kyal",
                "pickupLastName": "Scarlett",
                "pickupCompanyName": "Scarlett Music",
                "pickupEmail": "kyal@scarlettmusic.com.au",
                "pickupAddress1": "286-288 Ballarat Rd",
                "pickupAddress2": "",
                "pickupSuburb": "Footscray",
                "pickupState": "VIC",
                "pickupPostcode": "3011",
                "pickupPhone": "0393185751",
                "pickupIsBusiness": "true",
                "destinationFirstName": first_name,
                "destinationLastName": last_name,
                "destinationCompanyName": company,
                "destinationEmail": email,
                "destinationAddress1": address1,
                "destinationAddress2": address2,
                "destinationSuburb": city,
                "destinationState": state,
                "destinationPostcode": postcode,
                "destinationPhone": phone,
                "destinationIsBusiness": "false",
                "contactFirstName": "Kyal",
                "contactLastName": "Scarlett",
                "contactCompanyName": "Scarlett Music",
                "contactEmail": "kyal@scarlettmusic.com.au",
                "contactAddress1": "286-288 Ballarat Rd",
                "contactAddress2": "",
                "contactSuburb": "Footscray",
                "contactState": "VIC",
                "contactPostcode": "3011",
                "contactPhone": "0393185751",
                "contactIsBusiness": "true",
                "referenceNumber": item_string_counter,
                "termsAccepted": "true",
                "dangerousGoods": "false",
                "rateCardId": lowest_cp_service,
                "specialInstruction": "",
                "isATL": "false",
                "readyDateTime": next_cp_day,
                "items": [{
                    "length": math.ceil(float(length)),
                    "height": math.ceil(float(height)),
                    "width": math.ceil(float(width)),
                    "physicalWeight": float(weightvalue),
                    "quantity": 1
                }]}

            r = requests.post(cp_url, headers=cp_headers, json=cp_validate_body)
            response = r.text
            response = json.loads(response)
            # pprint.pprint(response)

            if response['responseCode'] != 'SUCCESS':
                ##create shipment / print label
                cp_price = 1000

            # print(f"Courier's Please: {courierprice}")
        except:

            try:

                if 'ebay' in address1:
                    address = address2 + " " + city + " " + state + " " + postcode

                    # get_google_address()

                    dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    qstr = urllib.parse.urlencode(dict1)
                    URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                    URL = URL + qstr
                    response = urllib.request.urlopen(URL)
                    data = json.load(response)
                    # pprint.pprint(data['candidates'])
                    placeid = data['candidates'][0]['place_id']
                    payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    r = requests.get(
                        'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                    #  pprint.pprint(r.json())
                    data = r.json()
                    #   pprint.pprint(data['result']['address_components'])
                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'locality':
                            suburb = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'postal_code':
                            postal_code = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][
                            0] == 'administrative_area_level_1':
                            google_state = data['result']['address_components'][t]['short_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'subpremise':
                            google_street0 = data['result']['address_components'][t]['short_name']
                            break
                        else:
                            google_street0 = ''

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'street_number':
                            google_street1 = data['result']['address_components'][t]['short_name']
                            break

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'route':
                            google_street2 = data['result']['address_components'][t]['long_name']
                            break

                    if google_street0 == '':

                        backup_google_street_address = google_street1 + ' ' + google_street2

                    else:

                        backup_google_street_address = google_street0 + '/' + google_street1 + ' ' + google_street2

                    postcode = postal_code
                    city = suburb
                    state = google_state

                else:
                    address = address1 + " " + address2 + " " + city + " " + state + " " + postcode

                    # get_google_address()

                    dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    qstr = urllib.parse.urlencode(dict1)
                    URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                    URL = URL + qstr
                    response = urllib.request.urlopen(URL)
                    data = json.load(response)
                    # pprint.pprint(data['candidates'])
                    placeid = data['candidates'][0]['place_id']
                    payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                    r = requests.get(
                        'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                    #    pprint.pprint(r.json())
                    data = r.json()
                    #  pprint.pprint(data['result']['address_components'])
                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'locality':
                            suburb = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'postal_code':
                            postal_code = data['result']['address_components'][t]['long_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][
                            0] == 'administrative_area_level_1':
                            google_state = data['result']['address_components'][t]['short_name']
                            break

                        else:
                            continue

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'subpremise':
                            google_street0 = data['result']['address_components'][t]['short_name']
                            break
                        else:
                            google_street0 = ''

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'street_number':
                            google_street1 = data['result']['address_components'][t]['short_name']
                            break

                    for t in range(20):
                        if data['result']['address_components'][t]['types'][0] == 'route':
                            google_street2 = data['result']['address_components'][t]['long_name']
                            break

                    if google_street0 == '':

                        backup_google_street_address = google_street1 + ' ' + google_street2

                    else:

                        backup_google_street_address = google_street0 + '/' + google_street1 + ' ' + google_street2

                    city = suburb
                    state = google_state

                cp_body = {
                    "fromSuburb": "Footscray",
                    "fromPostcode": 3011,
                    "toSuburb": suburb,
                    "toPostcode": postcode,
                    "items": [{
                        "length": math.ceil(float(length)),
                        "height": math.ceil(float(height)),
                        "width": math.ceil(float(width)),
                        "physicalWeight": float(weightvalue),
                        "quantity": 1
                    }]
                }

                lowest_cp_service = ''

                cp_validate_body = {
                    "pickupFirstName": "Kyal",
                    "pickupLastName": "Scarlett",
                    "pickupCompanyName": "Scarlett Music",
                    "pickupEmail": "kyal@scarlettmusic.com.au",
                    "pickupAddress1": "286-288 Ballarat Rd",
                    "pickupAddress2": "",
                    "pickupSuburb": "Footscray",
                    "pickupState": "VIC",
                    "pickupPostcode": "3011",
                    "pickupPhone": "0393185751",
                    "pickupIsBusiness": "true",
                    "destinationFirstName": first_name,
                    "destinationLastName": last_name,
                    "destinationCompanyName": company,
                    "destinationEmail": email,
                    "destinationAddress1": address1,
                    "destinationAddress2": address2,
                    "destinationSuburb": city,
                    "destinationState": state,
                    "destinationPostcode": postcode,
                    "destinationPhone": phone,
                    "destinationIsBusiness": "false",
                    "contactFirstName": "Kyal",
                    "contactLastName": "Scarlett",
                    "contactCompanyName": "Scarlett Music",
                    "contactEmail": "kyal@scarlettmusic.com.au",
                    "contactAddress1": "286-288 Ballarat Rd",
                    "contactAddress2": "",
                    "contactSuburb": "Footscray",
                    "contactState": "VIC",
                    "contactPostcode": "3011",
                    "contactPhone": "0393185751",
                    "contactIsBusiness": "true",
                    "referenceNumber": item_string_counter,
                    "termsAccepted": "true",
                    "dangerousGoods": "false",
                    "rateCardId": lowest_cp_service,
                    "specialInstruction": "",
                    "isATL": "false",
                    "readyDateTime": next_cp_day,
                    "items": [{
                        "length": math.ceil(float(length)),
                        "height": math.ceil(float(height)),
                        "width": math.ceil(float(width)),
                        "physicalWeight": float(weightvalue),
                        "quantity": 1
                    }]}

                additional_costs = 0
                package = [math.ceil(float(length)), math.ceil(float(height)),
                           math.ceil(float(width))]
                sorted_package = sorted(package)
                if sorted_package[0] > 105 or sorted_package[1] > 105 or sorted_package[
                    2] > 105:
                    additional_costs += 15

                cp_url = 'https://api.couriersplease.com.au/v2/domestic/quote'

                r = requests.post(cp_url, headers=cp_headers, json=cp_body)

                response = r.text
                response = json.loads(response)
                # pprint.pprint(response)

                quotelen = len(response['data'])
                # print('Amount of quotes: ' + str(quotelen))
                # namesofquotes = list(response['quotes'])
                # print(namesofquotes)

                quotes = {}

                for x in range(quotelen):
                    quotes[str(response['data'][x]['RateCardCode'])] = str((float(
                        response['data'][x]['CalculatedFreightCharge']) + float(
                        response['data'][x]['CalculatedFuelCharge'])) * 1.1)

                # pprint.pprint(quotes)

                intquotes = dict(
                    (k, float(v)) for k, v in quotes.items())  ### Converting all values into FLOAT and then into INT
                lowest_cp_service = min(intquotes, key=intquotes.get)

                courierprice = float(quotes[lowest_cp_service])*1.025

                if 'wa' in state.lower():
                    courierprice = courierprice*1.20

                courierprice += additional_costs

                cp_url = 'https://api.couriersplease.com.au/v1/domestic/shipment/validate'
                # print(lowest_cp_service)

                r = requests.post(cp_url, headers=cp_headers, json=cp_validate_body)
                response = r.text
                response = json.loads(response)
                # pprint.pprint(response)

                if response['responseCode'] != 'SUCCESS':
                    ##create shipment / print label
                    cp_price = 1000

            # print(f"Courier's Please: {courierprice}")
            except:
                courierprice = 1000

    if weightvalue == 0:
        transdirectprice = 1000

    try:
        dai_price
    except:
        dai_price = 1000

    try:
        if float(dai_price) == float(fastwayprice):
            dai_price = dai_price + 0.01
    except:
        pass

    try:
        if float(dai_price) == float(sendleprice):
            dai_price = dai_price + 0.01
    except:
        pass

    lowestcourier = {'auspostprice': float(auspostprice), 'sendleprice': float(sendleprice),
                     'transdirectprice': float(transdirectprice), 'fastwayprice': float(fastwayprice),
                     'courierprice': float(courierprice), 'freightsterprice': float(freightster_price),
                     'cbdprice': float(cbdprice), 'alliedprice': float(alliedprice), 'tollprice': float(tollprice),
                     'daiprice': float(dai_price)}
    lowestcourier = min(lowestcourier, key=lowestcourier.get)
    print(lowestcourier + ' booked!')

    if lowestcourier == 'sendleprice':
        finalcourier = 'Sendle'

    if lowestcourier == 'auspostprice':
        finalcourier = 'Australia Post'

    if lowestcourier == 'fastwayprice':
        finalcourier = 'Fastway'

    if lowestcourier == 'courierprice':
        finalcourier = 'Couriers Please'

    if lowestcourier == 'freightsterprice':
        finalcourier = 'Freightster'

    if lowestcourier == 'alliedprice':
        finalcourier = 'Allied Express'

    if lowestcourier == 'tollprice':
        finalcourier = 'Toll'

    if lowestcourier == 'daiprice':
        finalcourier = 'Dai Post'

    if lowestcourier == 'cbdprice':
        finalcourier = 'CBDExpress'

    if lowestcourier == 'transdirectprice':
        finalcourier = 'Transdirect'

        lowesttranscourier = lowesttranscourier.replace('_', ' ')
        lowesttranscourier = lowesttranscourier.title()
        lowesttranscourier = lowesttranscourier.replace('Tnt', 'TNT')

    try:
        id
    except NameError:
        id = ''

    try:
        fastway_data
    except NameError:
        fastway_data = 'NA'

    try:
        lowesttranscourier
    except NameError:
        lowesttranscourier = 'NA'

    try:
        Job
    except:
        Job = ''

    try:
        JobIDs
    except:
        JobIDs = ''

    try:
        cp_validate_body
    except NameError:
        cp_validate_body = ''

    return finalcourier, lowesttranscourier, id, SatchNumber, fastway_data, cp_validate_body, Job, JobIDs


with concurrent.futures.ThreadPoolExecutor(max_workers=None) as executor:
    results = executor.map(SatchelQuote, Satchel_Position)

    for g in results:

        ebayresults[int(g[3])].update({'Final_Courier': g[0]})

        if g[0] == 'Transdirect':
            ebayresults[g[3]].update({'Final_Transdirect_Courier': g[1]})
            ebayresults[g[3]].update({'Transdirect_ID': g[2]})

        if g[0] == 'Fastway':
            ebayresults[g[3]].update({'fastway_data': g[4]})
        if g[0] == 'Couriers Please':
            ebayresults[g[3]].update({'cp_validate_body': g[5]})

        if g[0] == 'Allied Express':
            ebayresults[g[3]].update({'Job': g[6]})
            ebayresults[g[3]].update({'JobIDs': g[7]})

# pprint.pprint(ebayresults)

for x in ebayresults:

    if 'satchel' in x['Final_Postage'].lower():

        if x['Final_Courier'] == 'Fastway':

            try:

                x['Phone'] = x['Phone'].replace('+61', '0')
                x['Phone'] = x['Phone'].replace(' ', '')

                fastway_data = x['fastway_data']

                fastway_attempts = 0

                while True:

                    try:

                        r = requests.post(base_url + '/api/consignments', headers=Fastway_Headers, json=fastway_data,
                                          timeout=120)
                        response = r.text
                        response = json.loads(response)

                    except:
                        fastway_attempts += 1
                        if fastway_attempts > 5:
                            break

                        continue

                    if 'errors' in response:
                        break

                    else:
                        break
                # pprint.pprint(response)

                id = response['data']['conId']
                tracking_number = response['data']['items'][0]['label']

                r = requests.get(base_url + '/api/consignments/' + str(id) + '/labels?pageSize=4x6',
                                 headers=Fastway_Headers)
                response = r.content

                label_name = label_location + "\\" + str(x['String_SalesRecordNumber']) + ".pdf"

                Label_Location.append(label_name)

                with open(label_name, 'wb') as f:
                    f.write(response)

                for lines in range(len(x['SKU'])):
                    x['ItemTitle'][f'Name (Item {str(lines + 1)})'] = x['ItemTitle'][
                        f'Name (Item {str(lines + 1)})'].replace("'", "")

                    x['BuyerName'] = x['BuyerName'].replace("'", "")
                    x['Company'] = x['Company'].replace("'", "")
                    x['AddressLine1'] = x['AddressLine1'].replace("'", "")
                    x['AddressLine2'] = x['AddressLine2'].replace("'", "")
                    x['City'] = x['City'].replace("'", "")
                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()

                    try:

                        cursor.execute(
                            fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order,Sales_Record_Number , Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{x['String_OrderID'][0:50]}', '{x['String_PurchaseOrderNumber'][0:50]}', '{x['String_SalesRecordNumber'][0:50]}', 'ebay', '{x['BuyerName'][0:50]}', '{x['Company'][0:50]}', '{x['AddressLine1'][0:50]}', '{x['AddressLine2'][0:50]}', '{x['City'][0:50]}', '{x['State']}', '{x['Postcode']}', 'AU', '{x['Phone'][0:50]}', '{x['Email'][0:50]}', '{x['date_placed']}','' , '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '{x['OrderLineID'][f'Orderline (Item {str(lines + 1)})'][0:50]}', '{x['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{x['ItemTitle'][f'Name (Item {str(lines + 1)})'][0:50]}', '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '{x['String_OrderID'][0:50]}', '{x['Final_Courier']}','{tracking_number}' , '{str(x['item_count'])}', '{str(x['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                    except:
                        pass
            except:
                pass

        if x['Final_Courier'] == 'Transdirect':

            try:

                id = x['Transdirect_ID']
                Final_Transdirect_Courier = x['Final_Transdirect_Courier']

                if Final_Transdirect_Courier == 'Couriers Please Multi 21' or Final_Transdirect_Courier == 'Fastway Multi 7':

                    if Final_Transdirect_Courier == 'Couriers Please Multi 21':
                        courier = 'couriers_please_domestic_proirity_authority'
                        tier = 2
                    elif Final_Transdirect_Courier == 'Fastway Multi 7':
                        courier = 'fastway'
                        tier = 2

                    # print(tomorrow)
                    payload = {
                        "courier": courier,
                        "pickup-date": str(next_day),
                        "tier": tier
                    }
                    headers = {'Api-key': '74a242a1f4b8ba6ea5e011a80bdd2732', 'Content-Type': 'application/json',
                               "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
                               "Accept-Encoding": "*",
                               "Connection": "keep-alive"}

                    r = requests.post('https://www.transdirect.com.au/api/bookings/v4/' + str(id) + '/confirm',
                                      auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'), headers=headers,
                                      json=payload)
                    response = r.text
                    # print(response)
                    if response == '':
                        pass
                    # print('Success!')
                    elif response == '{"errors":["Pickup dates invalid."]}':
                        tomorrow = next_business_day()
                        while response == '{"errors":["Pickup dates invalid."]}':
                            tomorrow = tomorrow + datetime.timedelta(days=1)
                            #  print(tomorrow)

                            payload = {"courier": str(courier), "pickup-date": str(tomorrow), "tier": tier}
                            r = requests.post('https://www.transdirect.com.au/api/bookings/v4/' + str(id) + '/confirm',
                                              auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'),
                                              headers=headers,
                                              json=payload)
                            response = r.text
                            # print(response)
                            if response == '':
                                # print('Success!')
                                break
                            continue

                else:
                    x['Phone'] = x['Phone'].replace('+61', '0')
                    x['Phone'] = x['Phone'].replace(' ', '')

                    maxrow = wsTransdirect.max_row
                    maxrow = maxrow + 1
                    wsTransdirect['A' + str(maxrow)].value = x['String_SalesRecordNumber']
                    wsTransdirect['B' + str(maxrow)].value = 'Scarlett Music Order ' + x['String_SalesRecordNumber']
                    wsTransdirect['C' + str(maxrow)].value = x['Final_Transdirect_Courier']
                    wsTransdirect['D' + str(maxrow)].value = next_day
                    wsTransdirect['E' + str(maxrow)].value = 'Private'
                    wsTransdirect['F' + str(maxrow)].value = x['BuyerName']
                    wsTransdirect['G' + str(maxrow)].value = x['Email']
                    wsTransdirect['H' + str(maxrow)].value = x['Phone']
                    wsTransdirect['I' + str(maxrow)].value = x['Company']

                    if 'ebay' in x['AddressLine1']:
                        wsTransdirect['J' + str(maxrow)].value = x['AddressLine2'].strip()
                    else:
                        wsTransdirect['J' + str(maxrow)].value = x['AddressLine1'].strip() + ' ' + x[
                            'AddressLine2'].strip()

                    wsTransdirect['K' + str(maxrow)].value = x['City']
                    wsTransdirect['L' + str(maxrow)].value = x['Postcode']
                    wsTransdirect['M' + str(maxrow)].value = x['State']  # 'State'
                    wsTransdirect['N' + str(maxrow)].value = 'AU'  # 'Country' # 'AU' always
                    wsTransdirect['O' + str(maxrow)].value = x['Weight']  # 'Weight' #kg
                    wsTransdirect['P' + str(maxrow)].value = x['Length']  # 'Length' #cm
                    wsTransdirect['Q' + str(maxrow)].value = x['Width']  # 'Width' #cm
                    wsTransdirect['R' + str(maxrow)].value = x['Height']  # 'Height' #cm
                    wsTransdirect['S' + str(maxrow)].value = '1'  # 'Quantity' #Always 1 i think?
                    wsTransdirect['T' + str(maxrow)].value = 'Carton'  # 'Packaging' #Always 'Carton'

            except KeyError:
                pass

        if x['Final_Courier'] == 'Toll':

            x['Phone'] = x['Phone'].replace('+61', '0')
            x['Phone'] = x['Phone'].replace(' ', '')
            x['BuyerName'] = x['BuyerName'].replace("'", "")
            x['BuyerName'] = x['BuyerName'].replace("’", "")

            # volumevalue = ((float(x['Length']) / 100) * (float(x['Width']) / 100) * (
            #         float(x['Height']) / 100)) * 250
            volumevalue = (float(x['Length']) * float(x['Width']) * float(x['Height'])) * 0.000001

            message_identifier = str(uuid.uuid4())  # getting unique string for toll order
            ###Getting current date / time
            now = datetime.datetime.now()
            current_time = f"{now.strftime('%Y-%m-%d')}T{now.strftime('%H:%M:%S')}+10:00"

            next_day = next_business_day()
            next_toll_day = f"{next_day.strftime('%Y-%m-%d')}T09:00:00+10:00"

            try:
                cursor = connection.cursor()
            except:
                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                cursor = connection.cursor()
            cursor.execute("SELECT MAX(ShipmentID), MAX(SSCC) FROM Toll;")
            results = cursor.fetchall()
            pprint.pprint(results)
            max_shipment_id = int(results[0][0])
            max_sscc = int(results[0][1])
            environment = "PRD"

            manifest_test_url = "https://api-uat.teamglobalexp.com:6930/gateway/TollMessageManifestRestService/1.0/tom/receiveManifest"
            manifest_prod_url = "https://api.teamglobalexp.com:6930/gateway/TollMessageManifestRestService/1.0/tom/receiveManifest"

            headers = {'Content-Type': 'application/json', 'Accept': 'application/json'}


            def round_up_to_nearest_10(num):
                return math.ceil(num / 10) * 10


            toll_item_array = []
            toll_sql_insert = []

            if x['AddressLine2'] == '':
                x['AddressLine2'] = x['AddressLine1']

            total_toll_information = {
                "ShipmentFinancials": {
                    "ExtraServicesAmount": {
                        "Currency": "AUD",
                        "Value": str(round(float(x['TotalPrice'])))}},
                "ShipmentID": str(max_shipment_id),
                "CreateDateTime": current_time,
                "ConsigneeParty": {
                    "PartyName": x['BuyerName'],
                    "PhysicalAddress": {
                        "AddressLine1": x['AddressLine1'],
                        "AddressLine2": x['AddressLine2'],

                        "Suburb": x['City'],
                        "PostalCode": x['Postcode'],
                        "StateCode": x['State'],
                        "CountryName": "Australia",
                        "CountryCode": "AU"
                    },
                    "Contact": {
                        "Name": x['BuyerName'],
                        "Phone": {
                            "Number": x['Phone']
                        }
                    }
                },

                "BillToParty": {
                    "AccountCode": "80119621",
                    "PartyName": "SCARLETTMUSIC",
                    "PhysicalAddress": {
                        "AddressLine1": "288 Ballarat Rd",

                        "Suburb": "Footscray",
                        "PostalCode": "3011",
                        "StateCode": "VIC",
                        "CountryName": "Australia",
                        "CountryCode": "AU"
                    },
                    "Contact": {
                        "Name": "Kyal Scarlett",
                        "Phone": {
                            "Number": "0382563460"
                        }
                    }
                },
                "ShipmentItemCollection": {
                    "ShipmentItem": []
                }
            }

            toll_item_array = {
                "ShipmentFinancials": {
                    "ExtraServicesAmount": {
                        "Currency": "AUD",
                        "Value": str(round(float(x['TotalPrice'])))}},
                "ShipmentID": str(max_shipment_id),
                "CreateDateTime": current_time,
                "ConsigneeParty": {
                    "PartyName": x['BuyerName'],
                    "PhysicalAddress": {
                        "AddressLine1": x['AddressLine1'],
                        "AddressLine2": x['AddressLine2'],

                        "Suburb": x['City'],
                        "PostalCode": x['Postcode'],
                        "StateCode": x['State'],
                        "CountryName": "Australia",
                        "CountryCode": "AU"
                    },
                    "Contact": {
                        "Name": x['BuyerName'],
                        "Phone": {
                            "Number": x['Phone']
                        }
                    }
                },

                "BillToParty": {
                    "AccountCode": "80119621",
                    "PartyName": "SCARLETTMUSIC",
                    "PhysicalAddress": {
                        "AddressLine1": "288 Ballarat Rd",

                        "Suburb": "Footscray",
                        "PostalCode": "3011",
                        "StateCode": "VIC",
                        "CountryName": "Australia",
                        "CountryCode": "AU"
                    },
                    "Contact": {
                        "Name": "Kyal Scarlett",
                        "Phone": {
                            "Number": "0382563460"
                        }
                    }
                },
                "ShipmentItemCollection": {
                    "ShipmentItem": []
                }
            }

            sscc_array = []

            max_shipment_id += 1
            max_sscc += 1

            sscc = str(max_sscc)
            sscc_equation_1 = (int(sscc[16]) + int(sscc[14]) + int(sscc[12]) + int(sscc[10]) + int(
                sscc[8]) + int(
                sscc[6]) + int(sscc[4]) + int(sscc[2]) + int(sscc[0])) * 3
            sscc_equation_2 = int(sscc[15]) + int(sscc[13]) + int(sscc[11]) + int(sscc[9]) + int(
                sscc[7]) + int(
                sscc[5]) + int(sscc[3]) + int(sscc[1])

            sscc_equation_3 = sscc_equation_1 + sscc_equation_2

            sscc_equation_4 = round_up_to_nearest_10(sscc_equation_3)

            sscc_check_digit = sscc_equation_4 - sscc_equation_3

            sscc = '00' + sscc + str(sscc_check_digit)

            sscc_array.append({
                "Value": sscc,
                "SchemeName": "SSCC"
            })

            toll_item_array['ShipmentItemCollection']['ShipmentItem'].append(
                {
                    "IDs": {
                        "ID": sscc_array
                    },

                    "ShipmentItemTotals": {
                        "ShipmentItemCount": '1'
                    },
                    "ShipmentService": {
                        "ServiceCode": "X",
                        "ServiceDescription": "ROAD EXPRESS",
                        "ShipmentProductCode": "1"
                    },
                    "Description": "Carton",
                    "Dimensions": {
                        "Volume": str(volumevalue),
                        "Weight": str(round(float(x['Weight'])))
                    },
                    "References": {
                        "Reference": [
                            {
                                "ReferenceType": "ConsignorItemReference",
                                "ReferenceValue": x['String_SalesRecordNumber']
                            },
                            {
                                "ReferenceType": "ConsigneeItemReference",
                                "ReferenceValue": x['String_SalesRecordNumber']
                            }
                        ]
                    }
                })
            ########################## Need more information in print API, so making seperate list
            total_toll_information['ShipmentItemCollection']['ShipmentItem'].append({
                "IDs": {
                    "ID": sscc_array
                },

                "ShipmentItemTotals": {
                    "ShipmentItemCount": '1'
                },
                "ShipmentService": {
                    "ServiceCode": "X",
                    "ServiceDescription": "ROAD EXPRESS",
                    "ShipmentProductCode": "1"
                },
                "Description": "Carton",
                "Dimensions": {
                    "Volume": str(volumevalue),
                    "Weight": str(round(float(x['Weight']), 1)),
                    "Length": str(math.ceil(float(x['Length']))),
                    "Width": str(math.ceil(float(x['Width']))),
                    "Height": str(math.ceil(float(x['Height'])))
                },
                "References": {
                    "Reference": [
                        {
                            "ReferenceType": "ConsignorItemReference",
                            "ReferenceValue": x['String_SalesRecordNumber']
                        },
                        {
                            "ReferenceType": "ConsigneeItemReference",
                            "ReferenceValue": x['String_SalesRecordNumber']
                        }
                    ]
                }
            })

            toll_sql_insert.append(
                f"INSERT INTO Toll(ShipmentID, SSCC, Name, Address1, Address2, Suburb, Postcode, State, Reference, phone_number, item_count, current_sscc, length, width, weight, height, volume) VALUES ('{max_shipment_id}', '{max_sscc}', '{x['BuyerName']}', '{x['AddressLine1']}', '{x['AddressLine2']}', '{x['City']}', '{x['Postcode']}', '{x['State']}', '{x['String_SalesRecordNumber']}', '{x['Phone']}', '1', '{sscc}', '{str(x['Length'])}', '{str(x['Width'])}', '{str(x['Width'])}', '{str(x['Height'])}', '{str(volumevalue)}'); COMMIT;")

            try:
                cursor = connection.cursor()
            except:
                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                cursor = connection.cursor()
            cursor.execute(
                f"INSERT INTO Toll(ShipmentID, SSCC, Name, Address1, Address2, Suburb, Postcode, State, Reference, phone_number, item_count, current_sscc, length, width, weight, height, volume) VALUES ('{max_shipment_id}', '{max_sscc}', '{x['BuyerName']}', '{x['AddressLine1']}', '{x['AddressLine2']}', '{x['City']}', '{x['Postcode']}', '{x['State']}', '{x['String_SalesRecordNumber']}', '{x['Phone']}', '1', '{sscc}', '{str(x['Length'])}', '{str(x['Width'])}', '{str(x['Width'])}', '{str(x['Height'])}', '{str(volumevalue)}'); COMMIT;")

            message_identifier = str(uuid.uuid4())

            test_rate_enquiry_headers = {
                "MessageVersion": "3.1",
                "MessageIdentifier": message_identifier,
                "CreateTimestamp": current_time,
                "DocumentType": "Manifest",
                "Environment": environment,
                "MessageSender": "SCARLETTMUSIC",
                "MessageReceiver": "TOLL",
                'SourceSystemCode': 'XH56'
            }

            toll_message = {"@version": "3.1",
                            "@encoding": "utf-8",
                            "TollMessage": {"Header": test_rate_enquiry_headers,
                                            "Manifest": {
                                                "BusinessID": "IPEC",
                                                "CreateDateTime": current_time,
                                                "DatePeriodCollection": {
                                                    "DatePeriod": [{
                                                        "DateType": "DespatchDate",
                                                        "DateTime": next_toll_day
                                                    }]},

                                                "ConsignorParty": {
                                                    "PartyName": "Scarlett Music",
                                                    "PhysicalAddress": {
                                                        "AddressLine1": "288 Ballarat Rd",

                                                        "Suburb": "FOOTSCRAY",
                                                        "PostalCode": "3011",
                                                        "StateCode": "VIC",
                                                        "CountryName": "Australia",
                                                        "CountryCode": "AU"
                                                    },
                                                    "Contact": {
                                                        "Name": "Michael Demetre",
                                                        "Phone": {
                                                            "Number": "1800688586"
                                                        }
                                                    }
                                                },
                                                "ShipmentCollection": {
                                                    "Shipment": [toll_item_array]}}}}

            message_identifier = str(uuid.uuid4())

            r = requests.post(url=manifest_prod_url, auth=('accounts@scarlettmusic.com.au', 't2TrAPsNTB'),
                              json=toll_message, headers=headers)

            response = r.text

            response = json.loads(response)

            try:

                if str(response['TollMessage']['ResponseMessages']['ResponseMessage'][0]['ResponseID'][
                           'Value']) == '200':
                    # for shipment_lines in payload[]
                    pass

                else:
                    print(response['TollMessage']['ErrorMessages']['ErrorMessage'][0]['ErrorMessage'])
                    input('see above error and press enter')
            except:
                print(response)

            ####Will need to change environment in prod

            ############## ABOVE IS THE CREATE SHIPMENT API, BELOW IS THE PRINT ORDER API ############

            # message_identifier = str(uuid.uuid4())  # getting unique string for toll order
            ###Getting current date / time
            now = datetime.datetime.now()
            current_time = now.strftime("%Y-%m-%dT%H:%M:%SZ")

            next_day = next_business_day()
            next_toll_day = next_day.strftime("%Y-%m-%dT09:00:00Z")

            environment = "PRD"

            headers = {"Content-Type": "application/json",
                       "Accept": "application/json",
                       "x-api-key": "0O87zfzRVu4yf84mxgRiQ2uw9Q86Xb1Z8JrcnJfL",
                       "callId": "SCARLETTMUSIC",
                       "x-mytoll-identity": "727aba8b-0807-4a90-93be-0231d61d4806",
                       "Channel": "CES",
                       "x-mytoll-token": "eyJhbGciOiJSUzUxMiJ9.eyJ1c2VySWQiOiI3MjdhYmE4Yi0wODA3LTRhOTAtOTNiZS0wMjMxZDYxZDQ4MDYiLCJUJlQiOnRydWUsImdlbmVyYXRlRGF0ZSI6MTY4NjcxNjAxMTM3MSwiY3VzdG9tTmFtZSI6IjE0LTA2LTIzX015VGVhbUdFVG9rZW4iLCJjZXJ0aWZpY2F0ZU5hbWUiOiJsb2NhbGhjbCIsIkMmQyI6dHJ1ZSwidW5pcXVlSWQiOiJjNDcxNzgyOTZkMGQzM2RjY2NlMWY4MjQwODFhYjQzYTk5MzY0NGMxMDY1OGFkZWE2YjhlNGI5OGFkNGEzMGZmIiwiZXhwIjoxNzQ5ODc0NDExfQ.A-gOQU6Pc1_yuFkHTqQ219So4lkeoRI0CxtQrZlAsF9VBgqt085lffV_QRGDBPeogjLL5bae-XloKfPO-Ah23HErGHh_oXw_9CkRg8mcG7tkBZsf8StPPN-6HD1i-9iFioJvRE6d9njkVdePapet1FkBuWVg9WOKp8ft_516XR_pok1JmG_fnA55nDBADMDvUHFPW_YUqaoNbJmLpjf7CV0RGiT4pASilzQ4Ut4cuZ0NxQ3d-bQXBQetL5BxQzYNfANsxRD25icSGmi06alngfIFFxoCqBnuxYs_QCT1BvJHJw5e9LUMnXEGzuNAwx_6baRta7Fjq6UsuQPZ8zU-VA"}
            # Need to change x-api-key to 0O87zfzRVu4yf84mxgRiQ2uw9Q86Xb1Z8JrcnJfL in prod
            test_rate_enquiry_headers = {
                "MessageVersion": "1.0",
                "MessageIdentifier": message_identifier,
                "CreateTimestamp": current_time,
                "Environment": environment,
                "MessageSender": "SCARLETTMUSIC",
                "MessageReceiver": "Toll",
                "SourceSystemCode": "XH56"
            }
            manifest_test_url = "https://au-print-sit-apigw.internal.mytoll.com/printDocument"
            manifest_prod_url = "https://au-print-prod-apigw.internal.myteamge.com/printDocument"

            toll_item_count = 0
            total_volume_count = 0
            total_weight_count = 0

            shipment_id = total_toll_information['ShipmentID']
            name = total_toll_information['ConsigneeParty']['PartyName']
            address1 = total_toll_information['ConsigneeParty']['PhysicalAddress']['AddressLine1']
            address2 = total_toll_information['ConsigneeParty']['PhysicalAddress']['AddressLine2']
            suburb = total_toll_information['ConsigneeParty']['PhysicalAddress']['Suburb']
            postcode = total_toll_information['ConsigneeParty']['PhysicalAddress']['PostalCode']
            state = total_toll_information['ConsigneeParty']['PhysicalAddress']['StateCode']
            phone = total_toll_information['ConsigneeParty']['Contact']['Phone']['Number']
            address2 = address2[0:50]

            address_type = 'Residential'

            payload = {
                "BusinessID": "IPEC",
                "PrintDocumentType": "Label",
                "PrintSettings": {
                    "IsLabelThermal": "false",
                    "IsZPLRawResponseRequired": "false",
                    "PDF": {
                        "IsPDFA4": "false",
                        "PDFSettings": {
                            "StartQuadrant": "1"
                        }
                    }
                },
                "ConsignorParty": {
                    "Contact": {
                        "Name": "Kyal",
                        "Phone": {
                            "Number": "61422747033"
                        }
                    },
                    "PartyName": "Scarlett Music",
                    "PhysicalAddress": {
                        "AddressLine1": "288 Ballarat Rd",
                        "AddressType": "Business",
                        "CountryCode": "AU",
                        "PostalCode": "3011",
                        "StateCode": "VIC",
                        "Suburb": "FOOTSCRAY"
                    }
                },
                "CreateDateTime": current_time,
                "ShipmentCollection": {
                    "Shipment": [
                        {
                            "BillToParty": {
                                "AccountCode": "80119621"
                            },
                            "ConsigneeParty": {

                                "Contact": {
                                    "Name": "Kyal",
                                    "Phone": {
                                        "Number": "0382563460"
                                    }
                                },
                                "PartyName": name,
                                "PhysicalAddress": {
                                    "AddressType": address_type,  ##Business / residential
                                    "AddressLine1": address1,
                                    "AddressLine2": address2,

                                    "Suburb": suburb,
                                    "PostalCode": postcode,
                                    "StateCode": state,
                                    "CountryName": "Australia",
                                    "CountryCode": "AU"
                                }
                            },
                            "CreateDateTime": current_time,
                            "DatePeriodCollection": {
                                "DatePeriod": [
                                    {
                                        "DateTime": next_toll_day,
                                        "DateType": "DespatchDate"
                                    }
                                ]
                            },
                            "Orders": {
                                "Order": [
                                    {}
                                ]
                            },
                            "References": {
                                "Reference": [
                                    {
                                        "ReferenceType": "ShipmentReference1",
                                        "ReferenceValue": "ShipmentReference1"
                                    }
                                ]
                            },
                            "ShipmentID": shipment_id,
                            "ShipmentItemCollection": {
                                "ShipmentItem": []
                            },
                            "ShipmentTotals": {
                                "MiscellaneousItemCount": 0,
                                "Volume": {
                                    "UOM": "m3",
                                    "Value": 1
                                },
                                "Weight": {
                                    "UOM": "kg",
                                    "Value": 1
                                }
                            },

                        }
                    ]
                }
            }

            for t in total_toll_information['ShipmentItemCollection']["ShipmentItem"]:
                message_identifier = str(uuid.uuid4())

                sscc = t['IDs']['ID']
                item_count = t['ShipmentItemTotals']['ShipmentItemCount']
                volume = t['Dimensions']['Volume']
                weight = round(float(t['Dimensions']['Weight']))
                length = t['Dimensions']['Length']
                width = t['Dimensions']['Width']
                height = t['Dimensions']['Height']

                total_volume = float(volume) * float(item_count)
                total_weight = float(weight) * float(item_count)

                toll_item_count += int(item_count)
                total_volume_count += float(total_volume)
                total_weight_count += float(total_weight)

                payload['ShipmentCollection']['Shipment'][0]['ShipmentItemCollection']['ShipmentItem'].append({
                    "Commodity": {
                        "CommodityCode": "Z",
                        "CommodityDescription": "ALL FREIGHT"
                    },
                    "Description": "Item- Carton",
                    "Dimensions": {
                        "Height": math.ceil(float(height)),
                        "HeightUOM": "cm3",
                        "Length": math.ceil(float(length)),
                        "LengthUOM": "cm3",
                        "Volume": round(float(volume), 4),
                        "VolumeUOM": "m3",
                        "Weight": round(float(weight)),
                        "WeightUOM": "kg",
                        "Width": math.ceil(float(width)),
                        "WidthUOM": "cm3"
                    },
                    "IDs": {
                        "ID": sscc
                    },
                    "References": {
                        "Reference": [
                            {
                                "ReferenceType": "ConsignorItemReference",
                                "ReferenceValue": x['String_SalesRecordNumber']
                            }
                        ]
                    },
                    "ShipmentItemTotals": {
                        "MiscellaneousItemQuantity": 0,
                        "ShipmentItemCount": str(item_count)
                    },
                    "ShipmentService": {
                        "ServiceCode": "X",
                        "ShipmentProductCode": "1"
                    }
                })

            payload['ShipmentCollection']['Shipment'][0]['ShipmentTotals']['Volume'][
                'Value'] = total_volume_count

            payload['ShipmentCollection']['Shipment'][0]['ShipmentTotals']['Weight']['Value'] = str(
                round(float(total_weight_count), 1))
            toll_message = {"TollMessage": {
                "Header": test_rate_enquiry_headers,
                "Print": payload}}

            # print(toll_message)
            cert_file_path = r"\\SERVER\Python\TOLL\my_client.cert"
            key_file_path = r"\\SERVER\Python\TOLL\my_client.key"
            cert = (cert_file_path, key_file_path)

            r = requests.post(url=manifest_prod_url, auth=("accounts@scarlettmusic.com.au", "t2TrAPsNTB"),
                              json=toll_message, headers=headers, cert=cert)
            # r = requests.post(url=manifest_test_url, json=toll_message, headers = headers)
            # print(curlify.to_curl(r.request))
            response = r.text

            response = json.loads(response)

            try:
                pdf = response['TollMessage']['ResponseMessages']['ResponseMessage'][0]['ResponseMessage']
            except:
                continue
            print(pdf)
            pdf = base64.b64decode(pdf)

            label_name = label_location + "\\" + str(y['String_OrderID']) + ".pdf"

            Label_Location.append(label_name)

            with open(label_name, 'wb') as f:
                f.write(pdf)
            ######^^^^^^^ WILL ONLY SPIT OUT PDF IN QUADRANT 1 ^^^^^^^^

            ######vvvvvvvv CROPPING PDF TO FULL SIZE vvvvvvvvvvvvvv

            images = convert_from_path(label_name)

            for page in images:
                page.save(imagelocations + '\python2.jpg', 'JPEG')

                im = Image.open(imagelocations + '\python2.jpg')

                # Size of the image in pixels (size of original image)
                # (This is not mandatory)
                # width, height = im.size

                # Setting the points for cropped image

                # Cropped image of above dimension
                # (It will not change original image)
                im1 = im.crop((0, 0, 900, 1200))
                im1.save(label_name, 'PDF')

            #### TOLL PRINTING HAS FINISHED

            ### BELOW IS FOR BOOKING A TOLL PICKUP

            message_identifier = str(uuid.uuid4())

            now = datetime.datetime.now()
            current_time = now.strftime("%Y-%m-%dT%H:%M:%SZ")

            next_day = next_business_day()
            next_toll_day = next_day.strftime("%Y-%m-%dT09:00:00Z")
            business_close_time = next_day.strftime("%Y-%m-%dT16:00:00Z")

            test_booking_url = 'https://api-uat.teamglobalexp.com:6930/gateway/TollMessageBookingRestService/1.0/tom/bookingRequest'
            production_booking_url = 'https://api.teamglobalexp.com:6930/gateway/TollMessageBookingRestService/1.0/tom/bookingRequest'

            environment = "PRD"

            address_type = 'Residential'

            for t in total_toll_information['ShipmentItemCollection']["ShipmentItem"]:
                message_identifier = str(uuid.uuid4())

                sscc = t['IDs']['ID']
                item_count = t['ShipmentItemTotals']['ShipmentItemCount']
                volume = t['Dimensions']['Volume']
                weight = round(float(t['Dimensions']['Weight']))
                length = t['Dimensions']['Length']
                width = t['Dimensions']['Width']
                height = t['Dimensions']['Height']

                total_volume = float(volume) * float(item_count)
                total_weight = float(weight) * float(item_count)

                toll_item_count += int(item_count)
                total_volume_count += float(total_volume)
                total_weight_count += float(total_weight)

                payload = {
                    "@version": "1.0",
                    "@encoding": "utf-8",
                    "TollMessage": {
                        "Header": {
                            "MessageVersion": "1.0",
                            "MessageIdentifier": message_identifier,
                            "CreateTimestamp": current_time,
                            "DocumentType": "Booking",
                            "Environment": environment,
                            "SourceSystemCode": "XH56",
                            "MessageSender": "SCARLETTMUSIC",
                            "MessageReceiver": "TOLL"
                        },
                        "Bookings": {
                            "Booking": [
                                {
                                    "BusinessID": "IPEC",
                                    "BookingParty": {
                                        "PartyName": "Scarlett Music",
                                        "AccountCode": "80119621",
                                        "PhysicalAddress": {
                                            "AddressType": "Business",
                                            "AddressLine1": "286-288 Ballarat Rd",
                                            "AddressLine2": "Scarlett Music",
                                            "Suburb": "Footscray",
                                            "PostalCode": "3011",
                                            "StateCode": "VIC",
                                            "CountryCode": "AU"
                                        },
                                        "Contact": {
                                            "Name": "Kyal",
                                            "Phone": {
                                                "CountryCode": "+61",
                                                "AreaCode": "03",
                                                "Number": "82563460"
                                            },
                                            "EMail": "kyal@scarlettmusic.com.au"
                                        }
                                    },
                                    "SystemFields": {
                                        "RecordCreateDateTime": current_time,
                                        "PickupDateTime": next_toll_day,
                                        "OpenDateTime": next_toll_day,
                                        "CloseDateTime": business_close_time
                                    },
                                    "References": {
                                        "Reference": [
                                            {
                                                "ReferenceType": "PickupLocation",
                                                "ReferenceValue": "Footscray"
                                            }
                                        ]
                                    },
                                    "Contact": {
                                        "Name": "Kyal Scarlett",
                                        "Phone": {
                                            "CountryCode": "+61",
                                            "AreaCode": "03",
                                            "Number": "82563460"
                                        },
                                        "EMail": "kyal@scarlettmusic.com.au",
                                        "Note": ""
                                    },
                                    "BookingFlags": {
                                        "BookingInstruction": "The music shop, open 9am-6pm. Best parking is at The Palms across the road.",
                                        "SameLocationFlag": "true",
                                    },
                                    "BookingItems": {
                                        "BookingItem": [
                                            {
                                                "ShipmentService": {
                                                    "ServiceCode": "X",
                                                    "ServiceDescription": "Road Express"
                                                },
                                                "ItemQuantity": str(item_count),
                                                #####Plug item details in here
                                                "Description": "Parcel",
                                                "ItemTypeCode": "BXCA",
                                                "ItemTypeName": "Box/Carton",
                                                "Dimensions": {
                                                    "Length": str(math.ceil(float(length))),
                                                    "LengthUOM": "CMT",
                                                    "Width": str(math.ceil(float(width))),
                                                    "WidthUOM": "CMT",
                                                    "Height": str(math.ceil(float(height))),
                                                    "HeightUOM": "CMT",
                                                    "Volume": str(round(float(volume), 3)),
                                                    "VolumeUOM": "MTQ",
                                                    "Weight": str(round(float(weight))),
                                                    "WeightUOM": "KGM"
                                                },
                                                "BookingFinancials": {
                                                    "ChargeCode": "Sender",
                                                    "AccountCode": "80119621"
                                                },
                                                "DeliveryParty": {
                                                    "PhysicalAddress": {
                                                        "AddressType": address_type,

                                                        "AddressLine1": address1,
                                                        #####Plug delivery address details in here
                                                        "Suburb": suburb,
                                                        "PostalCode": str(postcode),
                                                        "StateCode": state,
                                                        "CountryCode": "AU"
                                                    }
                                                },
                                                "DatePeriodCollection": {
                                                    "DatePeriod": [
                                                        {
                                                            "DateType": "OrderDate",
                                                            "DateTime": current_time
                                                        }
                                                    ]
                                                },
                                                "BookingItemFlags": {
                                                    "DangerousGoodsFlag": "false"
                                                }
                                            }
                                        ]
                                    }
                                }
                            ]
                        }
                    }
                }

                r = requests.post(url=production_booking_url,
                                  auth=("accounts@scarlettmusic.com.au", "t2TrAPsNTB"),
                                  json=payload)

                response = r.text

                response = json.loads(response)

            for lines in range(len(x['SKU'])):
                x['ItemTitle'][f'Name (Item {str(lines + 1)})'] = x['ItemTitle'][
                    f'Name (Item {str(lines + 1)})'].replace("'", "")

                x['BuyerName'] = x['BuyerName'].replace("'", "")
                x['Company'] = x['Company'].replace("'", "")
                x['AddressLine1'] = x['AddressLine1'].replace("'", "")
                x['AddressLine2'] = x['AddressLine2'].replace("'", "")
                x['City'] = x['City'].replace("'", "")
                try:
                    cursor = connection.cursor()
                except:
                    connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                    cursor = connection.cursor()

                try:

                    cursor.execute(
                        fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order,Sales_Record_Number , Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{x['String_OrderID'][0:50]}', '{x['String_PurchaseOrderNumber'][0:50]}', '{x['String_SalesRecordNumber'][0:50]}', 'ebay', '{x['BuyerName'][0:50]}', '{x['Company'][0:50]}', '{x['AddressLine1'][0:50]}', '{x['AddressLine2'][0:50]}', '{x['City'][0:50]}', '{x['State']}', '{x['Postcode']}', 'AU', '{x['Phone'][0:50]}', '{x['Email'][0:50]}', '{x['date_placed']}','' , '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '{x['OrderLineID'][f'Orderline (Item {str(lines + 1)})'][0:50]}', '{x['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{x['ItemTitle'][f'Name (Item {str(lines + 1)})'][0:50]}', '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '{x['String_OrderID'][0:50]}', '{x['Final_Courier']}','{tracking_number}' , '{str(x['item_count'])}', '{str(x['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                except:
                    pass

        try:
            if x['Final_Courier'] == 'Couriers Please':

                try:
                    x['Phone'] = x['Phone'].replace('+61', '0')
                    x['Phone'] = x['Phone'].replace(' ', '')

                    cp_url = 'https://api.couriersplease.com.au/v2/domestic/shipment/create'

                    r = requests.post(cp_url, headers=cp_headers, json=x['cp_validate_body'])
                    response = r.text
                    response = json.loads(response)
                    # pprint.pprint(response)

                    cp_tracking_number = response['data']['consignmentCode']

                    ###########BELOW IS CREATING LABEL FOR ORDER########

                    cp_url = f'https://api.couriersplease.com.au/v1/domestic/shipment/label?consignmentNumber={cp_tracking_number}'
                    r = requests.get(cp_url, headers=cp_headers, json=cp_validate_body)
                    response = r.text
                    response = json.loads(response)
                    # pprint.pprint(response)

                    base64_label = response['data']['label']
                    pdf = base64.b64decode(base64_label)

                    label_name = label_location + "\\" + str(y['String_OrderID']) + ".pdf"

                    Label_Location.append(label_name)

                    with open(label_name, 'wb') as f:
                        f.write(pdf)

                    couriers_please_item_count += 1
                    couriers_please_weight += float(x['weight'])

                    for lines in range(len(x['SKU'])):
                        x['ItemTitle'][f'Name (Item {str(lines + 1)})'] = x['ItemTitle'][
                            f'Name (Item {str(lines + 1)})'].replace("'", "")

                        x['BuyerName'] = x['BuyerName'].replace("'", "")
                        x['Company'] = x['Company'].replace("'", "")
                        x['AddressLine1'] = x['AddressLine1'].replace("'", "")
                        x['AddressLine2'] = x['AddressLine2'].replace("'", "")
                        x['City'] = x['City'].replace("'", "")
                        try:
                            cursor = connection.cursor()
                        except:
                            connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                           "5432")
                            cursor = connection.cursor()

                        try:

                            cursor.execute(
                                fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order,Sales_Record_Number , Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{x['String_OrderID'][0:50]}', '{x['String_PurchaseOrderNumber'][0:50]}', '{x['String_SalesRecordNumber'][0:50]}', 'ebay', '{x['BuyerName'][0:50]}', '{x['Company'][0:50]}', '{x['AddressLine1'][0:50]}', '{x['AddressLine2'][0:50]}', '{x['City'][0:50]}', '{x['State'][0:50]}', '{x['Postcode'][0:50]}', 'AU', '{x['Phone'][0:50]}', '{x['Email'][0:50]}', '{x['date_placed']}','' , '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})'][0:50]}', '{x['OrderLineID'][f'Orderline (Item {str(lines + 1)})'][0:50]}', '{x['Quantity'][f'Quantity (Item {str(lines + 1)})']}', '{x['ItemTitle'][f'Name (Item {str(lines + 1)})'][0:200]}', '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['String_OrderID'][0:50]}', '{x['Final_Courier']}','{cp_tracking_number}' , '{str(x['item_count'])}', '{str(x['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                        except:
                            pass

                except KeyError:
                    pass
        except:
            pass

        if x['Final_Courier'] == 'Dai Post':
            x['Phone'] = x['Phone'].replace('+61', '0')
            x['Phone'] = x['Phone'].replace(' ', '')

            weightvalue = x['Weight']

            dai_volume_value = ((float(x['Length']) / 100) * (float(x['Width']) / 100) * (
                        float(x['Height']) / 100)) * 250

            # if float(dai_volume_value) > float(weightvalue):
            #     final_dai_weight = dai_volume_value
            # elif float(weightvalue) > dai_volume_value:
            #     final_dai_weight = weightvalue

            final_dai_weight = weightvalue

            # final_dai_weight = round(float(final_dai_weight), 2)

            try:
                final_dai_weight = round(float(final_dai_weight), 2)
            except:
                final_dai_weight = round(float(weightvalue), 2)

            try:
                cursor = connection.cursor()
            except:
                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                cursor = connection.cursor()
            cursor.execute("SELECT MAX(job_number) FROM dai_post;")
            results = cursor.fetchall()
            #pprint.pprint(results)
            max_job_id = int(results[0][0])

            name = x['BuyerName']
            name = name.replace("'", "")
            name = name.replace("’", "")
            ###Getting current date / time
            now = datetime.datetime.now()
            current_time = f"{now.strftime('%Y-%m-%d')}T{now.strftime('%H:%M:%S')}"

            signature_required = '0'

            if round(float(x['TotalPrice'])) > 200:
                signature_required = '1'

            payload = {
                "shipment": {
                    "service": "Parcel Right",
                    "labelformat": "PDF",
                    "account": "SCA",
                    "datetime": current_time,
                    "reference": f"{x['String_PurchaseOrderNumber']} {x['item_count_string']}",
                    "jobnumber": max_job_id,
                    "signature": signature_required,
                    "value": str(round(float(x['TotalPrice']))),
                    "currency": "AUD",
                    "uom": "kg",
                    "weight": final_dai_weight,
                    "originterminal": "TME",
                    "shipper": {
                        "name": "Scarlett Music",
                        "attention": "Kyal",
                        "addr1": "286-288 Ballarat Rd",
                        "addr2": "",
                        "city": "Footscray",
                        "state": "VIC",
                        "country": "AU",
                        "postal": "3011",
                        "phone": "+61 03 8256 3460",
                        "email": "kyal@scarlettmusic.com.au"
                    },
                    "consignee": {
                        "name": x['BuyerName'],
                        "attention": x['BuyerName'],
                        "addr1": x['AddressLine1'],
                        "addr2": x['AddressLine2'],
                        "city": x['City'],
                        "state": x['State'],
                        "country": "AU",
                        "postal": x['Postcode'],
                        "phone": x['Phone'],
                        "email": x['Email']
                    },
                    "item": [
                        {
                            "description": f"Scarlett Music Order {x['item_count_string']}",
                            "qty": "1",
                            "unit": "pc",
                            "value": str(round(float(x['TotalPrice'])))
                        }
                    ]
                }
            }

            r = requests.post('https://daiglobaltrack.com/prod/serviceconnect',
                              auth=HTTPBasicAuth('ScarlettMusic', 'D5es4stu!'), json=payload)

            response = r.text


            try:
                dai_response = json.loads(response)

                dai_tracking_number = dai_response['shipmentresponse']['tracknbr']
                dai_pdf = dai_response['shipmentresponse']['label']
            except:
                continue

            dai_pdf = base64.b64decode(dai_pdf)

            label_name = label_location + "\\" + str(x['String_SalesRecordNumber']) + ".pdf"

            Label_Location.append(label_name)

            with open(label_name, 'wb') as f:
                f.write(dai_pdf)

            x['BuyerName'] = x['BuyerName'].replace("'", "")

            try:
                cursor = connection.cursor()
            except:
                connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                cursor = connection.cursor()
            cursor.execute(
                fr"INSERT INTO dai_post(customer_name, tracking_number, time_created, job_number) VALUES ('{x['BuyerName']}', '{dai_tracking_number}', '{now}', '{max_job_id}'); COMMIT;")

            for lines in range(len(x['SKU'])):
                x['ItemTitle'][f'Name (Item {str(lines + 1)})'] = x['ItemTitle'][
                    f'Name (Item {str(lines + 1)})'].replace("'", "")

                x['BuyerName'] = x['BuyerName'].replace("'", "")
                x['Company'] = x['Company'].replace("'", "")
                x['AddressLine1'] = x['AddressLine1'].replace("'", "")
                x['AddressLine2'] = x['AddressLine2'].replace("'", "")
                x['City'] = x['City'].replace("'", "")
                try:
                    cursor = connection.cursor()
                except:
                    connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                    cursor = connection.cursor()

                try:
                    cursor.execute(
                        fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order,Sales_Record_Number , Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{x['String_OrderID'][0:50]}', '{x['String_PurchaseOrderNumber'][0:50]}', '{x['String_SalesRecordNumber'][0:50]}', 'ebay', '{x['BuyerName'][0:50]}', '{x['Company'][0:50]}', '{x['AddressLine1'][0:50]}', '{x['AddressLine2'][0:50]}', '{x['City'][0:50]}', '{x['State'][0:50]}', '{x['Postcode'][0:50]}', 'AU', '{x['Phone'][0:50]}', '{x['Email'][0:50]}', '{x['date_placed'][0:50]}','' , '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['OrderLineID'][f'Orderline (Item {str(lines + 1)})']}', '{x['Quantity'][f'Quantity (Item {str(lines + 1)})']}', '{x['ItemTitle'][f'Name (Item {str(lines + 1)})'][0:200]}', '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['String_OrderID'][0:50]}', '{x['Final_Courier']}','{dai_tracking_number}' , '{str(x['item_count'])}', '{str(x['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                except:
                    pass

        if x['Final_Courier'] == 'Freightster':

            try:

                x['Phone'] = x['Phone'].replace('+61', '0')
                x['Phone'] = x['Phone'].replace(' ', '')

                freightster_volume_value = ((float(x['Length']) / 100) * (float(x['Width']) / 100) * (
                        float(x['Height']) / 100)) * 250

                # print(f'l : {length}')
                # print(f'w : {width}')
                # print(f'l : {height}')
                if float(freightster_volume_value) > float(weightvalue):
                    final_freightster_weight = freightster_volume_value
                elif float(weightvalue) > freightster_volume_value:
                    final_freightster_weight = weightvalue

                if float(x['Length']) <= 4.0 and float(x['Width']) <= 23.0 and float(
                        x['Height']) <= 23.0 and float(
                    weightvalue) <= 0.25:
                    final_freightster_weight = 0.250

                elif float(x['Length']) <= 18.0 and float(x['Width']) <= 23.0 and float(
                        x['Height']) <= 4.0 and float(
                    weightvalue) <= 0.25:
                    final_freightster_weight = 0.250

                elif float(weightvalue) <= 0.25 and freightster_volume_value < 0.331:
                    final_freightster_weight = 0.250

                elif float(x['Length']) <= 18.0 and float(x['Width']) <= 23.0 and float(
                        x['Height']) <= 4.0 and float(
                    weightvalue) <= 0.5:
                    final_freightster_weight = 0.500

                payload = {"order": {"serviceCode": 12,
                                     "consignee": {"company": x['Company'],
                                                   "name": x['BuyerName'],
                                                   "address1": x['AddressLine1'],
                                                   "address2": x['AddressLine2'],
                                                   "city": x['City'],
                                                   "postcode": x['Postcode'],
                                                   "state": x['State'],
                                                   "phone": x['Phone'],
                                                   "email": x['Email']},
                                     "sender": {"name": "Kyal Scarlett",
                                                "address1": "286-288 Ballarat Rd",
                                                "address2": "",
                                                "city": "Footscray",
                                                "postcode": "3011",
                                                "state": "VIC",
                                                "phone": "0382563460",
                                                "email": "kyal@scarlettmusic.com.au"},
                                     "shipment": {"reference": x['item_count_string'],
                                                  "description": x['item_count_string'],
                                                  "weight": str(final_freightster_weight)}}}

                while True:

                    r = requests.post('https://freightster.com.au/api/v1/shippingAPI/create', json=payload,
                                      headers=freightster_headers)

                    if r.status_code == 429:
                        continue
                    else:
                        break
                freightster_response = json.loads(r.text)
                freightster_orderid = freightster_response['response_data']['order_id']
                freightster_trackingnumber = freightster_response['response_data']['tracking_number']
                payload = {"order": {"orderIds": [freightster_orderid]}}

                while True:

                    r = requests.post('https://freightster.com.au/api/v1/shippingAPI/print', json=payload,
                                      headers=freightster_headers)

                    if r.status_code == 429:
                        continue
                    else:
                        break
                freightster_response = json.loads(r.text)

                freightster_pdf = freightster_response['response_data']['labels'][0]['label']
                freightster_pdf = base64.b64decode(freightster_pdf)
                buyer_name = x['BuyerName']

                label_name = label_location + "\\" + str(x['String_SalesRecordNumber']) + ".pdf"

                Label_Location.append(label_name)

                with open(label_name, 'wb') as f:
                    f.write(freightster_pdf)

                buyer_name = buyer_name.replace("'", "")

                try:
                    cursor = connection.cursor()
                except:
                    connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                    cursor = connection.cursor()
                cursor.execute(
                    f"INSERT INTO Freightster(order_id, tracking_number, name) VALUES ('{freightster_orderid}', '{freightster_trackingnumber}', '{buyer_name}'); COMMIT;")

                for lines in range(len(x['SKU'])):
                    x['ItemTitle'][f'Name (Item {str(lines + 1)})'] = x['ItemTitle'][
                        f'Name (Item {str(lines + 1)})'].replace("'", "")

                    x['BuyerName'] = x['BuyerName'].replace("'", "")
                    x['Company'] = x['Company'].replace("'", "")
                    x['AddressLine1'] = x['AddressLine1'].replace("'", "")
                    x['AddressLine2'] = x['AddressLine2'].replace("'", "")
                    x['City'] = x['City'].replace("'", "")
                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()

                    try:
                        cursor.execute(
                            fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order,Sales_Record_Number , Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{x['String_OrderID'][0:50]}', '{x['String_PurchaseOrderNumber'][0:50]}', '{x['String_SalesRecordNumber'][0:50]}', 'ebay', '{x['BuyerName'][0:50]}', '{x['Company'][0:50]}', '{x['AddressLine1'][0:50]}', '{x['AddressLine2'][0:50]}', '{x['City'][0:50]}', '{x['State'][0:50]}', '{x['Postcode'][0:50]}', 'AU', '{x['Phone'][0:50]}', '{x['Email'][0:50]}', '{x['date_placed'][0:50]}','' , '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['OrderLineID'][f'Orderline (Item {str(lines + 1)})']}', '{x['Quantity'][f'Quantity (Item {str(lines + 1)})']}', '{x['ItemTitle'][f'Name (Item {str(lines + 1)})'][0:200]}', '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['String_OrderID'][0:50]}', '{x['Final_Courier']}','{freightster_trackingnumber}' , '{str(x['item_count'])}', '{str(x['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                    except:
                        pass

            except KeyError:
                pass

        if x['Final_Courier'] == 'Allied Express':

            try:

                today = datetime.date.today()
                if today.isoweekday() in set((6, 7)):
                    today += datetime.timedelta(days=today.isoweekday() % 5)
                next_day_allied = str(today.day) + '/' + str(today.month) + '/' + str(today.year) + " 10:00:00"

                history = HistoryPlugin()
                session = Session()
                transport = Transport(session=session)
                wsdl = 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS?wsdl'

                try:
                    allied_client = zeep.Client(wsdl=wsdl, transport=transport, plugins=[history])

                    allied_client.transport.session.proxies = {
                        # Utilize for all http/https connections
                        'http': 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS', }
                    allied_account = allied_client.service.getAccountDefaults('755cf13abb3934695f03bd4a75cfbca7',
                                                                              "SCAMUS", "VIC",
                                                                              "AOE")

                except:
                    allied_failed = 'true'

                ONE_DAY = datetime.timedelta(days=1)  ###### Getting next business day for pickups
                HOLIDAYS_AU = holidays.AU(prov='VIC')

                pickup_address = {'address1': "286-288 Ballarat Rd",
                                  'address2': "",
                                  'country': "Australia",
                                  'postCode': '3011',
                                  'state': 'VIC',
                                  'suburb': 'Footscray'}

                reciever_address = {'address1': x['AddressLine1'],
                                    'address2': x['AddressLine2'],
                                    'country': "Australia",
                                    'postCode': x['Postcode'],
                                    'state': x['State'],
                                    'suburb': x['City']}

                Jobstop_pickupstop = {'companyName': 'Scarlett Music',
                                      'contact': 'Kyal Scarlett',
                                      'emailAddress': 'info@scarlettmusic.com.au',
                                      'geographicAddress': pickup_address,
                                      'phoneNumber': '03 9318 5751',
                                      'stopNumber': 1,
                                      'stopType': 'P'}

                Jobstop_deliverystop = {'companyName': x['BuyerName'],
                                        'contact': x['Company'],
                                        'emailAddress': x['Email'],
                                        'geographicAddress': reciever_address,
                                        'phoneNumber': x['Phone'],
                                        'stopNumber': 2,
                                        'stopType': 'D'}

                Jobstop_final = [Jobstop_pickupstop, Jobstop_deliverystop]

                allied_item_array = []
                allied_total_volume = 0
                allied_total_weight = 0
                allied_total_items = 0
                #   multi_items_to_ship.append((weightvalue, length, height, width,volumevalue, cubicweight, quantity, satchel))

                volumevalue = str(
                    float(float(x['Length']) / 100) * float(float(x['Width']) / 100) * float(float(x['Height']) / 100))

                cubicweight = round(((float(x['Length']) / 100) * (float(x['Width']) / 100) * (float(x['Height']) / 100)) * 250, 2)

                allied_item_array.append({'dangerous': 'false',
                                          'height': str(x['Height']),
                                          'itemCount': str(1),
                                          'length': str(x['Length']),
                                          'volume': str(volumevalue),
                                          'weight': str(x['Weight']),
                                          'width': str(x['Width'])})

                allied_total_volume = str(volumevalue)
                allied_total_weight = str(x['Weight'])
                allied_total_items = '1'

                # item = {'dangerous': 'false',
                #         'height': float(height),
                #         'itemCount': 1,
                #         'length': float(length),
                #         'volume': float(volumevalue),
                #         'weight': float(weightvalue),
                #         'width': float(width)}

                ##IF MULTIPLE ITEMS, easily put into dic, item = [{item1, item2}]

                Service_Level = 'R'  # Overnight Express

                pickup_instructions = "The music shop, open 9am-6pm. Best parking is at The Palms across the road."
                reference_number = x['String_SalesRecordNumber']
                booked_by = 'Kyal Scarlett'

                job_number = int(re.sub("[^0-9]", "", reference_number))

                Job = {'account': allied_account,
                       'cubicWeight': cubicweight,
                       'Docket': "SCM",
                       'instructions': pickup_instructions,
                       'cubedItems': allied_item_array,
                       'itemCount': allied_total_items,
                       'weight': allied_total_weight,
                       'volume': allied_total_volume,
                       'items': allied_item_array,
                       'jobStops': Jobstop_final,
                       'serviceLevel': Service_Level,
                       'referenceNumbers': reference_number,
                       'bookedBy': booked_by,
                       'readyDate': today,
                       'jobNumber': job_number,
                       'vehicle': {'vehicleID': 1}}

                JobIDs = {'jobIds': job_number}

                try:

                    Job = allied_client.service.validateBooking('755cf13abb3934695f03bd4a75cfbca7', Job)

                    job_price = allied_client.service.calculatePrice('755cf13abb3934695f03bd4a75cfbca7', Job)
                    alliedprice = job_price['totalCharge']
                    alliedprice = round((float(alliedprice) * 1.269) * 1.1, 2)
                    if alliedprice == 0.0:
                        alliedprice = 1000
                    # print('Allied price: $' + str(alliedprice))

                except zeep.exceptions.Fault:
                    alliedprice = 1000



                ########################ALLIED EXPRESS AUTHENTICATION

                today = datetime.date.today()
                if today.isoweekday() in set((6, 7)):
                    today += datetime.timedelta(days=today.isoweekday() % 5)
                next_day_allied = str(today.day) + '/' + str(today.month) + '/' + str(today.year) + " 10:00:00"

                history = HistoryPlugin()
                session = Session()
                transport = Transport(session=session)
                wsdl = 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS?wsdl'

                try:
                    allied_client = zeep.Client(wsdl=wsdl, transport=transport, plugins=[history])

                    allied_client.transport.session.proxies = {
                        # Utilize for all http/https connections
                        'http': 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS', }
                    allied_account = allied_client.service.getAccountDefaults('755cf13abb3934695f03bd4a75cfbca7',
                                                                              "SCAMUS", "VIC",
                                                                              "AOE")

                except:
                    allied_failed = 'true'

                with allied_client.settings(strict=False):
                    allied_client.service.savePendingJob('755cf13abb3934695f03bd4a75cfbca7', Job)
                    dispatch_jobs = allied_client.service.dispatchPendingJobs('755cf13abb3934695f03bd4a75cfbca7',
                                                                              JobIDs)
                    xml = etree.tostring(history.last_received["envelope"], encoding="unicode")
                    xml = xmltodict.parse(xml)

                connote_number = \
                xml['soapenv:Envelope']['soapenv:Body']['ns1:dispatchPendingJobsResponse']['result']['item'][
                    'docketNumber']
                reference = \
                xml['soapenv:Envelope']['soapenv:Body']['ns1:dispatchPendingJobsResponse']['result']['item'][
                    'referenceNumbers']

                pdf = allied_client.service.getLabel('755cf13abb3934695f03bd4a75cfbca7', "AOE", connote_number,
                                                     reference,
                                                     '3011', 1)
                xml = etree.tostring(history.last_received["envelope"], encoding="unicode")
                xml = xmltodict.parse(xml)

                pdf = xml['soapenv:Envelope']['soapenv:Body']['ns1:getLabelResponse']['result']
                pdf = base64.b64decode(pdf)

                label_name = label_location + "\\" + str(x['String_SalesRecordNumber']) + ".pdf"

                # new_image.save(label_name, 'PDF')

                with open(label_name, 'wb') as f:
                    f.write(pdf)

                Label_Location.append(label_name)

                for lines in range(len(x['SKU'])):
                    x['ItemTitle'][f'Name (Item {str(lines + 1)})'] = x['ItemTitle'][
                        f'Name (Item {str(lines + 1)})'].replace("'", "")

                    x['BuyerName'] = x['BuyerName'].replace("'", "")
                    x['Company'] = x['Company'].replace("'", "")
                    x['AddressLine1'] = x['AddressLine1'].replace("'", "")
                    x['AddressLine2'] = x['AddressLine2'].replace("'", "")
                    x['City'] = x['City'].replace("'", "")
                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()

                    try:
                        cursor.execute(
                            fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order,Sales_Record_Number , Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{x['String_OrderID'][0:50]}', '{x['String_PurchaseOrderNumber'][0:50]}', '{x['String_SalesRecordNumber'][0:50]}', 'ebay', '{x['BuyerName'][0:50]}', '{x['Company'][0:50]}', '{x['AddressLine1'][0:50]}', '{x['AddressLine2'][0:50]}', '{x['City'][0:50]}', '{x['State'][0:50]}', '{x['Postcode'][0:50]}', 'AU', '{x['Phone'][0:50]}', '{x['Email'][0:50]}', '{x['date_placed']}','' , '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['OrderLineID'][f'Orderline (Item {str(lines + 1)})'][0:50]}', '{x['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{x['ItemTitle'][f'Name (Item {str(lines + 1)})'][0:200]}', '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['String_OrderID'][0:50]}', '{x['Final_Courier']}','{tracking}' , '{str(x['item_count'])}', '{str(x['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                    except:
                        pass
            except:
                pass

        if x['Final_Courier'] == 'CBDExpress':

            try:

                x['Phone'] = x['Phone'].replace('+61', '0')
                x['Phone'] = x['Phone'].replace(' ', '')

                # instructions = 'Authority to leave'
                #
                # if float(totalprice) > 200 and volumevalue > 0.0001:
                #     instructions = 'Signature on Delivery'

                volumevalue = str(
                    float(float(x['Length']) / 100) * float(float(x['Width']) / 100) * float(float(x['Height']) / 100))

                instructions = 'Authority to leave'

                if float(x['TotalPrice']) > 150 and volumevalue > 0.0001:
                    instructions = 'Signature on Delivery'

                try:
                    cursor = connection.cursor()
                except:
                    connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                    cursor = connection.cursor()
                cursor.execute(
                    fr"SELECT MAX(consignment_number) FROM CDBExpress")
                results = cursor.fetchall()
                max_consignment = results[0][0]
                new_consignment = str((int(max_consignment) + 1)).zfill(9)
                barcode_param = f'SCARLET{new_consignment}'

                now = datetime.datetime.now()
                current_time = now.strftime("%Y-%m-%dT%H:%M:%SZ")

                try:
                    cursor = connection.cursor()
                except:
                    connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                    cursor = connection.cursor()
                cursor.execute(
                    f"INSERT INTO CDBExpress(order_id, consignment_number, name, time_created) VALUES ('{x['String_SalesRecordNumber']}', '{new_consignment}', '{x['BuyerName']}', '{current_time}' ); COMMIT;")

                payload = {"barcode": barcode_param,
                           "description": "Scarlett Music Order " + x['String_SalesRecordNumber'],
                           "weight": {"value": x['Weight'], "units": "kg"},
                           "volume": {"value": volumevalue, "units": "m3"},
                           "customer_reference": x['item_count_string'], "sender": {
                        "contact": {
                            "name": "Scarlett Music",
                            "phone": "(03) 9318 5751",
                            "company": "Scarlett Music"
                        },
                        "address": {
                            "address_line1": "286-288 Ballarat Rd",
                            "suburb": "Maidstone",
                            "state_name": "VIC",
                            "postcode": "3012",
                            "country": "Australia"
                        },
                        "instructions": "The music shop, open 9am-6pm. Best parking is at The Palms across the road."
                    }, "receiver": {
                        "contact": {
                            "name": x['BuyerName'],
                            "email": x['Email'],
                            "phone": x['Phone'],
                            "company": x['Company']
                        },
                        "address": {
                            "address_line1": x['AddressLine1'],
                            "address_line2": x['AddressLine2'],
                            "suburb": x['City'],
                            "state_name": x['State'],
                            "postcode": x['Postcode'],
                            "country": "Australia"
                        },
                        "instructions": instructions
                    }
                           }

                cbdurl = 'https://apis.hubsystems.com.au/booking/'
                cbdauth = ('code-scarlett', 'syY55DxG41sd8')
                cbdheaders = {'Content-Type': 'text/scarlett'}
                data = payload

                response = requests.post(cbdurl, auth=cbdauth, headers=cbdheaders, json=data)

                # Create barcode image
                barcode_image = code128.image(barcode_param, height=100)

                # Create empty image for barcode + text
                top_bott_margin = 70
                l_r_margin = 10
                new_height = barcode_image.height + (2 * top_bott_margin)
                new_width = barcode_image.width + (2 * l_r_margin)
                new_image = Image.new('RGB', (new_width, 700), (255, 255, 255))

                # put barcode on new image
                barcode_y = 525
                new_image.paste(barcode_image, (0, barcode_y))

                # object to draw text
                draw = ImageDraw.Draw(new_image)

                # Define custom text size and font
                h1_size = 28
                h2_size = 28
                h3_size = 16
                footer_size = 21

                h1_font = ImageFont.truetype("r'C:\Windows\Fonts\calibri.ttf", h1_size)
                h2_font = ImageFont.truetype("r'C:\Windows\Fonts\calibri.ttf", h2_size)
                h3_font = ImageFont.truetype("r'C:\Windows\Fonts\calibri.ttf", h3_size)
                footer_font = ImageFont.truetype("r'C:\Windows\Fonts\calibri.ttf", footer_size)

                # Define custom text
                To = 'To:'

                product_type = "Courier's By Demand Express"
                center_product_type = (barcode_image.width / 2) - len(product_type) * 5
                center_barcode_value = (barcode_image.width / 2) - len(barcode_param) * 8

                # Draw text on picture
                draw.text((l_r_margin, 200), To, fill=(0, 0, 0), font=h2_font)
                draw.text((40, 250), x['BuyerName'], fill=(0, 0, 0), font=h2_font)
                draw.text((40, 275), x['Company'], fill=(0, 0, 0), font=h2_font)
                draw.text((40, 300), x['AddressLine1'], fill=(0, 0, 0), font=h2_font)
                draw.text((40, 325), x['AddressLine2'], fill=(0, 0, 0), font=h2_font)
                draw.text((40, 350), x['City'], fill=(0, 0, 0), font=h2_font)
                draw.text((40, 375), x['State'], fill=(0, 0, 0), font=h2_font)
                draw.text((40, 400), x['Postcode'], fill=(0, 0, 0), font=h2_font)
                draw.text((40, 425), x['Phone'], fill=(0, 0, 0), font=h2_font)
                draw.text((40, 490), x['item_count_string'], fill=(0, 0, 0), font=h2_font)

                draw.text((center_product_type, (490)), product_type, fill=(0, 0, 0), font=footer_font)
                draw.text((center_barcode_value, (640)), barcode_param, fill=(0, 0, 0), font=h2_font)

                scarlett_music_logo = Image.open(rf"\\SERVER\Python\website_logo.png")

                half = 0.2
                out = scarlett_music_logo.resize([int(half * s) for s in scarlett_music_logo.size])

                new_image.paste(out, (170, 10))

                # save in file

                label_name = label_location + "\\" + str(x['String_SalesRecordNumber']) + ".pdf"

                new_image.save(label_name, 'PDF')

                Label_Location.append(label_name)

                for lines in range(len(x['SKU'])):
                    x['ItemTitle'][f'Name (Item {str(lines + 1)})'] = x['ItemTitle'][
                        f'Name (Item {str(lines + 1)})'].replace("'", "")

                    x['BuyerName'] = x['BuyerName'].replace("'", "")
                    x['Company'] = x['Company'].replace("'", "")
                    x['AddressLine1'] = x['AddressLine1'].replace("'", "")
                    x['AddressLine2'] = x['AddressLine2'].replace("'", "")
                    x['City'] = x['City'].replace("'", "")
                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()

                    try:
                        cursor.execute(
                            fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order,Sales_Record_Number , Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{x['String_OrderID'][0:50]}', '{x['String_PurchaseOrderNumber'][0:50]}', '{x['String_SalesRecordNumber'][0:50]}', 'ebay', '{x['BuyerName'][0:50]}', '{x['Company'][0:50]}', '{x['AddressLine1'][0:50]}', '{x['AddressLine2'][0:50]}', '{x['City'][0:50]}', '{x['State'][0:50]}', '{x['Postcode'][0:50]}', 'AU', '{x['Phone'][0:50]}', '{x['Email'][0:50]}', '{x['date_placed']}','' , '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['OrderLineID'][f'Orderline (Item {str(lines + 1)})'][0:50]}', '{x['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{x['ItemTitle'][f'Name (Item {str(lines + 1)})'][0:200]}', '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['String_OrderID'][0:50]}', '{x['Final_Courier']}','{barcode_param}' , '{str(x['item_count'])}', '{str(x['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                    except:
                        pass

            except (KeyError, TypeError):
                pass

        if x['Final_Courier'] == 'Sendle':
            try:

                x['Phone'] = x['Phone'].replace('+61', '0')
                x['Phone'] = x['Phone'].replace(' ', '')

                # instructions = 'Authority to leave'
                #
                # if float(totalprice) > 200 and volumevalue > 0.0001:
                #     instructions = 'Signature on Delivery'

                volumevalue = str(
                    float(float(x['Length']) / 100) * float(float(x['Width']) / 100) * float(float(x['Height']) / 100))

                payload = {"description": "Scarlett Music Order " + x['String_SalesRecordNumber'],
                           "weight": {"value": x['Weight'], "units": "kg"},
                           "volume": {"value": volumevalue, "units": "m3"},
                           "customer_reference": x['item_count_string'], "sender": {
                        "contact": {
                            "name": "Scarlett Music",
                            "phone": "(03) 9318 5751",
                            "company": "Scarlett Music"
                        },
                        "address": {
                            "address_line1": "286-288 Ballarat Rd",
                            "suburb": "Maidstone",
                            "state_name": "VIC",
                            "postcode": "3012",
                            "country": "Australia"
                        },
                        "instructions": "The music shop, open 9am-6pm. Best parking is at The Palms across the road."
                    }, "receiver": {
                        "contact": {
                            "name": x['BuyerName'],
                            "email": x['Email'],
                            "phone": x['Phone'],
                            "company": x['Company']
                        },
                        "address": {
                            "address_line1": x['AddressLine1'],
                            "address_line2": x['AddressLine2'],
                            "suburb": x['City'],
                            "state_name": x['State'],
                            "postcode": x['Postcode'],
                            "country": "Australia"
                        },
                        "instructions": ''
                    }
                           }

                r = requests.post('https://api.sendle.com/api/orders/',
                                  auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'), json=payload)

                # pprint.pprint(r.text)
                # pprint.pprint(r.json())
                response = r.text
                response = json.loads(response)

                # pprint.pprint(response)
                tracking = response['sendle_reference']
                orderurl = response['order_url']
                price = response['price']['gross']['amount']
                croppedpdfurl = response['labels'][1]['url']

                r = requests.get(croppedpdfurl, auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'),
                                 stream=True)

                label_name = label_location + "\\" + str(x['String_SalesRecordNumber']) + ".pdf"

                Label_Location.append(label_name)

                with open(label_name, "wb") as pdf:
                    for chunk in r.iter_content(chunk_size=1024):

                        # writing one chunk at a time to pdf file
                        if chunk:
                            pdf.write(chunk)

                for lines in range(len(x['SKU'])):
                    x['ItemTitle'][f'Name (Item {str(lines + 1)})'] = x['ItemTitle'][
                        f'Name (Item {str(lines + 1)})'].replace("'", "")

                    x['BuyerName'] = x['BuyerName'].replace("'", "")
                    x['Company'] = x['Company'].replace("'", "")
                    x['AddressLine1'] = x['AddressLine1'].replace("'", "")
                    x['AddressLine2'] = x['AddressLine2'].replace("'", "")
                    x['City'] = x['City'].replace("'", "")
                    try:
                        cursor = connection.cursor()
                    except:
                        connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")
                        cursor = connection.cursor()

                    try:
                        cursor.execute(
                            fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order,Sales_Record_Number , Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{x['String_OrderID'][0:50]}', '{x['String_PurchaseOrderNumber'][0:50]}', '{x['String_SalesRecordNumber'][0:50]}', 'ebay', '{x['BuyerName'][0:50]}', '{x['Company'][0:50]}', '{x['AddressLine1'][0:50]}', '{x['AddressLine2'][0:50]}', '{x['City'][0:50]}', '{x['State'][0:50]}', '{x['Postcode'][0:50]}', 'AU', '{x['Phone'][0:50]}', '{x['Email'][0:50]}', '{x['date_placed']}','' , '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['OrderLineID'][f'Orderline (Item {str(lines + 1)})'][0:50]}', '{x['Quantity'][f'Quantity (Item {str(lines + 1)})'][0:50]}', '{x['ItemTitle'][f'Name (Item {str(lines + 1)})'][0:200]}', '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['String_OrderID'][0:50]}', '{x['Final_Courier']}','{tracking}' , '{str(x['item_count'])}', '{str(x['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")
                    except:
                        pass

                maxrow = wsSendle.max_row
                maxrow = maxrow + 1
                wsSendle['A' + str(maxrow)].value = 'Lorelle Scarlett'  # 'sender_name'
                wsSendle['B' + str(maxrow)].value = 'Scarlett Music'  # 'sender_company'
                wsSendle['C' + str(maxrow)].value = '284-288 Ballarat Rd'  # 'sender_address_line1'
                wsSendle['E' + str(maxrow)].value = 'Footscray'  # 'sender_suburb'
                wsSendle['F' + str(maxrow)].value = 'VIC'  # 'sender_state_name'
                wsSendle['G' + str(maxrow)].value = '3011'  # 'sender_postcode'
                wsSendle['H' + str(maxrow)].value = 'Australia'  # 'sender_country'
                wsSendle['I' + str(maxrow)].value = '0417557472'  # 'sender_contact_number'
                wsSendle['J' + str(
                    maxrow)].value = 'At the music shop, open 9am-6pm. Parking is best across the road at The Palms.'  # 'pickup_instructions'  ##At the music shop, open 9am-6pm. Parking is best across the road at The Palms.
                wsSendle['K' + str(maxrow)].value = x['BuyerName']  # 'receiver_name'
                wsSendle['L' + str(maxrow)].value = x['Email']  # 'receiver_email'
                wsSendle['M' + str(maxrow)].value = x['Company']  # 'receiver_company'
                wsSendle['N' + str(maxrow)].value = x['AddressLine1']  # 'receiver_address_line1'
                wsSendle['O' + str(maxrow)].value = x['AddressLine2']  # 'receiver_address_line2'
                wsSendle['P' + str(maxrow)].value = x['City']  # 'receiver_suburb'
                wsSendle['Q' + str(maxrow)].value = x['State']  # 'receiver_state_name'
                wsSendle['R' + str(maxrow)].value = x['Postcode']  # 'receiver_postcode'
                wsSendle['S' + str(maxrow)].value = 'Australia'  # 'receiver_country'
                wsSendle['T' + str(maxrow)].value = x['Phone']  # 'receiver_contact_number'
                wsSendle['W' + str(maxrow)].value = x['Weight']  # 'kilogram_weight'
                wsSendle['X' + str(maxrow)].value = str(
                    float(float(x['Length']) / 100) * float(float(x['Width']) / 100) * float(
                        float(x['Height']) / 100))  # 'cubic_metre_volume' ### l x w x h (in meters!)
                wsSendle['Y' + str(maxrow)].value = 'Scarlett Music Order ' + x[
                    'String_SalesRecordNumber']  # 'description'
                wsSendle['Z' + str(maxrow)].value = x['String_SalesRecordNumber']  # 'customer_reference'

            except (KeyError, TypeError):
                pass

        if x['Final_Courier'] == 'Australia Post':

            try:

                x['Phone'] = x['Phone'].replace('+61', '0')
                x['Phone'] = x['Phone'].replace(' ', '')

                headers = {'Content-Type': 'application/json', 'Accept': 'application/json',
                           'Account-Number': accountnumber,
                           "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
                           "Accept-Encoding": "*",
                           "Connection": "keep-alive"}

                payload = {
                    "shipments": [
                        {
                            "shipment_reference": x['item_count_string'],
                            "customer_reference_1": x['item_count_string'],
                            "customer_reference_2": "SKU-1",
                            "email_tracking_enabled": 'true',
                            "from": {
                                "name": "Lorelle Scarlett",
                                "business_name": "Scarlett Music",
                                "lines": [
                                    "268-288 Ballarat Rd"
                                ],
                                "suburb": "FOOTSCRAY",
                                "state": "VIC",
                                "postcode": "3011",
                                "phone": "0422747033",
                                "email": "lorelle@scarlettmusic.com.au"
                            },
                            "to": {
                                "name": x['BuyerName'],
                                "business_name": x['Company'][0:50],
                                "lines": [
                                    x['AddressLine1'], x['AddressLine2']
                                ],
                                "suburb": x['City'],
                                "state": x['State'],
                                "postcode": x['Postcode'],
                                "phone": x['Phone'],
                                "email": x['Email']
                            },
                            "items": [{
                                "item_reference": x['item_count_string'] + '-1',
                                "product_id": '3D55',
                                "length": x['Length'],
                                "height": x['Height'],
                                "width": x['Width'],
                                "weight": x['Weight'],
                                "authority_to_leave": 'false',
                                "allow_partial_delivery": 'true',

                            },

                            ]
                        }
                    ]
                }

                try:

                    r = requests.post('https://digitalapi.auspost.com.au/shipping/v1/shipments', headers=headers,
                                      auth=HTTPBasicAuth(username, secret),
                                      json=payload)

                    response = r.text
                    response = json.loads(response)

                    itemID = response['shipments'][0]['items'][0]['item_id']
                    shipmentID = response['shipments'][0]['shipment_id']
                    totalcost = response['shipments'][0]['shipment_summary']['total_cost']
                    ausposttracking = response['shipments'][0]['items'][0]['tracking_details']['consignment_id']

                    payload = {
                        "wait_for_label_url": 'true',
                        "preferences": [
                            {
                                "type": "PRINT",
                                "format": "PDF",
                                "groups": [
                                    {
                                        "group": "Parcel Post",
                                        "layout": "THERMAL-LABEL-A6-1PP",
                                        "branded": 'true',
                                        "left_offset": 0,
                                        "top_offset": 0
                                    },
                                    {
                                        "group": "Express Post",
                                        "layout": "THERMAL-LABEL-A6-1PP",
                                        "branded": 'false',
                                        "left_offset": 0,
                                        "top_offset": 0
                                    }
                                ]
                            }
                        ],
                        "shipments": [
                            {
                                "shipment_id": shipmentID,
                                "items": [
                                    {
                                        "item_id": itemID
                                    }
                                ]
                            }
                        ]
                    }

                    r = requests.post('https://digitalapi.auspost.com.au/shipping/v1/labels', headers=headers,
                                      auth=HTTPBasicAuth(username, secret),
                                      json=payload)

                    response = r.text
                    response = json.loads(response)
                    # pprint.pprint(response)

                    url = response['labels'][0]['url']

                    r = requests.get(url, stream=True)

                    label_name = label_location + "\\" + str(x['String_SalesRecordNumber']) + ".pdf"

                    Label_Location.append(label_name)

                    with open(label_name, "wb") as pdf:
                        pdf.write(r.content)

                    for lines in range(len(x['SKU'])):
                        x['ItemTitle'][f'Name (Item {str(lines + 1)})'] = x['ItemTitle'][
                            f'Name (Item {str(lines + 1)})'].replace("'", "")

                        x['BuyerName'] = x['BuyerName'].replace("'", "")
                        x['Company'] = x['Company'].replace("'", "")
                        x['AddressLine1'] = x['AddressLine1'].replace("'", "")
                        x['AddressLine2'] = x['AddressLine2'].replace("'", "")
                        x['City'] = x['City'].replace("'", "")
                        try:
                            cursor = connection.cursor()
                        except:
                            connection = create_connection("postgres", "kyal", "123456789Kyal", "143.244.166.150",
                                                           "5432")
                            cursor = connection.cursor()
                        try:

                            cursor.execute(
                                fr"INSERT INTO dailyorders(ORDER_ID, Purchase_Order,Sales_Record_Number , Sales_Channel, Name, Company, Address1, Address2, City, State, Postcode, Country, Phone, Email, Date_Placed, Date_Completed, Item_Sku, Item_number_eBay, Item_Quantity, Item_Name, Image_URL, Multiple_Orders, Postage_Service, Tracking_Number, Misc1, Misc2, Misc3, Misc4, Misc5, Misc6, Misc7, Misc8, Misc9, Misc10) VALUES ('{x['String_OrderID'][0:50]}', '{x['String_PurchaseOrderNumber'][0:50]}', '{x['String_SalesRecordNumber'][0:50]}', 'ebay', '{x['BuyerName'][0:50]}', '{x['Company'][0:50]}', '{x['AddressLine1'][0:50]}', '{x['AddressLine2'][0:50]}', '{x['City'][0:50]}', '{x['State'][0:50]}', '{x['Postcode'][0:50]}', 'AU', '{x['Phone'][0:50]}', '{x['Email'][0:50]}', '{x['date_placed']}','' , '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['OrderLineID'][f'Orderline (Item {str(lines + 1)})'][0:50]}', '{x['Quantity'][f'Quantity (Item {str(lines + 1)})']}', '{x['ItemTitle'][f'Name (Item {str(lines + 1)})'][0:200]}', '{x['ItemSKU'][f'SKU (Item {str(lines + 1)})']}', '{x['String_OrderID'][0:50]}', '{x['Final_Courier']}','{ausposttracking}' , '{str(x['item_count'])}', '{str(x['item_count_string'])}','' ,'' ,'' , '', '', '','' ,'' ); COMMIT;")

                        except:
                            pass
                except:
                    pass

                maxrow = wsAustraliaPost.max_row
                maxrow = maxrow + 1
                wsAustraliaPost['B' + str(maxrow)].value = x['Email']  # 'C_CONSIGNEE_EMAIL' ### Email for reciever
                wsAustraliaPost[
                    'C' + str(maxrow)].value = 'Y'  # 'C_EMAIL_NOTIFICATION' ### Email notifications on? Always make 'Y'
                wsAustraliaPost['D' + str(maxrow)].value = str((float(x['Length']) * float(x['Width']) * float(x[
                                                                                                                   'Height'])) / 6000)  # 'A_ACTUAL_CUBIC_WEIGHT' ### Formula for this is length (cm) x width (cm) x height (cm) / 6000
                wsAustraliaPost['E' + str(maxrow)].value = x['Length']  # 'A_LENGTH' ### Length (cm)
                wsAustraliaPost['F' + str(maxrow)].value = x['Width']  # 'A_WIDTH' ### Width (cm)
                wsAustraliaPost['G' + str(maxrow)].value = x['Height']  # 'A_HEIGHT' ##Height (cm)
                wsAustraliaPost['H' + str(maxrow)].value = 'Scarlett Music Order ' + x[
                    'String_SalesRecordNumber']  # 'G_DESCRIPTION' ### Scarlett music order xxx
                wsAustraliaPost['I' + str(maxrow)].value = x['Weight']  # 'G_WEIGHT' ### Weight in kg
                wsAustraliaPost['J' + str(maxrow)].value = '3D55'  # 'CHRG_CODE'### Always 7D55 (Standard eParcel)
                wsAustraliaPost['K' + str(maxrow)].value = x['BuyerName']  # 'CNSGNEE_NAME' ### Customer Name
                wsAustraliaPost['L' + str(maxrow)].value = x['Company']  # 'CNSGNEE_BUS_NAME'###Company
                wsAustraliaPost['M' + str(maxrow)].value = x['AddressLine1']  # 'CNSGNEE_ADDR_LINE1' ### AddressLine1
                wsAustraliaPost['N' + str(maxrow)].value = x['AddressLine2']  # 'CNSGNEE_ADDR_LINE2' ###AddressLine2
                wsAustraliaPost['O' + str(maxrow)].value = x['City']  # 'CNSGNEE_SUBURB'###SUBURB
                wsAustraliaPost['P' + str(maxrow)].value = x['State']  # 'CNSGNEE_STATE_CODE' ##Short State Code
                wsAustraliaPost['Q' + str(maxrow)].value = 'AU'  # 'CNSGNEE_CNTRY_CODE' ##AU
                wsAustraliaPost['R' + str(maxrow)].value = x['Phone']  # 'CNSGNEE_PHONE_NBR' ##Phone
                wsAustraliaPost['S' + str(maxrow)].value = 'Y'  # 'IS_PHONE_PRNT_REQD' ##Show phone number, always 'Y'
                wsAustraliaPost['T' + str(maxrow)].value = 'Y'  # 'IS_SIGNTR_REQD' #Y always i guess
                wsAustraliaPost['U' + str(maxrow)].value = x['String_SalesRecordNumber']  # 'REF' ##Reference / Order ID
                wsAustraliaPost['V' + str(maxrow)].value = 'Y'  # 'IS_REF_PRINT_REQD' ##Show reference, always Y
                wsAustraliaPost['W' + str(maxrow)].value = x['Postcode']  # 'EMAIL_NOTIFICATION' ##always Y

            except KeyError:
                pass
#### Need to have it so that if weight = 0, item gets put back into label
### Need to add Items to master picking sheet. Just enought so the VBA script can take over.


wb.save(r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\MasterShippingSheet.xlsx')

wsMasterPickSheet['D1'].value = 'Areas'

connection = create_connection(
    "postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432")

for x in range(2, wsMasterPickSheet.max_row + 1):
    ItemTitle = wsMasterPickSheet['A' + str(x)].value
    SKU = wsMasterPickSheet['B' + str(x)].value
    Quantity = wsMasterPickSheet['C' + str(x)].value

    if SKU == '':
        wsMasterPickSheet['D' + str(x)].value = '5) Undefined'
        continue

    select_item = f"SELECT item_location FROM item_data WHERE sku = '{SKU}' ;"
    item = execute_read_query(connection, select_item)
    # print(item)

    if item == []:

        while True:  ### Will have to have item title also here below vvvvvv
            location = str(input(f'''\n\nLocation not found for '{ItemTitle}', (SKU:{SKU}). Where is it located?
        1) String Room
        2) Back Area
        3) Cage
        4) Out the Front
        5) I don't know!
        Enter Response:'''))

            if location == '1':
                location = '1) String Room'
                add_item(SKU, location)
                break

            if location == '2':
                location = '2) Back Area'
                add_item(SKU, location)
                break

            if location == '3':
                location = '3) Cage'
                add_item(SKU, location)
                break

            if location == '4':
                location = '4) Out the Front'
                add_item(SKU, location)
                break

            if location == '5':
                location = '5) Undefined'
                break

            else:
                print('Valid input not found. Try again.')
                time.sleep(1)
                continue

    else:
        location = item[0][0]

    wsMasterPickSheet['D' + str(x)].value = location

    if 'plectrums' in ItemTitle.lower() or 'pick' in ItemTitle.lower() or 'australasian' in ItemTitle.lower() or 'bridge pins' in ItemTitle.lower() or 'jim root' in ItemTitle.lower():
        pass


    else:

        try:

            multiple = re.search(r'^x?\s?\d+\s?x', ItemTitle, re.I).group()
            multiple = multiple.replace('x', '')
            multiple = multiple.replace('X', '')
            multiple = multiple.strip()

            newquantity = int(multiple) * int(Quantity)

            wsMasterPickSheet['C' + str(x)].value = str(newquantity)
        except AttributeError:
            continue

wb.save(r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\MasterShippingSheet.xlsx')

read_file = pd.read_excel(r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\MasterShippingSheet.xlsx',
                          sheet_name='MasterPickSheet')
read_file.to_csv(r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\MasterPickSheet.csv', index=None, header=True)

try:
    read_file.to_csv(r'\\SERVER\Python\MasterPickSheet.csv', index=None, header=True)

except:
    pass


if wsAustraliaPost['B3'].value is not None:

    read_file = pd.read_excel(
        r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\MasterShippingSheet.xlsx',
        sheet_name='AustraliaPost')
    read_file.to_csv(r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\AustraliaPost.csv', index=None, header=True)

else:
    del wb['AustraliaPost']

if wsSendle['A2'].value is not None:

    read_file = pd.read_excel(
        r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\MasterShippingSheet.xlsx',
        sheet_name='Sendle')
    read_file.to_csv(r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Sendle.csv', index=None, header=True)

else:
    del wb['Sendle']

if wsTransdirect['A2'].value is not None:

    read_file = pd.read_excel(
        r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\MasterShippingSheet.xlsx',
        sheet_name='Transdirect')
    read_file.to_csv(r'C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Transdirect.csv', index=None, header=True)

else:
    del wb['Transdirect']

# mergeDevilopes.pop(0)
# mergeMinilopes.pop(0)
# mergeLabel.pop(0)

if wsMini['A2'].value is not None:

    template = r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\Minilope Template JAN 2022.docx"

    document = MailMerge(template)
    # print(document.get_merge_fields())

    document.merge_templates(mergeMinilopes, separator='page_break')

    document.write(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Minilopes.docx")

else:
    del wb['Mini.']

if wsDevil['A2'].value is not None:

    template = r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\Devilope Template JAN 2022.docx"

    document = MailMerge(template)
    # print(document.get_merge_fields())

    document.merge_templates(mergeDevilopes, separator='page_break')

    document.write(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Devilopes.docx")

else:
    del wb['Devil.']

if wsLabel['A2'].value is not None:

    template = r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\Label Template JAN 2022.docx"
    template = r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\Label Template JAN 2022.docx"

    document = MailMerge(template)
    # print(document.get_merge_fields())

    document.merge_templates(mergeLabel, separator='continuous_section')

    document.write(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\Labels.docx")

    convert(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\Labels.docx",
            r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\Labels_Intermediate.pdf")

    generateNup(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Envelope Templates\Labels_Intermediate.pdf", 4,
                r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Labels.pdf")

else:
    del wb['Label.']

merger = PdfFileMerger()

for x in Label_Location:
    merger.append(x)

intermediate_label_location = label_location + r"/Intermediate_Labels.pdf"
merger.write(intermediate_label_location)

merger.close()

final_label_location = label_location + r"/Final_Labels.pdf"

generateNup(intermediate_label_location, 4, final_label_location)

with open(r"C:\Users\lorel\Documents\Python\DailyLabels\Final_Labels.pdf", "rb") as pdf_file:
    pdf_reader = PdfFileReader(pdf_file)
    total_pages = pdf_reader.numPages
writer = PyPDF2.PdfFileWriter()

pdf = PyPDF2.PdfFileReader(final_label_location)
for x in range(0, total_pages):
    page0 = pdf.getPage(x)
    page0.scaleBy(1.9)  # float representing scale factor - this happens in-place
    # create a writer to save the updated results
    writer.addPage(page0)
with open(r"C:\Users\lorel\Documents\Python\DailyLabels\Final_LabelsV2.pdf", "wb+") as f:
    writer.write(f)

if os.path.isfile(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Minilopes.docx") == True:
    input('\nPress Enter when Minilopes are in the printer.')

    os.startfile(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Minilopes.docx", "print")

    input('Press Enter when printing is finished.')

if os.path.isfile(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Devilopes.docx") == True:
    input('\nPress Enter when Devilopes are in the printer.')

    os.startfile(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Devilopes.docx", "print")

    input('Press Enter when printing is finished.')

if os.path.isfile(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Labels.pdf") == True:
    input('\nPress Enter when Label paper is in the printer.')

    os.startfile(r"C:\Users\lorel\Desktop\Daily Envelopes & CSVs\Labels.pdf", "print")
    os.startfile(r"C:\Users\lorel\Documents\Python\DailyLabels\Final_LabelsV2.pdf", "print")

    input('Press Enter when printing is finished.')

if couriers_please_item_count > 1:

    cp_url = "https://api.couriersplease.com.au/v2/domestic/bookPickup"

    cp_body = {
        "accountName": "Scarlett Music",
        "contactName": "Kyal Scarlett",
        "contactEmail": "kyal@scarlettmusic.com.au",
        "readyDateTime": next_cp_day,
        "specialInstructions": "Use front door",
        "consignmentCount": str(couriers_please_item_count),
        "totalItemCount": str(couriers_please_item_count),
        "totalWeight": str(math.ceil(float(couriers_please_weight))),
        "pickup": {
            "phoneNumber": "0393185751",
            "companyName": "Scarlett Music",
            "address1": "286-288",
            "address2": "Ballarat Rd",
            "address3": "",
            "postcode": "3011",
            "suburb": "Footscray"
        }
    }

    try:
        r = requests.post(cp_url, headers=cp_headers, json=cp_body)
    except:
        courierprice = 1000
    response = r.text
    response = json.loads(response)
    pprint.pprint(response)

input('\n\nScript finished successfully! Praise Durst! (Press Enter to exit.)')
