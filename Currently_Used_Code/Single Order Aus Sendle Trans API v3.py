from ebaysdk.exception import ConnectionError  ##Stuff for Ebay API
from ebaysdk.trading import Connection as Trading
import pprint, psycopg2, time, sys, pyperclip, re, math, holidays, zeep, base64, uuid
from zeep.transports import Transport
from requests import Session
from zeep.plugins import HistoryPlugin
from lxml import etree
import xmltodict
import ctypes.wintypes
from psycopg2 import OperationalError
import concurrent.futures
import openpyxl
import code128

from PyPDF2 import PdfWriter, PdfReader

import urllib, json, requests  # Stuff for Google API
from urllib.parse import urlparse
from time import sleep
from oauthlib.oauth2 import BackendApplicationClient
from requests_oauthlib import OAuth2Session

from requests.auth import HTTPBasicAuth
from pdf2image import convert_from_path, convert_from_bytes
import image_slicer  # Stuff for Sendle API

try:
    from pil import Image, ImageDraw, ImageFont, ImageChops
    from pil import ImageOps

except ModuleNotFoundError as err:
    from PIL import Image, ImageDraw, ImageFont, ImageChops
    from PIL import ImageOps

import datetime

import brother_ql, os
from brother_ql.raster import BrotherQLRaster
from brother_ql.backends.helpers import send  # Stuff for label printer

# from selenium import webdriver
from bs4 import BeautifulSoup

# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC  # Stuff for Transdirect Tracking
# from selenium.common.exceptions import TimeoutException
import re
# from selenium.webdriver.chrome.options import Options

from rich.console import Console

console = Console(force_terminal=True)


def create_connection(db_name, db_user, db_password, db_host, db_port):
    connection = None
    try:
        connection = psycopg2.connect(
            database=db_name,
            user=db_user,
            password=db_password,
            host=db_host,
            port=db_port,
        )
    # print("Connection to PostgreSQL DB successful")
    except OperationalError as e:
        print(f"The error '{e}' occurred")
    return connection


connection = create_connection(
    "postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432"
)


def execute_read_query(connection, item_check):
    needs_weight_check = 'false'
    cursor1 = connection.cursor()
    result = None
    try:
        result = cursor1.execute(item_check)
        result = cursor1.fetchall()

        if result == []:
            needs_weight_check = 'true'
            return needs_weight_check

        else:
            return needs_weight_check

    except OperationalError as e:
        print(e)
        sys.exit()


def execute_insert_query(connection, item_check):
    cursor1 = connection.cursor()
    result = None
    try:
        result = cursor1.execute(item_check)
        # result = cursor1.fetchall()

    except OperationalError as e:
        print(e)
        sys.exit()


def item_insert(sku):
    item_check = f"INSERT INTO sendle_weight_test(item_number) VALUES('{sku}');COMMIT;"
    execute_insert_query(connection, item_check)


def item_check(sku):
    item_check = f"SELECT 1 FROM sendle_weight_test WHERE item_number = '{sku}';"
    results = execute_read_query(connection, item_check)
    return results


def item_thickness_check(sku):
    item_check = f"SELECT 1 FROM fastwaythicknesstest WHERE item_number = '{sku}';"
    results = execute_read_query(connection, item_check)
    return results


def item_thickness_insert(sku):
    item_check = f"INSERT INTO fastwaythicknesstest(item_number) VALUES('{sku}');COMMIT;"
    execute_insert_query(connection, item_check)


connection = create_connection(
    "postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432"
)


def get_google_address():  # inital definiton for google places api
    global suburb
    global postal_code
    global state
    dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
    qstr = urllib.parse.urlencode(dict1)
    URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
    URL = URL + qstr
    response = urllib.request.urlopen(URL)
    data = json.load(response)
    # pprint.pprint(data['candidates'])
    placeid = data['candidates'][0]['place_id']
    payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
    r = requests.get(
        'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
    # pprint.pprint(r.json())
    data = r.json()
    #  pprint.pprint(data['result']['address_components'])
    for t in range(20):
        if data['result']['address_components'][t]['types'][0] == 'locality':
            suburb = data['result']['address_components'][t]['long_name']
            break

        else:
            continue

    for t in range(20):
        if data['result']['address_components'][t]['types'][0] == 'postal_code':
            postal_code = data['result']['address_components'][t]['long_name']
            break

        else:
            continue

    for t in range(20):
        if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
            google_state = data['result']['address_components'][t]['short_name']
            break

        else:
            continue

    for t in range(20):
        if data['result']['address_components'][t]['types'][0] == 'street_number':
            google_street1 = data['result']['address_components'][t]['short_name']
            break

    for t in range(20):
        if data['result']['address_components'][t]['types'][0] == 'route':
            google_street2 = data['result']['address_components'][t]['long_name']
            break
    backup_google_street_address = google_street1 + ' ' + google_street2

    postcode = postal_code
    city = suburb
    state = google_state


try:
    r = requests.get('https://uselessfacts.jsph.pl/random.json?language=en')

    fun_response = r.text

    fun_response = json.loads(fun_response)

    fun_response = fun_response['text']

    # print(f"\n{response['text']}\n")


except:
    pass


def GetQuote(courier):
    global srn
    global name
    global email
    global address1
    global address2
    global city
    global state
    global postcode
    global height
    global length, mailplus_address, dai_failed, po_box, final_dai_weight
    global weightvalue, backup_google_street_address, id, Job, JobIDs, lowesttranscourier, trans_response, quotes, final_freightster_weight, mailplusfailed, mailplusreference
    global phone, company, sendlefailed, toll_failed, auspostfailed, transdirectfailed, freightsterfailed, fastwayfailed, allied_failed, aus_post_item_array, fastway_data, couriers_failed, next_cp_day, cp_auth, cp_auth_encoded, cp_headers, cp_validate_body, cp_body, cp_item_array, cbdexpressfailed, bondsfailed

    try:

        split_name = name.split(' ')

        first_name = split_name[0]

        last_name = ''
        for x in range(1, len(split_name)):
            last_name = f'{last_name} {split_name[x]}'

        last_name = last_name.strip()
    except:
        first_name = name
        last_name = ''

    # print(courier)

    if courier == 'Australia Post':
        try:

            if auspostfailed == 'true':
                courierprice = 10000


            else:

                product_id = "3D55"
                #print(multi_items_to_ship)

                if answer == '3':
                    product_id = '3J55'

                cover_amount = float(totalprice) / 2
                if cover_amount >= 499:
                    cover_amount = 499

                aus_post_item_array = []

                aus_post_parcel_number = 0
                for all_items in multi_items_to_ship:
                    aus_post_parcel_number = aus_post_parcel_number + 1
                    quantity = all_items[6]
                    for t in range(quantity):
                        aus_post_item_array.append({
                            "item_reference": OrderID + '-' + str(aus_post_parcel_number),
                            "product_id": product_id,
                            "length": all_items[1],
                            "height": all_items[2],
                            "width": all_items[3],
                            "weight": all_items[0],
                            "authority_to_leave": 'false',
                            "allow_partial_delivery": 'true',
                        })

                headers = {'Content-Type': 'application/json', 'Accept': 'application/json',
                           'Account-Number': accountnumber}

                payload = {"shipments": [{
                    "from": {
                        "suburb": "FOOTSCRAY",
                        "state": "VIC",
                        "postcode": "3011"
                    },
                    "to": {
                        "suburb": str(city),
                        "state": str(state),
                        "postcode": str(postcode)
                    },
                    "items":
                        aus_post_item_array}]}

                r = requests.post('https://digitalapi.auspost.com.au/shipping/v1/prices/shipments', headers=headers,
                                  auth=HTTPBasicAuth(username, secret), json=payload)
                response = r.text
                response = json.loads(response)
                # pprint.pprint(response)
                courierprice = response['shipments'][0]['shipment_summary']['total_cost']

                print('Aus Post price: ' + str(courierprice))

        except KeyError:
            courierprice = 1000

    if courier == 'Bonds':

        if bondsfailed == 'true':
            courierprice = 1000

        else:
            bonds_item_string = ''

            now = datetime.datetime.now()

            next_day = next_business_day()
            next_bonds_day = f"{next_day.strftime('%Y-%m-%d')}"

            bonds_total_items = 0
            bonds_total_weight = 0
            for all_items in multi_items_to_ship:
                quantity = all_items[6]

                package = [float(all_items[1]), float(all_items[2]), float(all_items[3])]
                sorted_package = sorted(package)

                for t in range(quantity):
                    bonds_total_items = bonds_total_items + 1
                    bonds_total_weight += float(all_items[0])

                bonds_item_string = f'''{bonds_item_string}<dimension>
                                    <qty>{quantity}</qty>
                                    <length>{sorted_package[2]}</length>
                                    <width>{sorted_package[1]}</width>
                                    <height>{sorted_package[0]}</height>
                                    </dimension>'''

            service_vehicle_tuples = [('C', 'CAR'), ('C', 'SW'), ('C', 'SV'), ('TTK', '')]

            #need to iterate through service_vehicle_types to get correct vehicle
            #Need to set date to next day

            for service_vehicle_groups in service_vehicle_tuples:

                xml_payload = f'''
                <job xmlns:xi="http://www.w3.org/2001/XInclude" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="job-bonds.xsd">
                <job_action>QUOTE</job_action>
                <notifications>
                <notification>
                <notify_type>DELIVERY</notify_type>
                <notify_target>{email}</notify_target>
                </notification>
                <notification>
                <notify_type>DELIVERY</notify_type>
                <notify_target>{phone}</notify_target>
                </notification>
                </notifications>
                <job_id/>
                <account>V01523</account>
                <authorization_code>@WV6mSH4NByW</authorization_code>
                <containsDangerousGoods>false</containsDangerousGoods>
                <branch>MEL</branch>
                <job_date>{next_bonds_day}</job_date>
                <time_ready>09:00:00</time_ready>
                <deliver_by_time xsi:nil="true"/>
                <deliver_by_time_reason xsi:nil="true"/>
                <order_number>{OrderID}</order_number>
                <contact>Kyal</contact>
                <insurance>true</insurance>
                <references>
                <reference>{OrderID}</reference>
                </references>
                <service_code>{service_vehicle_groups[0]}</service_code>
                <vehicle_code>{service_vehicle_groups[1]}</vehicle_code>
                <goods_description/>
                <instructions></instructions>
                <pallets/>
                <cubic/>
                <job_legs>
                <job_leg>
                <action>P</action>
                <service_qual/>
                <suburb>Footscray</suburb>
                <state>VIC</state>
                <company>SCARLETT MUSIC</company>
                <address1>286-288 Ballarat Rd</address1>
                <address2></address2>
                <contact>Kyal</contact>
                <items>{bonds_total_items}</items>
                <weight>{bonds_total_weight}</weight>
                <dimensions>
                {bonds_item_string}
                </dimensions>
                <references>
                <reference/>
                </references>
                </job_leg>
                <job_leg>
                <action>D</action>
                <service_qual></service_qual>
                <suburb>{city}</suburb>
                <state>{state}</state>
                <company>{company}</company>
                <address1>{address1}</address1>
                <address2>{address2}</address2>
                <contact/>
                <items>{bonds_total_items}</items>
                <weight>{bonds_total_weight}</weight>
                <dimensions>
                {bonds_item_string}
                </dimensions>
                <references>
                <reference/>
                </references>
                </job_leg>
                </job_legs>
                </job>
                '''

                url = 'https://appsrv.bondscouriers.com.au/bondsweb/api/upload-xml-job.htm'  # Replace with the actual API endpoint URL

                headers = {
                    'Content-Type': 'application/xml'
                }
                try:
                    response = requests.post(url, data=xml_payload, headers=headers, timeout=5)

                    # print(response.status_code)
                    xml_response = response.text


                # Parse the XML and convert it to a Python dictionary
                    data_dict = xmltodict.parse(xml_response)

                except:
                    continue
                # pprint.pprint(data_dict)

                try:
                    msg_status = data_dict['job_message']['msg_status']
                    if msg_status == 'ERROR':
                        continue
                except:
                    continue

                else:
                    fuel_charge = float(data_dict['job_message']['job_details']['fuel_charge'])
                    job_charge = float(data_dict['job_message']['job_details']['job_charge'])
                    gst_charge = float(data_dict['job_message']['job_details']['gst'])
                    quote_price = fuel_charge + job_charge + gst_charge

                    if service_vehicle_groups[0] == "C":
                        service_code = service_vehicle_groups[0]
                        vehicle_code = service_vehicle_groups[1]
                    else:
                        service_code = service_vehicle_groups[0]
                        vehicle_code = data_dict['job_message']['job_details']['vehicle_code']
                    break

            try:

                courierprice = quote_price

            except:
                courierprice = 1000

            print(f"Bonds Transport Price: {courierprice}")

    if courier == 'CBDExpress':

        if len(multi_items_to_ship) > 1:
            cbdexpressfailed = 'true'

        # print(cbdexpressfailed)

        if 'vic' not in state.lower():
            cbdexpressfailed = 'true'
        cbdexpressfailed = 'true'

        if ((float(length) * float(width) * float(height)) / 4000) > 25 or float(
                weightvalue) > 25 or length > 120 or width > 120 or height > 120:
            cbdexpressfailed = 'true'

        if cbdexpressfailed == 'true':
            courierprice = 1000

        else:
            cbd_workbook = openpyxl.load_workbook(
                rf'\\SERVER\Project Folder\Python\Courier Info\CBD_Express_Areas.xlsx')

            cbd_sheet = cbd_workbook['Sheet']
            courierprice = 1000

            for xx in range(2, cbd_sheet.max_row + 1):
                cbd_suburb_check = cbd_sheet['A' + str(xx)].value.strip().lower()
                # print(cbd_suburb_check)
                if cbd_suburb_check == city.lower().strip():
                    courierprice = float(cbd_sheet['C' + str(xx)].value)

        print(f"Courier's By Demand Price: {courierprice}")

    if courier == 'Mailplus':

        # mailplusfailed = 'true'

        if len(multi_items_to_ship) > 1:
            mailplusfailed = 'true'

        print(mailplusfailed)

        if mailplusfailed == 'true':
            courierprice = 1000

        else:

            if 'ebay' in address1:
                mailplus_address = address2
            else:
                mailplus_address = (str(address1) + ', ' + str(address2)).strip()

            try:

                cbd_workbook = openpyxl.load_workbook(
                    rf'\\SERVER\Project Folder\Python\Courier Info\Toll_Surcharges.xlsx')

                cbd_sheet = cbd_workbook['Toll_Surcharges']
                surcharge = 0
                for xx in range(2, cbd_sheet.max_row + 1):
                    cbd_suburb_check = cbd_sheet['A' + str(xx)].value
                    # print(cbd_suburb_check)
                    if cbd_suburb_check is None:
                        continue
                    if str(cbd_suburb_check) == postcode.strip():
                        surcharge = float(cbd_sheet['D' + str(xx)].value)

                mailplus_service = 'express'

                headers = {'x-api-key': 'CSKNIYW1oa5R0fXhG7gzmxwlxW2zunE8tAT4doW6', 'Content-Type': 'application/json',
                           'Accept': 'application/json'}
                payload = {
                    "service": mailplus_service,
                    "receiver": {
                        "address": {
                            "address_line1": mailplus_address,
                            "suburb": city,
                            "postcode": postcode,
                            "state": state
                        }
                    },
                    "sender": {
                        "address": {
                            "address_line1": "288 Ballarat Rd",
                            "suburb": "Footscray",
                            "postcode": "3011",
                            "state": "VIC"
                        }
                    },
                    "volume": {
                        "units": "m3",
                        "value": str(volumevalue)
                    },
                    "weight": {
                        "units": "kg",
                        "value": str(weightvalue)
                    },
                    "dimension": {
                        "length": {
                            "units": "cm",
                            "value": str(length)
                        },
                        "width": {
                            "units": "cm",
                            "value": str(width)
                        },
                        "height": {
                            "units": "cm",
                            "value": str(height)
                        }
                    }
                }

                r = requests.post('https://papi.mailplus.com.au/api/quote',
                                  headers=headers,
                                  json=payload)
                response = r.text

                # print(r)
                response = r.text
                response = json.loads(response)

                # pprint.pprint(response)

                courierprice = round(float(response['amount'] * 1.1) * 1.3116, 2)+surcharge+1.65
                if float(response['amount']) == 0:
                    courierprice = 1000

                print('Mailplus price: ' + str(courierprice))

            except Exception as e:
                #print(e)
                courierprice = 1000

    if courier == 'Sendle':

        sendlefailed = 'true'
        # if answer == '3':
        #     sendlefailed = 'true'

        if len(multi_items_to_ship) > 1:
            sendlefailed = 'true'

        if sendlefailed == 'true' or float(totalprice) > 200:
            courierprice = 1000
            pass

        else:

            # sendle_pickup_suburb = 'Footscray'
            # sendle_pickup_postcode = '3011'
            #
            # if float(weightvalue) <= 0.25:
            #     sendle_pickup_suburb = 'Maidstone'
            #     sendle_pickup_postcode = '3012'

            try:

                data = {
                    "pickup_suburb": 'Maidstone',
                    "pickup_postcode": '3012',
                    "delivery_suburb": city,
                    "delivery_postcode": str(postcode),
                    "weight_value": str(weightvalue),
                    'weight_units': 'kg',
                    'volume_value': str(volumevalue),
                    'volume_units': 'm3'}

                if answer == '3':
                    r = requests.get(
                        f'https://api.sendle.com/api/products?sender_suburb=Maidstone&sender_postcode=3012&sender_country=AU&weight_value={str(weightvalue)}&weight_units=kg&volume_value={str(volumevalue)}&volume_units=m3&receiver_country=AU&receiver_postcode={str(postcode)}&receiver_suburb={city}',
                        auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'), timeout=30)

                    sendleresponse = json.loads(r.text)

                    courierprice = 1000

                    for t in sendleresponse:
                        if t['product']['code'] == 'EXPRESS-PICKUP':
                            courierprice = t['quote']['gross']['amount']



                else:

                    r = requests.get('https://api.sendle.com/api/quote',
                                     auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'), json=data, timeout=30)
                    sendleresponse = json.loads(r.text)
                    courierprice = sendleresponse[0]['quote']['gross']['amount']

            except:
                courierprice = 1000

        print('Sendle price: ' + str(courierprice))

    if courier == 'Toll':

        if answer == '13':
            toll_failed = 'true'

        if toll_failed == 'true':
            courierprice = 1000
            pass

        else:

            toll_item_array = []
            additional_costs = 0

            for all_items in multi_items_to_ship:

                # for _ in range(int(all_items[6])):

                toll_item_array.append({"Commodity": {"CommodityCode": "Z",
                                                      "CommodityDescription": "ALL FREIGHT"},
                                        "ShipmentItemTotals": {
                                            "ShipmentItemCount": str(all_items[6])},
                                        "Dimensions": {
                                            "Width": str(math.ceil(float(all_items[3]))),
                                            "Length": str(math.ceil(float(all_items[1]))),
                                            "Height": str(math.ceil(float(all_items[2]))),
                                            "Volume": str(all_items[4]),
                                            "Weight": str(all_items[0])
                                        }})

                package = [math.ceil(float(all_items[3])), math.ceil(float(all_items[1])),
                           math.ceil(float(all_items[2]))]
                sorted_package = sorted(package)

                for _ in range(int(all_items[6])):

                    if float(all_items[0]) > 35:
                        additional_costs += 77.55

                    if float(all_items[0]) > 35 or sorted_package[0] > 180 or sorted_package[1] > 180 or sorted_package[
                        2] > 180 or float(all_items[4]) > 0.7:
                        additional_costs += 50

                    elif float(all_items[0]) > 30 or sorted_package[0] > 60 or sorted_package[1] > 80 or \
                            sorted_package[2] > 120 or float(all_items[4]) > 0.7:
                        additional_costs += 12

                if float(totalprice) < 500:
                    additional_costs += 6.95
                else:
                    additional_costs += float(totalprice) * 0.02

            surcharge = 0

            try:
                cbd_workbook = openpyxl.load_workbook(
                    rf'\\SERVER\Project Folder\Python\Courier Info\Toll_Road_Express_Surcharges.xlsx')

                cbd_sheet = cbd_workbook['Sheet1']
                surcharge = 0
                for xx in range(2, cbd_sheet.max_row + 1):
                    cbd_suburb_check = cbd_sheet['A' + str(xx)].value
                    # print(cbd_suburb_check)
                    if cbd_suburb_check is None:
                        continue
                    if str(cbd_suburb_check) == postcode.strip():
                        surcharge = float(cbd_sheet['E' + str(xx)].value)

            except:
                pass

            additional_costs += surcharge

            message_identifier = str(uuid.uuid4())
            now = datetime.datetime.now()
            current_time = f"{now.strftime('%Y-%m-%d')}T{now.strftime('%H:%M:%S')}+10:00"

            next_day = next_business_day()
            next_toll_day = f"{next_day.strftime('%Y-%m-%d')}T09:00:00+10:00"

            environment = 'PRD'  #######################Will have to change this in prod

            headers = {'Content-Type': 'application/json', 'Accept': 'application/json'}

            test_rate_enquiry_headers = {
                "MessageVersion": "3.1",
                "MessageIdentifier": message_identifier,
                "CreateTimestamp": current_time,
                "DocumentType": "RateEnquiry",
                "Environment": environment,
                "MessageSender": "SCARLETTMUSIC",
                "MessageReceiver": "TOLL",
                'SourceSystemCode': 'SCAR'
            }
            # rate_enquiry_test_url = "https://api-uat.teamglobalexp.com:6930/gateway/TollMessageRateEnquiryRestService/1.0/tom/rateEnquiry"
            rate_enquiry_prod_url = "https://api.teamglobalexp.com:6930/gateway/TollMessageRateEnquiryRestService/1.0/tom/rateEnquiry"
            price = totalprice
            payload = {
                "Request": {
                    "BusinessID": "IPEC",
                    "SystemFields": {"PickupDateTime": next_toll_day},
                    "ShipmentService": {
                        "ServiceCode": "X",
                        "ServiceDescription": "Road Express"
                    },
                    "ShipmentFlags": {
                        "ExtraServiceFlag": "true"},
                    "ShipmentFinancials": {
                        "ExtraServicesAmount": {
                            "Currency": "AUD",
                            "Value": str(round(float(totalprice)))}},
                    "BillToParty": {
                        "AccountCode": "80119621"
                    },
                    "ConsignorParty": {
                        "PhysicalAddress": {
                            "Suburb": "FOOTSCRAY",
                            "StateCode": "VIC",
                            "PostalCode": "3011",
                            "CountryCode": "AU"
                        }
                    },
                    "ConsigneeParty": {
                        "PhysicalAddress": {
                            "Suburb": city,
                            "StateCode": state,
                            "PostalCode": str(postcode),
                            "CountryCode": "AU"
                        }
                    },
                    "ShipmentItems": {
                        "ShipmentItem": toll_item_array
                    }
                }
            }

            toll_message = {"@version": "3.1",
                            "@encoding": "utf-8",
                            "TollMessage": {"Header": test_rate_enquiry_headers,
                                            "RateEnquiry": payload}}

            attempts = 0

            while attempts < 6:

                r = requests.post(url=rate_enquiry_prod_url, auth=('accounts@scarlettmusic.com.au', 't2TrAPsNTB'),
                                  json=toll_message, headers=headers, timeout=100)

                response = r.text

                try:

                    response = json.loads(response)

                    if 'timeout' in response['TollMessage']['ErrorMessages']['ErrorMessage'][0]['ErrorMessage'].lower():
                        message_identifier = str(uuid.uuid4())
                        attempts += 1
                        continue

                    else:
                        courierprice = 1000
                        break

                except:
                    break

            try:
                courierprice = float(
                    response['TollMessage']['RateEnquiry']['Response']['TotalChargeAmount']['Value']) + additional_costs

                # if 'wa' in state.lower():
                #     courierprice = courierprice*1.15 #Remove once disaster levy has been lifted

                print('Toll price: ' + str(courierprice))

            except:
                courierprice = 1000

    if courier == 'Dai Post':

        if answer == '15':
            dai_failed = 'true'

        if len(multi_items_to_ship) > 1:
            dai_failed = 'true'

        if dai_failed == 'true' or float(totalprice) > 200:
            courierprice = 1000
            pass

        else:
            dai_volume_value = ((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250
            # if float(dai_volume_value) > float(weightvalue):
            #     final_dai_weight = dai_volume_value
            # elif float(weightvalue) > float(dai_volume_value):
            #     final_dai_weight = weightvalue
            # else:
            #     final_dai_weight = weightvalue
            final_dai_weight = weightvalue
            final_dai_weight = round(float(final_dai_weight), 2)

            dai_postcode_wb = openpyxl.load_workbook(rf"\\SERVER\Project Folder\Python\Courier Info\Dai_Postcodes.xlsx")
            # dai_postcode_wb = openpyxl.load_workbook(rf"\\SERVER\Project Folder\Python\Courier Info\Dai_Postcodes2.xlsx") CHANGE ON RATES ROLLOVER


            dai_postcode_sheet = dai_postcode_wb['Sheet']

            dai_zone_sheet1 = ''

            dai_rates_wb = openpyxl.load_workbook(rf"\\SERVER\Project Folder\Python\Courier Info\Dai_Rates.xlsx")

            dai_rates_sheet = dai_rates_wb['Sheet']
            # dai_rates_sheet = dai_rates_wb['Sheet2'] CHANGE ON RATES ROLLOVER


            for x in range(2,
                           dai_postcode_sheet.max_row + 1):  ###ITERATING THROUGH THE DAIPOST POSTCODE SHEET TO FIND REGION

                sheet_postcode = str(dai_postcode_sheet['A' + str(x)].value)

                if str(len(sheet_postcode)) == '3':
                    sheet_postcode = f'0{sheet_postcode}'

                if postcode == sheet_postcode:
                    dai_zone_sheet1 = str(dai_postcode_sheet['B' + str(x)].value)

                    for xx in range(2, dai_rates_sheet.max_row + 1):

                        dai_zone_sheet2 = str(dai_rates_sheet['A' + str(xx)].value)

                        if dai_zone_sheet1.lower() == dai_zone_sheet2.lower():  ###Two zones match

                            if final_dai_weight <= 0.5:
                                dai_price = float(dai_rates_sheet['B' + str(xx)].value) * 1.1

                            elif 0.501 <= final_dai_weight <= 1.00:
                                dai_price = float(dai_rates_sheet['C' + str(xx)].value) * 1.1

                            elif 1.01 <= final_dai_weight <= 1.5:
                                dai_price = float(dai_rates_sheet['D' + str(xx)].value) * 1.1

                            elif 1.501 <= final_dai_weight <= 2.0:
                                dai_price = float(dai_rates_sheet['D' + str(xx)].value) * 1.1

                            elif 2.01 <= final_dai_weight <= 2.5:
                                dai_price = float(dai_rates_sheet['E' + str(xx)].value) * 1.1

                            elif 2.501 <= final_dai_weight <= 3.0:
                                dai_price = float(dai_rates_sheet['E' + str(xx)].value) * 1.1

                            elif 3.01 <= final_dai_weight <= 3.50:
                                dai_price = float(dai_rates_sheet['F' + str(xx)].value) * 1.1

                            elif 3.501 <= final_dai_weight <= 4.0:
                                dai_price = float(dai_rates_sheet['F' + str(xx)].value) * 1.1

                            elif 4.01 <= final_dai_weight <= 4.50:
                                dai_price = float(dai_rates_sheet['G' + str(xx)].value) * 1.1

                            elif 4.501 <= final_dai_weight <= 5.00:
                                dai_price = float(dai_rates_sheet['G' + str(xx)].value) * 1.1

                            elif 5.01 <= final_dai_weight <= 7.0:
                                dai_price = float(dai_rates_sheet['H' + str(xx)].value) * 1.1

                            elif 7.01 <= final_dai_weight <= 10.0:
                                dai_price = float(dai_rates_sheet['I' + str(xx)].value) * 1.1

                            elif 10.01 <= final_dai_weight <= 15.00:
                                dai_price = float(dai_rates_sheet['J' + str(xx)].value) * 1.1

                            elif 15.01 <= final_dai_weight <= 22.00:
                                dai_price = float(dai_rates_sheet['K' + str(xx)].value) * 1.1

                            else:
                                dai_price = 1000

                            if po_box == 'true':
                                dai_price += 0.8

                            courierprice = round(dai_price, 2)

                            print('Dai Price: $' + str(courierprice))

        try:
            courierprice
        except:
            courierprice = 1000

        try:
            dai_price
        except:
            dai_price = 1000
            courierprice = 1000
            print('Dai Price: $' + str(courierprice))

    if courier == 'Freightster':

        freightsterfailed = 'true'

        if len(multi_items_to_ship) > 1:
            freightsterfailed = 'true'

        if freightsterfailed == 'true' or float(totalprice) > 200:
            courierprice = 1000
            pass

        else:

            try:

                for x in range(2, sheet.max_row + 1):
                    if sheet['A' + str(x)].value.lower() == city.lower():
                        if int(sheet['B' + str(x)].value) == int(postcode):
                            freightster_pricing_zone = sheet['D' + str(x)].value

                freightster_volume_value = ((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250

                # print(f'l : {length}')
                # print(f'w : {width}')
                # print(f'l : {height}')
                if float(freightster_volume_value) > float(weightvalue):
                    final_freightster_weight = freightster_volume_value
                elif float(weightvalue) > freightster_volume_value:
                    final_freightster_weight = weightvalue

                if float(length) <= 4.0 and float(width) <= 23.0 and float(height) <= 23.0 and float(
                        weightvalue) <= 0.25:
                    final_freightster_weight = 0.250

                elif float(length) <= 18.0 and float(width) <= 23.0 and float(height) <= 4.0 and float(
                        weightvalue) <= 0.25:
                    final_freightster_weight = 0.250

                elif float(weightvalue) <= 0.25 and freightster_volume_value < 0.331:
                    final_freightster_weight = 0.250

                elif float(length) <= 18.0 and float(width) <= 23.0 and float(height) <= 4.0 and float(
                        weightvalue) <= 0.5:
                    final_freightster_weight = 0.500

                if freightster_pricing_zone == 'ACT' or freightster_pricing_zone == 'CBR' or freightster_pricing_zone == 'NSW':
                    if float(final_freightster_weight) <= 0.250:
                        freightster_price = 4.60

                    elif 0.251 <= float(final_freightster_weight) <= 0.5:
                        freightster_price = 5.5

                    elif 0.501 <= float(final_freightster_weight) <= 1.000:
                        freightster_price = 6.5

                    elif 1.01 <= float(final_freightster_weight) <= 3.000:
                        freightster_price = 7.5

                    elif 3.01 <= float(final_freightster_weight) <= 5.000:
                        freightster_price = 8.5

                    elif 5.01 <= float(final_freightster_weight) <= 15:
                        freightster_price = 22.5

                    elif 15.01 <= float(final_freightster_weight) <= 20:
                        freightster_price = 28.5

                    else:
                        freightster_price = 1000

                elif freightster_pricing_zone == 'ADL':
                    if float(final_freightster_weight) <= 0.250:
                        freightster_price = 4.60

                    elif 0.251 <= float(final_freightster_weight) <= 0.5:
                        freightster_price = 5

                    elif 0.501 <= float(final_freightster_weight) <= 1.000:
                        freightster_price = 5.5

                    elif 1.01 <= float(final_freightster_weight) <= 3.000:
                        freightster_price = 6.4

                    elif 3.01 <= float(final_freightster_weight) <= 5.000:
                        freightster_price = 7.5

                    elif 5.01 <= float(final_freightster_weight) <= 15:
                        freightster_price = 19

                    elif 15.01 <= float(final_freightster_weight) <= 20:
                        freightster_price = 28

                    else:
                        freightster_price = 1000

                elif freightster_pricing_zone == 'BNE':
                    if float(final_freightster_weight) <= 0.250:
                        freightster_price = 4.60

                    elif 0.251 <= float(final_freightster_weight) <= 0.5:
                        freightster_price = 5

                    elif 0.501 <= float(final_freightster_weight) <= 1.000:
                        freightster_price = 5

                    elif 1.01 <= float(final_freightster_weight) <= 3.000:
                        freightster_price = 6.25

                    elif 3.01 <= float(final_freightster_weight) <= 5.000:
                        freightster_price = 7.3

                    elif 5.01 <= float(final_freightster_weight) <= 15:
                        freightster_price = 18

                    elif 15.01 <= float(final_freightster_weight) <= 20:
                        freightster_price = 27

                    else:
                        freightster_price = 1000

                elif freightster_pricing_zone == 'MEL':
                    if float(final_freightster_weight) <= 0.250:
                        freightster_price = 3.75

                    elif 0.251 <= float(final_freightster_weight) <= 0.5:
                        freightster_price = 4.2

                    elif 0.501 <= float(final_freightster_weight) <= 1.000:
                        freightster_price = 4.5

                    elif 1.01 <= float(final_freightster_weight) <= 3.000:
                        freightster_price = 4.85

                    elif 3.01 <= float(final_freightster_weight) <= 5.000:
                        freightster_price = 5.5

                    elif 5.01 <= float(final_freightster_weight) <= 15:
                        freightster_price = 7

                    elif 15.01 <= float(final_freightster_weight) <= 20:
                        freightster_price = 7.5

                    else:
                        freightster_price = 1000

                elif freightster_pricing_zone == 'PER':
                    if float(final_freightster_weight) <= 0.250:
                        freightster_price = 4.6

                    elif 0.251 <= float(final_freightster_weight) <= 0.5:
                        freightster_price = 5.6

                    elif 0.501 <= float(final_freightster_weight) <= 1.000:
                        freightster_price = 6.45

                    elif 1.01 <= float(final_freightster_weight) <= 3.000:
                        freightster_price = 8.15

                    elif 3.01 <= float(final_freightster_weight) <= 5.000:
                        freightster_price = 8.35

                    elif 5.01 <= float(final_freightster_weight) <= 15:
                        freightster_price = 22

                    elif 15.01 <= float(final_freightster_weight) <= 20:
                        freightster_price = 30.5

                    else:
                        freightster_price = 1000

                elif freightster_pricing_zone == 'QLD':
                    if float(final_freightster_weight) <= 0.250:
                        freightster_price = 4.6

                    elif 0.251 <= float(final_freightster_weight) <= 0.5:
                        freightster_price = 5.5

                    elif 0.501 <= float(final_freightster_weight) <= 1.000:
                        freightster_price = 7

                    elif 1.01 <= float(final_freightster_weight) <= 3.000:
                        freightster_price = 8

                    elif 3.01 <= float(final_freightster_weight) <= 5.000:
                        freightster_price = 9

                    elif 5.01 <= float(final_freightster_weight) <= 15:
                        freightster_price = 22.5

                    elif 15.01 <= float(final_freightster_weight) <= 20:
                        freightster_price = 28.5

                    else:
                        freightster_price = 1000

                elif freightster_pricing_zone == 'SA':
                    if float(final_freightster_weight) <= 0.250:
                        freightster_price = 4.6

                    elif 0.251 <= float(final_freightster_weight) <= 0.5:
                        freightster_price = 5.5

                    elif 0.501 <= float(final_freightster_weight) <= 1.000:
                        freightster_price = 5

                    elif 1.01 <= float(final_freightster_weight) <= 3.000:
                        freightster_price = 8

                    elif 3.01 <= float(final_freightster_weight) <= 5.000:
                        freightster_price = 9

                    elif 5.01 <= float(final_freightster_weight) <= 15:
                        freightster_price = 22.5

                    elif 15.01 <= float(final_freightster_weight) <= 20:
                        freightster_price = 28.5

                    else:
                        freightster_price = 1000

                elif freightster_pricing_zone == 'SYD':
                    if float(final_freightster_weight) <= 0.250:
                        freightster_price = 4.6

                    elif 0.251 <= float(final_freightster_weight) <= 0.5:
                        freightster_price = 4.6

                    elif 0.501 <= float(final_freightster_weight) <= 1.000:
                        freightster_price = 5

                    elif 1.01 <= float(final_freightster_weight) <= 3.000:
                        freightster_price = 5.95

                    elif 3.01 <= float(final_freightster_weight) <= 5.000:
                        freightster_price = 7.15

                    elif 5.01 <= float(final_freightster_weight) <= 15:
                        freightster_price = 17

                    elif 15.01 <= float(final_freightster_weight) <= 20:
                        freightster_price = 18

                    else:
                        freightster_price = 1000

                elif freightster_pricing_zone == 'VIC':
                    if float(final_freightster_weight) <= 0.250:
                        freightster_price = 4.6

                    elif 0.251 <= float(final_freightster_weight) <= 0.5:
                        freightster_price = 4.75

                    elif 0.501 <= float(final_freightster_weight) <= 1.000:
                        freightster_price = 4.95

                    elif 1.01 <= float(final_freightster_weight) <= 3.000:
                        freightster_price = 5.85

                    elif 3.01 <= float(final_freightster_weight) <= 5.000:
                        freightster_price = 8.5

                    elif 5.01 <= float(final_freightster_weight) <= 15:
                        freightster_price = 18.5

                    elif 15.01 <= float(final_freightster_weight) <= 20:
                        freightster_price = 20.9

                    else:
                        freightster_price = 1000

                elif freightster_pricing_zone == 'WA':
                    if float(final_freightster_weight) <= 0.250:
                        freightster_price = 4.6

                    elif 0.251 <= float(final_freightster_weight) <= 0.5:
                        freightster_price = 5.5

                    elif 0.501 <= float(final_freightster_weight) <= 1.000:
                        freightster_price = 7.12

                    elif 1.01 <= float(final_freightster_weight) <= 3.000:
                        freightster_price = 8.26

                    elif 3.01 <= float(final_freightster_weight) <= 5.000:
                        freightster_price = 10.26

                    elif 5.01 <= float(final_freightster_weight) <= 15:
                        freightster_price = 25.5

                    elif 15.01 <= float(final_freightster_weight) <= 20:
                        freightster_price = 31.5

                    else:
                        freightster_price = 1000

                else:
                    freightster_price = 1000

                payload = {"order": {"serviceCode": 12,
                                     "consignee": {"company": company,
                                                   "name": name,
                                                   "address1": address1,
                                                   "address2": address2,
                                                   "city": city,
                                                   "postcode": postcode,
                                                   "state": state,
                                                   "phone": phone,
                                                   "email": email},
                                     "sender": {"name": "Kyal Scarlett",
                                                "address1": "286-288 Ballarat Rd",
                                                "address2": "",
                                                "city": "Footscray",
                                                "postcode": "3011",
                                                "state": "VIC",
                                                "phone": "0382563460",
                                                "email": "kyal@scarlettmusic.com.au"},
                                     "shipment": {"reference": OrderID,
                                                  "description": OrderID,
                                                  "weight": str(final_freightster_weight)}}}

                r = requests.post('https://freightster.com.au/api/v1/shippingAPI/create', json=payload,
                                  headers=freightster_headers)
                freightster_response = json.loads(r.text)
                freightster_price = freightster_price * 1.1
                if freightster_response['status'] is False:
                    freightster_price = 1000

                if 'NEX' in freightster_response['response_data']['tracking_number']:
                    freightster_price = 1000

                #####Adding GST
                courierprice = round(freightster_price, 2)

                print('Freightster Price: $' + str(courierprice))


            except:
                courierprice = 1000
    if courier == 'Transdirect':
        transdirectfailed = 'true'
        if float(weightvalue) > 25:
            tailgate = 'true'
        else:
            tailgate = 'false'

        if answer == '9':
            transdirectfailed = 'true'

        if transdirectfailed == 'true':
            courierprice = 1000
            pass

        else:

            transdirect_item_array = []

            for all_items in multi_items_to_ship:
                transdirect_item_array.append({
                    "description": "carton",
                    "length": all_items[1],
                    "height": all_items[2],
                    "width": all_items[3],
                    "weight": all_items[0],
                    "quantity": int(all_items[6])
                })

            if email == '':
                email = 'info@scarlettmusic.com.au'

            if phone is None:
                phone = '0417557472'

            try:
                int(phone)

            except ValueError:
                phone = '0417557472'

            phone = phone.replace(' ', '')

            headers = {'Api-key': '74a242a1f4b8ba6ea5e011a80bdd2732', 'Content-Type': 'application/json'}

            try:

                if 'ebay' in address1:
                    payload = {
                        "declared_value": "0",
                        "referrer": "API",
                        "requesting_site": "www.scarlettmusic.com.au",
                        "tailgate_pickup": "false",
                        "tailgate_delivery": str(tailgate),
                        "items": transdirect_item_array,
                        "sender": {
                            "address": "288 Ballarat Rd",
                            "company_name": "Scarlett Music",
                            "email": "info@scarlettmusic.com.au",
                            "name": "Lorelle Scarlett",
                            "postcode": "3011",
                            "phone": "0417557472",
                            "state": "VIC",
                            "suburb": "FOOTSCRAY",
                            "type": "business",
                            "country": "AU"
                        },
                        "receiver": {
                            "address": str(address2),
                            "company_name": str(company),
                            "email": str(email),
                            "name": str(name),
                            "postcode": str(postcode),
                            "phone": int(phone),
                            "state": str(state),
                            "suburb": str(city),
                            "type": address_type,
                            "country": "AU"
                        }
                    }

                else:

                    payload = {
                        "declared_value": "0",
                        "referrer": "API",
                        "requesting_site": "www.scarlettmusic.com.au",
                        "tailgate_pickup": "false",
                        "tailgate_delivery": str(tailgate),
                        "items": transdirect_item_array,
                        "sender": {
                            "address": "288 Ballarat Rd",
                            "company_name": "Scarlett Music",
                            "email": "info@scarlettmusic.com.au",
                            "name": "Lorelle Scarlett",
                            "postcode": "3011",
                            "phone": "0417557472",
                            "state": "VIC",
                            "suburb": "FOOTSCRAY",
                            "type": "business",
                            "country": "AU"
                        },
                        "receiver": {
                            "address": str(address1) + ', ' + str(address2),
                            "company_name": str(company),
                            "email": str(email),
                            "name": str(name),
                            "postcode": str(postcode),
                            "phone": int(phone),
                            "state": str(state),
                            "suburb": str(city),
                            "type": address_type,
                            "country": "AU"
                        }
                    }
                try:
                    r = requests.post('https://www.transdirect.com.au/api/bookings/v4',
                                      auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'),
                                      headers=headers,
                                      json=payload)

                    ##Mock API https://private-anon-2b516758f2-transdirectapiv4.apiary-mock.com/api/bookings/v4
                    ###Legit API https://www.transdirect.com.au/api/bookings/v4

                    trans_response = r.text
                    trans_response = json.loads(trans_response)
                    # pprint.pprint(trans_response)
                    if 'errors' in trans_response:
                        try:

                            if 'ebay' in address1:
                                address = address2 + " " + city + " " + state + " " + postcode

                                # get_google_address()

                                dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                                qstr = urllib.parse.urlencode(dict1)
                                URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                                URL = URL + qstr
                                response = urllib.request.urlopen(URL)
                                data = json.load(response)
                                # pprint.pprint(data['candidates'])
                                placeid = data['candidates'][0]['place_id']
                                payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                                r = requests.get(
                                    'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                                #  pprint.pprint(r.json())
                                data = r.json()
                                #   pprint.pprint(data['result']['address_components'])
                                for t in range(20):
                                    if data['result']['address_components'][t]['types'][0] == 'locality':
                                        suburb = data['result']['address_components'][t]['long_name']
                                        break

                                    else:
                                        continue

                                for t in range(20):
                                    if data['result']['address_components'][t]['types'][0] == 'postal_code':
                                        postal_code = data['result']['address_components'][t]['long_name']
                                        break

                                    else:
                                        continue

                                for t in range(20):
                                    if data['result']['address_components'][t]['types'][
                                        0] == 'administrative_area_level_1':
                                        google_state = data['result']['address_components'][t]['short_name']
                                        break

                                    else:
                                        continue

                                for t in range(20):
                                    if data['result']['address_components'][t]['types'][0] == 'subpremise':
                                        google_street0 = data['result']['address_components'][t]['short_name']
                                        break
                                    else:
                                        google_street0 = ''

                                for t in range(20):
                                    if data['result']['address_components'][t]['types'][0] == 'street_number':
                                        google_street1 = data['result']['address_components'][t]['short_name']
                                        break

                                for t in range(20):
                                    if data['result']['address_components'][t]['types'][0] == 'route':
                                        google_street2 = data['result']['address_components'][t]['long_name']
                                        break

                                if google_street0 == '':

                                    backup_google_street_address = google_street1 + ' ' + google_street2

                                else:

                                    backup_google_street_address = google_street0 + '/' + google_street1 + ' ' + google_street2

                                postcode = postal_code
                                city = suburb
                                state = google_state

                            else:
                                address = address1 + " " + address2 + " " + city + " " + state + " " + postcode

                                # get_google_address()

                                dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                                qstr = urllib.parse.urlencode(dict1)
                                URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                                URL = URL + qstr
                                response = urllib.request.urlopen(URL)
                                data = json.load(response)
                                # pprint.pprint(data['candidates'])
                                placeid = data['candidates'][0]['place_id']
                                payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                                r = requests.get(
                                    'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                                #    pprint.pprint(r.json())
                                data = r.json()
                                #  pprint.pprint(data['result']['address_components'])
                                for t in range(20):
                                    if data['result']['address_components'][t]['types'][0] == 'locality':
                                        suburb = data['result']['address_components'][t]['long_name']
                                        break

                                    else:
                                        continue

                                for t in range(20):
                                    if data['result']['address_components'][t]['types'][0] == 'postal_code':
                                        postal_code = data['result']['address_components'][t]['long_name']
                                        break

                                    else:
                                        continue

                                for t in range(20):
                                    if data['result']['address_components'][t]['types'][
                                        0] == 'administrative_area_level_1':
                                        google_state = data['result']['address_components'][t]['short_name']
                                        break

                                    else:
                                        continue

                                for t in range(20):
                                    if data['result']['address_components'][t]['types'][0] == 'subpremise':
                                        google_street0 = data['result']['address_components'][t]['short_name']
                                        break
                                    else:
                                        google_street0 = ''

                                for t in range(20):
                                    if data['result']['address_components'][t]['types'][0] == 'street_number':
                                        google_street1 = data['result']['address_components'][t]['short_name']
                                        break

                                for t in range(20):
                                    if data['result']['address_components'][t]['types'][0] == 'route':
                                        google_street2 = data['result']['address_components'][t]['long_name']
                                        break

                                if google_street0 == '':

                                    backup_google_street_address = google_street1 + ' ' + google_street2

                                else:

                                    backup_google_street_address = google_street0 + ' ' + google_street1 + ' ' + google_street2

                                postcode = postal_code
                                city = suburb
                                state = google_state

                            payload = {
                                "declared_value": "0",
                                "referrer": "API",
                                "requesting_site": "www.scarlettmusic.com.au",
                                "tailgate_pickup": "false",
                                "tailgate_delivery": str(tailgate),
                                "items": transdirect_item_array,
                                "sender": {
                                    "address": "288 Ballarat Rd",
                                    "company_name": "Scarlett Music",
                                    "email": "info@scarlettmusic.com.au",
                                    "name": "Lorelle Scarlett",
                                    "postcode": "3011",
                                    "phone": "0417557472",
                                    "state": "VIC",
                                    "suburb": "FOOTSCRAY",
                                    "type": "business",
                                    "country": "AU"
                                },
                                "receiver": {
                                    "address": str(backup_google_street_address),
                                    "company_name": str(company),
                                    "email": str(email),
                                    "name": str(name),
                                    "postcode": str(postcode),
                                    "phone": int(phone),
                                    "state": str(state),
                                    "suburb": str(city),
                                    "type": address_type,
                                    "country": "AU"
                                }
                            }

                            try:
                                r = requests.post('https://www.transdirect.com.au/api/bookings/v4',
                                                  auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'),
                                                  headers=headers,
                                                  json=payload)

                                ##Mock API https://private-anon-2b516758f2-transdirectapiv4.apiary-mock.com/api/bookings/v4
                                ###Legit API https://www.transdirect.com.au/api/bookings/v4

                                trans_response = r.text
                                trans_response = json.loads(trans_response)

                                if 'id' in trans_response:
                                    google = 'true'
                            # pprint.pprint(trans_response)

                            except json.decoder.JSONDecodeError:
                                courierprice = 1000
                        except:
                            courierprice = 1000


                except json.decoder.JSONDecodeError:

                    courierprice = 1000

                id = trans_response['id']

                quotelen = len(trans_response['quotes'])
                # print('Amount of quotes: ' + str(quotelen))
                namesofquotes = list(trans_response['quotes'])
                # print(namesofquotes)

                quotes = {}

                for x in range(quotelen):
                    quotes[str(namesofquotes[x])] = str(trans_response['quotes'][namesofquotes[x]]['total'])

                # print(quotes)

                if 'tiers' in trans_response['quotes']['couriers_please_domestic_proirity_authority']:
                    quotes['couriers_please_multi_21'] = \
                        trans_response['quotes']['couriers_please_domestic_proirity_authority']['tiers'][1]['total']

                    if answer == '2':
                        quotes['couriers_please_multi_21'] = '1000'

                else:
                    quotes['couriers_please_multi_21'] = '1000'

                if 'fastway' in trans_response['quotes']:

                    if 'tiers' in trans_response['quotes']['fastway']:

                        quotes['fastway_multi_7'] = trans_response['quotes']['fastway']['tiers'][1]['total']

                        if answer == '2':
                            quotes['fastway_multi_7'] = '1000'

                    else:
                        quotes['fastway_multi_7'] = '1000'

               # print(quotes)

                try:
                    quotes['hunter_road_freight'] = '1000'
                except:
                    pass
                try:
                    quotes['parca'] = '1000'
                except:
                    pass

                if float(totalprice) > 200:
                    quotes['couriers_please_domestic_proirity_authority'] = '1000'
                    quotes['fastway_multi_7'] = '1000'
                    quotes['couriers_please_multi_21'] = '1000'
                    quotes['fastway'] = '1000'

                intquotes = dict(
                    (k, float(v)) for k, v in
                    quotes.items())  ### Converting all values into FLOAT and then into INT
                lowesttranscourier = min(intquotes, key=intquotes.get)

                courierprice = quotes[lowesttranscourier]

                print('Transdirect Price: $' + str(courierprice))



            except (KeyError, TypeError, IndexError):

                try:

                    if 'ebay' in address1:
                        address = address2 + " " + city + " " + state + " " + postcode

                        # get_google_address()

                        dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                        qstr = urllib.parse.urlencode(dict1)
                        URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                        URL = URL + qstr
                        response = urllib.request.urlopen(URL)
                        data = json.load(response)
                        # pprint.pprint(data['candidates'])
                        placeid = data['candidates'][0]['place_id']
                        payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                        r = requests.get(
                            'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                        #  pprint.pprint(r.json())
                        data = r.json()
                        #   pprint.pprint(data['result']['address_components'])
                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'locality':
                                suburb = data['result']['address_components'][t]['long_name']
                                break

                            else:
                                continue

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'postal_code':
                                postal_code = data['result']['address_components'][t]['long_name']
                                break

                            else:
                                continue

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                                google_state = data['result']['address_components'][t]['short_name']
                                break

                            else:
                                continue

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'subpremise':
                                google_street0 = data['result']['address_components'][t]['short_name']
                                break
                            else:
                                google_street0 = ''

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'street_number':
                                google_street1 = data['result']['address_components'][t]['short_name']
                                break

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'route':
                                google_street2 = data['result']['address_components'][t]['long_name']
                                break

                        if google_street0 == '':

                            backup_google_street_address = google_street1 + ' ' + google_street2

                        else:

                            backup_google_street_address = google_street0 + '/' + google_street1 + ' ' + google_street2

                        postcode = postal_code
                        city = suburb
                        state = google_state

                    else:
                        address = address1 + " " + address2 + " " + city + " " + state + " " + postcode

                        # get_google_address()

                        dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                        qstr = urllib.parse.urlencode(dict1)
                        URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                        URL = URL + qstr
                        response = urllib.request.urlopen(URL)
                        data = json.load(response)
                        # pprint.pprint(data['candidates'])
                        placeid = data['candidates'][0]['place_id']
                        payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                        r = requests.get(
                            'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                        #    pprint.pprint(r.json())
                        data = r.json()
                        #  pprint.pprint(data['result']['address_components'])
                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'locality':
                                suburb = data['result']['address_components'][t]['long_name']
                                break

                            else:
                                continue

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'postal_code':
                                postal_code = data['result']['address_components'][t]['long_name']
                                break

                            else:
                                continue

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'administrative_area_level_1':
                                google_state = data['result']['address_components'][t]['short_name']
                                break

                            else:
                                continue

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'subpremise':
                                google_street0 = data['result']['address_components'][t]['short_name']
                                break
                            else:
                                google_street0 = ''

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'street_number':
                                google_street1 = data['result']['address_components'][t]['short_name']
                                break

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'route':
                                google_street2 = data['result']['address_components'][t]['long_name']
                                break

                        if google_street0 == '':

                            backup_google_street_address = google_street1 + ' ' + google_street2

                        else:

                            backup_google_street_address = google_street0 + ' ' + google_street1 + ' ' + google_street2

                        postcode = postal_code
                        city = suburb
                        state = google_state

                    payload = {
                        "declared_value": "0",
                        "referrer": "API",
                        "requesting_site": "www.scarlettmusic.com.au",
                        "tailgate_pickup": "false",
                        "tailgate_delivery": str(tailgate),
                        "items": transdirect_item_array,
                        "sender": {
                            "address": "288 Ballarat Rd",
                            "company_name": "Scarlett Music",
                            "email": "info@scarlettmusic.com.au",
                            "name": "Lorelle Scarlett",
                            "postcode": "3011",
                            "phone": "0417557472",
                            "state": "VIC",
                            "suburb": "FOOTSCRAY",
                            "type": "business",
                            "country": "AU"
                        },
                        "receiver": {
                            "address": str(backup_google_street_address),
                            "company_name": str(company),
                            "email": str(email),
                            "name": str(name),
                            "postcode": str(postcode),
                            "phone": int(phone),
                            "state": str(state),
                            "suburb": str(city),
                            "type": address_type,
                            "country": "AU"
                        }
                    }
                    try:
                        r = requests.post('https://www.transdirect.com.au/api/bookings/v4',
                                          auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'),
                                          headers=headers,
                                          json=payload)

                        ##Mock API https://private-anon-2b516758f2-transdirectapiv4.apiary-mock.com/api/bookings/v4
                        ###Legit API https://www.transdirect.com.au/api/bookings/v4

                        trans_response = r.text
                        trans_response = json.loads(trans_response)
                    # pprint.pprint(response)

                    except json.decoder.JSONDecodeError:

                        courierprice = 1000

                    id = trans_response['id']

                    quotelen = len(trans_response['quotes'])
                    # print('Amount of quotes: ' + str(quotelen))
                    namesofquotes = list(trans_response['quotes'])
                    # print(namesofquotes)

                    quotes = {}

                    for x in range(quotelen):
                        quotes[str(namesofquotes[x])] = str(trans_response['quotes'][namesofquotes[x]]['total'])

                    # print(quotes)
                    if 'tiers' in trans_response['quotes']['couriers_please_domestic_proirity_authority']:
                        quotes['couriers_please_multi_21'] = \
                            trans_response['quotes']['couriers_please_domestic_proirity_authority']['tiers'][1][
                                'total']

                        if answer == '2':
                            quotes['couriers_please_multi_21'] = '1000'

                    else:
                        quotes['couriers_please_multi_21'] = '1000'

                    if 'fastway' in trans_response['quotes']:

                        if 'tiers' in trans_response['quotes']['fastway']:

                            quotes['fastway_multi_7'] = trans_response['quotes']['fastway']['tiers'][1]['total']

                            if answer == '2':
                                quotes['fastway_multi_7'] = '1000'

                        else:
                            quotes['fastway_multi_7'] = '1000'
                    try:
                        quotes['hunter_road_freight'] = '1000'
                    except:
                        pass
                    if float(totalprice) > 200:
                        quotes['couriers_please_domestic_proirity_authority'] = '1000'

                    intquotes = dict(
                        (k, float(v)) for k, v in
                        quotes.items())  ### Converting all values into FLOAT and then into INT
                    lowesttranscourier = min(intquotes, key=intquotes.get)

                    courierprice = quotes[lowesttranscourier]

                    print('Transdirect Price: $' + courierprice)


                except (KeyError, TypeError, IndexError):

                    courierprice = 1000

    if courier == 'Fastway':

        if answer == str(10):
            fastwayfailed = 'true'

        signature = "ATL"

        if float(totalprice) > 200:
            signature = "SGR"

        if fastwayfailed == 'true':
            courierprice = 1000

        else:

            satchel = 'false'

            if email == '':
                email = 'info@scarlettmusic.com.au'

            if phone is None:
                phone = '0417557472'

            phone = phone.replace(' ', '')

            ####### INITIAL FASTWAY VALIDATION ######

            if 'ebay' in address1:

                street_address = str(address2)

                data = {
                    "streetAddress": street_address,
                    "additionalDetails": "",
                    "locality": str(city),
                    "stateOrProvince": str(state),
                    "postalCode": str(postcode),
                    "country": "AU",
                    "lat": 0,
                    "lng": 0,
                    "userCreated": False,
                }


            else:

                street_address = str(address1) + ', ' + str(address2)

                data = {
                    "streetAddress": street_address,
                    "additionalDetails": "",
                    "locality": str(city),
                    "stateOrProvince": str(state),
                    "postalCode": str(postcode),
                    "country": "AU",
                    "lat": 0,
                    "lng": 0,
                    "userCreated": False,
                }

            r = requests.post(base_url + '/api/addresses/validate', headers=Fastway_Headers, json=data)
            try:

                response = r.text
                response = json.loads(response)
                # pprint.pprint(response)

                if 'errors' in response:
                    courierprice = 1000


                elif response['data']['stateOrProvince'].lower() != state.lower():
                    courierprice = 1000
            except:
                courierprice = 1000

            else:

                try:
                    country = response['data']['country']
                    addressid = response['data']['addressId']
                    latitude = response['data']['lat']
                    longitude = response['data']['lng']
                    suburb = response['data']['locality']
                    placeId = response['data']['placeId']
                    postcode = response['data']['postalCode']
                    state = response['data']['stateOrProvince']
                    street_address = response['data']['streetAddress']
                    hash = response['data']['hash']

                    # package = [float(length), float(width), float(height)]
                    # sorted_package = sorted(package)

                    fastway_item_array = []

                    for all_items in multi_items_to_ship:
                        # multi_items_to_ship.append((weightvalue, length, height, width,volumevalue, cubicweight, quantity, satchel))

                        if all_items[7] == 'false':
                            fastway_item_array.append({
                                "Quantity": int(all_items[6]),
                                "Reference": srn,
                                "PackageType": "P",
                                "WeightDead": all_items[0],
                                "WeightCubic": all_items[5],
                                "Length": all_items[1],
                                "Width": all_items[2],
                                "Height": all_items[3]
                            })
                        else:
                            fastway_item_array.append({
                                "Quantity": int(all_items[6]),
                                "Reference": srn,
                                "PackageType": "S",
                                "SatchelSize": all_items[7]
                            })

                    fastway_data = {
                        "To": {
                            "ContactName": name,
                            "BusinessName": company,
                            "PhoneNumber": phone,
                            "Email": email,
                            "Address": {
                                "StreetAddress": street_address,
                                "Locality": suburb,
                                "StateOrProvince": state,
                                "PostalCode": postcode,
                                "Country": country
                            }
                        }, "Services": [
                            {
                                "ServiceCode": "DELOPT",
                                "ServiceItemCode": signature
                            }
                        ],
                        "Items": fastway_item_array,
                        "ExternalRef1": srn,
                        "ExternalRef2": srn,
                    }

                    r = requests.post(base_url + '/api/consignments/quote', headers=Fastway_Headers,
                                      json=fastway_data)
                    response = r.text
                    response = json.loads(response)

                    courierprice = response['data']['total']

                    print('Fastway price: $' + str(courierprice))

                except (KeyError, TypeError, IndexError):

                    try:

                        if 'ebay' in address1:
                            address = address2 + " " + city + " " + state + " " + postcode

                            # get_google_address()

                            dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                            qstr = urllib.parse.urlencode(dict1)
                            URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                            URL = URL + qstr
                            response = urllib.request.urlopen(URL)
                            data = json.load(response)
                            # pprint.pprint(data['candidates'])
                            placeid = data['candidates'][0]['place_id']
                            payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                            r = requests.get(
                                'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                            #  pprint.pprint(r.json())
                            data = r.json()
                            #   pprint.pprint(data['result']['address_components'])
                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'locality':
                                    suburb = data['result']['address_components'][t]['long_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'postal_code':
                                    postal_code = data['result']['address_components'][t]['long_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][
                                    0] == 'administrative_area_level_1':
                                    google_state = data['result']['address_components'][t]['short_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'subpremise':
                                    google_street0 = data['result']['address_components'][t]['short_name']
                                    break
                                else:
                                    google_street0 = ''

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'street_number':
                                    google_street1 = data['result']['address_components'][t]['short_name']
                                    break

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'route':
                                    google_street2 = data['result']['address_components'][t]['long_name']
                                    break

                            if google_street0 == '':

                                backup_google_street_address = google_street1 + ' ' + google_street2

                            else:

                                backup_google_street_address = google_street0 + '/' + google_street1 + ' ' + google_street2

                            postcode = postal_code
                            city = suburb
                            state = google_state

                        else:
                            address = address1 + " " + address2 + " " + city + " " + state + " " + postcode

                            # get_google_address()

                            dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                            qstr = urllib.parse.urlencode(dict1)
                            URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                            URL = URL + qstr
                            response = urllib.request.urlopen(URL)
                            data = json.load(response)
                            # pprint.pprint(data['candidates'])
                            placeid = data['candidates'][0]['place_id']
                            payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                            r = requests.get(
                                'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                            #    pprint.pprint(r.json())
                            data = r.json()
                            #  pprint.pprint(data['result']['address_components'])
                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'locality':
                                    suburb = data['result']['address_components'][t]['long_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'postal_code':
                                    postal_code = data['result']['address_components'][t]['long_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][
                                    0] == 'administrative_area_level_1':
                                    google_state = data['result']['address_components'][t]['short_name']
                                    break

                                else:
                                    continue

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'subpremise':
                                    google_street0 = data['result']['address_components'][t]['short_name']
                                    break
                                else:
                                    google_street0 = ''

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'street_number':
                                    google_street1 = data['result']['address_components'][t]['short_name']
                                    break

                            for t in range(20):
                                if data['result']['address_components'][t]['types'][0] == 'route':
                                    google_street2 = data['result']['address_components'][t]['long_name']
                                    break

                            if google_street0 == '':

                                backup_google_street_address = google_street1 + ' ' + google_street2

                            else:

                                backup_google_street_address = google_street0 + '/' + google_street1 + ' ' + google_street2

                            city = suburb
                            state = google_state

                        new_data = {
                            "streetAddress": str(backup_google_street_address),
                            "additionalDetails": "",
                            "locality": str(city),
                            "stateOrProvince": str(state),
                            "postalCode": str(postcode),
                            "country": "AU",
                            "lat": 0,
                            "lng": 0,
                            "userCreated": False,
                        }

                        r2 = requests.post(base_url + '/api/addresses/validate', headers=Fastway_Headers,
                                           json=new_data)
                        response = r2.text
                        response = json.loads(response)
                        # pprint.pprint(response)

                        try:
                            country = response['data']['country']
                            addressid = response['data']['addressId']
                            latitude = response['data']['lat']
                            longitude = response['data']['lng']
                            suburb = response['data']['locality']
                            placeId = response['data']['placeId']
                            Postcode = response['data']['postalCode']
                            State = response['data']['stateOrProvince']
                            AddressLine1 = response['data']['streetAddress']
                            hash = response['data']['hash']

                            fastway_data = {
                                "To": {
                                    "ContactName": name,
                                    "BusinessName": company,
                                    "PhoneNumber": phone,
                                    "Email": email,
                                    "Address": {
                                        "StreetAddress": street_address,
                                        "Locality": suburb,
                                        "StateOrProvince": state,
                                        "PostalCode": postcode,
                                        "Country": country
                                    }
                                }, "Services": [
                                    {
                                        "ServiceCode": "DELOPT",
                                        "ServiceItemCode": signature
                                    }
                                ],
                                "Items": fastway_item_array,
                                "ExternalRef1": srn,
                                "ExternalRef2": srn,
                            }

                            r = requests.post(base_url + '/api/consignments/quote', headers=Fastway_Headers,
                                              json=fastway_data)
                            response = r.text
                            response = json.loads(response)

                            courierprice = response['data']['total']

                            print('Fastway price: $' + str(courierprice))

                        except (KeyError, TypeError, IndexError):

                            courierprice = 1000

                    except:

                        courierprice = 1000

    if courier == 'Allied':

        if allied_failed == 'true':
            courierprice = 1000

        else:

            pickup_address = {'address1': "286-288 Ballarat Rd",
                              'address2': "",
                              'country': "Australia",
                              'postCode': '3011',
                              'state': 'VIC',
                              'suburb': 'Footscray'}

            reciever_address = {'address1': address1,
                                'address2': address2,
                                'country': "Australia",
                                'postCode': postcode,
                                'state': state,
                                'suburb': city}

            Jobstop_pickupstop = {'companyName': 'Scarlett Music',
                                  'contact': 'Kyal Scarlett',
                                  'emailAddress': 'info@scarlettmusic.com.au',
                                  'geographicAddress': pickup_address,
                                  'phoneNumber': '03 9318 5751',
                                  'stopNumber': 1,
                                  'stopType': 'P'}

            Jobstop_deliverystop = {'companyName': company,
                                    'contact': name,
                                    'emailAddress': email,
                                    'geographicAddress': reciever_address,
                                    'phoneNumber': phone,
                                    'stopNumber': 2,
                                    'stopType': 'D'}

            Jobstop_final = [Jobstop_pickupstop, Jobstop_deliverystop]

            allied_item_array = []
            allied_total_volume = 0
            allied_total_weight = 0
            allied_total_items = 0
            #   multi_items_to_ship.append((weightvalue, length, height, width,volumevalue, cubicweight, quantity, satchel))

            for all_items in multi_items_to_ship:
                allied_item_array.append({'dangerous': 'false',
                                          'height': all_items[2],
                                          'itemCount': all_items[6],
                                          'length': all_items[1],
                                          'volume': all_items[4],
                                          'weight': all_items[0],
                                          'width': all_items[3]})

                allied_total_volume = allied_total_volume + all_items[4]
                allied_total_weight = float(allied_total_weight) + float(all_items[0])
                allied_total_items = allied_total_items + all_items[6]

            # item = {'dangerous': 'false',
            #         'height': float(height),
            #         'itemCount': 1,
            #         'length': float(length),
            #         'volume': float(volumevalue),
            #         'weight': float(weightvalue),
            #         'width': float(width)}

            ##IF MULTIPLE ITEMS, easily put into dic, item = [{item1, item2}]

            Service_Level = 'R'  # Overnight Express

            pickup_instructions = "The music shop, open 9am-6pm. Best parking is at The Palms across the road."
            reference_number = srn
            booked_by = 'Kyal Scarlett'

            job_number = int(re.sub("[^0-9]", "", reference_number))

            Job = {'account': allied_account,
                   'cubicWeight': cubicweight,
                   'Docket': "SCM",
                   'instructions': pickup_instructions,
                   'cubedItems': allied_item_array,
                   'itemCount': allied_total_items,
                   'weight': allied_total_weight,
                   'volume': allied_total_volume,
                   'items': allied_item_array,
                   'jobStops': Jobstop_final,
                   'serviceLevel': Service_Level,
                   'referenceNumbers': reference_number,
                   'bookedBy': booked_by,
                   'readyDate': today,
                   'jobNumber': job_number,
                   'vehicle': {'vehicleID': 1}}

            JobIDs = {'jobIds': job_number}

            try:

                Job = allied_client.service.validateBooking('755cf13abb3934695f03bd4a75cfbca7', Job)

                job_price = allied_client.service.calculatePrice('755cf13abb3934695f03bd4a75cfbca7', Job)
                courierprice = job_price['totalCharge']
                courierprice = round((float(courierprice) * 1.269) * 1.1, 2)
                if courierprice == 0.0:
                    courierprice = 1000
                print('Allied price: $' + str(courierprice))

            except zeep.exceptions.Fault:
                courierprice = 1000

    if courier == 'Couriers Please':

        couriers_failed = 'true'

        if couriers_failed == 'true':
            courierprice = 1000

        else:

            cp_item_array = []

            for all_items in multi_items_to_ship:
                cp_item_array.append({
                    "length": math.ceil(int(all_items[1])),
                    "height": math.ceil(int(all_items[2])),
                    "width": math.ceil(int(all_items[3])),
                    "physicalWeight": float(all_items[0]),
                    "quantity": int(all_items[6])
                })

            cp_body = {
                "fromSuburb": "Footscray",
                "fromPostcode": 3011,
                "toSuburb": city,
                "toPostcode": postcode,
                "items": cp_item_array
            }

            # pprint.pprint(cp_body)

            additional_costs = 0

            for all_items in multi_items_to_ship:

                # for _ in range(int(all_items[6])):

                package = [math.ceil(float(all_items[3])), math.ceil(float(all_items[1])),
                           math.ceil(float(all_items[2]))]
                sorted_package = sorted(package)

                for _ in range(int(all_items[6])):

                    if sorted_package[0] > 120 or sorted_package[1] > 120 or sorted_package[
                        2] > 120:
                        additional_costs += 10

            cp_url = 'https://api.couriersplease.com.au/v2/domestic/quote'

            try:
                r = requests.post(cp_url, headers=cp_headers, json=cp_body)
                response = r.text
                response = json.loads(response)
                # pprint.pprint(response)

                quotelen = len(response['data'])
                # print('Amount of quotes: ' + str(quotelen))
                # namesofquotes = list(response['quotes'])
                # print(namesofquotes)

                quotes = {}

                for x in range(quotelen):
                    quotes[str(response['data'][x]['RateCardCode'])] = str((float(
                        response['data'][x]['CalculatedFreightCharge']) + float(
                        response['data'][x]['CalculatedFuelCharge'])) * 1.1)

                # pprint.pprint(quotes)

                intquotes = dict(
                    (k, float(v)) for k, v in quotes.items())  ### Converting all values into FLOAT and then into INT
                lowest_cp_service = min(intquotes, key=intquotes.get)

                courierprice = float(quotes[lowest_cp_service]) * 1.025

                courierprice = courierprice + additional_costs

                cp_url = 'https://api.couriersplease.com.au/v1/domestic/shipment/validate'
                # print(lowest_cp_service)

                cp_validate_body = {
                    "pickupFirstName": "Kyal",
                    "pickupLastName": "Scarlett",
                    "pickupCompanyName": "Scarlett Music",
                    "pickupEmail": "kyal@scarlettmusic.com.au",
                    "pickupAddress1": "286-288 Ballarat Rd",
                    "pickupAddress2": "",
                    "pickupSuburb": "Footscray",
                    "pickupState": "VIC",
                    "pickupPostcode": "3011",
                    "pickupPhone": "0393185751",
                    "pickupIsBusiness": "true",
                    "destinationFirstName": first_name,
                    "destinationLastName": last_name,
                    "destinationCompanyName": company,
                    "destinationEmail": email,
                    "destinationAddress1": address1,
                    "destinationAddress2": address2,
                    "destinationSuburb": city,
                    "destinationState": state,
                    "destinationPostcode": postcode,
                    "destinationPhone": phone,
                    "destinationIsBusiness": "false",
                    "contactFirstName": "Kyal",
                    "contactLastName": "Scarlett",
                    "contactCompanyName": "Scarlett Music",
                    "contactEmail": "kyal@scarlettmusic.com.au",
                    "contactAddress1": "286-288 Ballarat Rd",
                    "contactAddress2": "",
                    "contactSuburb": "Footscray",
                    "contactState": "VIC",
                    "contactPostcode": "3011",
                    "contactPhone": "0393185751",
                    "contactIsBusiness": "true",
                    "referenceNumber": srn,
                    "termsAccepted": "true",
                    "dangerousGoods": "false",
                    "rateCardId": lowest_cp_service,
                    "specialInstruction": "",
                    "isATL": "false",
                    "readyDateTime": next_cp_day,
                    "items": cp_item_array}

                r = requests.post(cp_url, headers=cp_headers, json=cp_validate_body, timeout=10)
                response = r.text
                response = json.loads(response)
                # pprint.pprint(response)

                if response['responseCode'] != 'SUCCESS':
                    ##create shipment / print label
                    cp_price = 1000

                if 'wa' in state.lower():
                    courierprice = courierprice*1.20 #Remove once disaster levy has been lifted

                print(f"Courier's Please: {courierprice}")
            except:

                try:

                    if 'ebay' in address1:
                        address = address2 + " " + city + " " + state + " " + postcode

                        # get_google_address()

                        dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                        qstr = urllib.parse.urlencode(dict1)
                        URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                        URL = URL + qstr
                        response = urllib.request.urlopen(URL)
                        data = json.load(response)
                        # pprint.pprint(data['candidates'])
                        placeid = data['candidates'][0]['place_id']
                        payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                        r = requests.get(
                            'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                        #  pprint.pprint(r.json())
                        data = r.json()
                        #   pprint.pprint(data['result']['address_components'])
                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'locality':
                                suburb = data['result']['address_components'][t]['long_name']
                                break

                            else:
                                continue

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'postal_code':
                                postal_code = data['result']['address_components'][t]['long_name']
                                break

                            else:
                                continue

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][
                                0] == 'administrative_area_level_1':
                                google_state = data['result']['address_components'][t]['short_name']
                                break

                            else:
                                continue

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'subpremise':
                                google_street0 = data['result']['address_components'][t]['short_name']
                                break
                            else:
                                google_street0 = ''

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'street_number':
                                google_street1 = data['result']['address_components'][t]['short_name']
                                break

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'route':
                                google_street2 = data['result']['address_components'][t]['long_name']
                                break

                        if google_street0 == '':

                            backup_google_street_address = google_street1 + ' ' + google_street2

                        else:

                            backup_google_street_address = google_street0 + '/' + google_street1 + ' ' + google_street2

                        postcode = postal_code
                        city = suburb
                        state = google_state

                    else:
                        address = address1 + " " + address2 + " " + city + " " + state + " " + postcode

                        # get_google_address()

                        dict1 = {'input': address, 'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                        qstr = urllib.parse.urlencode(dict1)
                        URL = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?inputtype=textquery&fields=formatted_address,place_id,name,geometry&'
                        URL = URL + qstr
                        response = urllib.request.urlopen(URL)
                        data = json.load(response)
                        # pprint.pprint(data['candidates'])
                        placeid = data['candidates'][0]['place_id']
                        payload = {'key': 'AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc'}
                        r = requests.get(
                            'https://maps.googleapis.com/maps/api/place/details/json?place_id=' + placeid + '&fields=address_component&key=AIzaSyBQRdE8bz9WK0mCQir7oB41ADTbK9DVoMc')
                        #    pprint.pprint(r.json())
                        data = r.json()
                        #  pprint.pprint(data['result']['address_components'])
                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'locality':
                                suburb = data['result']['address_components'][t]['long_name']
                                break

                            else:
                                continue

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'postal_code':
                                postal_code = data['result']['address_components'][t]['long_name']
                                break

                            else:
                                continue

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][
                                0] == 'administrative_area_level_1':
                                google_state = data['result']['address_components'][t]['short_name']
                                break

                            else:
                                continue

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'subpremise':
                                google_street0 = data['result']['address_components'][t]['short_name']
                                break
                            else:
                                google_street0 = ''

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'street_number':
                                google_street1 = data['result']['address_components'][t]['short_name']
                                break

                        for t in range(20):
                            if data['result']['address_components'][t]['types'][0] == 'route':
                                google_street2 = data['result']['address_components'][t]['long_name']
                                break

                        if google_street0 == '':

                            backup_google_street_address = google_street1 + ' ' + google_street2

                        else:

                            backup_google_street_address = google_street0 + '/' + google_street1 + ' ' + google_street2

                        city = suburb
                        state = google_state

                    cp_body = {
                        "fromSuburb": "Footscray",
                        "fromPostcode": 3011,
                        "toSuburb": city,
                        "toPostcode": postcode,
                        "items": cp_item_array
                    }

                    lowest_cp_service = ''

                    cp_validate_body = {
                        "pickupFirstName": "Kyal",
                        "pickupLastName": "Scarlett",
                        "pickupCompanyName": "Scarlett Music",
                        "pickupEmail": "kyal@scarlettmusic.com.au",
                        "pickupAddress1": "286-288 Ballarat Rd",
                        "pickupAddress2": "",
                        "pickupSuburb": "Footscray",
                        "pickupState": "VIC",
                        "pickupPostcode": "3011",
                        "pickupPhone": "0393185751",
                        "pickupIsBusiness": "true",
                        "destinationFirstName": first_name,
                        "destinationLastName": last_name,
                        "destinationCompanyName": company,
                        "destinationEmail": email,
                        "destinationAddress1": address1,
                        "destinationAddress2": address2,
                        "destinationSuburb": city,
                        "destinationState": state,
                        "destinationPostcode": postcode,
                        "destinationPhone": phone,
                        "destinationIsBusiness": "false",
                        "contactFirstName": "Kyal",
                        "contactLastName": "Scarlett",
                        "contactCompanyName": "Scarlett Music",
                        "contactEmail": "kyal@scarlettmusic.com.au",
                        "contactAddress1": "286-288 Ballarat Rd",
                        "contactAddress2": "",
                        "contactSuburb": "Footscray",
                        "contactState": "VIC",
                        "contactPostcode": "3011",
                        "contactPhone": "0393185751",
                        "contactIsBusiness": "true",
                        "referenceNumber": srn,
                        "termsAccepted": "true",
                        "dangerousGoods": "false",
                        "rateCardId": lowest_cp_service,
                        "specialInstruction": "",
                        "isATL": "false",
                        "readyDateTime": next_cp_day,
                        "items": cp_item_array}

                    cp_url = 'https://api.couriersplease.com.au/v2/domestic/quote'

                    r = requests.post(cp_url, headers=cp_headers, json=cp_body)

                    response = r.text
                    response = json.loads(response)
                    # pprint.pprint(response)

                    quotelen = len(response['data'])
                    # print('Amount of quotes: ' + str(quotelen))
                    # namesofquotes = list(response['quotes'])
                    # print(namesofquotes)

                    quotes = {}

                    for x in range(quotelen):
                        quotes[str(response['data'][x]['RateCardCode'])] = str((float(
                            response['data'][x]['CalculatedFreightCharge']) + float(
                            response['data'][x]['CalculatedFuelCharge'])) * 1.1)

                    # pprint.pprint(quotes)

                    intquotes = dict(
                        (k, float(v)) for k, v in
                        quotes.items())  ### Converting all values into FLOAT and then into INT
                    lowest_cp_service = min(intquotes, key=intquotes.get)

                    courierprice = float(quotes[lowest_cp_service]) * 1.025

                    cp_url = 'https://api.couriersplease.com.au/v1/domestic/shipment/validate'
                    # print(lowest_cp_service)

                    r = requests.post(cp_url, headers=cp_headers, json=cp_validate_body, timeout=10)
                    response = r.text
                    response = json.loads(response)
                    # pprint.pprint(response)

                    if response['responseCode'] != 'SUCCESS':
                        ##create shipment / print label
                        cp_price = 1000

                    print(f"Courier's Please: {courierprice}")
                except:
                    courierprice = 1000

    return courierprice


today = datetime.date.today()
if today.isoweekday() in set((6, 7)):
    today += datetime.timedelta(days=today.isoweekday() % 5)
next_day_allied = str(today.day) + '/' + str(today.month) + '/' + str(today.year) + " 10:00:00"

history = HistoryPlugin()
session = Session()
transport = Transport(session=session)

#######NEED TO UNCOMMENT BELOW IN PROD

wsdl = 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS?wsdl'
# wsdl = 'http://triton.alliedexpress.com.au:8080/ttws-ejb/TTWS'
allied_failed = ''

try:
    allied_client = zeep.Client(wsdl=wsdl, transport=transport, plugins=[history])

    allied_client.transport.session.proxies = {
        # Utilize for all http/https connections
        'http': 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS', }
    allied_account = allied_client.service.getAccountDefaults('755cf13abb3934695f03bd4a75cfbca7', "SCAMUS", "VIC",
                                                              "AOE")

except:
    allied_failed = 'true'

################################^^^^^^^^^ALLIED ACCOUNT^^^^^^^##########################################

client = BackendApplicationClient(client_id='fw-fl2-MEL0900146-d2f1bfc5a108')
oauth = OAuth2Session(client=client)

fastwayfailed = ''

try:

    token = oauth.fetch_token(token_url='https://identity.fastway.org/connect/token',
                              client_id='fw-fl2-MEL0900146-d2f1bfc5a108',
                              client_secret='4c6483bb-5994-4773-b263-e0dfb7b29edf', scope='fw-fl2-api-au')

    bearer_token = token['access_token']
    bearer_token = 'bearer ' + bearer_token  # Authorization token

    base_url = 'https://api.myfastway.com.au'

    Fastway_Headers = {"Authorization": bearer_token}  # Authorization Header

except:

    fastwayfailed = 'true'

#######^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^  Fastway Authorization ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^###############

GetItemheaders = {'Content-Type': 'application/json', 'Accept': 'application/json', 'NETOAPI_ACTION': 'GetItem',
                  'NETOAPI_USERNAME': 'kyalAPI', 'NETOAPI_KEY': 'HSZ3EyI9ki6sseqRIkwy6tL6yFNPEUyM', }

#######^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^  Neto Get Item Info ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^###############

UpdateItemheaders = {'Content-Type': 'application/json', 'Accept': 'application/json', 'NETOAPI_ACTION': 'UpdateItem',
                     'NETOAPI_USERNAME': 'kyalAPI', 'NETOAPI_KEY': 'HSZ3EyI9ki6sseqRIkwy6tL6yFNPEUyM', }

#######^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^  Neto Update Item Info ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^###############

freightster_headers = {'Content-Type': 'application/json', 'Accept': 'application/json',
                       'Authorization': 'S1OL4341HUTIDPKRS0LMOZ9QHANPC1L5'}

#######^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^  Freighster Headers ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^###############

wb = openpyxl.load_workbook(r"\\SERVER\Python\Freightster.xlsx")
sheet = wb['Freightster']


def sendToPrinter():

    if filename2 == '':
        while True:

            try:
                print_data = brother_ql.brother_ql_create.convert(printer, [filename], '62', dither=True)
                send(print_data, PRINTER_IDENTIFIER)
                break

            except Exception as e:
                print(e)
                input('Printing Failed. Press Enter to try again')
    else:

        while True:

             try:
                print_data = brother_ql.brother_ql_create.convert(printer, [filename, filename2], '62', dither=True)
                send(print_data, PRINTER_IDENTIFIER)
                break

             except:
                 input('Printing Failed. Press Enter to try again')


ONE_DAY = datetime.timedelta(days=1)  ###### Getting next business day for pickups
HOLIDAYS_AU = holidays.AU(prov='VIC')


def next_business_day():
    next_day = datetime.date.today() + ONE_DAY
    while next_day.weekday() in holidays.WEEKEND or next_day in HOLIDAYS_AU:
        next_day += ONE_DAY
    return next_day


CSIDL_PERSONAL = 5  # My Documents
SHGFP_TYPE_CURRENT = 0  # Get current, not default value
buf = ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_PERSONAL, None, SHGFP_TYPE_CURRENT, buf)
documents_folder = buf.value

next_day = next_business_day()
next_cp_day = next_day.strftime('%Y-%m-%d')
next_day = next_day.strftime('%d/%m/%Y')

next_cp_day = f"{next_cp_day} 04:00 PM"

cp_auth = '113153878:CBBC1CAD2F335C9CEEA1D0BC63C7056ADBC9C36DD54A9A4530A4380D9AAC5FE0'
### This is for sandbox, when ready, replace 2nd half with CBBC1CAD2F335C9CEEA1D0BC63C7056ADBC9C36DD54A9A4530A4380D9AAC5FE0

cp_auth_encoded = cp_auth.encode()
cp_auth_encoded = base64.b64encode(cp_auth_encoded)
cp_auth_encoded = cp_auth_encoded.decode()

cp_headers = {'Host': 'api.couriersplease.com.au',
              'Accept': 'application/json',
              'Content-Type': 'application/json',
              'Authorization': f'Basic {cp_auth_encoded}',
              'Content-Length': '462'}

google = ''

accountnumber = '0007805312'
username = '966afe1c-c07b-4902-b1d7-77ad1aac0915'
secret = 'x428429fabeb99f420f1'
documentsaddress = documents_folder
imagelocations = documents_folder + r"\Python"

UpdateItemheaders = {'Content-Type': 'application/json', 'Accept': 'application/json', 'NETOAPI_ACTION': 'UpdateItem',
                     'NETOAPI_USERNAME': 'kyalAPI', 'NETOAPI_KEY': 'HSZ3EyI9ki6sseqRIkwy6tL6yFNPEUyM', }

#######^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^  Neto Update Item Info ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^###############

############################################################


console.print('''Hello valued [cyan]Scarlett Music[/] employee! What would you like to do today?

1) Create & Print Shipping Label (Recommended, but may have to leave item in packing area)

2) Create & Print Shipping Label URGENTLY

3) Create & Print [underline yellow]EXPRESS[/] Shipping Label (BETA)

4) Get Quote (Make sure you have the order ID copied!)

5) Delete Shipment

6) Create and Print Australia Post API Manifest

7) Create Aus Post Label & Save to Documents (Mainly for dropships)

8) Update Item Location

9) Create & Print Shipping Label URGENTLY (No Trandirect)

10) Create & Print Shipping Label URGENTLY (No Aramex)

11) Book in all pending Freightster Orders (DO AFTER FREIGHTSTER LABELS ARE DONE)

12) Find the tracking number of a past Freightster order (Need customer's name)

13) Create & Print Shipping Label URGENTLY (No Toll)

14) Create Return Label (Copy order number)

15) Create & Print Shipping Label URGENTLY (No DAI)

''')

answer = input('Enter response here (Number 1-7):')

try:
    print(f'\n{fun_response}\n')
    time.sleep(1)


except:
    pass

if answer == str(1) or answer == str(2) or answer == str(3) or answer == str(4) or answer == str(9) or answer == str(
        10) or answer == str(13) or answer == str(15):

    mailplusfailed = ''
    auspostfailed = ''
    sendlefailed = ''
    transdirectfailed = ''
    address_type = 'residential'
    couriers_failed = ''
    freightsterfailed = ''
    toll_failed = ''
    dai_failed = ''
    po_box = ''
    bondsfailed = ''

    cbdexpressfailed = ''

    backend = 'pyusb'
    while True:

        try:

            print(brother_ql.backends.helpers.discover('pyusb')[0]['identifier'])

        except (IndexError):

            print('No printer found. Press enter when printer is connected.')
            input()
            continue

        if brother_ql.backends.helpers.discover('pyusb') == []:
            print('No printer found. Press enter when printer is connected.')
            input()
            continue
        else:

            PRINTER_IDENTIFIER = brother_ql.backends.helpers.discover('pyusb')[0]['identifier']

            if brother_ql.backends.helpers.discover('pyusb')[0]['identifier'] == 'usb://0x04f9:0x2042':
                PRINTER_IDENTIFIER = 'usb://0x04F9:0x2042'

                printer = BrotherQLRaster('QL-700')
                printermodel = 'QL-700'

            elif brother_ql.backends.helpers.discover('pyusb')[0]['identifier'] == 'usb://0x04f9:0x2042_Љ':
                PRINTER_IDENTIFIER = 'usb://0x04F9:0x2042'

                printer = BrotherQLRaster('QL-700')
                printermodel = 'QL-700'

            elif brother_ql.backends.helpers.discover('pyusb')[0]['identifier'] == 'usb://0x04f9:0x2028':
                PRINTER_IDENTIFIER = 'usb://0x04F9:0x2028'

                printer = BrotherQLRaster('QL-570')
                printermodel = 'QL-570'

            elif brother_ql.backends.helpers.discover('pyusb')[0]['identifier'] == 'usb://0x04f9:0x2028_Љ':
                PRINTER_IDENTIFIER = 'usb://0x04F9:0x2028'

                printer = BrotherQLRaster('QL-570')
                printermodel = 'QL-570'

            break

    reNeto = re.compile(r'N\d+')
    reEbay = re.compile(r'\d+-\d\d\d\d\d-\d\d\d\d\d')
    OrderID = str(pyperclip.paste())
    OrderID = OrderID.strip()

    if reNeto.match(OrderID):

        data = {'Filter': {'OrderID': OrderID,
                           'OutputSelector': ["ID", "ShippingOption", "Email", "GrandTotal",
                                              'OrderLine.ProductName',
                                              'OrderLine.Quantity', ' OrderLine.Weight', 'OrderLine.Cubic',
                                              'OrderLine.ExtraOptions', 'StickyNotes', 'ShipAddress',
                                              'OrderLine.WarehouseName', 'SalesChannel']}}

        headers = {'Content-Type': 'application/json', 'Accept': 'application/json', 'NETOAPI_ACTION': 'GetOrder',
                   'NETOAPI_USERNAME': 'kyalAPI', 'NETOAPI_KEY': 'HSZ3EyI9ki6sseqRIkwy6tL6yFNPEUyM', }

        r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=headers, json=data)

        response = r.text
        response = json.loads(response)
        #  pprint.pprint(response)

        orderlinelen = len(response['Order'][0]['OrderLine'])
        quantity = response['Order'][0]['OrderLine'][0]['Quantity']
        totalprice = response['Order'][0]['GrandTotal']
        SKU = response['Order'][0]['OrderLine'][0]['SKU']
        email = response['Order'][0]['Email']
        orderid = response['Order'][0]['OrderID']
        city = response['Order'][0]['ShipCity']
        if 'Express' in response['Order'][0]['ShippingOption']:
            if answer != str(3):

                while True:

                    print('Oh no Rijan')
                    time.sleep(1)
                    print('You silly billy')
                    time.sleep(2)

                    shipping_answer = input("""
Was this supposed to be Express by any chance? (Y/N)

Input Answer:"""
                                            )

                    if shipping_answer.lower() == 'y':
                        answer = '3'
                        print('Changing to express...')
                        time.sleep(1)
                        break

                    elif shipping_answer.lower() == 'n':
                        break

                    else:
                        print('Valid input not found, try again')
                        time.sleep(1)
                        contine

        try:
            company = response['Order'][0]['ShipCompany']
        except KeyError:
            company = ''
        name = response['Order'][0]['ShipFirstName'] + " " + response['Order'][0]['ShipLastName']

        if 'ShipPhone' in response['Order'][0]:
            phone = response['Order'][0]['ShipPhone']

        else:
            phone = '0417557472'
        postcode = response['Order'][0]['ShipPostCode']
        state = response['Order'][0]['ShipState']

        address1 = response['Order'][0]['ShipStreetLine1']
        try:

            address2 = response['Order'][0]['ShipStreetLine2']
        except KeyError:
            address2 = ''

        if 'ShipCompany' in response['Order'][0]:

            while True:

                address_type = input('Is this going to a business address? (Y/N)')

                if address_type.lower() == 'y':
                    address_type = 'business'
                    company = response['Order'][0]['ShipCompany']
                    break

                elif address_type.lower() == 'n':
                    address_type = 'residential'
                    break

                else:
                    print('Valid input not found. Try again.')
                    time.sleep(1)
                    continue

        else:
            company = ''

        srn = orderid

        platform = 'Neto'

        GetItemdata = {'Filter': {'SKU': str(SKU),
                                  'OutputSelector': ["Name", 'PrimarySupplier', 'ShippingHeight', 'ShippingLength',
                                                     'ShippingWidth', 'ShippingWeight', 'Misc06', 'Categories',
                                                     'CategoryID', 'WarehouseQuantity']}}

        r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=GetItemheaders,
                          json=GetItemdata)

        response = r.text
        response = json.loads(response)

        try:

            if response['Item'][0]['Misc06'] == 'e-parcel':
                response['Item'][0]['Misc06'] = 'Satchel'

            existing_item = 'true'

        except IndexError:
            existing_item = 'false'

        if int(orderlinelen) == 1 and int(quantity) == 1:

            try:

                if response['Item'][0]['Misc06'] == 'Satchel' and response['Item'][0]['ShippingHeight'] != '0.000':
                    existing_dimensions = 'true'
                    weightvalue = round(float(response['Item'][0]['ShippingWeight']), 2)
                    length = round(float(response['Item'][0]['ShippingLength']) * 100, 2)
                    width = round(float(response['Item'][0]['ShippingWidth']) * 100, 2)
                    height = round(float(response['Item'][0]['ShippingHeight']) * 100, 2)
                    volumevalue = (length * width * height) * 0.000001

                    data = {
                        "length": length,
                        "width": width,
                        "height": height
                    }
                    try:
                        r = requests.post(base_url + '/api/utils/calc-cubic', headers=Fastway_Headers, json=data)

                        response = r.text
                        response = json.loads(response)
                        if 'errors' in response:
                            fastwayfailed = 'true'

                            cubicweight = round(
                                ((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250, 2)


                        else:

                            cubicweight = response['data']['cubicWeight']

                    except:
                        fastwayfailed = 'true'

                        cubicweight = round(
                            ((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250, 2)

                    if ((float(length) * float(width) * float(height)) / 4000) > 25 or float(
                            weightvalue) > 25 or length > 120 or width > 120 or height > 120:
                        cbdexpressfailed = 'true'

                    if length > 105 or width > 105 or height > 105 or float(weightvalue) > 22:
                        auspostfailed = 'true'
                        mailplusfailed = 'true'


                    if length > 105 or width > 105 or height > 105 or float(weightvalue) > 20:
                        freightsterfailed = 'true'

                    if float(cubicweight) > 40 or float(weightvalue) > 25:
                        fastwayfailed = 'true'
                        

                    if length > 105 or width > 105 or height > 105 or float(weightvalue) > 22:
                        dai_failed = 'true'

                    if float(cubicweight) > 40 or length > 180 or width > 180 or height > 180:
                        couriers_failed = 'true'

                    if (
                            length * width * height) / 1000 > 100 or float(
                        weightvalue) > 25 or length > 120 or width > 120 or height > 120:
                        sendlefailed = 'true'

                else:
                    existing_dimensions = 'false'

            except IndexError:
                existing_dimensions = 'false'

        else:
            existing_dimensions = 'false'

    if reEbay.match(OrderID):

        platform = 'ebay'
        try:
            api = Trading(appid="KyalScar-Awaiting-PRD-aca833663-f54c920e", timeout=120, config_file=None,
                          devid="5c97c2a8-eae7-49cc-ae81-c3ed1bbab335", certid="PRD-ca8336639ad2-75af-45bc-992a-d0bc",
                          token="v^1.1#i^1#r^1#f^0#I^3#p^3#t^Ul4xMF84OkUzOEU3NzVFQzA3NDUwNjUyMkY0NTc4QjlENjRFRDVFXzFfMSNFXjI2MA==")
            response = api.execute('GetOrders', {"OrderIDArray": {"OrderID": OrderID}})
            #  pprint.pprint(response.dict())
            unformatteddic = response.dict()

            name = unformatteddic['OrderArray']['Order'][0]['ShippingAddress']['Name']
            phone = unformatteddic['OrderArray']['Order'][0]['ShippingAddress']['Phone']
            city = unformatteddic['OrderArray']['Order'][0]['ShippingAddress']['CityName']
            state = unformatteddic['OrderArray']['Order'][0]['ShippingAddress']['StateOrProvince']
            address1 = unformatteddic['OrderArray']['Order'][0]['ShippingAddress']['Street1']
            address2 = unformatteddic['OrderArray']['Order'][0]['ShippingAddress']['Street2']
            srn = unformatteddic['OrderArray']['Order'][0]['ShippingDetails'][
                'SellingManagerSalesRecordNumber']
            email = unformatteddic['OrderArray']['Order'][0]['TransactionArray']['Transaction'][0]['Buyer']['Email']

            shipping_service = unformatteddic['OrderArray']['Order'][0]['ShippingServiceSelected']['ShippingService']

            if 'express' in shipping_service.lower():
                if answer != str(3):

                    while True:

                        shipping_answer = input("""
Uh oh! Was this supposed to be Express by any chance? (Y/N)

                Input Answer:"""
                                                )

                        if shipping_answer.lower() == 'y':
                            answer = '3'
                            print('Changing to express...')
                            time.sleep(1)
                            break

                        elif shipping_answer.lower() == 'n':
                            break

                        else:
                            print('Valid input not found, try again')
                            time.sleep(1)
                            contine

            if email == 'Invalid Request':
                email = 'info@scarlettmusic.com.au'

            postcode = unformatteddic['OrderArray']['Order'][0]['ShippingAddress']['PostalCode']
            company = ''
            itemid = unformatteddic['OrderArray']['Order'][0]['TransactionArray']['Transaction'][0]['Item']['ItemID']
            orderlinelen = len(unformatteddic['OrderArray']['Order'][0]['TransactionArray']['Transaction'])
            quantity = unformatteddic['OrderArray']['Order'][0]['TransactionArray']['Transaction'][0][
                'QuantityPurchased']
            totalprice = unformatteddic['OrderArray']['Order'][0]['AmountPaid']['value']

            GetItemAPI = api.execute('GetItem', {"ItemID": itemid})
            GetItemAPIdic = GetItemAPI.dict()

            if int(orderlinelen) == 1 and int(quantity) == 1:

                if 'PackageDepth' not in GetItemAPIdic['Item']['ShippingPackageDetails']:
                    existing_dimensions = 'false'

                else:
                    existing_dimensions = 'true'

                    length = round(float(GetItemAPIdic['Item']['ShippingPackageDetails']['PackageLength']['value']), 2)
                    width = round(float(GetItemAPIdic['Item']['ShippingPackageDetails']['PackageWidth']['value']), 2)
                    height = round(float(GetItemAPIdic['Item']['ShippingPackageDetails']['PackageDepth']['value']), 2)
                    weightmajor = GetItemAPIdic['Item']['ShippingPackageDetails']['WeightMajor']['value']
                    weightminor = GetItemAPIdic['Item']['ShippingPackageDetails']['WeightMinor']['value']
                    weightvalue = str(round(float(weightmajor) + float(float(weightminor) / 1000), 2))
                    volumevalue = (float(length) * float(width) * float(height)) * 0.000001

                    data = {
                        "length": length,
                        "width": width,
                        "height": height
                    }

                    try:

                        r = requests.post(base_url + '/api/utils/calc-cubic', headers=Fastway_Headers, json=data)

                        response = r.text
                        response = json.loads(response)

                        if 'errors' in response:
                            fastwayfailed = 'true'

                            cubicweight = round(
                                ((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250, 2)
                        else:

                            cubicweight = response['data']['cubicWeight']

                    except:
                        cubicweight = round(
                            ((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250, 2)

                    if ((float(length) * float(width) * float(height)) / 4000) > 25 or float(
                            weightvalue) > 25 or length > 120 or width > 120 or height > 120:
                        cbdexpressfailed = 'true'

                    if float(cubicweight) > 40 or float(weightvalue) > 25:
                        fastwayfailed = 'true'

                    if length > 105 or width > 105 or height > 105 or float(weightvalue) > 22:
                        auspostfailed = 'true'

                    if length > 105 or width > 105 or height > 105 or float(weightvalue) > 22:
                        dai_failed = 'true'

                    if float(cubicweight) > 40 or length > 180 or width > 180 or height > 180:
                        couriers_failed = 'true'

                    if (
                            length * width * height) / 1000 > 100 or float(
                        weightvalue) > 25 or length > 120 or width > 120 or height > 120:
                        sendlefailed = 'true'

            else:
                existing_dimensions = 'false'

        except ConnectionError as e:
            print(e)
            print(e.response.dict())
            sys.exit()

    if address2 is None:
        address2 = ''

    if state is None:
        state = ''

    try:
        if 'ebay' in address1:
            address = address2 + " " + city + " " + state + " " + postcode
            #   print(srn + " " + city + " " + address)

            get_google_address()
            # postcode = postal_code
            # city = suburb
        # print(srn + " " + city + " " + address)

        else:
            address = address1 + " " + address2 + " " + city + " " + state + " " + postcode
            #    print(newSalesRecordNumber + " " + newBuyerCity + " " + address)

            get_google_address()
            # postcode = postal_code
            # city = suburb
            #    print(newSalesRecordNumber + " " + newBuyerCity + " " + address)

    except (IndexError, NameError):
        pass

    if 'po box' in address1.lower() or 'po box' in address2.lower() or 'care po' in address1.lower() or 'care po' in address2.lower() or 'p o box' in address1.lower() or 'p o box' in address2.lower() or 'parcel locker' in address1.lower() or 'parcel locker' in address2.lower() or 'p.o' in address1.lower() or 'p.o' in address2.lower() or 'parcel collect' in address1.lower() or 'parcel collect' in address2.lower() or 'pobox' in address1.lower() or 'pobox' in address2.lower() or 'locker' in address1.lower() or 'locker' in address2.lower() or 'collect' in address1.lower() or 'collect' in address2.lower() or 'parcel' in address1.lower() or 'parcel' in address2.lower() or 'pmb' in address1.lower() or 'pmb' in address2.lower() or 'p/o' in address1.lower() or 'p/o' in address2.lower() or 'post office box' in address1.lower() or 'post office box' in address2.lower() or 'lpo' in address1.lower() or 'lpo' in address2.lower() or 'post office' in address1.lower() or 'post office' in address2.lower() or 'australia post' in address1.lower() or 'australia post' in address2.lower() or 'australia post' in company.lower() or 'post box' in address1.lower() or 'post box' in address2.lower():
        sendlefailed = 'true'
        transdirectfailed = 'true'
        fastwayfailed = 'true'
        allied_failed = 'true'
        couriers_failed = 'true'
        freightsterfailed = 'true'
        mailplusfailed = 'true'
        toll_failed = 'true'
        cbdexpressfailed = 'true'
        po_box = 'true'
        bondsfailed = 'true'
        dai_failed = 'true'

    if str(answer) == '3':
        transdirectfailed = 'true'

        fastwayfailed = 'true'

        couriers_failed = 'true'

        freightsterfailed = 'true'

        dai_failed = 'true'

    # elif state.lower == 'northern territory':
    #     state = 'NT'

    #################  BEGINNING OF SENDLE API ########################

    # print('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n')
    passcondition1 = ''
    passcondition2 = ''
    australiapostfailed = ''

    while True:

        try:
            print('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\nCURRENT ORDER - ' + name + '\n')

        except NameError:
            print('No order found. Are you sure you copied the order ID?')
            time.sleep(2)

            sys.exit()

        existing_dimensions_answer = ''

        if existing_dimensions == 'true':

            while True:
                existing_dimensions_answer = input(
                    "Existing dimensions of " + str(length) + "cm x " + str(width) + "cm x " + str(
                        height) + "cm, " + str(
                        weightvalue) + "kg found. \n\nWould you like to use these? (Enter Number 1-2) \n\n1) Yes\n2) No\n\nEnter Response:")

                if platform == 'Neto':

                    sendle_weight_results = item_check(SKU)
                    item_thickness_results = item_thickness_check(SKU)

                elif platform == 'ebay':

                    sendle_weight_results = item_check(itemid)
                    item_thickness_results = item_thickness_check(itemid)

                if existing_dimensions_answer.lower() == 'y' or existing_dimensions_answer == '1':

                    if sendle_weight_results == 'true' and float(weightvalue) == 0.5:

                        while True:
                            sendleweighttest = input('''

Hold up there cowboy - any chance this here package would be less than 250g?

1) Ohhh boy yessir!
2) No

Enter response here:''')

                            if sendleweighttest == str(1):
                                weightvalue = 0.25

                                if platform.lower() == 'ebay':

                                    t = math.modf(weightvalue)  # (0.5678000000000338, 1234.0)

                                    minorvalue = round((t[0]) * 1000)
                                    majorvalue = round(t[1])

                                    try:

                                        dimensionupdate = api.execute('ReviseItem', {'Item': {"ItemID": itemid,
                                                                                              'ShippingPackageDetails': {
                                                                                                  'MeasurementUnit': 'Metric',
                                                                                                  'WeightMajor': majorvalue,
                                                                                                  'WeightMinor': minorvalue}}})

                                        item_insert(itemid)
                                        break

                                    except:

                                        break

                                if platform.lower() == 'neto':
                                    UpdateItemData = {'Item': {'SKU': SKU,
                                                               'ShippingWeight': weightvalue}}

                                    r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI',
                                                      headers=UpdateItemheaders,
                                                      json=UpdateItemData)
                                    item_insert(SKU)
                                    break

                            elif sendleweighttest == str(2):
                                if platform.lower() == 'ebay':
                                    item_insert(itemid)

                                if platform.lower() == 'neto':
                                    item_insert(SKU)
                                break

                            else:
                                print('Valid input not found, try again')
                                time.sleep(1)
                                continue



                else:
                    pass

                if item_thickness_results == 'true' and float(weightvalue) <= 0.3:

                    while True:
                        sendleweighttest = console.input('''

        ___
     __|___|__
      ('o_o')                             
      _\~-~/_    ______.                  STOP RIGHT THERE. 
     //\__/\ \ ~(_]---'                   
    / )O  O( .\/_)                        I AM A [red]TERRORIST[/].
    \ \    / \_/            
    )/_|  |_\\                             WOULD THIS PACKAGE BE [red]LESS THAN 3cm THICK[/]?
   // /(\/)\ \\     
   /_/      \_\\                          
  (_||      ||_)                       
    \| |__| |/
     | |  | |
     | |  | |                             1) Yes Mr Terrorist.
     |_|  |_|                             2) No sir
     /_\  /_\\                            
                                          Enter response here:''')

                        if sendleweighttest == str(1):
                            height = 21
                            width = 21
                            length = 3
                            weightvalue = float(weightvalue)

                            if platform.lower() == 'ebay':

                                t = math.modf(weightvalue)  # (0.5678000000000338, 1234.0)

                                minorvalue = round((t[0]) * 1000)
                                majorvalue = round(t[1])

                                try:

                                    dimensionupdate = api.execute('ReviseItem', {'Item': {"ItemID": itemid,
                                                                                          'ShippingPackageDetails': {
                                                                                              'MeasurementUnit': 'Metric',
                                                                                              'PackageDepth': height,
                                                                                              'PackageLength': length,
                                                                                              'PackageWidth': width, }}})

                                    item_thickness_insert(itemid)
                                    break

                                except:

                                    break

                            if platform.lower() == 'neto':
                                UpdateItemData = {'Item': {'SKU': SKU,
                                                           'ShippingHeight': height / 100,
                                                           'ShippingLength': length / 100,
                                                           'ShippingWidth': width / 100}}

                                r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI',
                                                  headers=UpdateItemheaders,
                                                  json=UpdateItemData)
                                item_thickness_insert(SKU)
                                break

                        elif sendleweighttest == str(2):
                            if platform.lower() == 'ebay':
                                item_thickness_insert(itemid)

                            if platform.lower() == 'neto':
                                item_thickness_insert(SKU)
                            break

                        else:
                            print('Valid input not found, try again')
                            time.sleep(1)
                            continue



                    else:
                        pass

                if existing_dimensions_answer == '1' or existing_dimensions_answer.lower() == 'y':

                    package = [float(length), float(width), float(height)]
                    sorted_package = sorted(package)

                    if float(weightvalue) <= 0.3 and float(sorted_package[0]) <= 3 and float(
                            sorted_package[1]) <= 21 and float(sorted_package[2]) <= 21:
                        satchel = '300gm'


                    elif float(cubicweight) <= 0.5 and float(weightvalue) <= 0.5:
                        satchel = 'A5'

                    elif float(cubicweight) <= 1 and float(weightvalue) <= 1:
                        satchel = 'A4'

                    elif float(cubicweight) <= 3 and float(weightvalue) <= 3:
                        satchel = 'A3'

                    elif float(cubicweight) <= 5 and float(weightvalue) <= 5:
                        satchel = 'A2'
                    else:
                        satchel = 'false'
                    multi_items_to_ship = [(weightvalue, length, height, width, volumevalue, cubicweight, 1, satchel)]

                    break

                if existing_dimensions_answer == '2' or existing_dimensions_answer.lower() == 'n':
                    break

                else:
                    print('Valid input not found, try again.')
                    time.sleep(1)

        if existing_dimensions_answer == '1' or existing_dimensions_answer.lower() == 'y':
            break

        if existing_dimensions_answer != '1' or existing_dimensions_answer.lower() != 'y':

            multi_items_to_ship = []

            parcel_number = 0
            try_again = ''

            while True:

                if try_again == '2':
                    parcel_number = parcel_number - 1

                parcel_number = parcel_number + 1

                try_again = ''

                parcelsize = console.input('''\n[red underline]Article ''' + str(parcel_number) + '''[/]

What size is your parcel? 

[green bold]SINGLE ORDER[/] - locks in order straight away, use for single parcels

1) Bubble Mailer
2) Other

[green bold]MULTI ORDER[/] - use for multiple parcels

3) Bubble Mailer
4) Other


Alternatively, press [cyan underline]ENTER[/] to lock in order.

Enter response here:''')

                ##### Needs to ask whether the person wants to upload dimensions for future use. Warn user not to do it if the dimensions will change (ie frankenstien box)
                #####BUT. Can only do this for orders with 1 quantity item with 1 orderline

                uploadsize = 'No'

                if parcelsize == str(1) or parcelsize == str(3):

                    fastway_thickness_result = 'False'

                    weightvalue = 0.50

                    if parcelsize == str(1) and answer != str(3):

                        while True:
                            sendleweighttest = input('''What is the weight?

                            1) Less than 250g
                            2) Greater than 250g

                            Enter response here:''')

                            if sendleweighttest == str(1):
                                weightvalue = 0.25

                                while True:
                                    fastway_thickness_test = input('''What is the thickness?

                                    1) Less than 3cm
                                    2) Greater than 3cm

                                    Enter response here:''')

                                    if fastway_thickness_test == str(1):
                                        fastway_thickness_result = 'True'
                                        break
                                    elif fastway_thickness_test == str(2):

                                        break
                                break

                            elif sendleweighttest == str(2):
                                break

                            else:
                                print('Valid input not found, try again')
                                time.sleep(1)
                                continue

                    if fastway_thickness_result == 'False':
                        volumevalue = 0.0001
                        length = 18
                        width = 23
                        height = 4

                        passcondition1 = 'True'

                        cubicweight = 0.414

                        satchel = 'A5'

                    if fastway_thickness_result == 'True':
                        volumevalue = 0.0001
                        length = 21
                        width = 21
                        height = 3

                        passcondition1 = 'True'

                        cubicweight = 0.414

                        satchel = '300gm'

                    if int(quantity) > 1:
                        existing_dimensions_answer = '1'

                    if parcelsize == str(1):
                        quantity = 1

                    else:

                        while True:

                            try:

                                quantity = int(input('\nQuantity:'))

                                if quantity < 1:
                                    print('Integer must be greater than 0. Try again.')
                                    time.sleep(1)
                                    continue

                                if quantity > 1:
                                    sendlefailed = 'true'
                                    freightsterfailed = 'true'
                                    mailplusfailed = 'true'
                                    dai_failed = 'true'

                                break

                            except ValueError:
                                print('Invalid value, must be integer. Try again.')
                                time.sleep(1)
                                continue

                    multi_items_to_ship.append(
                        (weightvalue, length, height, width, volumevalue, cubicweight, quantity, satchel))






                elif parcelsize == str(2) or parcelsize == str(4):

                    while True:

                        dimensions_search = input('''\nWould you like to:

    1) Use an existing items dimensions
    2) Enter dimensions manually

    Enter Response:''')
                        if dimensions_search != '1' and dimensions_search != '2':
                            print('Valid input not found. Try Again.')
                            time.sleep(1)
                            continue
                        else:
                            break

                    satchel = 'false'

                    if dimensions_search == '1':

                        try_again = ''
                        while True:

                            if try_again == '2':
                                existing_dimensions_answer = '1'
                                break

                            try_again = ''

                            SKU = input('\nEnter SKU: ')

                            GetItemdata = {'Filter': {'SKU': str(SKU),
                                                      'OutputSelector': ["Name", 'PrimarySupplier', 'ShippingHeight',
                                                                         'ShippingLength',
                                                                         'ShippingWidth', 'ShippingWeight', 'Misc06',
                                                                         'Categories',
                                                                         'CategoryID', 'WarehouseQuantity']}}

                            r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=GetItemheaders,
                                              json=GetItemdata)

                            response = r.text
                            response = json.loads(response)

                            try:

                                if response['Item'][0]['Misc06'] == 'e-parcel':
                                    response['Item'][0]['Misc06'] = 'Satchel'

                                existing_item = 'true'


                            except IndexError:
                                existing_item = 'false'
                                while True:

                                    try_again = input(f'''\n{SKU} dimensions not found. Would you like to try again?

1) Yes
2) No

Enter Response:''')

                                    if try_again != '1' and try_again != '2':
                                        print('Valid input not found. Try again.')
                                        time.sleep(1)
                                        continue
                                    else:
                                        break
                            if try_again != '1' and try_again != '2':
                                pass
                            else:
                                continue

                            try:

                                if response['Item'][0][
                                    'ShippingHeight'] != '0.000':
                                    existing_dimensions = 'true'
                                    weightvalue = round(float(response['Item'][0]['ShippingWeight']), 2)
                                    length = round(float(response['Item'][0]['ShippingLength']) * 100, 2)
                                    width = round(float(response['Item'][0]['ShippingWidth']) * 100, 2)
                                    height = round(float(response['Item'][0]['ShippingHeight']) * 100, 2)
                                    volumevalue = (length * width * height) * 0.000001

                                    print(
                                        f"Existing dimensions of {length}cm x {width}cm x {height}cm, {weightvalue} kg to be used")
                                    existing_dimensions_answer = '1'
                                    time.sleep(1)
                                    break

                                else:
                                    while True:

                                        try_again = input(f'''\n{SKU} dimensions not found. Would you like to try again?

                                    1) Yes
                                    2) No

                                    Enter Response:''')

                                        if try_again != '1' and try_again != '2':
                                            print('Valid input not found. Try again.')
                                            time.sleep(1)
                                            continue
                                        else:
                                            break

                            except:

                                while True:

                                    try_again = input(f'''\n{SKU} dimensions not found. Would you like to try again?

                                1) Yes
                                2) No

                                Enter Response:''')

                                    if try_again != '1' and try_again != '2':
                                        print('Valid input not found. Try again.')
                                        time.sleep(1)
                                        continue
                                    else:
                                        break

                    if try_again == '2':
                        continue
                    if dimensions_search == '2':
                        weightvalue = float(input('Enter item weight (kg)'))

                        length = float(input('Enter item length (cm)'))
                        width = float(input('Enter item width (cm)'))
                        height = float(input('Enter item height (cm)'))

                        volumevalue = (length * width * height) * 0.000001

                    data = {
                        "length": length,
                        "width": width,
                        "height": height
                    }

                    try:
                        r = requests.post(base_url + '/api/utils/calc-cubic', headers=Fastway_Headers, json=data)

                        response = r.text
                        response = json.loads(response)

                        cubicweight = response['data']['cubicWeight']

                        if float(cubicweight) > 40 or float(weightvalue) > 25:
                            fastwayfailed = 'true'

                    except:
                        cubicweight = round(
                            ((float(length) / 100) * (float(width) / 100) * (float(height) / 100)) * 250, 2)

                        fastwayfailed = 'true'

                    if length > 105 or width > 105 or height > 105 or weightvalue > 22:
                        auspostfailed = 'true'

                    if length > 105 or width > 105 or height > 105 or weightvalue > 20:
                        freightsterfailed = 'true'

                    if length > 105 or width > 105 or height > 105 or weightvalue > 22:
                        dai_failed = 'true'

                    if (
                            length * width * height) / 1000 > 100 or weightvalue > 25 or length > 120 or width > 120 or height > 120:
                        sendlefailed = 'true'

                    passcondition1 = 'True'

                    if int(quantity) > 1:
                        existing_dimensions_answer = '1'

                    if parcelsize == str(2):
                        quantity = 1

                    else:

                        while True:

                            try:

                                quantity = int(input('\nQuantity:'))

                                if quantity < 1:
                                    print('Integer must be greater than 0. Try again.')
                                    time.sleep(1)
                                    continue

                                if quantity > 1:
                                    sendlefailed = 'true'
                                    freightsterfailed = 'true'
                                    mailplusfailed = 'true'
                                    cbdexpressfailed = 'true'
                                    dai_failed = 'true'

                                break

                            except ValueError:
                                print('Invalid value, must be integer. Try again.')
                                time.sleep(1)
                                continue

                    package = [float(length), float(width), float(height)]
                    sorted_package = sorted(package)

                    if float(weightvalue) <= 0.3 and float(sorted_package[0]) <= 3 and float(
                            sorted_package[1]) <= 21 and float(sorted_package[2]) <= 21:
                        satchel = '300gm'

                    elif float(cubicweight) <= 0.5 and float(weightvalue) <= 0.5:
                        satchel = 'A5'

                    elif float(cubicweight) <= 1 and float(weightvalue) <= 1:
                        satchel = 'A4'

                    elif float(cubicweight) <= 3 and float(weightvalue) <= 3:
                        satchel = 'A3'

                    elif float(cubicweight) <= 5 and float(weightvalue) <= 5:
                        satchel = 'A2'

                    multi_items_to_ship.append(
                        (weightvalue, length, height, width, volumevalue, cubicweight, quantity, satchel))

                if parcelsize == '' or parcelsize == str(1) or parcelsize == str(2):
                    passcondition = 'true'
                    break

                if parcelsize != '' and parcelsize != str(1) and parcelsize != str(2) and parcelsize != str(
                        3) and parcelsize != str(4):
                    print('Number not found, select a number next time, ya dingus.')
                    parcel_number = parcel_number - 1
                    time.sleep(2)
                    continue

        if int(orderlinelen) == 1 and int(quantity) == 1 and existing_dimensions_answer != '1' and len(
                multi_items_to_ship) == 1 and int(answer) != 3:

            while True:

                uploadsize = input('''\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\nWould you like to upload these dimensions to eBay / Neto?

        NOTE: Don't upload dimensions for items whose size will dramatically differ each time.
        e.g. Anything that needs a Frankenstein box.

        1) Yes
        2) No

        Enter Response (1-2):''')

                if uploadsize.lower() == 'y' or str(uploadsize) == '1':

                    if platform.lower() == 'ebay':

                        ItemNote = 'Satchel.'

                        noteresponse = api.execute('SetUserNotes',
                                                   {"Action": "AddOrUpdate", "ItemID": int(itemid),
                                                    "NoteText": ItemNote})
                        usernotesdic = noteresponse.dict()

                        t = math.modf(weightvalue)  # (0.5678000000000338, 1234.0)

                        minorvalue = round((t[0]) * 1000)
                        majorvalue = round(t[1])

                        try:

                            dimensionupdate = api.execute('ReviseItem', {'Item': {"ItemID": itemid,
                                                                                  'ShippingPackageDetails': {
                                                                                      'MeasurementUnit': 'Metric',
                                                                                      'PackageDepth': height,
                                                                                      'PackageLength': length,
                                                                                      'PackageWidth': width,
                                                                                      'WeightMajor': majorvalue,
                                                                                      'WeightMinor': minorvalue}}})
                            passcondition2 = 'True'
                        except:
                            passcondition2 = 'True'

                            break

                    if platform.lower() == 'neto':
                        newPostageType = 'Satchel'

                        UpdateItemData = {'Item': {'SKU': SKU, 'Misc06': newPostageType,
                                                   'ShippingHeight': height / 100, 'ShippingLength': length / 100,
                                                   'ShippingWidth': width / 100, 'ShippingWeight': weightvalue}}

                        r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI',
                                          headers=UpdateItemheaders,
                                          json=UpdateItemData)

                    passcondition2 = 'True'

                    break

                elif uploadsize.lower() == 'n' or str(uploadsize) == '2':

                    passcondition2 = 'True'

                    uploadsize = 'No'
                    break

                else:

                    print('Valid selection not found, do better next time, ya dingus.')
                    time.sleep(2)
                    continue

        if passcondition1 == 'True':
            break

    #### Need to run Aus Post + Sendle post calculators

    couriers = ['Australia Post', 'Sendle', 'Transdirect', 'Fastway', 'Allied', 'Couriers Please', 'Freightster',
                'Mailplus', 'Toll', 'CBDExpress', 'Dai Post', 'Bonds']

    with concurrent.futures.ThreadPoolExecutor(max_workers=None) as executor:
        results = executor.map(GetQuote, couriers)

        # auspostprice = results[0]
        # sendleprice = results[1]
        # transdirectprice = results[2]
        # fastwayprice = results[3]
        # alliedprice = results[4]
    courier_prices = []
    for x in results:
        courier_prices.append(x)
    # print(x)

    auspostprice = courier_prices[0]
    sendleprice = courier_prices[1]
    transdirectprice = courier_prices[2]
    fastwayprice = courier_prices[3]
    alliedprice = courier_prices[4]
    courierspleaseprice = courier_prices[5]
    freightsterprice = courier_prices[6]
    mailplusprice = courier_prices[7]
    tollprice = courier_prices[8]
    cbdprice = courier_prices[9]
    daiprice = courier_prices[10]
    bondsprice = courier_prices[11]


    if float(auspostprice) == float(fastwayprice):
        fastwayprice = fastwayprice + 0.01

    if float(daiprice) == float(fastwayprice):
        daiprice = daiprice + 0.01

    if float(daiprice) == float(sendleprice):
        daiprice = daiprice + 0.01

    lowestcourier = {'auspostprice': float(auspostprice), 'sendleprice': float(sendleprice),
                     'transdirectprice': float(transdirectprice), 'fastwayprice': float(fastwayprice),
                     'alliedprice': float(alliedprice), 'courierspleaseprice': float(courierspleaseprice),
                     'freightsterprice': float(freightsterprice), 'mailplusprice': float(mailplusprice),
                     'tollprice': float(tollprice), 'cbdprice': float(cbdprice), 'daiprice': float(daiprice),
                      'bondsprice': float(bondsprice)}
    lowestcourier = min(lowestcourier, key=lowestcourier.get)
    # print(lowestcourier)

    if lowestcourier == 'sendleprice':
        finalcourier = 'Sendle'

    if lowestcourier == 'auspostprice':
        finalcourier = 'Australia Post'

    if lowestcourier == 'fastwayprice':
        finalcourier = 'Fastway'

    if lowestcourier == 'alliedprice':
        finalcourier = 'Allied Express'

    if lowestcourier == 'courierspleaseprice':
        finalcourier = 'Couriers Please'

    if lowestcourier == 'freightsterprice':
        finalcourier = 'Freightster'

    if lowestcourier == 'mailplusprice':
        finalcourier = 'Mailplus'

    if lowestcourier == 'tollprice':
        finalcourier = 'Toll'

    if lowestcourier == 'cbdprice':
        finalcourier = 'CBDExpress'

    if lowestcourier == 'daiprice':
        finalcourier = 'Dai Post'

    if lowestcourier == 'bondsprice':
        finalcourier = 'Bonds Transport'

    if lowestcourier == 'transdirectprice':
        finalcourier = 'Transdirect'

        lowesttranscourier = lowesttranscourier.replace('_', ' ')
        lowesttranscourier = lowesttranscourier.title()
        lowesttranscourier = lowesttranscourier.replace('Tnt', 'TNT')

    try:
        id
    except NameError:
        id = ''

    try:
        lowesttranscourier
    except NameError:
        lowesttranscourier = ''

    if email == '':
        email = 'info@scarlettmusic.com.au'

    # if parcelsize == str(2):
    #     if int(sendleprice) > 10 and int(auspostprice) > 10 and int(transdirectprice) > 10:
    #         print('All couriers are too expensive, please use a prepaid satchel.')
    #         input('Press Enter to exit')
    #         sys.exit()

    if answer == str(4) or answer == str(3):
        if lowestcourier == 'sendleprice':
            continueanswer = input('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\nThe cheapest courier is Sendle at $' + str(
                sendleprice) + '. Would you like to continue? (Y/N)')

            if continueanswer.lower() == 'n':
                sys.exit()
            elif continueanswer.lower() == 'y':
                pass
            else:
                print('No viable answer found. SELF DESTRUCTING.')
                time.sleep(2)
                sys.exit()

        elif lowestcourier == 'auspostprice':
            continueanswer = input('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\nThe cheapest courier is Australia Post at $' + str(
                auspostprice) + '. Would you like to continue? (Y/N)')

            if continueanswer.lower() == 'n':
                sys.exit()
            elif continueanswer.lower() == 'y':
                pass
            else:
                print('No viable answer found. SELF DESTRUCTING.')
                time.sleep(2)
                sys.exit()

        elif lowestcourier == 'transdirectprice':
            continueanswer = input('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\nThe cheapest courier is Transdirect at $' + str(
                transdirectprice) + '. Would you like to continue? (Y/N)')

            if continueanswer.lower() == 'n':
                sys.exit()
            elif continueanswer.lower() == 'y':
                pass
            else:
                print('No viable answer found. SELF DESTRUCTING.')
                time.sleep(2)
                sys.exit()

        elif lowestcourier == 'fastwayprice':
            continueanswer = input('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\nThe cheapest courier is Fastway at $' + str(
                fastwayprice) + '. Would you like to continue? (Y/N)')

            if continueanswer.lower() == 'n':
                sys.exit()
            elif continueanswer.lower() == 'y':
                pass
            else:
                print('No viable answer found. SELF DESTRUCTING.')
                time.sleep(2)
                sys.exit()

        elif lowestcourier == 'alliedprice':

            continueanswer = input('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\nThe cheapest courier is Allied Express at $' + str(
                alliedprice) + '. Would you like to continue? (Y/N)')

            if continueanswer.lower() == 'n':
                sys.exit()
            elif continueanswer.lower() == 'y':
                pass
            else:
                print('No viable answer found. SELF DESTRUCTING.')
                time.sleep(2)
                sys.exit()


        elif lowestcourier == 'courierspleaseprice':

            continueanswer = input('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\nThe cheapest courier is Couriers Please at $' + str(
                courierspleaseprice) + '. Would you like to continue? (Y/N)')

            if continueanswer.lower() == 'n':
                sys.exit()
            elif continueanswer.lower() == 'y':
                pass
            else:
                print('No viable answer found. SELF DESTRUCTING.')
                time.sleep(2)
                sys.exit()

        elif lowestcourier == 'freightsterprice':

            continueanswer = input('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\nThe cheapest courier is Freightster at $' + str(
                freightsterprice) + '. Would you like to continue? (Y/N)')

            if continueanswer.lower() == 'n':
                sys.exit()
            elif continueanswer.lower() == 'y':
                pass
            else:
                print('No viable answer found. SELF DESTRUCTING.')
                time.sleep(2)
                sys.exit()

        elif lowestcourier == 'cbdprice':

            continueanswer = input('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\nThe cheapest courier is CBD Express at $' + str(
                cbdprice) + '. Would you like to continue? (Y/N)')

            if continueanswer.lower() == 'n':
                sys.exit()
            elif continueanswer.lower() == 'y':
                pass
            else:
                print('No viable answer found. SELF DESTRUCTING.')
                time.sleep(2)
                sys.exit()

        elif lowestcourier == 'daiprice':

            continueanswer = input('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\nThe cheapest courier is Dai Post at $' + str(
                daiprice) + '. Would you like to continue? (Y/N)')

            if continueanswer.lower() == 'n':
                sys.exit()
            elif continueanswer.lower() == 'y':
                pass
            else:
                print('No viable answer found. SELF DESTRUCTING.')
                time.sleep(2)
                sys.exit()

        elif lowestcourier == 'mailplusprice':

            continueanswer = input('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\nThe cheapest courier is Mail Plus at $' + str(
                mailplusprice) + '. Would you like to continue? (Y/N)')

            if continueanswer.lower() == 'n':
                sys.exit()
            elif continueanswer.lower() == 'y':
                pass
            else:
                print('No viable answer found. SELF DESTRUCTING.')
                time.sleep(2)
                sys.exit()

        elif lowestcourier == 'tollprice':

            continueanswer = input('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\nThe cheapest courier is Toll at $' + str(
                tollprice) + '. Would you like to continue? (Y/N)')

            if continueanswer.lower() == 'n':
                sys.exit()
            elif continueanswer.lower() == 'y':
                pass
            else:
                print('No viable answer found. SELF DESTRUCTING.')
                time.sleep(2)
                sys.exit()

        elif lowestcourier == 'bondsprice':

            continueanswer = input('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\nThe cheapest courier is Bonds Transport at $' + str(
                bondsprice) + '. Would you like to continue? (Y/N)')

            if continueanswer.lower() == 'n':
                sys.exit()
            elif continueanswer.lower() == 'y':
                pass
            else:
                print('No viable answer found. SELF DESTRUCTING.')
                time.sleep(2)
                sys.exit()
    print('Printing label for ' + finalcourier + '. Please wait...')

    if finalcourier == 'CBDExpress':

        instructions = 'Authority to leave'

        if float(totalprice) > 150 and volumevalue > 0.0001:
            instructions = 'Signature on Delivery'

        cursor = connection.cursor()
        cursor.execute(
            fr"SELECT MAX(consignment_number) FROM CDBExpress")
        results = cursor.fetchall()
        max_consignment = results[0][0]
        new_consignment = str((int(max_consignment) + 1)).zfill(9)
        barcode_param = f'SCARLET{new_consignment}'

        now = datetime.datetime.now()
        current_time = now.strftime("%Y-%m-%dT%H:%M:%SZ")

        cursor = connection.cursor()
        cursor.execute(
            f"INSERT INTO CDBExpress(order_id, consignment_number, name, time_created) VALUES ('{OrderID}', '{new_consignment}', '{name}', '{current_time}' ); COMMIT;")

        pyperclip.copy(barcode_param)

        ###### Creating JSON / request

        payload = {"barcode": barcode_param,
                   "description": "Scarlett Music Order ", "weight": {"value": weightvalue, "units": "kg"},
                   "volume": {"value": volumevalue, "units": "m3"}, "customer_reference": OrderID, "sender": {
                "contact": {
                    "name": "Scarlett Music",
                    "phone": "(03) 9318 5751",
                    "company": "Scarlett Music"
                },
                "address": {
                    "address_line1": "286-288 Ballarat Rd",
                    "suburb": 'Maidstone',
                    "state_name": "VIC",
                    "postcode": '3012',
                    "country": "Australia"
                },
                "instructions": "The music shop, open 9am-6pm. Best parking is at The Palms across the road."
            }, "receiver": {
                "contact": {
                    "name": name,
                    "email": email,
                    "phone": phone,
                    "company": company
                },
                "address": {
                    "address_line1": address1,
                    "address_line2": address2,
                    "suburb": city,
                    "state_name": state,
                    "postcode": postcode,
                    "country": "Australia"
                },
                "instructions": instructions
            }
                   }

        url = 'https://apis.hubsystems.com.au/booking/'
        auth = ('code-scarlett', 'syY55DxG41sd8')
        headers = {'Content-Type': 'text/scarlett'}
        data = payload

        response = requests.post(url, auth=auth, headers=headers, json=data)

        print(response.text)

        ####Need to put in request here, And only continue if response is good

        # Create barcode image
        barcode_image = code128.image(barcode_param, height=100)

        # Create empty image for barcode + text
        top_bott_margin = 70
        l_r_margin = 10
        new_height = barcode_image.height + (2 * top_bott_margin)
        new_width = barcode_image.width + (2 * l_r_margin)
        new_image = Image.new('RGB', (new_width, 700), (255, 255, 255))

        # put barcode on new image
        barcode_y = 525
        new_image.paste(barcode_image, (0, barcode_y))

        # object to draw text
        draw = ImageDraw.Draw(new_image)

        # Define custom text size and font
        h1_size = 28
        h2_size = 28
        h3_size = 16
        footer_size = 21

        h1_font = ImageFont.truetype("r'C:\Windows\Fonts\calibri.ttf", h1_size)
        h2_font = ImageFont.truetype("r'C:\Windows\Fonts\calibri.ttf", h2_size)
        h3_font = ImageFont.truetype("r'C:\Windows\Fonts\calibri.ttf", h3_size)
        footer_font = ImageFont.truetype("r'C:\Windows\Fonts\calibri.ttf", footer_size)

        # Define custom text
        To = 'To:'

        product_type = "Courier's By Demand Express"
        center_product_type = (barcode_image.width / 2) - len(product_type) * 5
        center_barcode_value = (barcode_image.width / 2) - len(barcode_param) * 8

        # Draw text on picture
        draw.text((l_r_margin, 200), To, fill=(0, 0, 0), font=h2_font)
        draw.text((40, 250), name, fill=(0, 0, 0), font=h2_font)
        draw.text((40, 275), company, fill=(0, 0, 0), font=h2_font)
        draw.text((40, 300), address1, fill=(0, 0, 0), font=h2_font)
        draw.text((40, 325), address2, fill=(0, 0, 0), font=h2_font)
        draw.text((40, 350), suburb, fill=(0, 0, 0), font=h2_font)
        draw.text((40, 375), state, fill=(0, 0, 0), font=h2_font)
        draw.text((40, 400), postcode, fill=(0, 0, 0), font=h2_font)
        draw.text((40, 425), phone, fill=(0, 0, 0), font=h2_font)
        draw.text((40, 490), OrderID, fill=(0, 0, 0), font=h2_font)

        draw.text((center_product_type, (490)), product_type, fill=(0, 0, 0), font=footer_font)
        draw.text((center_barcode_value, (640)), barcode_param, fill=(0, 0, 0), font=h2_font)

        scarlett_music_logo = Image.open(rf"\\SERVER\Python\website_logo.png")

        half = 0.2
        out = scarlett_music_logo.resize([int(half * s) for s in scarlett_music_logo.size])

        new_image.paste(out, (170, 10))

        # save in file
        new_image.save(imagelocations + r'\pythontest2.pdf', 'PDF')

        images = convert_from_path(imagelocations + r'\pythontest2.pdf')


        def trim(im):
            bg = Image.new(im.mode, im.size, im.getpixel((0, 0)))
            diff = ImageChops.difference(im, bg)
            diff = ImageChops.add(diff, diff, 2.0, -100)
            bbox = diff.getbbox()
            if bbox:
                return im.crop(bbox)


        for page in images:
            printer = None
            printer = BrotherQLRaster(printermodel)
            page.save(imagelocations + '\python2.jpg', 'JPEG')

            colorImage = Image.open(imagelocations + '\python2.jpg')

            transposed = colorImage.transpose(Image.ROTATE_90)

            transposed.save(imagelocations + '\python2.jpg')

            image_slicer.slice(imagelocations + '\python2.jpg', 2, 0, 1)

            im = Image.open(imagelocations + "\python2_01_01.png")

            rgb_im = im.convert('RGB')
            rgb_im.save(imagelocations + '\python2_01_01.jpg')

            filename = imagelocations + '\python2_01_01.jpg'

            im2 = Image.open(imagelocations + "\python2_01_02.png")

            rgb_im2 = im2.convert('RGB')
            rgb_im2.save(imagelocations + '\python2_01_02.jpg')

            filename2 = imagelocations + '\python2_01_02.jpg'
            sendToPrinter()

    if finalcourier == 'Bonds Transport':
        bonds_item_string = ''

        now = datetime.datetime.now()

        next_day = next_business_day()
        next_bonds_day = f"{next_day.strftime('%Y-%m-%d')}"

        bonds_total_items = 0
        bonds_total_weight = 0
        for all_items in multi_items_to_ship:
            quantity = all_items[6]

            package = [float(all_items[1]), float(all_items[2]), float(all_items[3])]
            sorted_package = sorted(package)

            for t in range(quantity):
                bonds_total_items = bonds_total_items + 1
                bonds_total_weight += float(all_items[0])

            bonds_item_string = f'''{bonds_item_string}<dimension>
                                            <qty>{quantity}</qty>
                                            <length>{sorted_package[2]}</length>
                                            <width>{sorted_package[1]}</width>
                                            <height>{sorted_package[0]}</height>
                                            </dimension>'''

        service_vehicle_tuples = [('C', 'CAR'), ('C', 'SW'), ('C', 'SV'), ('TTK', '')]

        # need to iterate through service_vehicle_types to get correct vehicle
        # Need to set date to next day

        company = ''
        if company == '':
            bonds_company = f'<company>{name}</company>'
        else:
            bonds_company = f'<company>{company}</company>'

        for service_vehicle_groups in service_vehicle_tuples:

            xml_payload = f'''
                        <job xmlns:xi="http://www.w3.org/2001/XInclude" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="job-bonds.xsd">
                        <job_action>BOOKING</job_action>
                        <notifications>
                        <notification>
                        <notify_type>DELIVERY</notify_type>
                        <notify_target>{email}</notify_target>
                        </notification>
                        <notification>
                        <notify_type>DELIVERY</notify_type>
                        <notify_target>{phone}</notify_target>
                        </notification>
                        </notifications>
                        <job_id/>
                        <account>V01523</account>
                        <authorization_code>@WV6mSH4NByW</authorization_code>
                        <containsDangerousGoods>false</containsDangerousGoods>
                        <branch>MEL</branch>
                        <job_date>{next_bonds_day}</job_date>
                        <time_ready>09:00:00</time_ready>
                        <deliver_by_time xsi:nil="true"/>
                        <deliver_by_time_reason xsi:nil="true"/>
                        <order_number>{OrderID}</order_number>
                        <contact>Kyal</contact>
                        <insurance>true</insurance>
                        <references>
                        <reference>{OrderID}</reference>
                        </references>
                        <service_code>{service_vehicle_groups[0]}</service_code>
                        <vehicle_code>{service_vehicle_groups[1]}</vehicle_code>
                        <goods_description/>
                        <instructions></instructions>
                        <pallets/>
                        <cubic/>
                        <job_legs>
                        <job_leg>
                        <action>P</action>
                        <service_qual/>
                        <suburb>Footscray</suburb>
                        <state>VIC</state>
                        <company>SCARLETT MUSIC</company>
                        <address1>286-288 Ballarat Rd</address1>
                        <address2></address2>
                        <contact>Kyal</contact>
                        <items>{bonds_total_items}</items>
                        <weight>{bonds_total_weight}</weight>
                        <dimensions>
                        {bonds_item_string}
                        </dimensions>
                        <references>
                        <reference/>
                        </references>
                        </job_leg>
                        <job_leg>
                        <action>D</action>
                        <service_qual></service_qual>
                        <suburb>{city}</suburb>
                        <state>{state}</state>
                        {bonds_company}
                        <address1>{address1}</address1>
                        <address2>{address2}</address2>
                        <contact/>
                        <items>{bonds_total_items}</items>
                        <weight>{bonds_total_weight}</weight>
                        <dimensions>
                        {bonds_item_string}
                        </dimensions>
                        <references>
                        <reference/>
                        </references>
                        </job_leg>
                        </job_legs>
                        </job>
                        '''

            url = 'https://appsrv.bondscouriers.com.au/bondsweb/api/upload-xml-job.htm'  # Replace with the actual API endpoint URL

            headers = {
                'Content-Type': 'application/xml'
            }

            response = requests.post(url, data=xml_payload, headers=headers)

           #print(response.status_code)
            xml_response = response.text
            # Parse the XML and convert it to a Python dictionary
            data_dict = xmltodict.parse(xml_response)
            # pprint.pprint(data_dict)
            msg_status = data_dict['job_message']['msg_status']
            if msg_status == 'ERROR':
                continue

            else:


                job_id = data_dict['job_message']['job_details']['job_id']

                cursor = connection.cursor()
                now = datetime.datetime.now()
                current_time = f"{now.strftime('%Y-%m-%d')}T{now.strftime('%H:%M:%S')}"
                cursor.execute(
                    fr"INSERT INTO bonds(customer_name, time_created, job_number) VALUES ('{name}', '{now}', '{job_id}'); COMMIT;")

                pyperclip.copy(job_id)

                # if service_vehicle_groups[0] == "C":
                #     service_code = service_vehicle_groups[0]
                #     vehicle_code = service_vehicle_groups[1]
                # else:
                #     service_code = service_vehicle_groups[0]
                #     vehicle_code = data_dict['job_message']['job_details']['vehicle_code']
                break

        xml_payload = f'''<?xml version="1.0" encoding="UTF-8"?>
        <job xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
             xsi:noNamespaceSchemaLocation="job-bond-joblabel.xsd">
        <job_action>JOBLABEL</job_action>
        <job_id>{job_id}</job_id>
        <account>V01523</account>
        <authorization_code>@WV6mSH4NByW</authorization_code>
        </job>'''

        print(xml_payload)

        url = 'https://appsrv.bondscouriers.com.au/bondsweb/api/upload-xml-job.htm'  # Replace with the actual API endpoint URL

        headers = {
            'Content-Type': 'application/xml'
        }

        # time.sleep(3)

        r = requests.post(url, data=xml_payload, headers=headers)

        print(r.status_code)
        # print(response.json())
        content_type = r.headers.get('Content-Type')
        print(content_type)

        with open(imagelocations + r"\pythontest2.pdf", "wb") as pdf:
            for chunk in r.iter_content(chunk_size=1024):

                # writing one chunk at a time to pdf file
                if chunk:
                    pdf.write(chunk)

        # r = requests.delete(orderurl, auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'))

        images = convert_from_path(imagelocations + r'\pythontest2.pdf')


        def trim(im):
            bg = Image.new(im.mode, im.size, im.getpixel((0, 0)))
            diff = ImageChops.difference(im, bg)
            diff = ImageChops.add(diff, diff, 2.0, -100)
            bbox = diff.getbbox()
            if bbox:
                return im.crop(bbox)

        for page in images:
            printer = None
            printer = BrotherQLRaster(printermodel)
            page.save(imagelocations + '\python2.jpg', 'JPEG')

            colorImage = Image.open(imagelocations + '\python2.jpg')

            transposed = colorImage.transpose(Image.ROTATE_90)

            transposed.save(imagelocations + '\python2.jpg')

            image_slicer.slice(imagelocations + '\python2.jpg', 2, 0, 1)

            im = Image.open(imagelocations + "\python2_01_01.png")
            im = trim(im)

            rgb_im = im.convert('RGB')
            rgb_im.save(imagelocations + '\python2_01_01.jpg')

            filename = imagelocations + '\python2_01_01.jpg'

            im2 = Image.open(imagelocations + "\python2_01_02.png")
            im2 = trim(im2)

            rgb_im2 = im2.convert('RGB')
            rgb_im2.save(imagelocations + '\python2_01_02.jpg')

            filename2 = imagelocations + '\python2_01_02.jpg'

            if bonds_total_items < 2:
                filename2 = ''

            sendToPrinter()

    if finalcourier == 'Sendle':

        instructions = 'Authority to leave'

        # sendle_pickup_suburb = 'Footscray'
        # sendle_pickup_postcode = '3011'
        #
        # if float(weightvalue) <= 0.25:
        #     sendle_pickup_suburb = 'Maidstone'
        #     sendle_pickup_postcode = '3012'

        if float(totalprice) > 150 and volumevalue > 0.0001:
            instructions = 'Signature on Delivery'

        if answer == '3':
            product_code = 'EXPRESS-PICKUP'

        else:
            product_code = 'STANDARD-PICKUP'

        payload = {"description": "Scarlett Music Order " + OrderID, "weight": {"value": weightvalue, "units": "kg"},
                   "volume": {"value": volumevalue, "units": "m3"}, "customer_reference": str(srn), "sender": {
                "contact": {
                    "name": "Scarlett Music",
                    "phone": "(03) 9318 5751",
                    "company": "Scarlett Music"
                },
                "address": {
                    "address_line1": "286-288 Ballarat Rd",
                    "suburb": 'Maidstone',
                    "state_name": "VIC",
                    "postcode": '3012',
                    "country": "Australia"
                },
                "instructions": "The music shop, open 9am-6pm. Best parking is at The Palms across the road."
            }, "receiver": {
                "contact": {
                    "name": name,
                    "email": email,
                    "phone": phone,
                    "company": company
                },
                "address": {
                    "address_line1": address1,
                    "address_line2": address2,
                    "suburb": city,
                    "state_name": state,
                    "postcode": postcode,
                    "country": "Australia"
                },
                "instructions": instructions
            }, "product_code": product_code
                   }

        r = requests.post('https://api.sendle.com/api/orders/',
                          auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'), json=payload)

        # pprint.pprint(r.text)
        # pprint.pprint(r.json())
        response = r.text
        response = json.loads(response)

        # pprint.pprint(response)
        tracking = response['sendle_reference']
        orderurl = response['order_url']
        price = response['price']['gross']['amount']
        pyperclip.copy(tracking)
        print('Tracking number: ' + tracking)
        print('Price: $' + str(price))
        croppedpdfurl = response['labels'][1]['url']

        r = requests.get(croppedpdfurl, auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'),
                         stream=True)

        with open(imagelocations + r"\pythontest2.pdf", "wb") as pdf:
            for chunk in r.iter_content(chunk_size=1024):

                # writing one chunk at a time to pdf file
                if chunk:
                    pdf.write(chunk)

        # r = requests.delete(orderurl, auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'))

        images = convert_from_path(imagelocations + r'\pythontest2.pdf')

        for page in images:
            printer = None
            printer = BrotherQLRaster(printermodel)
            page.save(imagelocations + '\python2.jpg', 'JPEG')

            colorImage = Image.open(imagelocations + '\python2.jpg')

            transposed = colorImage.transpose(Image.ROTATE_90)

            transposed.save(imagelocations + '\python2.jpg')

            image_slicer.slice(imagelocations + '\python2.jpg', 2, 0, 1)

            im = Image.open(imagelocations + "\python2_01_01.png")
            rgb_im = im.convert('RGB')
            rgb_im.save(imagelocations + '\python2_01_01.jpg')

            filename = imagelocations + '\python2_01_01.jpg'

            im2 = Image.open(imagelocations + "\python2_01_02.png")
            rgb_im2 = im2.convert('RGB')
            rgb_im2.save(imagelocations + '\python2_01_02.jpg')

            filename2 = imagelocations + '\python2_01_02.jpg'
            sendToPrinter()

    if finalcourier == 'Toll':

        name = name.replace("'", "")
        name = name.replace("’", "")
        phone = phone.replace(" ", "")
        message_identifier = str(uuid.uuid4())  # getting unique string for toll order
        ###Getting current date / time
        now = datetime.datetime.now()
        current_time = f"{now.strftime('%Y-%m-%d')}T{now.strftime('%H:%M:%S')}+10:00"

        next_day = next_business_day()
        next_toll_day = f"{next_day.strftime('%Y-%m-%d')}T09:00:00+10:00"

        cursor = connection.cursor()
        cursor.execute("SELECT MAX(ShipmentID), MAX(SSCC) FROM Toll;")
        results = cursor.fetchall()
        pprint.pprint(results)
        max_shipment_id = int(results[0][0])
        max_sscc = int(results[0][1])
        environment = "PRD"

        manifest_test_url = "https://api-uat.teamglobalexp.com:6930/gateway/TollMessageManifestRestService/1.0/tom/receiveManifest"
        manifest_prod_url = "https://api.teamglobalexp.com:6930/gateway/TollMessageManifestRestService/1.0/tom/receiveManifest"

        headers = {'Content-Type': 'application/json', 'Accept': 'application/json'}


        def round_up_to_nearest_10(num):
            return math.ceil(num / 10) * 10


        toll_item_array = []
        toll_sql_insert = []

        if address2 == '':
            address2 = address1

        total_toll_information = {
            "ShipmentFinancials": {
                "ExtraServicesAmount": {
                    "Currency": "AUD",
                    "Value": str(round(float(totalprice)))}},
            "ShipmentID": str(max_shipment_id),
            "CreateDateTime": current_time,
            "ConsigneeParty": {
                "PartyName": name,
                "PhysicalAddress": {
                    "AddressLine1": address1,
                    "AddressLine2": address2,

                    "Suburb": suburb,
                    "PostalCode": postcode,
                    "StateCode": state,
                    "CountryName": "Australia",
                    "CountryCode": "AU"
                },
                "Contact": {
                    "Name": name,
                    "Phone": {
                        "Number": phone
                    }
                }
            },

            "BillToParty": {
                "AccountCode": "80119621",
                "PartyName": "SCARLETTMUSIC",
                "PhysicalAddress": {
                    "AddressLine1": "288 Ballarat Rd",

                    "Suburb": "Footscray",
                    "PostalCode": "3011",
                    "StateCode": "VIC",
                    "CountryName": "Australia",
                    "CountryCode": "AU"
                },
                "Contact": {
                    "Name": "Kyal Scarlett",
                    "Phone": {
                        "Number": "0382563460"
                    }
                }
            },
            "ShipmentItemCollection": {
                "ShipmentItem": []
            }
        }

        toll_item_array = {
            "ShipmentFinancials": {
                "ExtraServicesAmount": {
                    "Currency": "AUD",
                    "Value": str(round(float(totalprice)))}},
            "ShipmentID": str(max_shipment_id),
            "CreateDateTime": current_time,
            "ConsigneeParty": {
                "PartyName": name,
                "PhysicalAddress": {
                    "AddressLine1": address1,
                    "AddressLine2": address2,

                    "Suburb": suburb,
                    "PostalCode": postcode,
                    "StateCode": state,
                    "CountryName": "Australia",
                    "CountryCode": "AU"
                },
                "Contact": {
                    "Name": name,
                    "Phone": {
                        "Number": phone
                    }
                }
            },

            "BillToParty": {
                "AccountCode": "80119621",
                "PartyName": "SCARLETTMUSIC",
                "PhysicalAddress": {
                    "AddressLine1": "288 Ballarat Rd",

                    "Suburb": "Footscray",
                    "PostalCode": "3011",
                    "StateCode": "VIC",
                    "CountryName": "Australia",
                    "CountryCode": "AU"
                },
                "Contact": {
                    "Name": "Kyal Scarlett",
                    "Phone": {
                        "Number": "0382563460"
                    }
                }
            },
            "ShipmentItemCollection": {
                "ShipmentItem": []
            }
        }

        for all_items in multi_items_to_ship:

            sscc_array = []

            for _ in range(int(all_items[6])):
                max_shipment_id += 1
                max_sscc += 1

                sscc = str(max_sscc)
                sscc_equation_1 = (int(sscc[16]) + int(sscc[14]) + int(sscc[12]) + int(sscc[10]) + int(sscc[8]) + int(
                    sscc[6]) + int(sscc[4]) + int(sscc[2]) + int(sscc[0])) * 3
                sscc_equation_2 = int(sscc[15]) + int(sscc[13]) + int(sscc[11]) + int(sscc[9]) + int(sscc[7]) + int(
                    sscc[5]) + int(sscc[3]) + int(sscc[1])

                sscc_equation_3 = sscc_equation_1 + sscc_equation_2

                sscc_equation_4 = round_up_to_nearest_10(sscc_equation_3)

                sscc_check_digit = sscc_equation_4 - sscc_equation_3

                sscc = '00' + sscc + str(sscc_check_digit)

                sscc_array.append({
                    "Value": sscc,
                    "SchemeName": "SSCC"
                })

            toll_item_array['ShipmentItemCollection']['ShipmentItem'].append(
                {
                    "IDs": {
                        "ID": sscc_array
                    },

                    "ShipmentItemTotals": {
                        "ShipmentItemCount": str(all_items[6])
                    },
                    "ShipmentService": {
                        "ServiceCode": "X",
                        "ServiceDescription": "ROAD EXPRESS",
                        "ShipmentProductCode": "1"
                    },
                    "Description": "Carton",
                    "Dimensions": {
                        "Volume": str(all_items[4]),
                        "Weight": str(round(float(all_items[0])))
                    },
                    "References": {
                        "Reference": [
                            {
                                "ReferenceType": "ConsignorItemReference",
                                "ReferenceValue": str(srn)
                            },
                            {
                                "ReferenceType": "ConsigneeItemReference",
                                "ReferenceValue": str(srn)
                            }
                        ]
                    }
                })
            ########################## Need more information in print API, so making seperate list
            total_toll_information['ShipmentItemCollection']['ShipmentItem'].append({
                "IDs": {
                    "ID": sscc_array
                },

                "ShipmentItemTotals": {
                    "ShipmentItemCount": str(all_items[6])
                },
                "ShipmentService": {
                    "ServiceCode": "X",
                    "ServiceDescription": "ROAD EXPRESS",
                    "ShipmentProductCode": "1"
                },
                "Description": "Carton",
                "Dimensions": {
                    "Volume": str(all_items[4]),
                    "Weight": str(round(float(all_items[0]), 1)),
                    "Length": str(math.ceil(float(all_items[1]))),
                    "Width": str(math.ceil(float(all_items[3]))),
                    "Height": str(math.ceil(float(all_items[2])))
                },
                "References": {
                    "Reference": [
                        {
                            "ReferenceType": "ConsignorItemReference",
                            "ReferenceValue": str(srn)
                        },
                        {
                            "ReferenceType": "ConsigneeItemReference",
                            "ReferenceValue": str(srn)
                        }
                    ]
                }
            })

            suburb2 = suburb

            suburb2 = suburb2.replace("'", "")
            suburb2 = suburb2.replace("’", "")
            address1 = address1.replace("'", "")
            address1 = address1.replace("’", "")
            address2 = address2.replace("'", "")
            address2 = address2.replace("’", "")

            toll_sql_insert.append(
                f"INSERT INTO Toll(ShipmentID, SSCC, Name, Address1, Address2, Suburb, Postcode, State, Reference, phone_number, item_count, current_sscc, length, width, weight, height, volume) VALUES ('{max_shipment_id}', '{max_sscc}', '{name}', '{address1}', '{address2}', '{suburb2}', '{postcode}', '{state}', '{str(srn)}', '{phone}', '1', '{sscc}', '{str(all_items[1])}', '{str(all_items[3])}', '{str(all_items[3])}', '{str(all_items[2])}', '{str(all_items[4])}'); COMMIT;")

            cursor = connection.cursor()
            cursor.execute(
                f"INSERT INTO Toll(ShipmentID, SSCC, Name, Address1, Address2, Suburb, Postcode, State, Reference, phone_number, item_count, current_sscc, length, width, weight, height, volume) VALUES ('{max_shipment_id}', '{max_sscc}', '{name}', '{address1}', '{address2}', '{suburb2}', '{postcode}', '{state}', '{str(srn)}', '{phone}', '1', '{sscc}', '{str(all_items[1])}', '{str(all_items[3])}', '{str(all_items[3])}', '{str(all_items[2])}', '{str(all_items[4])}'); COMMIT;")

        message_identifier = str(uuid.uuid4())

        test_rate_enquiry_headers = {
            "MessageVersion": "3.1",
            "MessageIdentifier": message_identifier,
            "CreateTimestamp": current_time,
            "DocumentType": "Manifest",
            "Environment": environment,
            "MessageSender": "SCARLETTMUSIC",
            "MessageReceiver": "TOLL",
            'SourceSystemCode': 'XH56'
        }

        toll_message = {"@version": "3.1",
                        "@encoding": "utf-8",
                        "TollMessage": {"Header": test_rate_enquiry_headers,
                                        "Manifest": {
                                            "BusinessID": "IPEC",
                                            "CreateDateTime": current_time,
                                            "DatePeriodCollection": {
                                                "DatePeriod": [{
                                                    "DateType": "DespatchDate",
                                                    "DateTime": next_toll_day
                                                }]},

                                            "ConsignorParty": {
                                                "PartyName": "Scarlett Music",
                                                "PhysicalAddress": {
                                                    "AddressLine1": "288 Ballarat Rd",

                                                    "Suburb": "FOOTSCRAY",
                                                    "PostalCode": "3011",
                                                    "StateCode": "VIC",
                                                    "CountryName": "Australia",
                                                    "CountryCode": "AU"
                                                },
                                                "Contact": {
                                                    "Name": "Michael Demetre",
                                                    "Phone": {
                                                        "Number": "1800688586"
                                                    }
                                                }
                                            },
                                            "ShipmentCollection": {
                                                "Shipment": [toll_item_array]}}}}

        message_identifier = str(uuid.uuid4())

        r = requests.post(url=manifest_prod_url, auth=('accounts@scarlettmusic.com.au', 't2TrAPsNTB'),
                          json=toll_message, headers=headers)

        response = r.text

        response = json.loads(response)

        try:

            if str(response['TollMessage']['ResponseMessages']['ResponseMessage'][0]['ResponseID'][
                       'Value']) == '200':
                # for shipment_lines in payload[]
                pass

            else:
                print(response['TollMessage']['ErrorMessages']['ErrorMessage'][0]['ErrorMessage'])
                input('see above error and press enter')
        except:
            print(response)

        ####Will need to change environment in prod

        ############## ABOVE IS THE CREATE SHIPMENT API, BELOW IS THE PRINT ORDER API ############

        # message_identifier = str(uuid.uuid4())  # getting unique string for toll order
        ###Getting current date / time
        now = datetime.datetime.now()
        current_time = now.strftime("%Y-%m-%dT%H:%M:%SZ")

        next_day = next_business_day()
        next_toll_day = next_day.strftime("%Y-%m-%dT09:00:00Z")

        environment = "PRD"

        headers = {"Content-Type": "application/json",
                   "Accept": "application/json",
                   "x-api-key": "0O87zfzRVu4yf84mxgRiQ2uw9Q86Xb1Z8JrcnJfL",
                   "callId": "SCARLETTMUSIC",
                   "x-mytoll-identity": "727aba8b-0807-4a90-93be-0231d61d4806",
                   "Channel": "CES",
                   "x-mytoll-token": "eyJhbGciOiJSUzUxMiJ9.eyJ1c2VySWQiOiI3MjdhYmE4Yi0wODA3LTRhOTAtOTNiZS0wMjMxZDYxZDQ4MDYiLCJUJlQiOnRydWUsImdlbmVyYXRlRGF0ZSI6IjE3NTAxMzM4Njg3NzEiLCJjdXN0b21OYW1lIjoiMTctMDYtMjVfTXlUZWFtR0VUb2tlbiIsImNlcnRpZmljYXRlTmFtZSI6ImxvY2FsaGNsIiwiQyZDIjp0cnVlLCJ1bmlxdWVJZCI6ImM0NzE3ODI5NmQwZDMzZGNjY2UxZjgyNDA4MWFiNDNhOTkzNjQ0YzEwNjU4YWRlYTZiOGU0Yjk4YWQ0YTMwZmYiLCJleHAiOjE4MTMyMDU4Njh9.f-_21kIPQYnB1EQNvdhj-NA7k8_W0hvTgJUNf-PK3znbHh-eaE_4-AWkNzEsSlGgV64ZUKCoBTvVgPvQ3aWwP96cbmrwGkdvDISVXdUiEfLWsfIOuRBxgNln5e42zujrGqEl93wrBUFdEhwyvf37LNXtwGA9Qu8MRHuDAJsHyBMgr_p-1YbgglUYtRU_vyLd9KqEw0cxKPm3c1VqYxt8mBbbaOSyOV9KN_2bcOPU8SdzxTuH8AbrNjw2L329mQX3D2VPPZ6SlxGu8hKCqMXj4Y4e-ZEpq7S2-QKN6j0mEPouMZyt15Lj2Dn5a1kasqgVrnXx-3bCjSocbbhW49qk2Q"}
        # Need to change x-api-key to 0O87zfzRVu4yf84mxgRiQ2uw9Q86Xb1Z8JrcnJfL in prod
        test_rate_enquiry_headers = {
            "MessageVersion": "1.0",
            "MessageIdentifier": message_identifier,
            "CreateTimestamp": current_time,
            "Environment": environment,
            "MessageSender": "SCARLETTMUSIC",
            "MessageReceiver": "Toll",
            "SourceSystemCode": "XH56"
        }
        manifest_test_url = "https://au-print-sit-apigw.internal.mytoll.com/printDocument"
        manifest_prod_url = "https://au-print-prod-apigw.internal.myteamge.com/printDocument"

        toll_item_count = 0
        total_volume_count = 0
        total_weight_count = 0

        shipment_id = total_toll_information['ShipmentID']
        name = total_toll_information['ConsigneeParty']['PartyName']
        address1 = total_toll_information['ConsigneeParty']['PhysicalAddress']['AddressLine1']
        address2 = total_toll_information['ConsigneeParty']['PhysicalAddress']['AddressLine2']
        suburb = total_toll_information['ConsigneeParty']['PhysicalAddress']['Suburb']
        postcode = total_toll_information['ConsigneeParty']['PhysicalAddress']['PostalCode']
        state = total_toll_information['ConsigneeParty']['PhysicalAddress']['StateCode']
        phone = total_toll_information['ConsigneeParty']['Contact']['Phone']['Number']

        if address_type == 'residential':
            address_type = 'Residential'

        else:
            address_type = 'Business'

        payload = {
            "BusinessID": "IPEC",
            "PrintDocumentType": "Label",
            "PrintSettings": {
                "IsLabelThermal": "false",
                "IsZPLRawResponseRequired": "false",
                "PDF": {
                    "IsPDFA4": "false",
                    "PDFSettings": {
                        "StartQuadrant": "1"
                    }
                }
            },
            "ConsignorParty": {
                "Contact": {
                    "Name": "Kyal",
                    "Phone": {
                        "Number": "61422747033"
                    }
                },
                "PartyName": "Scarlett Music",
                "PhysicalAddress": {
                    "AddressLine1": "288 Ballarat Rd",
                    "AddressType": "Business",
                    "CountryCode": "AU",
                    "PostalCode": "3011",
                    "StateCode": "VIC",
                    "Suburb": "FOOTSCRAY"
                }
            },
            "CreateDateTime": current_time,
            "ShipmentCollection": {
                "Shipment": [
                    {
                        "BillToParty": {
                            "AccountCode": "80119621"
                        },
                        "ConsigneeParty": {

                            "Contact": {
                                "Name": "Kyal",
                                "Phone": {
                                    "Number": "0382563460"
                                }
                            },
                            "PartyName": name,
                            "PhysicalAddress": {
                                "AddressType": address_type,  ##Business / residential
                                "AddressLine1": address1,
                                "AddressLine2": address2,

                                "Suburb": suburb,
                                "PostalCode": postcode,
                                "StateCode": state,
                                "CountryName": "Australia",
                                "CountryCode": "AU"
                            }
                        },
                        "CreateDateTime": current_time,
                        "DatePeriodCollection": {
                            "DatePeriod": [
                                {
                                    "DateTime": next_toll_day,
                                    "DateType": "DespatchDate"
                                }
                            ]
                        },
                        "Orders": {
                            "Order": [
                                {}
                            ]
                        },
                        "References": {
                            "Reference": [
                                {
                                    "ReferenceType": "ShipmentReference1",
                                    "ReferenceValue": "ShipmentReference1"
                                }
                            ]
                        },
                        "ShipmentID": shipment_id,
                        "ShipmentItemCollection": {
                            "ShipmentItem": []
                        },
                        "ShipmentTotals": {
                            "MiscellaneousItemCount": 0,
                            "Volume": {
                                "UOM": "m3",
                                "Value": 1
                            },
                            "Weight": {
                                "UOM": "kg",
                                "Value": 1
                            }
                        },

                    }
                ]
            }
        }

        for t in total_toll_information['ShipmentItemCollection']["ShipmentItem"]:
            message_identifier = str(uuid.uuid4())

            sscc = t['IDs']['ID']
            item_count = t['ShipmentItemTotals']['ShipmentItemCount']
            volume = t['Dimensions']['Volume']
            weight = round(float(t['Dimensions']['Weight']))
            length = t['Dimensions']['Length']
            width = t['Dimensions']['Width']
            height = t['Dimensions']['Height']

            total_volume = float(volume) * float(item_count)
            total_weight = float(weight) * float(item_count)

            toll_item_count += int(item_count)
            total_volume_count += float(total_volume)
            total_weight_count += float(total_weight)

            payload['ShipmentCollection']['Shipment'][0]['ShipmentItemCollection']['ShipmentItem'].append({
                "Commodity": {
                    "CommodityCode": "Z",
                    "CommodityDescription": "ALL FREIGHT"
                },
                "Description": "Item- Carton",
                "Dimensions": {
                    "Height": math.ceil(float(height)),
                    "HeightUOM": "cm3",
                    "Length": math.ceil(float(length)),
                    "LengthUOM": "cm3",
                    "Volume": round(float(volume), 4),
                    "VolumeUOM": "m3",
                    "Weight": round(float(weight)),
                    "WeightUOM": "kg",
                    "Width": math.ceil(float(width)),
                    "WidthUOM": "cm3"
                },
                "IDs": {
                    "ID": sscc
                },
                "References": {
                    "Reference": [
                        {
                            "ReferenceType": "ConsignorItemReference",
                            "ReferenceValue": "Consignor-123"
                        }
                    ]
                },
                "ShipmentItemTotals": {
                    "MiscellaneousItemQuantity": 0,
                    "ShipmentItemCount": str(item_count)
                },
                "ShipmentService": {
                    "ServiceCode": "X",
                    "ShipmentProductCode": "1"
                }
            })

        payload['ShipmentCollection']['Shipment'][0]['ShipmentTotals']['Volume']['Value'] = total_volume_count

        payload['ShipmentCollection']['Shipment'][0]['ShipmentTotals']['Weight']['Value'] = str(
            round(float(total_weight_count), 1))
        toll_message = {"TollMessage": {
            "Header": test_rate_enquiry_headers,
            "Print": payload}}

        # print(toll_message)
        cert_file_path = r"\\SERVER\Python\TOLL\my_client.cert"
        key_file_path = r"\\SERVER\Python\TOLL\my_client.key"
        cert = (cert_file_path, key_file_path)

        r = requests.post(url=manifest_prod_url, auth=("accounts@scarlettmusic.com.au", "t2TrAPsNTB"),
                          json=toll_message, headers=headers, cert=cert)
        # r = requests.post(url=manifest_test_url, json=toll_message, headers = headers)
        # print(curlify.to_curl(r.request))
        response = r.text

        response = json.loads(response)
        pdf = response['TollMessage']['ResponseMessages']['ResponseMessage'][0]['ResponseMessage']
        print(pdf)
        pdf = base64.b64decode(pdf)

        pyperclip.copy(shipment_id)  ###Tracking number

        with open(imagelocations + "\pythontest2.pdf", 'wb') as f:
            f.write(pdf)
        ######^^^^^^^ WILL ONLY SPIT OUT PDF IN QUADRANT 1 ^^^^^^^^

        ######vvvvvvvv CROPPING PDF TO FULL SIZE vvvvvvvvvvvvvv

        images = convert_from_path(imagelocations + r'\pythontest2.pdf')

        for page in images:
            printer = None
            printer = BrotherQLRaster(printermodel)
            page.save(imagelocations + '\python2.jpg', 'JPEG')

            im = Image.open(imagelocations + '\python2.jpg')

            # Size of the image in pixels (size of original image)
            # (This is not mandatory)
            # width, height = im.size

            # Setting the points for cropped image

            # Cropped image of above dimension
            # (It will not change original image)
            im1 = im.crop((0, 0, 900, 1200))
            im1.save(imagelocations + '\python2.jpg', 'JPEG')

            colorImage = Image.open(imagelocations + '\python2.jpg')

            transposed = colorImage.transpose(Image.ROTATE_90)

            transposed.save(imagelocations + '\python2.jpg')

            image_slicer.slice(imagelocations + '\python2.jpg', 2, 0, 1)

            im = Image.open(imagelocations + "\python2_01_01.png")
            rgb_im = im.convert('RGB')
            rgb_im.save(imagelocations + '\python2_01_01.jpg')

            filename = imagelocations + '\python2_01_01.jpg'

            im2 = Image.open(imagelocations + "\python2_01_02.png")
            rgb_im2 = im2.convert('RGB')
            rgb_im2.save(imagelocations + '\python2_01_02.jpg')

            filename2 = imagelocations + '\python2_01_02.jpg'
            sendToPrinter()

        #### TOLL PRINTING HAS FINISHED

        ### BELOW IS FOR BOOKING A TOLL PICKUP

        message_identifier = str(uuid.uuid4())

        now = datetime.datetime.now()
        current_time = now.strftime("%Y-%m-%dT%H:%M:%SZ")

        next_day = next_business_day()
        next_toll_day = next_day.strftime("%Y-%m-%dT09:00:00Z")
        business_close_time = next_day.strftime("%Y-%m-%dT16:00:00Z")

        test_booking_url = 'https://api-uat.teamglobalexp.com:6930/gateway/TollMessageBookingRestService/1.0/tom/bookingRequest'
        production_booking_url = 'https://api.teamglobalexp.com:6930/gateway/TollMessageBookingRestService/1.0/tom/bookingRequest'

        environment = "PRD"

        if address_type == 'residential':
            address_type = 'Residential'

        else:
            address_type = 'Business'

        for t in total_toll_information['ShipmentItemCollection']["ShipmentItem"]:
            message_identifier = str(uuid.uuid4())

            sscc = t['IDs']['ID']
            item_count = t['ShipmentItemTotals']['ShipmentItemCount']
            volume = t['Dimensions']['Volume']
            weight = round(float(t['Dimensions']['Weight']))
            length = t['Dimensions']['Length']
            width = t['Dimensions']['Width']
            height = t['Dimensions']['Height']

            total_volume = float(volume) * float(item_count)
            total_weight = float(weight) * float(item_count)

            toll_item_count += int(item_count)
            total_volume_count += float(total_volume)
            total_weight_count += float(total_weight)

            payload = {
                "@version": "1.0",
                "@encoding": "utf-8",
                "TollMessage": {
                    "Header": {
                        "MessageVersion": "1.0",
                        "MessageIdentifier": message_identifier,
                        "CreateTimestamp": current_time,
                        "DocumentType": "Booking",
                        "Environment": environment,
                        "SourceSystemCode": "XH56",
                        "MessageSender": "SCARLETTMUSIC",
                        "MessageReceiver": "TOLL"
                    },
                    "Bookings": {
                        "Booking": [
                            {
                                "BusinessID": "IPEC",
                                "BookingParty": {
                                    "PartyName": "Scarlett Music",
                                    "AccountCode": "80119621",
                                    "PhysicalAddress": {
                                        "AddressType": "Business",
                                        "AddressLine1": "286-288 Ballarat Rd",
                                        "AddressLine2": "Scarlett Music",
                                        "Suburb": "Footscray",
                                        "PostalCode": "3011",
                                        "StateCode": "VIC",
                                        "CountryCode": "AU"
                                    },
                                    "Contact": {
                                        "Name": "Kyal",
                                        "Phone": {
                                            "CountryCode": "+61",
                                            "AreaCode": "03",
                                            "Number": "82563460"
                                        },
                                        "EMail": "kyal@scarlettmusic.com.au"
                                    }
                                },
                                "SystemFields": {
                                    "RecordCreateDateTime": current_time,
                                    "PickupDateTime": next_toll_day,
                                    "OpenDateTime": next_toll_day,
                                    "CloseDateTime": business_close_time
                                },
                                "References": {
                                    "Reference": [
                                        {
                                            "ReferenceType": "PickupLocation",
                                            "ReferenceValue": "Footscray"
                                        }
                                    ]
                                },
                                "Contact": {
                                    "Name": "Kyal Scarlett",
                                    "Phone": {
                                        "CountryCode": "+61",
                                        "AreaCode": "03",
                                        "Number": "82563460"
                                    },
                                    "EMail": "kyal@scarlettmusic.com.au",
                                    "Note": ""
                                },
                                "BookingFlags": {
                                    "BookingInstruction": "The music shop, open 9am-6pm. Best parking is at The Palms across the road.",
                                    "SameLocationFlag": "true",
                                },
                                "BookingItems": {
                                    "BookingItem": [
                                        {
                                            "ShipmentService": {
                                                "ServiceCode": "X",
                                                "ServiceDescription": "Road Express"
                                            },
                                            "ItemQuantity": str(item_count),  #####Plug item details in here
                                            "Description": "Parcel",
                                            "ItemTypeCode": "BXCA",
                                            "ItemTypeName": "Box/Carton",
                                            "Dimensions": {
                                                "Length": str(math.ceil(float(length))),
                                                "LengthUOM": "CMT",
                                                "Width": str(math.ceil(float(width))),
                                                "WidthUOM": "CMT",
                                                "Height": str(math.ceil(float(height))),
                                                "HeightUOM": "CMT",
                                                "Volume": str(round(float(volume), 3)),
                                                "VolumeUOM": "MTQ",
                                                "Weight": str(round(float(weight))),
                                                "WeightUOM": "KGM"
                                            },
                                            "BookingFinancials": {
                                                "ChargeCode": "Sender",
                                                "AccountCode": "80119621"
                                            },
                                            "DeliveryParty": {
                                                "PhysicalAddress": {
                                                    "AddressType": address_type,

                                                    "AddressLine1": address1,
                                                    #####Plug delivery address details in here
                                                    "Suburb": suburb,
                                                    "PostalCode": str(postcode),
                                                    "StateCode": state,
                                                    "CountryCode": "AU"
                                                }
                                            },
                                            "DatePeriodCollection": {
                                                "DatePeriod": [
                                                    {
                                                        "DateType": "OrderDate",
                                                        "DateTime": current_time
                                                    }
                                                ]
                                            },
                                            "BookingItemFlags": {
                                                "DangerousGoodsFlag": "false"
                                            }
                                        }
                                    ]
                                }
                            }
                        ]
                    }
                }
            }

            r = requests.post(url=production_booking_url, auth=("accounts@scarlettmusic.com.au", "t2TrAPsNTB"),
                              json=payload)

            response = r.text

            response = json.loads(response)

            pprint.pprint(response)

    if finalcourier == 'Mailplus':

        headers = {'x-api-key': 'CSKNIYW1oa5R0fXhG7gzmxwlxW2zunE8tAT4doW6', 'Content-Type': 'application/json',
                   'Accept': 'application/json'}

        mailplus_service = 'standard'

        if answer == str(3):
            mailplus_service = 'express'

        payload = {
            "service": mailplus_service,
            "receiver": {
                "contact": {
                    "name": name,
                    "company": company,
                    "email": email,
                    "phone_number": phone,
                    "instructions": ""
                },
                "address": {
                    "address_line1": mailplus_address,
                    "suburb": city,
                    "postcode": postcode,
                    "state": state
                }
            },
            "sender": {
                "contact": {
                    "name": "Kyal Scarlett",
                    "company": "Scarlett Music",
                    "instructions": "The music shop, open 9am-6pm. Best parking is at The Palms across the road.",
                    "email": "kyal@scarlettmusic.com.au",
                    "phone_number": "0382563460"
                },
                "address": {
                    "address_line1": "286-288 Ballarat Rd",
                    "address_line2": "",
                    "suburb": "Footscray",
                    "postcode": "3011",
                    "state": "VIC"
                }
            },
            "volume": {
                "units": "m3",
                "value": str(volumevalue)
            },
            "weight": {
                "units": "kg",
                "value": str(weightvalue)
            },
            "dimension": {
                "length": {
                    "units": "cm",
                    "value": str(length)
                },
                "width": {
                    "units": "cm",
                    "value": str(width)
                },
                "height": {
                    "units": "cm",
                    "value": str(height)
                }
            }
        }

        while True:

            try:

                r = requests.post('https://papi.mailplus.com.au/api/job',
                                  headers=headers,
                                  json=payload)

                if r.status_code != 200:
                    time.sleep(2)
                    response = r.text
                    response = json.loads(response)
                    continue
                else:
                    break

            except Exception as e:

                print('Error found. \n\n')

                print(repr(e))

                input('\n\n Press Enter to die')

                continue

        response = r.text
        response = json.loads(response)

        # pprint.pprint(response)

        reference_id = response['reference_id']

        barcode = response['barcode']  ##Possibly toll tracking numbeer ?

        tracking_number = response['connote']
        pyperclip.copy(tracking_number)
        label_pdf = response['label_url']

        headers = {'x-api-key': 'CSKNIYW1oa5R0fXhG7gzmxwlxW2zunE8tAT4doW6', 'Content-Type': 'application/json',
                   'Accept': 'application/pdf'}

        while True:

            r = requests.get(f'https://papi.mailplus.com.au/api/pdf_label?reference_id={reference_id}',
                             headers=headers, stream=True, timeout=30)

            if r.status_code != 200:
                time.sleep(2)
                continue
            else:
                try:

                    response = r.text
                    # response = json.loads(response)
                    # pprint.pprint(response)

                    pdf = r.content

                    # pdf = base64.b64encode(pdf)
                    with open(imagelocations + "\pythontest2.pdf", 'wb') as f:
                        f.write(pdf)

                    images = convert_from_path(imagelocations + r'\pythontest2.pdf')
                    break

                except:
                    continue

        for page in images:
            printer = None
            printer = BrotherQLRaster(printermodel)
            page.save(imagelocations + '\python2.jpg', 'JPEG')

            colorImage = Image.open(imagelocations + '\python2.jpg')

            transposed = colorImage.transpose(Image.ROTATE_90)

            transposed.save(imagelocations + '\python2.jpg')

            image_slicer.slice(imagelocations + '\python2.jpg', 2, 0, 1)

            im = Image.open(imagelocations + "\python2_01_01.png")
            rgb_im = im.convert('RGB')
            rgb_im.save(imagelocations + '\python2_01_01.jpg')

            filename = imagelocations + '\python2_01_01.jpg'

            im2 = Image.open(imagelocations + "\python2_01_02.png")
            rgb_im2 = im2.convert('RGB')
            rgb_im2.save(imagelocations + '\python2_01_02.jpg')

            filename2 = imagelocations + '\python2_01_02.jpg'
            sendToPrinter()

    if finalcourier == 'Australia Post':
        headers = {'Content-Type': 'application/json', 'Accept': 'application/json', 'Account-Number': accountnumber}

        payload = {
            "shipments": [
                {
                    "shipment_reference": OrderID,
                    "customer_reference_1": OrderID,
                    "customer_reference_2": "SKU-1",
                    "email_tracking_enabled": 'true',
                    "from": {
                        "name": "Lorelle Scarlett",
                        "business_name": "Scarlett Music",
                        "lines": [
                            "268-288 Ballarat Rd"
                        ],
                        "suburb": "FOOTSCRAY",
                        "state": "VIC",
                        "postcode": "3011",
                        "phone": "0422747033",
                        "email": "lorelle@scarlettmusic.com.au"
                    },
                    "to": {
                        "name": name,
                        "business_name": company[0:40],
                        "lines": [
                            address1[0:50], address2[0:50]
                        ],
                        "suburb": city,
                        "state": state,
                        "postcode": postcode,
                        "phone": phone,
                        "email": email
                    },
                    "items": aus_post_item_array
                }
            ]
        }

        r = requests.post('https://digitalapi.auspost.com.au/shipping/v1/shipments', headers=headers,
                          auth=HTTPBasicAuth(username, secret),
                          json=payload)

        response = r.text
        response = json.loads(response)

        if 'errors' in response:
            payload = {
                "shipments": [
                    {
                        "shipment_reference": OrderID,
                        "customer_reference_1": OrderID,
                        "customer_reference_2": "SKU-1",
                        "email_tracking_enabled": 'true',
                        "from": {
                            "name": "Lorelle Scarlett",
                            "business_name": "Scarlett Music",
                            "lines": [
                                "268-288 Ballarat Rd"
                            ],
                            "suburb": "FOOTSCRAY",
                            "state": "VIC",
                            "postcode": "3011",
                            "phone": "0422747033",
                            "email": "lorelle@scarlettmusic.com.au"
                        },
                        "to": {
                            "name": name,
                            "business_name": company,
                            "lines": [
                                address1, address2
                            ],
                            "suburb": suburb,
                            "state": state,
                            "postcode": postcode,
                            "phone": phone,
                            "email": email
                        },
                        "items": aus_post_item_array
                    }
                ]
            }

            r = requests.post('https://digitalapi.auspost.com.au/shipping/v1/shipments', headers=headers,
                              auth=HTTPBasicAuth(username, secret),
                              json=payload)

            response = r.text
            response = json.loads(response)

        pprint.pprint(response)
        itemID_array = []
        for x in response['shipments'][0]['items']:
            itemID_array.append({"item_id": x['item_id']})
        # itemID = response['shipments'][0]['items'][0]['item_id']
        shipmentID = response['shipments'][0]['shipment_id']
        totalcost = response['shipments'][0]['shipment_summary']['total_cost']
        ausposttracking = response['shipments'][0]['items'][0]['tracking_details']['consignment_id']
        pyperclip.copy(ausposttracking)

        payload = {
            "wait_for_label_url": 'true',
            "preferences": [
                {
                    "type": "PRINT",
                    "format": "PDF",
                    "groups": [
                        {
                            "group": "Parcel Post",
                            "layout": "THERMAL-LABEL-A6-1PP",
                            "branded": 'true',
                            "left_offset": 0,
                            "top_offset": 0
                        },
                        {
                            "group": "Express Post",
                            "layout": "THERMAL-LABEL-A6-1PP",
                            "branded": 'false',
                            "left_offset": 0,
                            "top_offset": 0
                        }
                    ]
                }
            ],
            "shipments": [
                {
                    "shipment_id": shipmentID,
                    "items": itemID_array
                }
            ]
        }

        r = requests.post('https://digitalapi.auspost.com.au/shipping/v1/labels', headers=headers,
                          auth=HTTPBasicAuth(username, secret),
                          json=payload)

        response = r.text
        response = json.loads(response)
        # pprint.pprint(response)

        url = response['labels'][0]['url']
        print(url)

        r = requests.get(url, stream=True)

        with open(imagelocations + "\pythontest2.pdf", "wb") as pdf:
            pdf.write(r.content)

        if answer == "3":
            images = convert_from_path(imagelocations + r'\pythontest2.pdf')

            for page in images:
                printer = None
                printer = BrotherQLRaster(printermodel)
                page.save(imagelocations + '\python2.jpg', 'JPEG')

                colorImage = Image.open(imagelocations + '\python2.jpg')

                # --- Start of new code to stretch the image ---

                # Get the original width and height
                width, height = colorImage.size

                # Define the new height (e.g., 50% taller)
                # You can change the 1.5 to any value greater than 1 to make it taller
                new_height = int(height * 1.5)

                # Resize the image to the new height, keeping the width the same
                stretched_image = colorImage.resize((width, new_height))

                # --- End of new code ---

                # Now, continue with your rotation logic on the stretched image
                transposed = stretched_image

                transposed.save(imagelocations + '\python2.jpg')

                filename = imagelocations + '\python2.jpg'
                filename2 = ''
                sendToPrinter()
        else:

            images = convert_from_path(imagelocations + r'\pythontest2.pdf')

            for page in images:
                printer = None
                printer = BrotherQLRaster(printermodel)
                page.save(imagelocations + '\python2.jpg', 'JPEG')

                colorImage = Image.open(imagelocations + '\python2.jpg')

                transposed = colorImage.transpose(Image.ROTATE_90)

                transposed.save(imagelocations + '\python2.jpg')

                image_slicer.slice(imagelocations + '\python2.jpg', 2, 0, 1)

                im = Image.open(imagelocations + "\python2_01_01.png")
                rgb_im = im.convert('RGB')
                rgb_im.save(imagelocations + '\python2_01_01.jpg')

                filename = imagelocations + '\python2_01_01.jpg'

                im2 = Image.open(imagelocations + "\python2_01_02.png")
                rgb_im2 = im2.convert('RGB')
                rgb_im2.save(imagelocations + '\python2_01_02.jpg')

                filename2 = imagelocations + '\python2_01_02.jpg'
                sendToPrinter()

    if finalcourier == 'Dai Post':
        cursor = connection.cursor()
        cursor.execute("SELECT MAX(job_number) FROM dai_post;")
        results = cursor.fetchall()
        pprint.pprint(results)
        max_job_id = int(results[0][0])

        name = name.replace("'", "")
        name = name.replace("’", "")
        ###Getting current date / time
        now = datetime.datetime.now()
        current_time = f"{now.strftime('%Y-%m-%d')}T{now.strftime('%H:%M:%S')}"

        signature_required = '0'

        if round(float(totalprice)) > 200:
            signature_required = '1'

        try:
            suburb
        except:
            suburb = city
        while True:
            try:
                payload = {
                    "shipment": {
                        "service": "Parcel Right",
                        "labelformat": "PDF",
                        "account": "SCA",
                        "datetime": current_time,
                        "reference": f'{OrderID} {name}',
                        "jobnumber": max_job_id,
                        "signature": signature_required,
                        "value": str(round(float(totalprice))),
                        "currency": "AUD",
                        "uom": "kg",
                        "weight": final_dai_weight,
                        "originterminal": "TME",
                        "shipper": {
                            "name": "Scarlett Music",
                            "attention": "Kyal",
                            "addr1": "286-288 Ballarat Rd",
                            "addr2": "",
                            "city": "Footscray",
                            "state": "VIC",
                            "country": "AU",
                            "postal": "3011",
                            "phone": "+61 03 8256 3460",
                            "email": "kyal@scarlettmusic.com.au"
                        },
                        "consignee": {
                            "name": name,
                            "attention": name,
                            "addr1": address1,
                            "addr2": address2,
                            "city": suburb,
                            "state": state,
                            "country": "AU",
                            "postal": postcode,
                            "phone": phone,
                            "email": email
                        },
                        "item": [
                            {
                                "description": f"Scarlett Music Order {OrderID}",
                                "qty": "1",
                                "unit": "pc",
                                "value": str(round(float(totalprice)))
                            }
                        ]
                    }
                }

                r = requests.post('https://daiglobaltrack.com/prod/serviceconnect',
                                  auth=HTTPBasicAuth('ScarlettMusic', 'D5es4stu!'), json=payload)

                response = r.text

                dai_response = json.loads(response)
                print(dai_response)

                dai_tracking_number = dai_response['shipmentresponse']['tracknbr']
                break
            except:
                OrderID = OrderID + "-1"
                continue

        dai_pdf = dai_response['shipmentresponse']['label']

        dai_pdf = base64.b64decode(dai_pdf)
        with open(imagelocations + "\pythontest2.pdf", 'wb') as f:
            f.write(dai_pdf)

        images = convert_from_path(imagelocations + r'\pythontest2.pdf')

        cursor = connection.cursor()
        cursor.execute(
            fr"INSERT INTO dai_post(customer_name, tracking_number, time_created, job_number) VALUES ('{name}', '{dai_tracking_number}', '{now}', '{max_job_id}'); COMMIT;")

        pyperclip.copy(dai_tracking_number)

        for page in images:
            printer = None
            printer = BrotherQLRaster(printermodel)
            page.save(imagelocations + '\python2.jpg', 'JPEG')

            im = Image.open(imagelocations + '\python2.jpg')

            top_bott_margin = 150
            l_r_margin = 10
            new_height = im.height + (2 * top_bott_margin)
            new_width = im.width + (2 * l_r_margin)
            new_image = Image.new('RGB', (new_width, 1250), (255, 255, 255))
            new_image.paste(im, (0, 0))

            new_image.save(imagelocations + '\python2.jpg', 'JPEG')

            colorImage = Image.open(imagelocations + '\python2.jpg')

            transposed = colorImage.transpose(Image.ROTATE_90)

            transposed.save(imagelocations + '\python2.jpg')

            image_slicer.slice(imagelocations + '\python2.jpg', 2, 0, 1)

            im = Image.open(imagelocations + "\python2_01_01.png")
            rgb_im = im.convert('RGB')
            rgb_im.save(imagelocations + '\python2_01_01.jpg')

            filename = imagelocations + '\python2_01_01.jpg'

            im2 = Image.open(imagelocations + "\python2_01_02.png")
            rgb_im2 = im2.convert('RGB')
            rgb_im2.save(imagelocations + '\python2_01_02.jpg')

            filename2 = imagelocations + '\python2_01_02.jpg'
            sendToPrinter()

    if finalcourier == 'Freightster':

        payload = {"order": {"serviceCode": 12,
                             "consignee": {"company": company,
                                           "name": name,
                                           "address1": address1,
                                           "address2": address2,
                                           "city": city,
                                           "postcode": postcode,
                                           "state": state,
                                           "phone": phone,
                                           "email": email},
                             "sender": {"name": "Kyal Scarlett",
                                        "address1": "286-288 Ballarat Rd",
                                        "address2": "",
                                        "city": "Footscray",
                                        "postcode": "3011",
                                        "state": "VIC",
                                        "phone": "0382563460",
                                        "email": "kyal@scarlettmusic.com.au"},
                             "shipment": {"reference": OrderID,
                                          "description": OrderID,
                                          "weight": str(final_freightster_weight)}}}

        r = requests.post('https://freightster.com.au/api/v1/shippingAPI/create', json=payload,
                          headers=freightster_headers)
        freightster_response = json.loads(r.text)
        # pprint.pprint(freightster_response)

        freightster_orderid = freightster_response['response_data']['order_id']
        freightster_trackingnumber = freightster_response['response_data']['tracking_number']
        pyperclip.copy(freightster_trackingnumber)

        # vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv payload for printing label vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv

        payload = {"order": {"orderIds": [freightster_orderid]}}

        r = requests.post('https://freightster.com.au/api/v1/shippingAPI/print', json=payload,
                          headers=freightster_headers)
        freightster_response = json.loads(r.text)

        freightster_pdf = freightster_response['response_data']['labels'][0]['label']

        # pprint.pprint(freightster_pdf)

        freightster_pdf = base64.b64decode(freightster_pdf)
        with open(imagelocations + "\pythontest2.pdf", 'wb') as f:
            f.write(freightster_pdf)

        images = convert_from_path(imagelocations + r'\pythontest2.pdf')
        name = name.replace("'", "")
        cursor = connection.cursor()
        cursor.execute(
            fr"INSERT INTO Freightster(order_id, tracking_number, name) VALUES ('{freightster_orderid}', '{freightster_trackingnumber}', '{name}'); COMMIT;")

        for page in images:
            printer = None
            printer = BrotherQLRaster(printermodel)
            page.save(imagelocations + '\python2.jpg', 'JPEG')

            colorImage = Image.open(imagelocations + '\python2.jpg')

            transposed = colorImage.transpose(Image.ROTATE_90)

            transposed.save(imagelocations + '\python2.jpg')

            image_slicer.slice(imagelocations + '\python2.jpg', 2, 0, 1)

            im = Image.open(imagelocations + "\python2_01_01.png")
            rgb_im = im.convert('RGB')
            rgb_im.save(imagelocations + '\python2_01_01.jpg')

            filename = imagelocations + '\python2_01_01.jpg'

            im2 = Image.open(imagelocations + "\python2_01_02.png")
            rgb_im2 = im2.convert('RGB')
            rgb_im2.save(imagelocations + '\python2_01_02.jpg')

            filename2 = imagelocations + '\python2_01_02.jpg'
            sendToPrinter()

    if finalcourier == 'Transdirect':

        transaddress = ''

        Final_Transdirect_Courier = lowesttranscourier

        if address2 == '':
            transaddress = address1
        else:
            transaddress = address1 + ', ' + address2

        if google == 'true':
            transaddress = backup_google_street_address

        headers = {'Api-key': '74a242a1f4b8ba6ea5e011a80bdd2732', 'Content-Type': 'application/json'}

        if answer == '1' or answer == '4':

            if Final_Transdirect_Courier == 'Couriers Please Multi 21' or Final_Transdirect_Courier == 'Fastway Multi 7':

                if Final_Transdirect_Courier == 'Couriers Please Multi 21':
                    courier = 'couriers_please_domestic_proirity_authority'
                    tier = 2
                elif Final_Transdirect_Courier == 'Fastway Multi 7':
                    courier = 'fastway'
                    tier = 2

                # print(tomorrow)
                payload = {
                    "courier": courier,
                    "pickup-date": str(next_day),
                    "tier": tier
                }
                headers = {'Api-key': '74a242a1f4b8ba6ea5e011a80bdd2732', 'Content-Type': 'application/json'}

                r = requests.post('https://www.transdirect.com.au/api/bookings/v4/' + str(id) + '/confirm',
                                  auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'), headers=headers,
                                  json=payload)
                trans_response = r.text
                # print(response)
                if trans_response == '':

                    input(
                        'Courier successfully booked for ' + courier + ' please put item in picking area for next batch of packing.')

                    sys.exit()

                    pass
                elif trans_response == '{"errors":["Pickup dates invalid."]}':
                    tomorrow = next_business_day()
                    while trans_response == '{"errors":["Pickup dates invalid."]}':
                        tomorrow = tomorrow + datetime.timedelta(days=1)
                        #  print(tomorrow)

                        payload = {"courier": str(courier), "pickup-date": str(tomorrow), "tier": tier}
                        r = requests.post('https://www.transdirect.com.au/api/bookings/v4/' + str(id) + '/confirm',
                                          auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'),
                                          headers=headers,
                                          json=payload)
                        trans_response = r.text
                        # print(trans_response)
                        if trans_response == '':
                            # print('Success!')

                            input(
                                Final_Transdirect_Courier + ''' successfully booked. Please put item in picking area for next batch of packing.

                                Press Enter to Exit''')

                            sys.exit()

                        continue

        if 'tiers' in trans_response['quotes']['couriers_please_domestic_proirity_authority']:
            quotes['couriers_please_multi_21'] = '1000'

        if 'fastway' in trans_response['quotes']:

            if 'tiers' in trans_response['quotes']['fastway']:
                quotes['fastway_multi_7'] = '1000'

        intquotes = dict(
            (k, float(v)) for k, v in quotes.items())  ### Converting all values into FLOAT and then into INT
        lowesttranscourier = min(intquotes, key=intquotes.get)

        # lowesttranscourier = lowesttranscourier.replace('_', ' ')
        # lowesttranscourier = lowesttranscourier.title()
        # lowesttranscourier = lowesttranscourier.replace('Tnt', 'TNT')

        # payload = {
        #     "declared_value": "0",
        #     "referrer": "API",
        #     "requesting_site": "www.scarlettmusic.com.au",
        #     "tailgate_pickup": "false",
        #     "tailgate_delivery": str(tailgate),
        #     "items": [
        #         {
        #             "weight": str(weightvalue),
        #             "height": str(height),
        #             "width": str(width),
        #             "length": str(length),
        #             "quantity": 1,
        #             "description": "carton"
        #         }
        #     ],
        #     "sender": {
        #         "address": "288 Ballarat Rd",
        #         "company_name": "Scarlett Music",
        #         "email": "info@scarlettmusic.com.au",
        #         "name": "Lorelle Scarlett",
        #         "postcode": "3011",
        #         "phone": "0417557472",
        #         "state": "VIC",
        #         "suburb": "FOOTSCRAY",
        #         "type": "business",
        #         "country": "AU"
        #     },
        #     "receiver": {
        #         "address": str(transaddress),
        #         "company_name": str(company),
        #         "email": str(email),
        #         "name": str(name),
        #         "postcode": str(postcode),
        #         "phone": int(phone),
        #         "state": str(state),
        #         "suburb": str(city),
        #         "type": "residential",
        #         "country": "AU"
        #     }
        # }

        # print(tomorrow)
        payload = {
            "courier": str(lowesttranscourier),
            "pickup-date": str(next_day)
        }

        r = requests.post('https://www.transdirect.com.au/api/bookings/v4/' + str(id) + '/confirm',
                          auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'), headers=headers, json=payload)
        trans_response = r.text
        # print(response)
        if trans_response == '':
            print('Success!')
        elif trans_response == '{"errors":["Pickup dates invalid."]}':
            tomorrow = next_business_day()

            while trans_response == '{"errors":["Pickup dates invalid."]}':
                tomorrow = tomorrow + datetime.timedelta(days=1)
                print(tomorrow)

                payload = {"courier": str(lowesttranscourier), "pickup-date": str(tomorrow)}
                r = requests.post('https://www.transdirect.com.au/api/bookings/v4/' + str(id) + '/confirm',
                                  auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'), headers=headers,
                                  json=payload)
                trans_response = r.text
                # print(trans_response)
                if trans_response == '':
                    print('Success!')
                    break
                continue

        headers = {'Api-key': '74a242a1f4b8ba6ea5e011a80bdd2732', 'Content-Type': 'application/pdf',
                   'Content-Disposition': 'attachment; filename="invoice.pdf"'
                   }
        label = 'https://www.transdirect.com.au/api/bookings/v4/' + str(id) + '/a6label'
        print(label)
        r = requests.get(label, auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'), headers=headers)
        # print(r.text)

        print('Printing Transdirect Label. Please wait 2-7 business winters.')

        while True:
            if 'Label not ready' in r.text:
                r = requests.get(label, auth=HTTPBasicAuth('lorelle@scarlettmusic.com.au', 'redd8484'), headers=headers)
                time.sleep(1)

            else:
                break

        with open(imagelocations + r"\pythontest2.pdf", 'wb') as f:
            f.write(r.content)

        images = convert_from_path(imagelocations + r'\pythontest2.pdf')

        for page in images:
            printer = None
            printer = BrotherQLRaster(printermodel)
            page.save(imagelocations + '\python2.jpg', 'JPEG')

            colorImage = Image.open(imagelocations + '\python2.jpg')

            transposed = colorImage.transpose(Image.ROTATE_90)

            transposed.save(imagelocations + '\python2.jpg')

            image_slicer.slice(imagelocations + '\python2.jpg', 2, 0, 1)

            im = Image.open(imagelocations + "\python2_01_01.png")
            rgb_im = im.convert('RGB')
            rgb_im.save(imagelocations + '\python2_01_01.jpg')

            filename = imagelocations + '\python2_01_01.jpg'

            im2 = Image.open(imagelocations + "\python2_01_02.png")
            rgb_im2 = im2.convert('RGB')
            rgb_im2.save(imagelocations + '\python2_01_02.jpg')

            filename2 = imagelocations + '\python2_01_02.jpg'
            sendToPrinter()

        print(
            '\n\n\nPlease wait, copying tracking number to clipboard. This may take a second,  sure beats reading though.')

        options = Options()
        options.headless = True

        driver = webdriver.Chrome(options=options)
        driver.get('https://www.transdirect.com.au/education/tracking/?reference=' + str(id))

        try:
            element = WebDriverWait(driver, 10).until(
                EC.text_to_be_present_in_element((By.ID, "modal-content"), 'Delivery Details'))

        except TimeoutException:
            pass

        finally:

            soup = BeautifulSoup(driver.page_source, 'html.parser')
            driver.quit()
            soup = soup.find(class_='modal-content')
            soup = soup.find(class_='col-lg-6').text.strip()

            mo = re.findall(r'Courier Reference:.*', soup)

            for x in mo:
                tracking = x

        tracking = tracking.replace('Courier Reference:', '').strip()

        print('Tracking number copied to clipboard: ' + tracking)

        time.sleep(2)

    if finalcourier == 'Fastway':

        while True:

            r = requests.post(base_url + '/api/consignments', headers=Fastway_Headers, json=fastway_data)
            response = r.text
            response = json.loads(response)

            if 'errors' in response:
                continue

            else:
                break
        # pprint.pprint(response)

        id = response['data']['conId']  ########  THIS IS WHERE THE SCRIPT FUCKS UP IF LABELS ARE WONK
        tracking_number = response['data']['items'][0]['label']
        pyperclip.copy(str(tracking_number))

        r = requests.get(base_url + '/api/consignments/' + str(id) + '/labels?pageSize=4x6', headers=Fastway_Headers)
        response = r.content

        with open(imagelocations + '\pythontest2.pdf', 'wb') as f:
            f.write(response)

        images = convert_from_path(imagelocations + r'\pythontest2.pdf')

        for page in images:
            printer = None
            printer = BrotherQLRaster(printermodel)
            page.save(imagelocations + '\python2.jpg', 'JPEG')

            colorImage = Image.open(imagelocations + '\python2.jpg')

            transposed = colorImage.transpose(Image.ROTATE_90)

            transposed.save(imagelocations + '\python2.jpg')

            image_slicer.slice(imagelocations + '\python2.jpg', 2, 0, 1)

            im = Image.open(imagelocations + "\python2_01_01.png")
            rgb_im = im.convert('RGB')
            rgb_im.save(imagelocations + '\python2_01_01.jpg')

            filename = imagelocations + '\python2_01_01.jpg'

            im2 = Image.open(imagelocations + "\python2_01_02.png")
            rgb_im2 = im2.convert('RGB')
            rgb_im2.save(imagelocations + '\python2_01_02.jpg')

            filename2 = imagelocations + '\python2_01_02.jpg'
            sendToPrinter()

    if finalcourier == 'Allied Express':

        today = datetime.date.today()
        if today.isoweekday() in set((6, 7)):
            today += datetime.timedelta(days=today.isoweekday() % 5)
        next_day_allied = str(today.day) + '/' + str(today.month) + '/' + str(today.year) + " 10:00:00"

        history = HistoryPlugin()
        session = Session()
        transport = Transport(session=session)
        wsdl = 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS?wsdl'

        try:
            allied_client = zeep.Client(wsdl=wsdl, transport=transport, plugins=[history])

            allied_client.transport.session.proxies = {
                # Utilize for all http/https connections
                'http': 'http://neptune.alliedexpress.com.au:8080/ttws-ejb/TTWS', }
            allied_account = allied_client.service.getAccountDefaults('755cf13abb3934695f03bd4a75cfbca7', "SCAMUS",
                                                                      "VIC",
                                                                      "AOE")

        except:
            allied_failed = 'true'

        with allied_client.settings(strict=False):
            allied_client.service.savePendingJob('755cf13abb3934695f03bd4a75cfbca7', Job)
            dispatch_jobs = allied_client.service.dispatchPendingJobs('755cf13abb3934695f03bd4a75cfbca7', JobIDs)
            xml = etree.tostring(history.last_received["envelope"], encoding="unicode")
            xml = xmltodict.parse(xml)

        connote_number = xml['soapenv:Envelope']['soapenv:Body']['ns1:dispatchPendingJobsResponse']['result']['item'][
            'docketNumber']
        reference = xml['soapenv:Envelope']['soapenv:Body']['ns1:dispatchPendingJobsResponse']['result']['item'][
            'referenceNumbers']

        pyperclip.copy(connote_number)

        pdf = allied_client.service.getLabel('755cf13abb3934695f03bd4a75cfbca7', "AOE", connote_number, reference,
                                             '3011', 1)
        xml = etree.tostring(history.last_received["envelope"], encoding="unicode")
        xml = xmltodict.parse(xml)

        pdf = xml['soapenv:Envelope']['soapenv:Body']['ns1:getLabelResponse']['result']
        pdf = base64.b64decode(pdf)

        with open(imagelocations + "\pythontest2.pdf", 'wb') as f:
            f.write(pdf)

        images = convert_from_path(imagelocations + r'\pythontest2.pdf')

        for page in images:
            printer = None
            printer = BrotherQLRaster(printermodel)
            page.save(imagelocations + '\python2.jpg', 'JPEG')

            colorImage = Image.open(imagelocations + '\python2.jpg')

            transposed = colorImage.transpose(Image.ROTATE_90)

            transposed.save(imagelocations + '\python2.jpg')

            image_slicer.slice(imagelocations + '\python2.jpg', 2, 0, 1)

            im = Image.open(imagelocations + "\python2_01_01.png")
            rgb_im = im.convert('RGB')
            rgb_im.save(imagelocations + '\python2_01_01.jpg')

            filename = imagelocations + '\python2_01_01.jpg'

            im2 = Image.open(imagelocations + "\python2_01_02.png")
            rgb_im2 = im2.convert('RGB')
            rgb_im2.save(imagelocations + '\python2_01_02.jpg')

            filename2 = imagelocations + '\python2_01_02.jpg'
            sendToPrinter()

    if finalcourier == 'Couriers Please':
        cp_url = 'https://api.couriersplease.com.au/v2/domestic/shipment/create'

        r = requests.post(cp_url, headers=cp_headers, json=cp_validate_body)
        response = r.text
        response = json.loads(response)
        # pprint.pprint(response)

        cp_tracking_number = response['data']['consignmentCode']
        pyperclip.copy(cp_tracking_number)

        ###########BELOW IS CREATING LABEL FOR ORDER########

        cp_url = f'https://api.couriersplease.com.au/v1/domestic/shipment/label?consignmentNumber={cp_tracking_number}'
        r = requests.get(cp_url, headers=cp_headers, json=cp_validate_body)
        response = r.text
        response = json.loads(response)
        pprint.pprint(response)

        base64_label = response['data']['label']
        pdf = base64.b64decode(base64_label)

        with open(imagelocations + "\pythontest2.pdf", 'wb') as f:
            f.write(pdf)

        images = convert_from_path(imagelocations + r'\pythontest2.pdf')

        for page in images:
            printer = None
            printer = BrotherQLRaster(printermodel)
            page.save(imagelocations + '\python2.jpg', 'JPEG')

            colorImage = Image.open(imagelocations + '\python2.jpg')

            transposed = colorImage.transpose(Image.ROTATE_90)

            transposed.save(imagelocations + '\python2.jpg')

            img = Image.open(imagelocations + '\python2.jpg')

            border = (63, 83, 220, 85)  # left, top, right, bottom
            cropped_img = ImageOps.crop(img, border)
            cropped_img.save(imagelocations + '\python2.jpg')

            image_slicer.slice(imagelocations + '\python2.jpg', 2, 0, 1)

            im = Image.open(imagelocations + "\python2_01_01.png")
            rgb_im = im.convert('RGB')
            rgb_im.save(imagelocations + '\python2_01_01.jpg')

            filename = imagelocations + '\python2_01_01.jpg'

            im2 = Image.open(imagelocations + "\python2_01_02.png")
            rgb_im2 = im2.convert('RGB')
            rgb_im2.save(imagelocations + '\python2_01_02.jpg')

            filename2 = imagelocations + '\python2_01_02.jpg'
            sendToPrinter()

            cp_url = "https://api.couriersplease.com.au/v2/domestic/bookPickup"

            total_weight = 0

            for x in cp_item_array:
                weight = int(x['physicalWeight'])
                total_weight += weight

            total_weight = math.ceil(total_weight)

            address1 = address1[0:18]
            address2 = address2[0:18]

            cp_body = {
                "accountName": "Scarlett Music",
                "contactName": "Kyal Scarlett",
                "contactEmail": "kyal@scarlettmusic.com.au",
                "readyDateTime": next_cp_day,
                "specialInstructions": "Use front door",
                "consignmentCount": "1",
                "consignmentCode": cp_tracking_number,
                "totalItemCount": str(len(cp_item_array)),
                "totalWeight": str(total_weight),
                "pickup": {
                    "phoneNumber": "0393185751",
                    "companyName": "Scarlett Music",
                    "address1": "286-288",
                    "address2": "Ballarat Rd",
                    "address3": "",
                    "postcode": "3011",
                    "suburb": "Footscray"
                },
                "delivery": {
                    "companyName": company,
                    "address1": address1,
                    "address2": address2,
                    "postcode": postcode,
                    "suburb": city
                }
            }
            r = requests.post(cp_url, headers=cp_headers, json=cp_body)
            response = r.text
            response = json.loads(response)
            pprint.pprint(response)

if answer == str(5):

    while True:
        courier = input('''\n\n\n\nWhich courier do you need to delete? (Enter number 1-2)


        1) Australia Post
        2) Fastway
        3) Allied Express
        4) Couriers Please
        5) Freightster
        6) Toll
        7) CBD Express
        8) Dai Post
        9) Bonds Transpost


        Enter Response: ''')

        if str(courier) == '1':
            courier = 'Australia Post'
            break

        elif str(courier) == '2':
            courier = 'Fastway'
            break

        elif str(courier) == '3':
            courier = 'Allied'
            break

        elif str(courier) == '4':
            courier = 'Couriers Please'
            break

        elif str(courier) == '5':
            courier = 'Freightster'
            break

        elif str(courier) == '6':
            courier = 'Toll'
            break

        elif str(courier) == '7':
            courier = 'CBD Express'
            break

        elif str(courier) == '8':
            courier = 'Dai Post'
            break

        elif str(courier) == '9':
            courier = 'Bonds'
            break

        else:
            print('Valid input not found. Please try again.')
            time.sleep(2)
            continue

    if courier == 'Australia Post':

        headers = {'Content-Type': 'application/json', 'Accept': 'application/json', 'Account-Number': accountnumber}
        account = accountnumber

        r = requests.get(
            'https://digitalapi.auspost.com.au/shipping/v1/shipments?offset=0&number_of_shipments=1000&status=Created',
            headers=headers, auth=HTTPBasicAuth(username, secret), timeout=100)

        response = r.text
        response = json.loads(response)
        # pprint.pprint(response)
        totalrecords = response['pagination']['total_number_of_records']

        print('-----Current Unmanifested Orders-----\n')

        id = {}
        for x in range(0, int(totalrecords)):
            id["{0}".format(x)] = response['shipments'][int(x)]['shipment_id']
            creationdate = response['shipments'][int(x)]['shipment_creation_date']
            date = re.search(r"\d\d\d\d-\d\d-\d\d", creationdate)
            time = re.search(r'\d\d:\d\d:\d\d', creationdate)
            print(str(x + 1) + ') ' + response['shipments'][int(x)]['to'][
                'name'] + ', Created: ' + time.group() + ' ' + date.group() + ', (Reference: ' +
                  response['shipments'][int(x)]['shipment_reference'] + ')')

        item_delete = input('''
        Which shipment would you like to delete? Enter number here:''')

        answer = input('Are you sure you want to delete this order? This is final. (Y/N)')

        if answer.lower() == 'n':
            sys.exit()
        elif answer.lower() == 'y':
            pass
        else:
            sys.exit()

        print('Deleting shipment. Please wait...')

        r = requests.delete(
            'https://digitalapi.auspost.com.au/shipping/v1/shipments/' + str(id[str(int(item_delete) - 1)]),
            headers=headers, auth=HTTPBasicAuth(username, secret))

        response = r.text

        print('Success!')

        sleep(2.1)

        if courier == 'CBD Express':
            cursor = connection.cursor()
            cursor.execute(
                "SELECT order_id, consignment_number, name, time_created FROM CDBExpress WHERE name NOT IN ('DELETED');")
            results = cursor.fetchall()
            # pprint.pprint(results)
            if results == []:
                input('No pending CBD orders')
            total_orders = len(results)
            # print(total_orders)

            print('-----Current Unmanifested CBD Express Orders-----\n')

            for x in range(int(total_orders)):
                print(
                    f'{x + 1}) SCARLET{results[x][1]} (Name: {results[x][2]}) Time: {results[x][3].replace("T", " ")}')

            item_delete = input('''

                    Which shipment would you like to delete? Enter number here:''')

            order_id = latest_SRN = results[int(item_delete) - 1][1]

            consignment_number = f"SCARLET{order_id}"

            answer = input('Are you sure you want to delete this order? This is final. (Y/N)')

            if answer.lower() == 'n':
                sys.exit()
            elif answer.lower() == 'y':
                pass
            else:
                sys.exit()

            print('Deleting shipment. Please wait...')

            reciever_email = 'kyal@scarlettmusic.com.au'

            subject = f'Delete Consignment {consignment_number}'

            body = f'''Hey Jarrod,

    Could we please cancel consignment {consignment_number} please?

    Sorry about that!

    Cheers,
    Kyal'''

            path_to_documents = ''
            m = Mail()
            m.send_email(reciever_email, subject, body, path_to_documents, attachment_file)

            cursor = connection.cursor()
            cursor.execute(
                f"UPDATE CDBExpress SET name = 'DELETED' WHERE consignment_number = '{order_id}'; COMMIT;")
            print('Success!')
            sleep(2.1)

    if courier == 'Toll':

        cursor = connection.cursor()
        cursor.execute(
            "SELECT Reference, ShipmentID, Name FROM toll WHERE Name IS NOT NULL AND manifest_number is NULL AND height is NOT NULL;")
        results = cursor.fetchall()
        # pprint.pprint(results)
        if results == []:
            input('No pending Toll orders')
        total_orders = len(results)
        # print(total_orders)

        print('-----Current Unmanifested Toll Orders-----\n')

        for x in range(int(total_orders)):
            print(f'{x + 1}) {results[x][1]} (Name: {results[x][2]})')

        item_delete = input('''

        Which shipment would you like to delete? Enter number here:''')

        order_id = latest_SRN = results[int(item_delete) - 1][1]

        answer = input('Are you sure you want to delete this order? This is final. (Y/N)')

        if answer.lower() == 'n':
            sys.exit()
        elif answer.lower() == 'y':
            pass
        else:
            sys.exit()

        print('Deleting shipment. Please wait...')

        cursor = connection.cursor()
        cursor.execute(
            f"UPDATE toll SET manifest_number = 'DELETED' WHERE ShipmentID = '{order_id}'; COMMIT;")
        print('Success!')
        sleep(2.1)

    if courier == 'Bonds':
        cursor = connection.cursor()
        cursor.execute(
            f"SELECT time_created, job_number, customer_name FROM bonds;")
        results = cursor.fetchall()

        if results == []:
            input('No pending Bonds orders')
        total_orders = len(results)

        print('-----Current Unmanifested Bonds Orders-----\n')

        for x in range(int(total_orders)):
            print(f'{x + 1}) {results[x][1]} (Name: {results[x][2]}), Time created: {results[x][0]}')

        item_delete = input('''

                        Which shipment would you like to delete? Enter number here:''')

        tracking_number = results[int(item_delete) - 1][1]

        answer = input('Are you sure you want to delete this order? This is final. (Y/N)')

        if answer.lower() == 'n':
            sys.exit()
        elif answer.lower() == 'y':
            pass
        else:
            sys.exit()

        print('Deleting shipment. Please wait...')

        xml_payload = f'''
        <job xmlns:xi="http://www.w3.org/2001/XInclude" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="job-bonds-delete.xsd">
        <job_action>DELETE</job_action>
        <delete_job_reason>No longer needed</delete_job_reason>
        <job_id>{tracking_number}</job_id>
        <account>V01523</account>
        <authorization_code>@WV6mSH4NByW</authorization_code>
        </job>
        '''

        url = 'https://appsrv.bondscouriers.com.au/bondsweb/api/upload-xml-job.htm'  # Replace with the actual API endpoint URL
        headers = {
            'Content-Type': 'application/xml'
        }

        response = requests.post(url, data=xml_payload, headers=headers)
        print(response.status_code)

        xml_response = response.text
        # Parse the XML and convert it to a Python dictionary
        data_dict = xmltodict.parse(xml_response)
        pprint.pprint(data_dict)


        cursor.execute(f"DELETE FROM bonds WHERE job_number = '{tracking_number}'; COMMIT;")
        input('Does the above mention an error? If not press Enter.')
        print('Success!')

    if courier == 'Dai Post':

        cursor = connection.cursor()
        cursor.execute("SELECT MAX(job_number) FROM dai_post;")
        results = cursor.fetchall()
        pprint.pprint(results)
        max_job_id = int(results[0][0])
        cursor.execute(
            f"SELECT time_created, tracking_number, customer_name FROM dai_post WHERE customer_name IS NOT NULL AND job_number = '{str(max_job_id)}';")
        # cursor.execute(
        #     f"SELECT time_created, tracking_number, customer_name, job_number FROM dai_post WHERE customer_name IS NOT NULL;")
        results = cursor.fetchall()
        # pprint.pprint(results)
        if results == []:
            input('No pending Dai orders')
        total_orders = len(results)

        print('-----Current Unmanifested Dai Orders-----\n')

        for x in range(int(total_orders)):
            print(f'{x + 1}) {results[x][1]} (Name: {results[x][2]}), Time created: {results[x][0]}')

        item_delete = input('''

                Which shipment would you like to delete? Enter number here:''')

        tracking_number = tracking_number = results[int(item_delete) - 1][1]

        answer = input('Are you sure you want to delete this order? This is final. (Y/N)')

        if answer.lower() == 'n':
            sys.exit()
        elif answer.lower() == 'y':
            pass
        else:
            sys.exit()

        print('Deleting shipment. Please wait...')

        payload = {
            "cancelshipment": {'tracknbr': str(tracking_number)}
        }

        r = requests.post('https://daiglobaltrack.com/prod/serviceconnect',
                          auth=HTTPBasicAuth('ScarlettMusic', 'D5es4stu!'), json=payload)

        response = r.text

        dai_response = json.loads(response)

        print(dai_response)

        cursor = connection.cursor()
        cursor.execute(f"DELETE FROM dai_post WHERE tracking_number = '{tracking_number}'; COMMIT;")

        print('Success!')

        sleep(2.1)

    if courier == 'Freightster':

        cursor = connection.cursor()
        cursor.execute("SELECT order_id, tracking_number, name FROM Freightster;")
        results = cursor.fetchall()
        # pprint.pprint(results)
        if results == []:
            input('No pending Freightster orders')
        total_orders = len(results)
        # print(total_orders)

        print('-----Current Unmanifested Freightster Orders-----\n')

        for x in range(int(total_orders)):
            print(f'{x + 1}) {results[x][1]} (Name: {results[x][2]})')

        item_delete = input('''

        Which shipment would you like to delete? Enter number here:''')

        order_id = latest_SRN = results[int(item_delete) - 1][0]

        answer = input('Are you sure you want to delete this order? This is final. (Y/N)')

        if answer.lower() == 'n':
            sys.exit()
        elif answer.lower() == 'y':
            pass
        else:
            sys.exit()

        print('Deleting shipment. Please wait...')

        cursor = connection.cursor()
        cursor.execute(f"DELETE FROM Freightster WHERE order_id = '{order_id}'; COMMIT;")

        print('Success!')

        sleep(2.1)

    if courier == 'Fastway':
        r = requests.get(base_url + '/api/consignments/pending', headers=Fastway_Headers)
        response = r.text
        response = json.loads(response)
        # pprint.pprint(response)

        number_of_records = len(response['data'])
        # pprint.pprint(response)
        # print(number_of_records)
        print('-----Current Unmanifested Orders-----\n')
        id = {}
        for x in range(0, int(number_of_records)):
            id["{0}".format(x)] = response['data'][int(x)]['conId']
            creationdate = response['data'][int(x)]['createdOn']
            date = re.search(r"\d\d\d\d-\d\d-\d\d", creationdate)
            time = re.search(r'\d\d:\d\d:\d\d', creationdate)
            print(str(x + 1) + ') ' + response['data'][int(x)][
                'toContactName'] + ', Created: ' + time.group() + ' ' + date.group())

        item_delete = input('''
            Which shipment would you like to delete? Enter number here:''')

        answer = input('Are you sure you want to delete this order? This is final. (Y/N)')

        if answer.lower() == 'n':
            sys.exit()
        elif answer.lower() == 'y':
            pass
        else:
            sys.exit()

        print('Deleting shipment. Please wait...')
        r = requests.delete(base_url + '/api/consignments/' + str(id[str(int(item_delete) - 1)]) + '/reason/Error',
                            headers=Fastway_Headers)

        response = r.text

        print('Success!')

        sleep(2.1)

    if courier == 'Allied':
        tracking_number = input("""What is the tracking number? (Don't include the \"-001\" at the end)

        Enter Response:""")

        suburb = input("""What is the 4 digit destination suburb?

        Enter Response:""")

        cancel = allied_client.service.cancelDispatchJob('755cf13abb3934695f03bd4a75cfbca7', tracking_number, suburb)

        if str(cancel) == '0':
            print('Success')
            time.sleep(2)
            sys.exit()

        elif str(cancel) == '-6':
            print('This has already been cancelled dingus')
            time.sleep(2)
            sys.exit()

        else:
            print("Fuck bro not sure if this cancelled, maybe tell Kyal?")
            time.sleep(2)
            sys.exit()

    if courier == 'Couriers Please':
        cp_auth = '113153878:CBBC1CAD2F335C9CEEA1D0BC63C7056ADBC9C36DD54A9A4530A4380D9AAC5FE0'
        ### This is for sandbox, when ready, replace 2nd half with CBBC1CAD2F335C9CEEA1D0BC63C7056ADBC9C36DD54A9A4530A4380D9AAC5FE0

        cp_auth_encoded = cp_auth.encode()
        cp_auth_encoded = base64.b64encode(cp_auth_encoded)
        cp_auth_encoded = cp_auth_encoded.decode()

        cp_headers = {'Host': 'api.couriersplease.com.au',
                      'Accept': 'application/json',
                      'Content-Type': 'application/json',
                      'Authorization': f'Basic {cp_auth_encoded}',
                      'Content-Length': '462'}

        cp_url = 'https://api.couriersplease.com.au/v1/domestic/shipment/cancel'
        consignment_number = input('What is the consignment number?')
        cp_body = {"consignmentCode": consignment_number}

        r = requests.post(cp_url, headers=cp_headers, json=cp_body)
        response = r.text
        response = json.loads(response)
        pprint.pprint(response)
        time.sleep(2)

if answer == str(6):

    while True:
        courier = input('''\n\n\n\nWhich courier do you want to manifest? (Enter number 1-2)


        1) Australia Post
        2) Toll

        Enter Response: ''')

        if str(courier) == '1':
            courier = 'Australia Post'
            break

        elif str(courier) == '2':
            courier = 'Toll'
            break

        else:
            print('Valid input not found. Please try again.')
            time.sleep(2)
            continue

    manifestanswer = input('Are you sure you want to manifest all pending orders? This is final. (Y/N)')

    if manifestanswer.lower() == 'n':
        sys.exit()
    elif manifestanswer.lower() == 'y':
        pass
    else:
        sys.exit()

    if courier == 'Australia Post':

        headers = {'Content-Type': 'application/json', 'Accept': 'application/json', 'Account-Number': accountnumber}
        account = accountnumber

        r = requests.get(
            'https://digitalapi.auspost.com.au/shipping/v1/shipments?offset=0&number_of_shipments=1000&status=Created',
            headers=headers, auth=HTTPBasicAuth(username, secret))

        response = r.text
        response = json.loads(response)
        # pprint.pprint(response)

        id = ''
        totalrecords = response['pagination']['total_number_of_records']

        # id = response['shipments'][0]['shipment_id']
        for x in range(0, int(totalrecords)):
            id = id + '{"shipment_id": "' + response['shipments'][int(x)]['shipment_id'] + '"}, '
            if response['shipments'][int(x)]['shipment_id'] == 'hBcK0E8cg_wAAAGFFmkCQzH7':
                print('Bing')
                print('bong')

        id = '[' + str(id[:-2]) + ']'
        id = json.loads(id)
        # print(id)

        payload = {
            "order_reference": "My order reference",
            "payment_method": "CHARGE_TO_ACCOUNT",
            "shipments": id
        }

        r = requests.put('https://digitalapi.auspost.com.au/shipping/v1/orders', headers=headers,
                         auth=HTTPBasicAuth(username, secret), json=(payload))

        createorderresponse = r.text
        createorderresponse = json.loads(createorderresponse)
        # pprint.pprint(createorderresponse)

        order_id = createorderresponse['order']['order_id']
        # print(order_id)

        headers = {"User-Agent": "PostmanRuntime/7.20.1",
                   "Accept": "*/*",
                   "Cache-Control": "no-cache",
                   "Postman-Token": "8eb5df70-4da6-4ba1-a9dd-e68880316cd9,30ac79fa-969b-4a24-8035-26ad1a2650e1",
                   "Host": "digitalapi.auspost.com.au",
                   "Accept-Encoding": "gzip, deflate",
                   "Connection": "keep-alive",
                   "cache-control": "no-cache", 'Account-Number': accountnumber}

        r = requests.get('https://digitalapi.auspost.com.au/shipping/v1/accounts/' + str(account) + '/orders/' + str(
            order_id) + '/summary', headers=headers,
                         auth=HTTPBasicAuth(username, secret), stream=True)
        time = str(datetime.datetime.now())
        time = time.replace('.', '-')
        time = time.replace(' ', '_')
        time = time.replace(':', '-')
        manifestfile = imagelocations + "\Australia_Post_Manifest_" + str(time) + ".pdf"

        with open(manifestfile, "wb") as pdf:
            pdf.write(r.content)
        os.startfile(manifestfile)

    if courier == 'Toll':

        ###Getting current date / time
        now = datetime.datetime.now()
        current_time = f"{now.strftime('%Y-%m-%d')}T{now.strftime('%H:%M:%S')}+10:00"

        next_day = next_business_day()
        next_toll_day = f"{next_day.strftime('%Y-%m-%d')}T09:00:00+10:00"

        environment = 'PRD'

        message_identifier = str(uuid.uuid4())

        manifest_test_url = "https://au-print-sit-apigw.internal.mytoll.com/printDocument"
        manifest_prod_url = "https://au-print-prod-apigw.internal.myteamge.com/printDocument"

        headers = {"Content-Type": "application/json",
                   "Accept": "application/json",
                   "x-api-key": "0O87zfzRVu4yf84mxgRiQ2uw9Q86Xb1Z8JrcnJfL",
                   "callId": "SCARLETTMUSIC",
                   "x-mytoll-identity": "727aba8b-0807-4a90-93be-0231d61d4806",
                   "Channel": "CES",
                   "x-mytoll-token": "eyJhbGciOiJSUzUxMiJ9.eyJ1c2VySWQiOiI3MjdhYmE4Yi0wODA3LTRhOTAtOTNiZS0wMjMxZDYxZDQ4MDYiLCJUJlQiOnRydWUsImdlbmVyYXRlRGF0ZSI6MTY4NjcxNjAxMTM3MSwiY3VzdG9tTmFtZSI6IjE0LTA2LTIzX015VGVhbUdFVG9rZW4iLCJjZXJ0aWZpY2F0ZU5hbWUiOiJsb2NhbGhjbCIsIkMmQyI6dHJ1ZSwidW5pcXVlSWQiOiJjNDcxNzgyOTZkMGQzM2RjY2NlMWY4MjQwODFhYjQzYTk5MzY0NGMxMDY1OGFkZWE2YjhlNGI5OGFkNGEzMGZmIiwiZXhwIjoxNzQ5ODc0NDExfQ.A-gOQU6Pc1_yuFkHTqQ219So4lkeoRI0CxtQrZlAsF9VBgqt085lffV_QRGDBPeogjLL5bae-XloKfPO-Ah23HErGHh_oXw_9CkRg8mcG7tkBZsf8StPPN-6HD1i-9iFioJvRE6d9njkVdePapet1FkBuWVg9WOKp8ft_516XR_pok1JmG_fnA55nDBADMDvUHFPW_YUqaoNbJmLpjf7CV0RGiT4pASilzQ4Ut4cuZ0NxQ3d-bQXBQetL5BxQzYNfANsxRD25icSGmi06alngfIFFxoCqBnuxYs_QCT1BvJHJw5e9LUMnXEGzuNAwx_6baRta7Fjq6UsuQPZ8zU-VA"}
        # manifestanswer = input('Are you sure you want to manifest all pending orders? This is final. (Y/N)')

        # if manifestanswer.lower() == 'n':
        #     sys.exit()
        # elif manifestanswer.lower() == 'y':
        #     pass
        # else:
        #     sys.exit()

        cursor = connection.cursor()

        cursor.execute("SELECT MAX(manifest_number) FROM Toll WHERE manifest_number NOT IN ('DELETED');")
        results = cursor.fetchall()
        manifest_number = int(results[0][0]) + 1

        cursor.execute(
            "SELECT ShipmentID, SSCC, Name, Address1, Address2, Suburb, Postcode, State, Reference, phone_number, item_count, current_sscc, length, width, weight, height, volume FROM toll WHERE Name IS NOT NULL AND manifest_number is NULL AND height is NOT NULL;")

        results = cursor.fetchall()

        if results == []:
            input('No pending Toll orders')

        manifest_orders = []

        for orders in results:
            print(orders)

            shipment_id = orders[0]
            pre_sscc = orders[1]
            name = orders[2]
            address1 = orders[3]
            address2 = orders[4]
            suburb = orders[5]
            postcode = orders[6]
            state = orders[7]
            reference = orders[8]
            phone = orders[9]
            item_count = orders[10]
            sscc = orders[11]
            length = orders[12]
            width = orders[13]
            weight = orders[14]
            height = orders[15]
            volume = orders[16]

            if address2 == '':
                address2 = address1

            manifest_orders.append({
                "BillToParty": {
                    "AccountCode": "80119621"
                },
                "ConsigneeParty": {
                    "ConsigneeID": "null",
                    "Contact": {
                        "Name": name,
                        "Phone": {
                            "CountryCode": "61",
                            "Number": str(phone),
                            "AreaCode": "03"
                        }
                    },
                    "PartyName": name,
                    "PhysicalAddress": {
                        "AddressLine1": address1,
                        "AddressLine2": address2,
                        "CountryCode": "AU",
                        "PostalCode": postcode,
                        "StateCode": state,
                        "Suburb": suburb
                    }
                },
                "CreateDateTime": current_time,
                "FreightMode": "Road",

                "Orders": {
                    "Order": [
                        {}
                    ]
                },
                "References": {
                    "Reference": [
                        {
                            "ReferenceType": "ShipmentReference1",
                            "ReferenceValue": reference
                        },

                    ]
                },

                "ShipmentID": str(shipment_id),
                "ShipmentItemCollection": {
                    "ShipmentItem": [
                        {
                            "Commodity": {
                                "CommodityCode": "Z",
                                "CommodityDescription": "ALL FREIGHT"
                            },
                            "Description": "Item- Carton",
                            "Dimensions": {
                                "Height": round(float(height), 1),
                                "HeightUOM": "cm3",
                                "Length": round(float(length), 1),
                                "LengthUOM": "cm3",
                                "Volume": round(float(volume), 1),
                                "VolumeUOM": "m3",
                                "Weight": round(float(weight)),
                                "WeightUOM": "kg",
                                "Width": round(float(width), 1),
                                "WidthUOM": "cm3"
                            },
                            "IDs": {
                                "ID": [
                                    {
                                        "SchemeName": "SSCC",
                                        "Value": str(sscc)
                                    }
                                ]
                            },

                            "References": {
                                "Reference": [
                                    {
                                        "ReferenceType": "ConsignorItemReference",
                                        "ReferenceValue": str(reference)
                                    }

                                ]
                            },
                            "ShipmentItemTotals": {
                                "MiscellaneousItemQuantity": 0,
                                "ShipmentItemCount": 1
                            },
                            "ShipmentService": {
                                "ServiceCode": "X",
                                "ServiceDescription": "General",
                                "ShipmentProductCode": "1"
                            }
                        }
                    ]
                },
                "ShipmentTotals": {
                    "MiscellaneousItemCount": 0,
                    "Volume": {
                        "UOM": "m3",
                        "Value": round(float(volume), 1)
                    },
                    "Weight": {
                        "UOM": "kg",
                        "Value": str(round(float(weight), 1))
                    }
                }})

        toll_message = {
            "TollMessage": {
                "Header": {
                    "CreateTimestamp": current_time,
                    "Environment": environment,
                    "MessageIdentifier": message_identifier,
                    "MessageReceiver": "TollMessage",
                    "MessageSender": "SCARLETTMUSIC",
                    "MessageVersion": "1.0",
                    "SourceSystemCode": "XH56"
                },
                "Print": {
                    "BusinessID": "IPEC",
                    "PrintDocumentType": "Manifest",
                    "ConsignorParty": {"Contact": {
                        "Name": "Kyal",
                        "Phone": {
                            "Number": "61422747033"
                        }
                    },
                        "PartyName": "Scarlett Music",
                        "PhysicalAddress": {
                            "AddressLine1": "288 Ballarat Rd",
                            "AddressType": "Business",
                            "CountryCode": "AU",
                            "PostalCode": "3011",
                            "StateCode": "VIC",
                            "Suburb": "FOOTSCRAY"
                        }},
                    "CreateDateTime": current_time,
                    "ManifestID": {
                        "Value": f"MYT-{str(manifest_number)}"
                    },
                    "DatePeriodCollection": {
                        "DatePeriod": [
                            {
                                "DateTime": next_toll_day,
                                "DateType": "DespatchDate"
                            }
                        ]
                    },
                    "ShipmentCollection": {
                        "Shipment": manifest_orders}}}}

        pprint.pprint(toll_message)

        cert_file_path = r"\\SERVER\Python\TOLL\my_client.cert"
        key_file_path = r"\\SERVER\Python\TOLL\my_client.key"
        cert = (cert_file_path, key_file_path)

        for orders in results:
            cursor.execute(
                f"UPDATE toll SET manifest_number = '{str(manifest_number)}' WHERE ShipmentID = '{orders[0]}'; COMMIT;")

        r = requests.post(url=manifest_prod_url, auth=("accounts@scarlettmusic.com.au", "t2TrAPsNTB"),
                          json=toll_message, headers=headers, cert=cert)

        response = r.text

        response = json.loads(response)

        pprint.pprint(response)

        try:

            if str(response['TollMessage']['ResponseMessages']['ResponseMessage'][0]['ResponseID']['Value']) == '200':
                # for shipment_lines in payload[]
                pass

            else:
                print('Manifest unsuccessful, reverting changes, maybe grab Kyal')
                for orders in results:
                    cursor.execute(f"UPDATE toll SET manifest_number = null WHERE ShipmentID = '{orders[0]}'; COMMIT;")
        except:

            print('Manifest unsuccessful, reverting changes, maybe grab Kyal')

            for orders in results:
                cursor.execute(f"UPDATE toll SET manifest_number = null WHERE ShipmentID = '{orders[0]}'; COMMIT;")

        pdf = response['TollMessage']['ResponseMessages']['ResponseMessage'][0]['ResponseMessage']
        print(pdf)
        pdf = base64.b64decode(pdf)

        time = str(datetime.datetime.now())
        time = time.replace('.', '-')
        time = time.replace(' ', '_')
        time = time.replace(':', '-')
        manifestfile = imagelocations + "\Toll_Manifest_" + str(time) + ".pdf"

        with open(manifestfile, 'wb') as f:
            f.write(pdf)

        os.startfile(manifestfile)

if answer == str(7) or answer == str(14):

    backend = 'pyusb'

    reNeto = re.compile(r'N\d+')
    reEbay = re.compile(r'\d+-\d\d\d\d\d-\d\d\d\d\d')
    OrderID = str(pyperclip.paste())

    if reNeto.match(OrderID):
        data = {'Filter': {'OrderID': OrderID, 'OutputSelector': ['ID', 'Email', 'ShipAddress']}}

        headers = {'Content-Type': 'application/json', 'Accept': 'application/json', 'NETOAPI_ACTION': 'GetOrder',
                   'NETOAPI_USERNAME': 'kyalAPI', 'NETOAPI_KEY': 'HSZ3EyI9ki6sseqRIkwy6tL6yFNPEUyM', }

        r = requests.post('https://www.scarlettmusic.com.au/do/WS/NetoAPI', headers=headers, json=data)

        response = r.text
        response = json.loads(response)

        email = response['Order'][0]['Email']
        orderid = response['Order'][0]['OrderID']
        city = response['Order'][0]['ShipCity']
        try:
            company = response['Order'][0]['ShipCompany']
        except KeyError:
            company = ''
        name = response['Order'][0]['ShipFirstName'] + " " + response['Order'][0]['ShipLastName']

        try:
            phone = response['Order'][0]['ShipPhone']
        except:
            phone = '0417557472'
        postcode = response['Order'][0]['ShipPostCode']
        State = response['Order'][0]['ShipState']

        address1 = response['Order'][0]['ShipStreetLine1']
        try:

            address2 = response['Order'][0]['ShipStreetLine2']
        except KeyError:
            address2 = ''
        srn = orderid

    if reEbay.match(OrderID):
        try:
            api = Trading(appid="KyalScar-Awaiting-PRD-aca833663-f54c920e", timeout=120, config_file=None,
                          devid="5c97c2a8-eae7-49cc-ae81-c3ed1bbab335", certid="PRD-ca8336639ad2-75af-45bc-992a-d0bc",
                          token="v^1.1#i^1#r^1#f^0#I^3#p^3#t^Ul4xMF84OkUzOEU3NzVFQzA3NDUwNjUyMkY0NTc4QjlENjRFRDVFXzFfMSNFXjI2MA==")
            response = api.execute('GetOrders', {"OrderIDArray": {"OrderID": OrderID}})
            #  pprint.pprint(response.dict())
            unformatteddic = response.dict()

            name = unformatteddic['OrderArray']['Order'][0]['ShippingAddress']['Name']
            phone = unformatteddic['OrderArray']['Order'][0]['ShippingAddress']['Phone']
            city = unformatteddic['OrderArray']['Order'][0]['ShippingAddress']['CityName']
            State = unformatteddic['OrderArray']['Order'][0]['ShippingAddress']['StateOrProvince']
            address1 = unformatteddic['OrderArray']['Order'][0]['ShippingAddress']['Street1']
            address2 = unformatteddic['OrderArray']['Order'][0]['ShippingAddress']['Street2']

            if address2 is None:
                address2 = ''

            srn = unformatteddic['OrderArray']['Order'][0]['ShippingDetails'][
                'SellingManagerSalesRecordNumber']
            email = unformatteddic['OrderArray']['Order'][0]['TransactionArray']['Transaction'][0]['Buyer']['Email']
            postcode = unformatteddic['OrderArray']['Order'][0]['ShippingAddress']['PostalCode']
            company = ''

        except ConnectionError as e:
            print(e)
            print(e.response.dict())
            sys.exit()

    try:
        if 'ebay' in address1:
            address = address2 + " " + city + " " + state + " " + postcode
            #   print(srn + " " + city + " " + address)

            get_google_address()
            postcode = postal_code
            city = suburb
        # print(srn + " " + city + " " + address)

        else:
            try:
                address2
            except:
                address2 = ''

            try:
                state
            except:
                state = ''

            address = address1 + " " + address2 + " " + city + " " + state + " " + postcode
            #    print(newSalesRecordNumber + " " + newBuyerCity + " " + address)

            get_google_address()
            postcode = postal_code
            city = suburb
            #    print(newSalesRecordNumber + " " + newBuyerCity + " " + address)
        sendlefailed = ''

    except:
        pass

    if 'po box' in address1.lower() or 'po box' in address2.lower() or 'care po' in address1.lower() or 'care po' in address2.lower() or 'parcel locker' in address1.lower() or 'parcel locker' in address2.lower() or 'p.o' in address1.lower() or 'p.o' in address2.lower() or 'parcel collect' in address1.lower() or 'parcel collect' in address2.lower() or 'pobox' in address1.lower() or 'pobox' in address2.lower() or 'locker' in address1.lower() or 'locker' in address2.lower() or 'collect' in address1.lower() or 'collect' in address2.lower() or 'parcel' in address1.lower() or 'parcel' in address2.lower() or 'pmb' in address1.lower() or 'pmb' in address2.lower() or 'p/o' in address1.lower() or 'p/o' in address2.lower() or 'post office box' in address1.lower() or 'post office box' in address2.lower() or 'lpo' in address1.lower() or 'lpo' in address2.lower() or 'post office' in address1.lower() or 'post office' in address2.lower() or 'australia post' in address1.lower() or 'australia post' in address2.lower() or 'australia post' in company.lower():
        sendlefailed = 'true'
        transdirectfailed = 'true'
        fastwayfailed = 'true'
        allied_failed = 'true'
        couriers_failed = 'true'
        freightsterfailed = 'true'
        mailplusfailed = 'true'
        toll_failed = 'true'
        cbdexpressfailed = 'true'
        bondsfailed = 'true'

    if 'victoria' in State.strip().lower() or 'v.i.c' in State.strip().lower() or 'vic' in State.strip().lower():
        State = 'VIC'
    if 'new south wales' in State.strip().lower() or 'n.s.w' in State.strip().lower() or 'nsw' in State.strip().lower():
        State = 'NSW'
    if 'tasmania' in State.strip().lower() or 't.a.s' in State.strip().lower() or 'tas' in State.strip().lower():
        State = 'TAS'
    if 'queensland' in State.strip().lower() or 'q.l.d' in State.strip().lower() or 'qld' in State.strip().lower():
        State = 'QLD'
    if 'northern territory' in State.strip().lower() or 'n.t' in State.strip().lower() or 'nt' in State.strip().lower():
        State = 'NT'
    if 'western' in State.strip().lower() or 'w.a' in State.strip().lower() or 'wa' in State.strip().lower():
        State = 'WA'
    if 'south australia' in State.strip().lower() or 's.a' in State.strip().lower() or 'sa' in State.strip().lower():
        State = 'SA'
    if 'australian capital' in State.strip().lower() or 'a.c.t' in State.strip().lower() or 'act' in State.strip().lower():
        State = 'ACT'
    if 'tasmania' in State.strip().lower() or 'tas' in State.strip().lower() or 't.a.s' in State.strip().lower():
        State = 'TAS'

    # elif state.lower == 'northern territory':
    #     state = 'NT'

    #################  BEGINNING OF SENDLE API ########################

    # print('\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n')
    auspostfailed = ''
    fastwayfailed = ''

    while True:

        try:
            print('\nCURRENT ORDER - ' + name + '\n')

        except NameError:
            print('No order found. Are you sure you copied the order ID?')
            time.sleep(2)
            sys.exit()

        if answer == str(14):
            break

        if answer == str(7):

            parcelsize = input('''What size is your parcel? (Enter number 1-8)
             1) Bubble Mailer
             2) Satchel
             3) Handbag
             4) Shoebox
             5) Briefcase
             6) Carry-on
             7) Check-in
             8) Unsure
             Enter response here:''')

            if parcelsize == str(1):

                weightvalue = 0.49
                volumevalue = 0.0001
                length = 18
                width = 23
                height = 4
                break

            elif parcelsize == str(2):

                weightvalue = 0.49
                volumevalue = 0.0001
                length = 14
                width = 13
                height = 11
                break

            elif parcelsize == str(3):

                weightvalue = 1
                volumevalue = 0.004
                length = 15.7
                width = 15.9
                height = 16
                break

            elif parcelsize == str(4):

                weightvalue = 3
                volumevalue = 0.012
                length = 25
                width = 20
                height = 24
                break

            elif parcelsize == str(5):

                weightvalue = 5
                volumevalue = 0.020
                length = 27
                width = 27
                height = 27.4
                break

            elif parcelsize == str(6):

                weightvalue = 10
                volumevalue = 0.040
                length = 34
                width = 34
                height = 34.6
                break

            elif parcelsize == str(7):

                weightvalue = 25
                volumevalue = 0.1
                length = 50
                width = 50
                height = 40
                break

            elif parcelsize == str(8):

                weightvalue = float(input('Enter item weight (kg)'))

                length = float(input('Enter item length (cm)'))
                width = float(input('Enter item width (cm)'))
                height = float(input('Enter item height (cm)'))

                volumevalue = (length * width * height) * 0.000001

                if length > 105 or width > 105 or height > 105 or weightvalue > 22:
                    auspostfailed = 'true'

                if (length * width * height) / 1000 > 100 or weightvalue > 25:
                    sendlefailed = 'true'

                break

            else:
                print('Number not found, select a number next time, ya dingus.')
                time.sleep(2)
                answer = input('Did you want to try again? (Y/N)')

                if answer == 'N':
                    sys.exit()
                elif answer == 'Y':
                    continue
                else:
                    sys.exit()

    #### Need to run Aus Post + Sendle post calculators
    finalcourier = 'Australia Post'

    print(finalcourier + ' IS THE WINNER BING BING BING')

    if finalcourier == 'Sendle':
        payload = {"description": "Scarlett Music Order " + OrderID, "weight": {"value": weightvalue, "units": "kg"},
                   "volume": {"value": volumevalue, "units": "m3"}, "customer_reference": str(srn), "sender": {
                "contact": {
                    "name": "Scarlett Music",
                    "phone": "(03) 9318 5751",
                    "company": "Scarlett Music"
                },
                "address": {
                    "address_line1": "286-288 Ballarat Rd",
                    "suburb": "Footscray",
                    "state_name": "VIC",
                    "postcode": "3011",
                    "country": "Australia"
                },
                "instructions": "The music shop, open 9am-6pm. Best parking is at The Palms across the road."
            }, "receiver": {
                "contact": {
                    "name": name,
                    "email": email,
                    "phone": phone,
                    "company": company
                },
                "address": {
                    "address_line1": address1,
                    "address_line2": address2,
                    "suburb": city,
                    "state_name": state,
                    "postcode": postcode,
                    "country": "Australia"
                },
                "instructions": "Authority to Leave"
            }
                   }

        r = requests.post('https://api.sendle.com/api/orders/',
                          auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'), json=payload)

        # pprint.pprint(r.text)
        # pprint.pprint(r.json())
        response = r.text
        response = json.loads(response)

        # pprint.pprint(response)
        tracking = response['sendle_reference']
        orderurl = response['order_url']
        price = response['price']['gross']['amount']
        pyperclip.copy(tracking)
        print('Tracking number: ' + tracking)
        print('Price: $' + str(price))
        croppedpdfurl = response['labels'][1]['url']

        r = requests.get(croppedpdfurl, auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'),
                         stream=True)

        with open(imagelocations + r"\pythontest2.pdf", "wb") as pdf:
            for chunk in r.iter_content(chunk_size=1024):

                # writing one chunk at a time to pdf file
                if chunk:
                    pdf.write(chunk)

        # r = requests.delete(orderurl, auth=HTTPBasicAuth('info_scarlettmusic_c', 'KscVMqvNh9VC6PYGn2Z5TwfZ'))

        images = convert_from_path(imagelocations + r'\pythontest2.pdf')

        for page in images:
            page.save(imagelocations + '\python2.jpg', 'JPEG')

        colorImage = Image.open(imagelocations + '\python2.jpg')

        transposed = colorImage.transpose(Image.ROTATE_90)

        transposed.save(imagelocations + '\python2.jpg')

        image_slicer.slice(imagelocations + '\python2.jpg', 2, 0, 1)

        im = Image.open(imagelocations + "\python2_01_01.png")
        rgb_im = im.convert('RGB')
        rgb_im.save(imagelocations + '\python2_01_01.jpg')

        filename = imagelocations + '\python2_01_01.jpg'

        im2 = Image.open(imagelocations + "\python2_01_02.png")
        rgb_im2 = im2.convert('RGB')
        rgb_im2.save(imagelocations + '\python2_01_02.jpg')

        filename2 = imagelocations + '\python2_01_02.jpg'
        sendToPrinter()

    if finalcourier == 'Australia Post':
        headers = {'Content-Type': 'application/json', 'Accept': 'application/json', 'Account-Number': accountnumber}
        if email == "":
            email_tracking = 'false'
        else:
            email_tracking = 'true'

        if answer == str(7):
            payload = {
                "shipments": [
                    {
                        "shipment_reference": OrderID,
                        "customer_reference_1": OrderID,
                        "customer_reference_2": "SKU-1",
                        "email_tracking_enabled": email_tracking,
                        "from": {
                            "name": "Lorelle Scarlett",
                            "business_name": "Scarlett Music",
                            "lines": [
                                "268-288 Ballarat Rd"
                            ],
                            "suburb": "FOOTSCRAY",
                            "state": "VIC",
                            "postcode": "3011",
                            "phone": "0422747033",
                            "email": "lorelle@scarlettmusic.com.au"
                        },
                        "to": {
                            "name": name,
                            "business_name": company,
                            "lines": [
                                address1, address2
                            ],
                            "suburb": city,
                            "state": State,
                            "postcode": postcode,
                            "phone": phone,
                            "email": email
                        },
                        "items": [{
                            "item_reference": OrderID + '-1',
                            "product_id": "3D55",
                            "length": str(length),
                            "height": str(height),
                            "width": str(width),
                            "weight": str(weightvalue),
                            "authority_to_leave": 'false',
                            "allow_partial_delivery": 'true',

                        },

                        ]
                    }
                ]
            }

        elif answer == str(14):
            payload = {
                "shipments": [
                    {
                        "shipment_reference": OrderID,
                        "customer_reference_1": OrderID,
                        "customer_reference_2": "SKU-1",
                        "movement_type": "RETURN",
                        "to": {
                            "name": "Lorelle Scarlett",
                            "business_name": "Scarlett Music",
                            "lines": [
                                "PO Box 6066"
                            ],
                            "suburb": "WEST FOOTSCRAY",
                            "state": "VIC",
                            "postcode": "3012",
                            "phone": "0422747033",
                            "email": "lorelle@scarlettmusic.com.au"
                        },
                        "from": {
                            "name": name,
                            "lines": [
                                address1[0:40], address2
                            ],
                            "suburb": city,
                            "state": State,
                            "postcode": postcode,
                        },
                        "items": [{
                            "item_reference": OrderID + '-1',
                            "product_id": "PR",
                            "authority_to_leave": 'false',

                        },

                        ]
                    }
                ]
            }

        r = requests.post('https://digitalapi.auspost.com.au/shipping/v1/shipments', headers=headers,
                          auth=HTTPBasicAuth(username, secret),
                          json=payload)

        response = r.text
        response = json.loads(response)
        # pprint.pprint(response)
        itemID = response['shipments'][0]['items'][0]['item_id']
        shipmentID = response['shipments'][0]['shipment_id']
        totalcost = response['shipments'][0]['shipment_summary']['total_cost']
        ausposttracking = response['shipments'][0]['items'][0]['tracking_details']['consignment_id']
        pyperclip.copy(ausposttracking)

        payload = {
            "wait_for_label_url": 'true',
            "preferences": [
                {
                    "type": "PRINT",
                    "format": "PDF",
                    "groups": [
                        {
                            "group": "Parcel Post",
                            "layout": "THERMAL-LABEL-A6-1PP",
                            "branded": 'true',
                            "left_offset": 0,
                            "top_offset": 0
                        },
                        {
                            "group": "Express Post",
                            "layout": "THERMAL-LABEL-A6-1PP",
                            "branded": 'false',
                            "left_offset": 0,
                            "top_offset": 0
                        }
                    ]
                }
            ],
            "shipments": [
                {
                    "shipment_id": shipmentID,
                    "items": [
                        {
                            "item_id": itemID
                        }
                    ]
                }
            ]
        }

        r = requests.post('https://digitalapi.auspost.com.au/shipping/v1/labels', headers=headers,
                          auth=HTTPBasicAuth(username, secret),
                          json=payload)

        response = r.text
        response = json.loads(response)
        # pprint.pprint(response)

        url = response['labels'][0]['url']
        print(url)

        r = requests.get(url, stream=True)

        with open(documentsaddress + '\\' + str(name) + '.pdf', 'wb') as pdf:
            pdf.write(r.content)

        images = convert_from_path(documentsaddress + '\\' + str(name) + '.pdf', 'wb')

        for page in images:
            page.save(documentsaddress + '\\' + str(name) + '.jpg', 'JPEG')

        if answer == str(14):
            os.startfile(documentsaddress + '\\' + str(name) + '.pdf')

if answer == str(8):
    def create_connection(db_name, db_user, db_password, db_host, db_port):
        connection = None
        try:
            connection = psycopg2.connect(
                database=db_name,
                user=db_user,
                password=db_password,
                host=db_host,
                port=db_port,
            )
        # print("Connection to PostgreSQL DB successful")
        except OperationalError as e:
            print(f"The error '{e}' occurred")
        return connection


    def execute_read_query(connection, item_check, query):
        cursor1 = connection.cursor()
        cursor2 = connection.cursor()
        result = None
        result1 = None
        try:
            result = cursor1.execute(item_check)
            result = cursor1.fetchall()

            if result == []:
                print(f"SKU '{sku}' was not found in the database and cannot be updated.")
                time.sleep(1)
                sys.exit()

            else:
                result1 = cursor2.execute(query)
                return result

        except OperationalError as e:
            print(e)
            sys.exit()


    def update_item(sku, location):
        item_check = f"SELECT 1 FROM item_data WHERE sku = '{sku}';"
        query = f"UPDATE item_data SET item_location = '{location}' WHERE sku = '{sku}' ; COMMIT;"
        execute_read_query(connection, item_check, query)


    connection = create_connection(
        "postgres", "kyal", "123456789Kyal", "143.244.166.150", "5432"
    )

    while True:

        typedanswer = input('''\n\nHave you copied the SKU? (Enter number 1-2)
        1) Yes
        2) No (will have to manually type SKU)
        Enter Response:''')

        if typedanswer == '1':
            sku = str(pyperclip.paste())
            break

        if typedanswer == '2':
            sku = str(input('''\n\nWhat is the SKU?
            Enter Response:'''))
            break

        else:
            print('Valid answer not found. Please try again.')
            time.sleep(1)

    while True:  ### Will have to have item title also here below vvvvvv
        location = str(input(f'''\n\n'Where is '{sku}' located?
    1) String Room
    2) Back Area
    3) Cage
    4) Out the Front
    5) Other
    Enter Response:'''))

        if location == '1':
            location = '1) String Room'
            update_item(sku, location)
            break

        if location == '2':
            location = '2) Back Area'
            update_item(sku, location)
            break

        if location == '3':
            location = '3) Cage'
            update_item(sku, location)
            break

        if location == '4':
            location = '4) Out the Front'
            update_item(sku, location)
            break

        if location == '5':
            location = input("""What is the location?

Input:""")
            update_item(sku, location)
            break

        else:
            print('Valid input not found. Try again.')
            time.sleep(1)
            continue

    print(f'{sku} updated successfully!')
    time.sleep(1)

if answer == str(11):

    manifestanswer = input('Are you sure you want to book in all pending Freightster orders? This is final. (Y/N)')

    if manifestanswer.lower() == 'n':
        sys.exit()
    elif manifestanswer.lower() == 'y':
        pass
    else:
        sys.exit()

    freightster_orders = []

    cursor = connection.cursor()
    cursor.execute(
        f"INSERT INTO Freightster_Tracking(order_id, tracking_number, name) SELECT order_id, tracking_number, name FROM Freightster; COMMIT;")

    cursor = connection.cursor()
    cursor.execute("SELECT order_id, tracking_number, name FROM Freightster;")
    results = cursor.fetchall()
    # pprint.pprint(results)
    total_orders = len(results)

    for x in range(int(total_orders)):
        freightster_orders.append(results[x][0])

    payload = {"order": {"orderIds": freightster_orders}}

    r = requests.post('https://freightster.com.au/api/v1/shippingAPI/pickup', json=payload,
                      headers=freightster_headers)
    freightster_response = json.loads(r.text)

    if freightster_response['status'] is True:

        cursor = connection.cursor()
        cursor.execute(f"TRUNCATE Freightster; COMMIT;")
        print('Success!')
        time.sleep(2)
    else:

        print('No pending Freightster orders')
        time.sleep(2)

if answer == str(12):
    tracking_number = input('''
    What is the persons name?

    Input here:''')
    cursor = connection.cursor()
    cursor.execute(
        f"SELECT order_id, tracking_number, name FROM Freightster_Tracking WHERE LOWER(name) LIKE LOWER('%{tracking_number}%');")
    results = cursor.fetchall()
    print('RESULTS FOUND')

    print('_____________________________________________________\n')
    print("  Freightster_ID, Tracking Number, Customer Name\n")
    pprint.pprint(results)

    input()

####JUST NEED TO BOOK PICKUP AT THE END OF COURIERS PLEASE LABEL PRINT